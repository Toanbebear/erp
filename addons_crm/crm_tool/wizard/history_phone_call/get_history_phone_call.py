import base64
import json
from datetime import timedelta
from io import BytesIO

import requests
from openpyxl import load_workbook
from openpyxl.styles import Font, borders, Alignment

from odoo import fields, api, models
from odoo.addons.report_sale.wizard.theme_report import ThemeReport

thin = borders.Side(style='thin')
double = borders.Side(style='double')
all_border_thin = borders.Border(thin, thin, thin, thin)


class HistoryPhoneCalInheritResPartner(models.Model):
    _inherit = 'res.partner'

    def open_form_get_history_phone_call(self):
        list_brand = list()
        for company in self.env.companies:
            if company.brand_id.code.lower() == 'kn':
                if ('kn', 'KangNam') not in list_brand:
                    list_brand.append(('kn', 'KangNam'))
            if company.brand_id.code.lower() == 'da':
                if ('da', 'Đông Á') not in list_brand:
                    list_brand.append(('da', 'Đông Á'))
            if company.brand_id.code.lower() == 'pr':
                if ('pr', 'Paris') not in list_brand:
                    list_brand.append(('pr', 'Paris'))
            if company.brand_id.code.lower() == 'hh':
                if ('hh', 'Hồng Hà') not in list_brand:
                    list_brand.append(('hh', 'Hồng Hà'))

        result_phone = list()
        if self.phone:
            result_phone.append(('phone', self.phone))
        if self.mobile:
            result_phone.append(('mobile', self.mobile))
        val = {
            'name': 'Lịch sử cuộc gọi',
            'type': 'ir.actions.act_window',
            'view_type': 'form',
            'view_mode': 'form',
            'view_id': self.env.ref('crm_tool.view_get_history_phone_call').id,
            'res_model': 'history.phone.call.form',
            'context': {
                'default_start_time_since': fields.Datetime.now(),
                'default_start_time_to': fields.Datetime.now() + timedelta(days=7),
                'default_partner_id': self.id,
                'phone': result_phone,
                'brand_id': list_brand
            },
            'target': 'new',
        }
        return val


class FormGetHistoryPhoneCall(models.TransientModel):
    _name = 'history.phone.call.form'
    _description = 'Lấy lịch sử dữ liệu cuộc gọi'

    def _default_partner_id(self):
        return self.env.context['active_id']

    def _get_phone_partner(self):
        return self._context.get('phone', None)

    def _get_brand_user(self):
        return self._context.get('brand_id', None)

    phone = fields.Selection(selection=_get_phone_partner, string='Số điện thoại', default='phone')

    @api.onchange('start_time_since')
    def _onchange_start_time_since(self):
        if self.start_time_since:
            self.start_time_to = self.start_time_since + timedelta(days=7)

    partner_id = fields.Many2one('res.partner', string="Khách hàng")
    start_time_since = fields.Datetime('Thời gian bắt đầu')
    start_time_to = fields.Datetime('Thời gian kết thúc')
    phone_call_line = fields.One2many('history.phone.call.line', 'parent_id', string='Lịch sử gọi')
    brand_id = fields.Selection(selection=_get_brand_user, string='Thương hiệu')

    def get_history_phone_call(self):
        phone_list = self._get_phone_partner()
        phone_dict = dict([(key, value) for key, value in phone_list])

        params = self.env['ir.config_parameter'].sudo()
        domain_config = 'domain_caresoft_%s' % self.brand_id
        token_config = 'domain_caresoft_token_%s' % self.brand_id
        url = params.get_param(domain_config)
        token = params.get_param(token_config)
        headers = {
            'Authorization': 'Bearer ' + token,
            'Content-Type': 'application/json'
        }
        # lấy danh sách agent
        agent_name = {}
        agent_email = {}
        response = requests.get('%s/api/v1/agents' % url, headers=headers)
        response = response.json()
        for rec in response['agents']:
            agent_name['%s' % rec['agent_id']] = rec['username']
            agent_email['%s' % rec['agent_id']] = rec['email']

        # lấy lịch sử gọi
        data = {
            "start_time_since": self.start_time_since.strftime('%Y-%m-%d %H:%M:%S'),
            "start_time_to": self.start_time_to.strftime('%Y-%m-%d %H:%M:%S'),
            "phone": phone_dict.get(self.phone),
        }
        # response = requests.get(url, headers=headers, data=json.dumps(data))
        response = requests.get('%s/api/v1/calls' % url, headers=headers, data=json.dumps(data))
        response = response.json()
        value_phone_line = []
        for rec in response['calls']:
            value_phone_line.append(
                (0, 0, {
                    'path': rec['path'] if 'path' in rec else '',
                    'path_download': rec['path_download'] if 'path_download' in rec else '',
                    'start_time': rec['start_time'],
                    'end_time': rec['end_time'],
                    'last_agent': agent_name['%s' % rec['agent_id']],
                    'last_email_agent': agent_email['%s' % rec['agent_id']],
                    'status': rec['call_status']
                })
            )
        val = {
            'name': 'Lịch sử cuộc gọi',
            'type': 'ir.actions.act_window',
            'view_type': 'form',
            'view_mode': 'form',
            'view_id': self.env.ref('crm_tool.view_history_phone_call').id,
            'res_model': 'history.phone.call',
            'context': {
                'default_partner_id': self.partner_id.id,
                'default_phone': phone_dict.get(self.phone),
                'default_phone_call_line': value_phone_line,
            },
            'target': 'new',
        }
        return val


class GetHistoryPhoneCall(models.TransientModel):
    _name = 'history.phone.call'
    _description = 'Lấy lịch sử dữ liệu cuộc gọi'

    partner_id = fields.Many2one('res.partner', string="Khách hàng")
    phone = fields.Char(string='Số điện thoại')

    phone_call_line = fields.One2many('history.phone.call.line', 'parent_id', string='Lịch sử gọi',
                                      tracking=True)

    def export_history_phone_call(self):
        report_brand_overview_attachment = self.env['ir.attachment'].browse(
            self.env.ref('crm_tool.lich_su_phone_call_attachment').id)
        decode = base64.b64decode(report_brand_overview_attachment.datas)
        wb = load_workbook(BytesIO(decode))
        ws = wb.active
        line_font = Font(name='Times New Roman', size=14)

        key_list = [
            'last_email_agent',
            'last_agent',
            'start_time',
            'end_time',
            'path',
            'path_download',
        ]
        key_col_list = list(range(2, len(key_list) + 2))

        row = 6
        index_row = 0

        datas = []
        for rec in self.phone_call_line:
            datas.append({
                'last_email_agent': rec.last_email_agent,
                'last_agent': rec.last_agent,
                'start_time': rec.start_time,
                'end_time': rec.end_time,
                'path': rec.path,
                'path_download': rec.path_download,
            })

        for data in datas:
            ws.cell(row, 1).border = ThemeReport.all_border_thin
            ws.cell(row, 1).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(row, 1).value = index_row + 1
            for col, k in zip(key_col_list, key_list):
                cell = ws.cell(row, col)
                cell.value = data[k]
                cell.font = line_font
                cell.border = ThemeReport.all_border_thin
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            row += 1
            index_row += 1

        ws['E2'].value = self.phone

        fp = BytesIO()
        wb.save(fp)
        fp.seek(0)
        report = base64.encodebytes((fp.read()))
        fp.close()
        attachment = self.env['ir.attachment'].sudo().create({
            'name': 'phone_call.xlsx',
            'datas': report,
            'res_model': 'temp.creation',
            'public': True,
        })
        url = "/web/content/?model=ir.attachment&id=%s&filename_field=name&field=datas&download=true" \
              % attachment.id
        return {'name': 'Lịch sử Phone Call',
                'type': 'ir.actions.act_url',
                'url': url,
                'target': 'self',
                }


class HistoryPhoneCallLine(models.TransientModel):
    _name = 'history.phone.call.line'
    _description = 'Dòng lịch sử dữ liệu cuộc gọi'

    path = fields.Char('Link nghe')
    path_download = fields.Char('Link download')
    last_agent = fields.Char('Nhân viên xử lý cuối cùng')
    last_email_agent = fields.Char('Email nhân viên xử lý cuối cùng')
    start_time = fields.Char('Thời điểm nhấc máy')
    end_time = fields.Char('Thời điểm cúp máy')
    status = fields.Char('Trạng thái')
    parent_id = fields.Many2one('history.phone.call', string='Partner')
