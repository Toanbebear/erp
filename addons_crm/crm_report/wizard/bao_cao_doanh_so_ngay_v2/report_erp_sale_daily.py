from odoo import fields, models, api, _
from odoo.exceptions import ValidationError, UserError
from datetime import date, datetime, time
from calendar import monthrange
import pandas as pd
import numpy as np
import base64
from io import BytesIO
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, borders, Alignment, PatternFill, NamedStyle
from openpyxl.utils.dataframe import dataframe_to_rows, expand_index
from openpyxl.worksheet.pagebreak import Break

TEMPLATE = [
    'payment_date',
    'name',
    'payment_type',
    'payment_method',
    'journal_id',
    'communication',
    'code_customer',
    'name_customer',
    'address_customer',
    'code_booking',
    'amount_total',
    'amount',
    'currency',
    'amount_vnd',
    'amount_used',
    'amount_paid',
    'amount_remain',
    'company',
    'source',
    'user',
    'note',
]


class ModelName(models.TransientModel):
    _name = 'report.sales.erp.daily'
    _description = 'Báo cáo doanh số ngày trên erp'

    start_date = fields.Date('Start date', default=date.today(), required=True)
    end_date = fields.Date('End date', default=date.today(), required=True)
    company_ids = fields.Many2many('res.company', string='Chi nhánh', domain=lambda self: [('id', 'in', self.env.user.company_ids.ids)], required=True)

    @api.onchange('start_date')
    def _onchange_start_date(self):
        if self.start_date:
            if self.start_date.month == fields.date.today().month:
                self.end_date = fields.date.today()
            else:
                self.end_date = date(self.start_date.year, self.start_date.month,
                                     monthrange(self.start_date.year, self.start_date.month)[1])

    @api.constrains('start_date', 'end_date')
    def check_dates(self):
        for record in self:
            start_date = fields.Date.from_string(record.start_date)
            end_date = fields.Date.from_string(record.end_date)

            if start_date > end_date:
                raise ValidationError(
                    _("Ngày kết thúc phải sau ngày bắt đầu."))
            elif (end_date - start_date).days > 31:
                raise ValidationError(
                    _("Chỉ tải được dữ liệu trong một tháng"))

    def report_daily_sales(self):
        # Lấy dữ liệu theo điều kiện
        select = """
        SELECT 
            payment.payment_date,
            payment.name,
            payment.payment_type,
            payment.payment_method,
            journal.name as journal_id,
            payment.communication,
            partner.code_customer,
            partner.name as name_customer,
            partner.street as address_customer,
            lead.name as code_booking,
            lead.amount_total,
            payment.amount,
            currency.name as currency,
            payment.amount_vnd,
            lead.amount_used,
            lead.amount_paid,
            lead.amount_remain,
            company.name as company,
            source.name as source,
            pr.name as user,
            '' as note
        FROM account_payment as payment
            LEFT JOIN res_partner as partner ON payment.partner_id = partner.id
            LEFT JOIN account_journal as journal ON payment.journal_id = journal.id
            LEFT JOIN res_currency as currency ON payment.currency_id = currency.id
            LEFT JOIN crm_lead as lead ON payment.crm_id = lead.id
            LEFT JOIN utm_source as source ON lead.source_id = source.id
            LEFT JOIN res_users as us ON us.id = payment.user
            LEFT JOIN res_partner as pr ON pr.id = us.partner_id
            LEFT JOIN res_company as company ON company.id = journal.company_id
        WHERE 
                (payment.payment_date between %s and %s) and
                payment.state not in ('draft', 'cancelled') and
                payment.payment_type != 'internal' and
                company.id in %s
        ORDER BY payment.payment_date"""
        self.env.cr.execute(select, (self.start_date.strftime('%Y-%m-%d'), self.end_date.strftime('%Y-%m-%d'), tuple(self.company_ids.ids)))
        data_select = self.env.cr.fetchall()
        if len(data_select) > 0:
            df = pd.DataFrame(data=data_select, columns=TEMPLATE)
            df['payment_date'] = pd.to_datetime(df['payment_date']).dt.strftime('%d/%m/%Y')
            df['stt'] = [i for i in range(1, len(df['payment_date']) + 1)]
            df.insert(0, 'stt', df.pop('stt'))
            df.loc[df['payment_type'] == 'inbound', 'payment_type'] = 'Nhận tiền'
            df.loc[df['payment_type'] == 'outbound', 'payment_type'] = 'Hoàn tiền'
            df.loc[df['payment_type'] == 'transfer', 'payment_type'] = 'Giao dịch nội bộ'

            df.loc[df['payment_method'] == 'tm', 'payment_method'] = 'Tiền mặt'
            df.loc[df['payment_method'] == 'ck', 'payment_method'] = 'Chuyển khoản'
            df.loc[df['payment_method'] == 'nb', 'payment_method'] = 'Thanh toán nội bộ'
            df.loc[df['payment_method'] == 'pos', 'payment_method'] = 'Quẹt thẻ qua POS'
            df.loc[df['payment_method'] == 'vdt', 'payment_method'] = 'Thanh toán qua ví điện tử'

            # --------------------------------------------
            # Khởi tạo đối tượng làm việc worksheet
            daily_sales_attachment = self.env['ir.attachment'].browse(
                self.env.ref('crm_report.report_sale_erp_daily_attachment').id)
            decode = base64.b64decode(daily_sales_attachment.datas)
            wb = load_workbook(BytesIO(decode))
            ws = wb.active
            # --------------------------------------------
            # In dữ liệu ra
            data = df.values.tolist()
            style = Font(color='00000000', bold=False, underline='none', size=14)
            thin = borders.Side(style='thin')
            all_border_thin = borders.Border(left=thin, right=thin, top=thin, bottom=thin)
            for row, line_values in zip(ws.iter_rows(min_row=6, max_row=6 + len(data), min_col=1, max_col=1 + len(data[0])), data):
                for cell, value in zip(row, line_values):
                    cell.value = value
                    cell.font = style
                    cell.border = all_border_thin
                    if cell.column in (12, 13, 15, 16, 17, 18):
                        cell.number_format = '_(* #,##0_);_(* (#,##0);_(* "-"??_);_(@_)'
            # --------------------------------------------
            last_row = len(data) + 6 + 5

            def f(row):
                if row['payment_type'] == 'Hoàn tiền':
                    val = -1 * row['amount_vnd']
                else:
                    val = row['amount_vnd']
                return val

            df_compute = df
            df_compute['amount_vnd'] = df_compute.apply(f, axis=1)

            table = pd.pivot_table(df_compute, index='payment_method', columns='payment_type', values='amount_vnd', fill_value=0, aggfunc='sum', margins=True, margins_name='Tổng cộng')
            table_index = table.index.get_level_values(0).tolist()
            table_column = table.columns.tolist()
            table_value = table.values.tolist()

            for col in range(1, len(table_column) + 1):
                cell = ws.cell(last_row - 1, col + 3)
                cell.value = table_column[col - 1]
                cell.font = style
                cell.border = all_border_thin

            temp = last_row
            for row in range(0, len(table_index)):
                cell = ws.cell(temp, 3)
                cell.value = table_index[row]
                cell.font = style
                cell.border = all_border_thin
                temp += 1

            for row, line_values in zip(ws.iter_rows(min_row=last_row, max_row=last_row + len(table_value), min_col=4, max_col=4 + len(table_value[0])), table_value):
                for cell, value in zip(row, line_values):
                    cell.value = value
                    cell.number_format = '_(* #,##0_);_(* (#,##0);_(* "-"??_);_(@_)'
                    cell.font = style
                    cell.border = all_border_thin
            # --------------------------------------------
            table = pd.pivot_table(df_compute, index=['payment_type', 'currency'], values=['amount_total', 'amount', 'amount_vnd', 'amount_used', 'amount_paid', 'amount_remain'],
                                   fill_value=0, aggfunc=sum,
                                   margins=True, margins_name='Tổng cộng')
            x_table = table[['amount_total', 'amount', 'amount_vnd', 'amount_used', 'amount_paid', 'amount_remain']]

            format_dict = {'amount_total': '{:,}', 'amount': '{:,}', 'amount_vnd': '{:,}', 'amount_used': '{:,}', 'amount_paid': '{:,}', 'amount_remain': '{:,}'}
            x_table.style.format(format_dict)
            x_table.rename(columns={'amount_total': 'Tổng tiền Booking', 'amount': 'Tổng tiền ngoại tệ', 'amount_vnd': 'Tổng tiền nguyên tệ',
                                    'amount_used': 'Tổng tiền khách sử dụng', 'amount_paid': 'Tổng tiền thanh toán', 'amount_remain': 'Tổng tiền còn lại'}, inplace=True)
            x_table.rename_axis(index={'payment_type': 'Loại thanh toán', 'currency': 'Loại tiền tệ'}, inplace=True)

            x_table_index_01 = x_table.index.get_level_values(0).tolist()
            x_table_index_02 = x_table.index.get_level_values(1).tolist()

            x_table_columns = x_table.columns.tolist()
            x_table_value = x_table.values.tolist()

            for col in range(1, len(x_table_columns) + 1):
                cell = ws.cell(last_row - 1, col + 9)
                cell.value = x_table_columns[col - 1]
                cell.font = style
                cell.border = all_border_thin

            temp = last_row
            for row in range(0, len(x_table_index_01)):
                cell = ws.cell(temp, 8)
                cell.value = x_table_index_01[row]
                cell.font = style
                cell.border = all_border_thin
                temp += 1

            temp = last_row
            for row in range(0, len(x_table_index_02)):
                cell = ws.cell(temp, 9)
                cell.value = x_table_index_02[row]
                cell.font = style
                cell.border = all_border_thin
                temp += 1

            for row, line_values in zip(ws.iter_rows(min_row=last_row, max_row=last_row + len(x_table_value), min_col=10, max_col=10 + len(x_table_value[0])), x_table_value):
                for cell, value in zip(row, line_values):
                    cell.value = value
                    cell.number_format = '_(* #,##0_);_(* (#,##0);_(* "-"??_);_(@_)'
                    cell.font = style
                    cell.border = all_border_thin

            # --------------------------------------------
            # Định dạng
            if self.start_date == self.end_date:
                ws['A3'].value = self.end_date.strftime('Ngày %d tháng %m năm %Y')
            else:
                ws['A3'].value = 'Từ ngày: %s đến ngày: %s' % (
                    self.start_date.strftime('%d/%m/%Y'), self.end_date.strftime('%d/%m/%Y'))

            fp = BytesIO()
            wb.save(fp)
            fp.seek(0)
            report = base64.encodebytes((fp.read()))
            fp.close()
            attachment = self.env['ir.attachment'].sudo().create({
                'name': 'bao_cao.xlsx',
                'datas': report,
                'res_model': 'temp.creation',
                'public': True,
            })
            url = "/web/content/?model=ir.attachment&id=%s&filename_field=name&field=datas&download=true" \
                  % attachment.id
            return {'name': 'BÁO CÁO DOANH SỐ BÁN SẢN PHẨM KANGNAM',
                    'type': 'ir.actions.act_url',
                    'url': url,
                    'target': 'self',
                    }
        else:
            context = dict(self._context or {})
            context['message'] = 'Không tồn tại dữ liệu từ ngày: %s tới ngày: %s' % (self.start_date.strftime('%d/%m/%Y'), self.end_date.strftime('%d/%m/%Y'))
            return {
                'name': _('Thông báo'),  # label
                'type': 'ir.actions.act_window',
                'view_mode': 'form',
                'view_id': self.env.ref('sh_message.sh_message_wizard').id,
                'res_model': 'sh.message.wizard',  # model want to display
                'target': 'new',  # if you want popup
                'context': context,
            }
