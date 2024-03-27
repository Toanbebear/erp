from odoo import fields, api, models, _
from odoo.exceptions import ValidationError, UserError
from datetime import date, datetime, time
from calendar import monthrange
from openpyxl import load_workbook
from openpyxl.styles import Font, borders, Alignment, PatternFill

import base64
from io import BytesIO
from dateutil.relativedelta import relativedelta
from pytz import timezone, utc
from copy import copy
import requests


class ReportCustomer(models.TransientModel):
    _name = 'report.loc.kh'
    _description = 'Báo cáo khách hàng'

    start_date = fields.Date('Start date', default=date.today())
    end_date = fields.Date('End date', default=date.today())
    start_datetime = fields.Datetime('Start datetime', compute='_compute_datetime')
    end_datetime = fields.Datetime('End datetime', compute='_compute_datetime')
    company_ids = fields.Many2many(string='Chi nhánh', comodel_name='res.company',
                                   domain="[('name', 'not like', 'KHÔNG XÁC ĐỊNH')]")
    location_ids = fields.Many2many(string='Tỉnh/Thành', comodel_name='res.country.state',
                                   domain="[('country_id.code', '=', 'VN')]")
    all_location = fields.Boolean(string='Tất cả Tỉnh/Thành', default=True)

    @api.depends('start_date', 'end_date')
    def _compute_datetime(self):
        self.start_datetime = False
        self.end_datetime = False
        if self.start_date and self.end_date:
            local_tz = timezone(self.env.user.tz or 'Etc/GMT+7')
            start_datetime = datetime(self.start_date.year, self.start_date.month, self.start_date.day, 0, 0, 0)
            end_datetime = datetime(self.end_date.year, self.end_date.month, self.end_date.day, 23, 59, 59)
            start_datetime = local_tz.localize(start_datetime, is_dst=None)
            end_datetime = local_tz.localize(end_datetime, is_dst=None)
            self.start_datetime = start_datetime.astimezone(utc).replace(tzinfo=None)
            self.end_datetime = end_datetime.astimezone(utc).replace(tzinfo=None)

    @api.constrains('start_date', 'end_date')
    def check_dates(self):
        for record in self:
            start_date = fields.Date.from_string(record.start_date)
            end_date = fields.Date.from_string(record.end_date)
            if start_date > end_date:
                raise ValidationError(
                    _("End Date cannot be set before Start Date."))

    def report(self):
        # Loại bỏ case
        # Loại bỏ booking bảo hành
        template = self.env['ir.attachment'].browse(self.env.ref('crm_report.report_loc_kh_attachment').id)
        decode = base64.b64decode(template.datas)
        wb = load_workbook(BytesIO(decode))
        ws = wb.get_sheet_by_name('Chi tiết')
        thin = borders.Side(style='thin')
        all_border_thin = borders.Border(left=thin, right=thin, top=thin, bottom=thin)
        line_font = Font(name='Times New Roman', size=13)

        self.env.cr.execute("select partner_id from crm_case")
        cases = self.env.cr.fetchall()
        partner_ids = []
        for c in cases:
            partner_ids.append(c[0])

        datas = []
        datas_2 = {}

        if not self.all_location and not self.location_ids:
            select = """select rp.id, cl.total_received, us.name, cld.name, cld.arrival_date, rc.name, cld.overseas_vietnamese,
                        rco.name, cld.phone, cld.code_customer, cld.contact_name, us2.name, rcs.name, cld.street, pt.name, pc.name, hd.name, cl.stage
                        from crm_line cl
                        left join crm_lead cld on cld.id = cl.crm_id
                        left join utm_source us on us.id = cl.source_extend_id
                        left join res_partner rp on rp.id = cld.partner_id
                        left join res_company rc on rc.id = cl.company_id
                        right join res_country rco on rco.id = cld.country_id
                        left join utm_source us2 on us2.id = cld.source_id
                        full join res_country_state rcs on rcs.id = rp.state_id
                        left join sh_medical_health_center_service smh on smh.id = cl.service_id
                        left join product_product pp on pp.id = smh.product_id
                        left join product_template pt on pt.id = pp.product_tmpl_id
                        left join sh_medical_health_center_service_category csc on csc.id = smh.service_category
                        LEFT JOIN product_category pc ON pc.id = csc.product_cat_id
                        left join hr_department hd on hd.id = cld.department_id
                        where cld.type_crm_id <> 3 and cld.stage_id = 4 and cld.arrival_date <= %s and cld.arrival_date >= %s and rc.id in %s and rp.state_id is Null
                        """
            self.env.cr.execute(select,
                                (self.end_datetime, self.start_datetime, tuple(self.company_ids.ids)))
            result = self.env.cr.fetchall()
            for c in result:
                datas.append(c)
        elif self.all_location:
            select = """select rp.id, cl.total_received, us.name, cld.name, cld.arrival_date, rc.name, cld.overseas_vietnamese,
                        rco.name, cld.phone, cld.code_customer, cld.contact_name, us2.name, rcs.name, cld.street, pt.name, pc.name, hd.name, cl.stage
                        from crm_line cl
                        left join crm_lead cld on cld.id = cl.crm_id
                        left join utm_source us on us.id = cl.source_extend_id
                        left join res_partner rp on rp.id = cld.partner_id
                        left join res_company rc on rc.id = cl.company_id
                        right join res_country rco on rco.id = cld.country_id
                        left join utm_source us2 on us2.id = cld.source_id
                        full join res_country_state rcs on rcs.id = rp.state_id
                        left join sh_medical_health_center_service smh on smh.id = cl.service_id
                        left join product_product pp on pp.id = smh.product_id
                        left join product_template pt on pt.id = pp.product_tmpl_id
                        left join sh_medical_health_center_service_category csc on csc.id = smh.service_category
                        LEFT JOIN product_category pc ON pc.id = csc.product_cat_id
                        left join hr_department hd on hd.id = cld.department_id
                        where cld.type_crm_id <> 3 and cld.stage_id = 4 and cld.arrival_date <= %s and cld.arrival_date >= %s and rc.id in %s
                        """
            self.env.cr.execute(select,
                                (self.end_datetime, self.start_datetime, tuple(self.company_ids.ids)))
            result = self.env.cr.fetchall()
            for c in result:
                datas.append(c)
        else:
            select = """select rp.id, cl.total_received, us.name, cld.name, cld.arrival_date, rc.name, cld.overseas_vietnamese,
                        rco.name, cld.phone, cld.code_customer, cld.contact_name, us2.name, rcs.name, cld.street, pt.name, pc.name, hd.name, cl.stage
                        from crm_line cl
                        left join crm_lead cld on cld.id = cl.crm_id
                        left join utm_source us on us.id = cl.source_extend_id
                        left join res_partner rp on rp.id = cld.partner_id
                        left join res_company rc on rc.id = cl.company_id
                        right join res_country rco on rco.id = cld.country_id
                        left join utm_source us2 on us2.id = cld.source_id
                        left join res_country_state rcs on rcs.id = rp.state_id
                        left join sh_medical_health_center_service smh on smh.id = cl.service_id
                        left join product_product pp on pp.id = smh.product_id
                        left join product_template pt on pt.id = pp.product_tmpl_id
                        left join sh_medical_health_center_service_category csc on csc.id = smh.service_category
                        LEFT JOIN product_category pc ON pc.id = csc.product_cat_id
                        left join hr_department hd on hd.id = cld.department_id
                        where cld.type_crm_id <> 3 and cld.stage_id = 4 and cld.arrival_date <= %s and cld.arrival_date >= %s and rc.id in %s and rcs.id in %s
                        """
            self.env.cr.execute(select,
                                (self.end_datetime, self.start_datetime, tuple(self.company_ids.ids), tuple(self.location_ids.ids)))
            result = self.env.cr.fetchall()
            for c in result:
                datas.append(c)

        index = 1
        index_2 = 1
        row = 2
        format_currency = '_(* #,##0_);_(* (#,##0);_(* "-"??_);_(@_)'
        for data in datas:
            if data[0] in partner_ids:
                continue

            if data[0] in datas_2:
                value = data[1] if data[1] else 0
                datas_2[data[0]][4] += value
            else:
                datas_2[data[0]] = [data[9], data[10], data[8], data[5], 0]
            col = 1
            for val in data:
                cell_val = ws.cell(row, col)
                if col == 7:
                    if val == 'yes':
                        cell_val.value = "Có"
                    else:
                        cell_val.value = "Không"
                elif col == 5:
                    if val:
                        cell_val.value = fields.Datetime.context_timestamp(self.with_context(tz=self.env.user.tz),
                                                                           val).strftime('%d/%m/%Y %H:%M:%S')
                elif col == 1:
                    cell_val.value = index
                elif col == 18:
                    if val == 'new':
                        cell_val.value = 'Được sử dụng'
                    elif val == 'processing':
                        cell_val.value = 'Đang xử trí'
                    elif val == 'done':
                        cell_val.value = 'Kết thúc'
                    elif val == 'waiting':
                        cell_val.value = 'Chờ phê duyệt'
                    elif val == 'cancel':
                        cell_val.value = 'Hủy'
                else:
                    cell_val.value = val
                cell_val.font = line_font
                cell_val.border = all_border_thin
                if col == 2:
                    cell_val.number_format = format_currency
                if col in [7, 8, 13, 18]:
                    cell_val.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                else:
                    cell_val.alignment = Alignment(wrap_text=True)
                col += 1
            row += 1
            index += 1
        ws2 = wb.get_sheet_by_name('Khách hàng')
        row_2 = 2
        for data in datas_2:
            col_2 = 1
            cell_stt = ws2.cell(row_2, col_2)
            cell_stt.value = index_2
            cell_stt.font = line_font
            cell_stt.border = all_border_thin
            cell_stt.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            col_2 = 2
            index_2 += 1
            for val in datas_2[data]:
                cell_val = ws2.cell(row_2, col_2)
                cell_val.value = val
                cell_val.font = line_font
                cell_val.border = all_border_thin
                if col_2 == 6:
                    cell_val.number_format = format_currency
                if col_2 in [1, 2]:
                    cell_val.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                else:
                    cell_val.alignment = Alignment(wrap_text=True)
                col_2 += 1
            row_2 += 1
        fp = BytesIO()
        wb.save(fp)
        fp.seek(0)
        report = base64.encodebytes((fp.read()))
        fp.close()
        attachment = self.env['ir.attachment'].sudo().create({
            'name': 'Lọc khách hàng.xlsx',
            'datas': report,
            'res_model': 'temp.creation',
            'public': True,
        })
        url = "/web/content/?model=ir.attachment&id=%s&filename_field=name&field=datas&download=true" \
              % attachment.id
        return {'name': 'Báo cáo lọc khách hàng.xlsx',
                'type': 'ir.actions.act_url',
                'url': url,
                'target': 'self',
                }
