from odoo import fields, api, models, _
from odoo.exceptions import ValidationError, UserError
from datetime import date, datetime, time
from calendar import monthrange
from openpyxl import load_workbook
from openpyxl.styles import Font, borders, Alignment, PatternFill

import base64
from io import BytesIO
from dateutil.relativedelta import relativedelta
from pytz import timezone, utc
from copy import copy
import requests


class ReportCustomer(models.TransientModel):
    _name = 'report.the.thanh.vien'
    _description = 'Báo cáo khách hàng'

    company_ids = fields.Many2many(string='Chi nhánh', comodel_name='res.company',
                                   domain="[('name', 'not like', 'KHÔNG XÁC ĐỊNH'), ('brand_id.name','=','Paris')]")
    all_company = fields.Boolean('Tất cả chi nhánh', default=False)
    rank = fields.Many2many(string='Hạng thẻ', comodel_name='crm.loyalty.rank', domain="[('brand_id.name','=','Paris')]")
    all_rank = fields.Boolean('Tất cả hạng thẻ',default=True)
    start_date = fields.Date('Ngày bắt đầu', default=date.today().replace(day=1))
    end_date = fields.Date('Ngày kết thúc')
    start_datetime = fields.Datetime('Start datetime', compute='_compute_datetime')
    end_datetime = fields.Datetime('End datetime', compute='_compute_datetime')

    @api.onchange('all_company')
    def onchange_all_company(self):
        if self.all_company:
            companies = self.env['res.company'].search(
                [('brand_id', '=', 3), ('name', 'not like', 'KHÔNG XÁC ĐỊNH'),
                 ('active', '=', True)])
            self.company_ids = companies.ids
        else:
            self.company_ids = None

    @api.onchange('all_rank')
    def onchange_all_rank(self):
        if self.all_rank:
            ranks = self.env['crm.loyalty.rank'].search(
                [('brand_id', '=', 3),('active', '=', True)])
            self.rank = ranks.ids
        else:
            self.rank = None

    @api.onchange('start_date')
    def _onchange_start_date(self):
        if self.start_date:
            if self.start_date.month == fields.date.today().month:
                self.end_date = fields.date.today()
            else:
                self.end_date = date(self.start_date.year, self.start_date.month,
                                     monthrange(self.start_date.year, self.start_date.month)[1])

    @api.depends('start_date', 'end_date')
    def _compute_datetime(self):
        self.start_datetime = False
        self.end_datetime = False
        if self.start_date and self.end_date:
            local_tz = timezone(self.env.user.tz or 'Etc/GMT+7')
            start_datetime = datetime(self.start_date.year, self.start_date.month, self.start_date.day, 0, 0, 0)
            end_datetime = datetime(self.end_date.year, self.end_date.month, self.end_date.day, 23, 59, 59)
            start_datetime = local_tz.localize(start_datetime, is_dst=None)
            end_datetime = local_tz.localize(end_datetime, is_dst=None)
            self.start_datetime = start_datetime.astimezone(utc).replace(tzinfo=None)
            self.end_datetime = end_datetime.astimezone(utc).replace(tzinfo=None)

    @api.constrains('start_date', 'end_date')
    def check_dates(self):
        for record in self:
            start_date = fields.Date.from_string(record.start_date)
            end_date = fields.Date.from_string(record.end_date)
            if start_date > end_date:
                raise ValidationError(
                    _("End Date cannot be set before Start Date."))

    def report(self):
        # Loại bỏ case
        template = self.env['ir.attachment'].browse(self.env.ref('crm_report.report_loc_the_thanh_vien_attachment').id)
        decode = base64.b64decode(template.datas)
        wb = load_workbook(BytesIO(decode))
        ws = wb.active
        thin = borders.Side(style='thin')
        all_border_thin = borders.Border(left=thin, right=thin, top=thin, bottom=thin)
        line_font = Font(name='Times New Roman', size=13)

        self.env.cr.execute("select partner_id from crm_case")
        cases = self.env.cr.fetchall()
        partner_ids = []
        for c in cases:
            partner_ids.append(c[0])

        datas = []
        datas_2 = {}
        if not self.all_rank:
            if self.rank:
                select = """select rp.id, clc.id, rp.code_customer, rc.name, rp.phone, rp.name, ut.name, pc.name, clr.name, rb.name, cld.type_crm_id, rp.id
                            from crm_loyalty_card clc
                            full join crm_loyalty_rank clr on clr.id = clc.rank_id
                            full join res_partner rp on rp.id = clc.partner_id
                            full join res_company rc on rc.id = clc.company_id
                            full join res_brand rb on rb.id = clc.brand_id
                            full join crm_lead cld on cld.loyalty_id = clc.id
                            full join utm_source ut on ut.id = cld.source_id
                            full join crm_line cln on cln.crm_id = cld.id
                            full join sh_medical_health_center_service smh on smh.id = cln.service_id
                            full join sh_medical_health_center_service_category csc on csc.id = smh.service_category
                            full join product_category pc on pc.id = csc.product_cat_id
                            where rb.id = 3 and rc.id in %s and clr.id in %s and cld.arrival_date >= %s and cld.arrival_date <= %s
                            """
                self.env.cr.execute(select,
                                    (tuple(self.company_ids.ids), tuple(self.rank.ids), self.start_datetime.strftime('%Y-%m-%d %H:%M:%S'), self.end_datetime.strftime('%Y-%m-%d %H:%M:%S')))
                result = self.env.cr.fetchall()
                for c in result:
                    datas.append(c)
            else:
                select = """select rp.id, clc.id, rp.code_customer, rc.name, rp.phone, rp.name, ut.name, pc.name, clr.name, rb.name, cld.type_crm_id, rp.id
                        from crm_loyalty_card clc
                        full join crm_loyalty_rank clr on clr.id = clc.rank_id
                        full join res_partner rp on rp.id = clc.partner_id
                        full join res_company rc on rc.id = clc.company_id
                        full join res_brand rb on rb.id = clc.brand_id
                        full join crm_lead cld on cld.loyalty_id = clc.id
                        full join utm_source ut on ut.id = cld.source_id
                        full join crm_line cln on cln.crm_id = cld.id
                        full join sh_medical_health_center_service smh on smh.id = cln.service_id
                        full join sh_medical_health_center_service_category csc on csc.id = smh.service_category
                        full join product_category pc on pc.id = csc.product_cat_id
                        where rb.id = 3 and rc.id in %s and clr.id is Null and cld.arrival_date >= %s and cld.arrival_date <= %s
                        """
                self.env.cr.execute(select,
                                    (tuple(self.company_ids.ids),
                                     self.start_datetime.strftime('%Y-%m-%d %H:%M:%S'),
                                     self.end_datetime.strftime('%Y-%m-%d %H:%M:%S')))
                result = self.env.cr.fetchall()
                for c in result:
                    datas.append(c)
        else:
            select = """select rp.id, clc.id, rp.code_customer, rc.name, rp.phone, rp.name, ut.name, pc.name, clr.name, rb.name, cld.type_crm_id, rp.id
                        from crm_loyalty_card clc
                        full join crm_loyalty_rank clr on clr.id = clc.rank_id
                        left join res_partner rp on rp.id = clc.partner_id
                        left join sh_medical_patient smp on smp.partner_id = rp.id
                        left join res_company rc on rc.id = clc.company_id
                        left join res_brand rb on rb.id = clc.brand_id
                        full join crm_lead cld on cld.loyalty_id = clc.id
                        left join utm_source ut on ut.id = cld.source_id
                        full join crm_line cln on cln.crm_id = cld.id
                        left join sh_medical_health_center_service smh on smh.id = cln.service_id
                        left join sh_medical_health_center_service_category csc on csc.id = smh.service_category
                        left join product_category pc on pc.id = csc.product_cat_id
                        where rb.id = 3 and rc.id in %s and cld.arrival_date >= %s and cld.arrival_date <= %s
                        """
            self.env.cr.execute(select,
                                (tuple(self.company_ids.ids), self.start_datetime.strftime('%Y-%m-%d %H:%M:%S'), self.end_datetime.strftime('%Y-%m-%d %H:%M:%S')))
            result = self.env.cr.fetchall()
            for c in result:
                datas.append(c)
        list_clc = []
        index = 1
        for data in datas:
            if data[1] not in list_clc:
                datas_2[data[1]] = [index, data[2], data[3],data[4],data[5],data[6],data[7],data[8],data[9],data[10],data[11]]
                list_clc.append(data[1])
                index += 1
            else:
                if str(data[6]) in str(datas_2[data[1]][5]):
                    pass
                else:
                    datas_2[data[1]][5] = str(datas_2[data[1]][5]) + '; ' + str(data[6])
                if str(data[7]) in str(datas_2[data[1]][6]):
                    pass
                else:
                    datas_2[data[1]][6] = str(datas_2[data[1]][6]) + '; ' + str(data[7])
                if data[10] != 3 and datas_2[data[1]][9] != 3:
                    pass
                elif data[10] != 3 and datas_2[data[1]][9] == 3:
                    datas_2[data[1]][9] = 3
                elif data[10] == 3:
                    datas_2[data[1]][9] = 3

        row = 2
        format_currency = '_(* #,##0_);_(* (#,##0);_(* "-"??_);_(@_)'
        for data in datas_2:
            col = 1
            for val in datas_2[data]:
                cell_val = ws.cell(row, col)
                if col == 10:
                    if val == 3:
                        cell_val.value = 'X'
                    else:
                        cell_val.value = ''
                elif col == 11:
                    if val in partner_ids:
                        cell_val.value = 'X'
                    else:
                        cell_val.value = ''
                elif col == 8:
                    if val:
                        cell_val.value = val
                    else:
                        cell_val.value = 'Không có hạng'
                else:
                    cell_val.value = val
                cell_val.font = line_font
                cell_val.border = all_border_thin
                if col in [1, 8, 9, 10, 11]:
                    cell_val.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                else:
                    cell_val.alignment = Alignment(wrap_text=True)
                col += 1
            row += 1


        fp = BytesIO()
        wb.save(fp)
        fp.seek(0)
        report = base64.encodebytes((fp.read()))
        fp.close()
        attachment = self.env['ir.attachment'].sudo().create({
            'name': 'Lọc khách hàng.xlsx',
            'datas': report,
            'res_model': 'temp.creation',
            'public': True,
        })
        url = "/web/content/?model=ir.attachment&id=%s&filename_field=name&field=datas&download=true" \
              % attachment.id
        return {'name': 'Báo cáo lọc khách hàng.xlsx',
                'type': 'ir.actions.act_url',
                'url': url,
                'target': 'self',
                }
