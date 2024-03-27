import base64
from calendar import monthrange
from copy import copy
from datetime import date, datetime, timedelta
from io import BytesIO

from openpyxl import load_workbook
from openpyxl.styles import Font, borders, Alignment, PatternFill
from pytz import timezone, utc

from odoo import fields, api, models
from odoo.exceptions import ValidationError

thin = borders.Side(style='thin')
double = borders.Side(style='double')
all_border_thin = borders.Border(thin, thin, thin, thin)

MONTHS = [('1', 'Tháng 1'), ('2', 'Tháng 2'), ('3', 'Tháng 3'), ('4', 'Tháng 4'), ('5', 'Tháng 5'), ('6', 'Tháng 6'),
          ('7', 'Tháng 7'), ('8', 'Tháng 8'), ('9', 'Tháng 9'), ('10', 'Tháng 10'), ('11', 'Tháng 11'),
          ('12', 'Tháng 12')]

YEARS = [('2021', '2021'), ('2022', '2022'), ('2023', '2023'), ('2024', '2024'), ('2025', '2025'), ('2026', '2026'),
         ('2027', '2027'), ('2028', '2028'), ('2029', '2029'), ('2030', '2030')]


class ReportTNKH(models.TransientModel):
    _name = 'report.tnkh'
    _description = 'Báo cao TNKH'

    start_year = fields.Selection(YEARS, 'Từ năm', default=lambda *a: str(date.today().year))
    start_month = fields.Selection(MONTHS, 'Từ tháng', default=lambda *a: '1')

    end_year = fields.Selection(YEARS, 'Tới năm', default=lambda *a: str(date.today().year))
    end_month = fields.Selection(MONTHS, 'Tới tháng', default=lambda *a: str(date.today().month))

    brand_id = fields.Many2one('res.brand', string='Thương hiệu')
    company_ids = fields.Many2many(string='Chi nhánh', comodel_name='res.company',
                                   domain="[('name', 'not like', 'KHÔNG XÁC ĐỊNH')]")
    fetch_all = fields.Boolean(string='Tất cả chi nhánh', default=True)

    @api.onchange('fetch_all', 'brand_id')
    def _onchange_companies(self):
        if self.fetch_all:
            if self.brand_id:
                companies = self.env['res.company'].search(
                    [('brand_id', '=', self.brand_id.id), ('name', 'not like', 'KHÔNG XÁC ĐỊNH'),
                     ('active', '=', True)])
                self.company_ids = companies.ids
        else:
            self.company_ids = None

    @api.onchange('start_year')
    def _onchange_start_year(self):
        if self.start_year:
            self.end_year = self.start_year

    @api.onchange('end_year')
    def _onchange_end_year(self):
        if self.end_year:
            self.start_year = self.end_year

    @api.constrains('start_year', 'end_year')
    def check_year(self):
        for record in self:
            start_year = int(record.start_year)
            end_year = int(record.end_year)

            if start_year != end_year:
                raise ValidationError("Hệ thống hỗ trợ báo cáo trong cùng 1 năm. "
                                      "Vui lòng chọn thời gian bắt đầu và thời gian kết thúc trong 1 năm.")

    @api.constrains('start_month', 'end_month')
    def check_month(self):
        for record in self:
            start_month = int(record.start_month)
            end_month = int(record.end_month)

            if start_month > end_month:
                raise ValidationError("Thời gian bắt đầu phải trước hoặc bằng thời gian kết thúc.")

    def report_1(self):
        start_year = int(self.start_year)
        start_month = int(self.start_month)
        end_year = int(self.end_year)
        end_month = int(self.end_month)
        number_company = 0
        data = []
        cols = []
        col_in_company = [
            'so_luong_kh_khong_hai_long',
            'so_luong_kh_chap_nhan_duoc',
            'so_luong_kh_dat_mong_doi',
            'so_luong_kh_thoa_man_uoc_muon',
            'so_luong_kh_tren_ca_uoc_muon',
            'so_luong_kh_qua_tuyet_voi',
            'tb_diem_tnnk',
            'tong_khao_sat',
            'ti_le_thuc_hien_ks',
            'ti_le_kh_chac_chan_gioi_thieu',
            'so_kh_thanh_cong',
            'ds_tong_cua_chi_nhanh',
            'so_luong_kh_duoc_gioi_thieu',
            'tong_dv_thanh_cong',
            'so_dv_tb_tren_dau_khach',
            'ti_le_kh_duoc_gt_tren_dau_khach',
            'ds_do_kh_duoc_gioi_thieu_mang_lai',
            'ti_le',
            'so_kh_quay_lai',
            'tong_so_dv_thanh_cong_kh_quay_lai',
            'ti_le_quay_lai_theo_dich_vu',
            'ti_le_quay_lai_theo_khach',
            'ds_kh_quay_lai',
            'ti_le_doanh_so',
            'kh_outsold',
            'so_kh_outsold_tc',
            'ds_kh_outsold_tc',
            'tong_case',
            'so_luong_case',
            'ti_le_case',
            'so_luong_phan_anh',
            'ti_le_phan_anh',
            'so_luong_canh_bao',
            'ti_le_canh_bao',
            'space'
        ]
        len_col_in_company = len(col_in_company)
        colname_in_company = [
            'Không hài lòng (1) (%)', 'Chấp nhận được (2) (%)', 'Đạt mong đợi (3) (%)', 'Thỏa mãn ước muốn (4) (%)',
            'Trên cả ước muốn (5) (%)', 'Quá tuyệt vời (6) (%)', 'TB điểm TNKH',
            'Tổng khảo sát', 'Tỉ lệ thực hiện khảo sát (%)', 'Tỉ lệ(NPS) (%)', 'Số KH thành công',
            'DS tổng của chi nhánh (VND)', 'Số lượng KH được giới thiệu', 'Tổng số dv thành công', 'Số DV TB/đẩu khách',
            'Tỉ lệ KH được gt/tổng khách (%)', 'DS do KH được giới thiệu mang lại (VND)', 'Tỉ lệ (%)', 'Số KH quay lại',
            'Tổng số dv thành công của KH quay lại', 'Tỉ lệ khách hàng quay lại theo dịch vụ',
            'Tỉ lệ quay lại theo khách (%)', 'DS KH quay lại (VND)',
            'Tỉ lệ doanh số (%)', 'Số KH outsold', 'Số KH thành công', 'DS KH thành công', 'Tổng case', 'Số lượng khiếu nại', 'Tỉ lệ khiếu nại(%)', 'Số lượng phản ánh',
            'Tỉ lệ phản ánh(%)', 'Số lượng cảnh báo', 'Tỉ lệ cảnh báo(%)', '-'
        ]
        colgroup_in_company = ['Khảo sát sự hài lòng của khách hàng', 'NPS', 'Chi nhánh', 'KH giới thiệu KH',
                               'KH quay lại', 'KH Outsold', 'CASE', 'Khiếu nại', 'Phản ánh', 'Cảnh báo', '']
        config = self.env['ir.config_parameter'].sudo()
        title_khl = config.get_param('title_answer_1')
        title_cnd = config.get_param('title_answer_2')
        title_dmd = config.get_param('title_answer_3')
        title_tm = config.get_param('title_answer_4')
        title_tum = config.get_param('title_answer_5')
        title_qtv = config.get_param('title_answer_6')
        cau_1 = config.get_param('id_cau_1')
        list_cau_1 = tuple(map(int, cau_1.split(', ')))
        list_khl = tuple(map(int, title_khl.split(', ')))
        list_cnd = tuple(map(int, title_cnd.split(', ')))
        list_dmd = tuple(map(int, title_dmd.split(', ')))
        list_tm = tuple(map(int, title_tm.split(', ')))
        list_tum = tuple(map(int, title_tum.split(', ')))
        list_qtv = tuple(map(int, title_qtv.split(', ')))

        tong_so_tl_cau_1 = """
            select sui.company_id,count(suil.id)
                from 
                    survey_user_input_line suil
                LEFT JOIN survey_user_input sui ON sui.id = suil.user_input_id
                LEFT JOIN survey_question sq ON sq.id = suil.question_id
                WHERE
                    sui.create_date >= %s and sui.create_date <= %s  and
                    sq.id in %s and sui.company_id in %s and sui.state = 'done'
                group by sui.company_id
        """
        select_so_luong_kh_khong_hai_long = """
                select sui.company_id,count(suil.id)
                from 
                    survey_user_input_line suil
                LEFT JOIN survey_user_input sui ON  sui.id = suil.user_input_id
                LEFT JOIN survey_label sl ON sl.id = suil.value_suggested
                WHERE
                    sui.create_date >= %s and sui.create_date <= %s  and
                    sl.id in %s and sui.company_id in %s and sui.state = 'done'
                group by sui.company_id
                """
        select_so_luong_kh_chap_nhan_duoc = """
                        select sui.company_id,count(suil.id)
                        from 
                            survey_user_input_line suil
                        LEFT JOIN survey_user_input sui ON  sui.id = suil.user_input_id
                        LEFT JOIN survey_label sl ON sl.id = suil.value_suggested
                        WHERE
                            sui.create_date >= %s and sui.create_date <= %s  and
                            sl.id in %s and sui.company_id in %s and sui.state = 'done'
                        group by sui.company_id
                        """
        select_so_luong_kh_dat_mong_doi = """
                        select sui.company_id,count(suil.id)
                        from 
                            survey_user_input_line suil
                        LEFT JOIN survey_user_input sui ON  sui.id = suil.user_input_id
                        LEFT JOIN survey_label sl ON sl.id = suil.value_suggested
                        WHERE
                            sui.create_date >= %s and sui.create_date <= %s  and
                            sl.id in %s and sui.company_id in %s and sui.state = 'done'
                        group by sui.company_id
                        """
        select_so_luong_kh_thoa_man = """
                        select sui.company_id,count(suil.id)
                        from 
                            survey_user_input_line suil
                        LEFT JOIN survey_user_input sui ON  sui.id = suil.user_input_id
                        LEFT JOIN survey_label sl ON sl.id = suil.value_suggested
                        WHERE
                            sui.create_date >= %s and sui.create_date <= %s  and
                            sl.id in %s and sui.company_id in %s and sui.state = 'done'
                        group by sui.company_id
                        """
        select_so_luong_kh_tren_uoc_muon = """
                        select sui.company_id,count(suil.id)
                        from 
                            survey_user_input_line suil
                        LEFT JOIN survey_user_input sui ON  sui.id = suil.user_input_id
                        LEFT JOIN survey_label sl ON sl.id = suil.value_suggested
                        WHERE
                            sui.create_date >= %s and sui.create_date <= %s  and
                            sl.id in %s and sui.company_id in %s and sui.state = 'done'
                        group by sui.company_id
                        """
        select_so_luong_kh_qua_tuyet_voi = """
                        select sui.company_id,count(suil.id)
                        from 
                            survey_user_input_line suil
                        LEFT JOIN survey_user_input sui ON  sui.id = suil.user_input_id
                        LEFT JOIN survey_label sl ON sl.id = suil.value_suggested
                        WHERE
                            sui.create_date >= %s and sui.create_date <= %s  and
                            sl.id in %s and sui.company_id in %s and sui.state = 'done'
                        group by sui.company_id
                        """
        select_tong_khao_sat = """
                        select sui.company_id,count(sui.id)
                        from 
                            survey_user_input sui
                        LEFT JOIN survey_survey ss ON ss.id = sui.survey_id
                        WHERE
                            sui.create_date >= %s and sui.create_date <= %s  and
                            sui.company_id in %s and ss.id = %s and sui.state = 'done'
                        group by sui.company_id
                        """
        select_tong_thuc_hien_khao_sat = """
                        select sui.company_id,count(sui.id)
                        from 
                            survey_user_input sui
                        LEFT JOIN survey_survey ss ON ss.id = sui.survey_id
                        WHERE
                            sui.create_date >= %s and sui.create_date <= %s  and
                            sui.company_id in %s and ss.id = %s
                            and sui.state = 'done'
                        group by sui.company_id
                        """
        select_ti_le_kh_ccgt = """
                        select sui.company_id,count(suil.id)
                        from 
                            survey_user_input_line suil
                        LEFT JOIN survey_label sl ON sl.id = suil.value_suggested
                        LEFT JOIN survey_user_input sui ON sui.id = suil.user_input_id
                        WHERE
                            sui.create_date >= %s and sui.create_date <= %s  and
                            sui.company_id in %s and sl.id in %s and sui.state = 'done'
                        group by sui.company_id
                        """
        select_ti_le_nps = """
                        select sui.company_id,count(suil.id)
                        from
                            survey_user_input_line suil
                        LEFT JOIN survey_label sl ON sl.id = suil.value_suggested
                        LEFT JOIN survey_user_input sui ON sui.id = suil.user_input_id
                        WHERE
                            sui.create_date >= %s and sui.create_date <= %s  and
                            sui.company_id in %s and sl.id in %s and sui.state = 'done'
                        group by sui.company_id
        """
        select = """
                select company_id,count(DISTINCT phone) 
                from crm_lead cl 
                where source_id in (85,86,137,138,125) 
                        and stage_id = 4 
                        and date_last_stage_update >= %s
                        and date_last_stage_update < %s
                        and company_id in %s
                group by company_id """

        select_kh_gioi_thieu = """
                select cl.company_id, sum(amount_proceeds) 
                from crm_sale_payment csp 
                left join crm_lead cl on cl.id = csp.booking_id 
                where payment_date >= %s and payment_date <= %s 
                        and cl.company_id in %s    
                        and cl.source_id in (85,86,137,138,125) 
                group by cl.company_id"""

        select_tong_cua_chi_nhanh = """
                select company_id, sum(amount_proceeds) 
                from crm_sale_payment 
                where payment_date >= %s and payment_date <= %s and company_id in %s    
                group by company_id"""

        # khách quay lại: số book thành công, có loại là khách hàng cũ, lọc trùng
        select_so_kh_quay_lai = """
                select cl.company_id, count(DISTINCT rp.id) from crm_lead cl
                left join res_partner rp on rp.id = cl.partner_id
                where stage_id = 4 and cl.type_data_partner = 'old'
                    and cl.date_last_stage_update >= %s and cl.date_last_stage_update < %s
                    and cl.company_id in %s
                group by cl.company_id"""

        # Tổng khách có booking thành công
        select_tong_khach_book_thanh_cong = """
                select cl.company_id, count(DISTINCT rp.id) from crm_lead cl
                left join res_partner rp on rp.id = cl.partner_id
                where cl.stage_id = 4 
                    and cl.date_last_stage_update >= %s and cl.date_last_stage_update < %s
                    and cl.company_id in %s
                group by cl.company_id"""

        # Tổng booking thành công
        select_tong_book_thanh_cong = """
                        select company_id, count(id) from crm_lead 
                        where stage_id = 4 
                            and date_last_stage_update >= %s and date_last_stage_update < %s
                            and company_id in %s
                        group by company_id"""

        # Tổng dịch vụ thành công
        # select_tong_dich_vu_thanh_cong = """
        #     select cld.company_id, count(cln.id) from crm_lead cld
        #     left join crm_line cln on cld.id = cln.crm_id
        #     where stage_id = 4 and type_data_partner = 'old' and cln.stage = 'done'
        #         and date_last_stage_update >= %s and date_last_stage_update < %s
        #         and cld.company_id in %s
        #     group by cld.company_id"""

        select_tong_dich_vu_thanh_cong = """
            select cld.company_id, count(cln.id) from crm_line cln 
            left join crm_lead cld on cld.id = cln.crm_id
            left join crm_category_source ccs on ccs.id = cld.category_source_id
            where cln.stage = 'done'
                and cln.date_done >= %s and cln.date_done < %s
                and cld.company_id in %s and cld.source_id in (85,86,137,138,125)
            group by cld.company_id """

        # Tổng dịch vụ thành công của khách hàng quay lại
        select_tong_dich_vu_thanh_cong_cua_kh_ql = """
                    select cld.company_id, count(cld.id) from crm_line cln
                    left join crm_lead cld on cld.id = cln.crm_id 
                    where cln.stage = 'done' and cld.type_data_partner = 'old'
                        and cln.date_done >= %s and cln.date_done < %s
                        and cld.company_id in %s
                    group by cld.company_id """

        # Doanh số khách hàng quay lại
        select_ds_kh_quay_lai = """
                        select cl.company_id, sum(amount_proceeds) 
                        from crm_sale_payment csp
                        left join crm_lead cl on cl.id = csp.booking_id
                        where payment_date >= %s and payment_date <= %s and cl.company_id in %s
                        and cl.type_data_partner = 'old'   
                        group by cl.company_id"""

        # Tổng booking thành công trong tháng
        select_tong_book_thanh_cong = """
            select company_id, count(id) from crm_lead 
            where stage_id = 4 and date_last_stage_update >= %s and date_last_stage_update < %s
                and company_id in %s
            group by company_id """

        # Số booking out sold
        select_tong_bk_out_sold = """
            select company_id, count(id) from crm_lead 
            where stage_id = 22 and date_last_stage_update >= %s and date_last_stage_update < %s
                and company_id in %s
            group by company_id
        """

        # Số phiếu khám thành công
        select_tong_phieu_kham = """
            select smh.his_company, count(sma.id) 
            from sh_medical_appointment_register_walkin sma
            left join sh_medical_health_center smh on smh.id = sma.institution
            where sma.state = 'Completed' and sma.create_date  >= %s and sma.create_date  < %s
                and smh.his_company in %s
            group by smh.his_company
        """

        # Số tái khám thành công
        select_tong_tai_kham = """
                    select smh.his_company, count(sme.id) 
                    from sh_medical_evaluation sme
                    left join sh_medical_health_center smh on smh.id = sme.institution
                    where sme.state = 'Completed' and sme.create_date >= %s and sme.create_date < %s
                        and smh.his_company in %s
                    group by smh.his_company
                """

        # Tổng case phát sinh trong tháng có phân loại là là ghi nhận hoặc khiếu nại
        select_tong_case_type = """
            select company_id, type_case, count(id) from crm_case 
            where type_case in ('complain', 'gop_y')
                and create_date >= %s and create_date < %s
                and company_id in  %s
            group by company_id, type_case"""

        # Tổng case khiếu nai trong tháng lọc trùng khách hàng
        select_tong_case_complain = """
        select company_id, count(id) from crm_case 
            where type_case = 'complain'
                and create_date >= %s and create_date < %s
                and company_id in  %s
            group by company_id
        """

        # Tổng case Cảnh báo trong tháng lọc trùng khách hàng
        select_tong_case_warning = """
        select company_id, count(id) from crm_case 
            where type_case = 'warning'
                and create_date >= %s and create_date < %s
                and company_id in  %s
            group by company_id
        """

        # Tổng case Góp ý-Phản ánh trong tháng lọc trùng khách hàng
        select_tong_case_gop_y = """
        select company_id, count(id) from crm_case 
            where type_case = 'gop_y'
                and create_date >= %s and create_date < %s
                and company_id in  %s
            group by company_id
        """

        # Tổng case phát sinh trong tháng
        select_tong_case = """
                select company_id, count(id) from crm_case 
                where create_date >= %s and create_date < %s
                    and company_id in  %s
                group by company_id"""

        # Số lượng bk thành công của khách hàng outsold trong 6 tháng gần nhất
        select_bk_tc_out_sold = """
        select count(DISTINCT partner_id)
        from crm_lead 
        where date_last_stage_update >= %s and date_last_stage_update < %s
            and company_id = %s and partner_id in %s and stage_id = 4"""
        select_bk_id_tc_out_sold = """
        select id
        from crm_lead 
        where date_last_stage_update >= %s and date_last_stage_update < %s
        and company_id = %s and partner_id in %s and stage_id = 4"""

        # Doanh số bk thành công của khách hàng outsold trong 6 tháng gần nhất
        select_ds_tc_out_sold = """
        select sum(amount_proceeds)
        from crm_sale_payment csp
        left join crm_lead cl on cl.id = csp.booking_id
        where payment_date >= %s and payment_date <= %s and cl.company_id = %s
        and cl.id in %s"""

        # Id partner có bk out sold trong 6 tháng
        select_partner_out_sold = """
        select DISTINCT partner_id 
        from crm_lead 
        where date_last_stage_update >= %s and date_last_stage_update < %s
        and company_id = %s and stage_id = 22"""

        # Số lượng khách hàng out sold
        select_partner_out_sold_bk = """
        select count(DISTINCT partner_id) 
        from crm_lead 
        where date_last_stage_update >= %s and date_last_stage_update < %s
        and company_id = %s and stage_id = 22"""

        # '2022-11-01 00:00:00.000'
        company_data = {}
        company_case = {}
        brand = ''
        for company_id in self.company_ids:

            company_month = {}
            company_case_month = {}
            for month in range(start_month, end_month + 1):
                line_data = {}
                for key in col_in_company:
                    line_data[key] = 0
                company_month[month] = line_data
                company_case_month[month] = {'complain': 0, 'gop_y': 0}

            company_data[company_id.id] = company_month
            company_case[company_id.id] = company_case_month
            brand = self.brand_id.code.lower()
            number_company += 1
        sum_kh_tm = {}
        sum_kh_tum = {}
        sum_kh_qtv = {}
        sum_kh_dmd = {}
        sum_kh_cnd = {}
        sum_kh_khl = {}
        tong_bk = {}
        tong_ccgt = {}
        # Query data
        for month in range(start_month, end_month + 1):
            today_last_day_in_start = monthrange(start_year, month)
            local_tz = timezone(self.env.user.tz or 'Etc/GMT+7')
            start_datetime = datetime(start_year, month, 1, 0, 0, 0)
            end_datetime = datetime(end_year, month, today_last_day_in_start[1], 23, 59, 59)
            start_datetime_6_month = datetime(end_year, month, 1, 0, 0, 0) - timedelta(
                days=180)
            end_datetime_6_month = datetime(end_year, month, today_last_day_in_start[1], 23, 59, 59) - timedelta(
                days=180)
            start_datetime = local_tz.localize(start_datetime, is_dst=None)
            end_datetime = local_tz.localize(end_datetime, is_dst=None)
            start_datetime_6_month = local_tz.localize(start_datetime_6_month, is_dst=None)
            end_datetime_6_month = local_tz.localize(end_datetime_6_month, is_dst=None)

            start_time_month = start_datetime.astimezone(utc).replace(tzinfo=None)
            end_time_month = end_datetime.astimezone(utc).replace(tzinfo=None)
            start_datetime_6 = start_datetime_6_month.astimezone(utc).replace(tzinfo=None)
            end_datetime_6 = end_datetime_6_month.astimezone(utc).replace(tzinfo=None)
            for company_id in self.company_ids:
                list_partner = [0]
                list_cl = [0]
                # Lấy id partner_id outsold
                self.env.cr.execute(select_partner_out_sold,
                                    (start_datetime_6, end_time_month, company_id.id))
                result = self.env.cr.fetchall()
                for c in result:
                    if c[0] not in list_partner:
                        list_partner.append(c[0])
                # Lấy số bk outsold tc
                self.env.cr.execute(select_bk_tc_out_sold,
                                    (start_time_month, end_time_month, company_id.id, tuple(list_partner)))
                result = self.env.cr.fetchall()
                for c in result:
                    if company_id.id in company_data:
                        if month in company_data[company_id.id]:
                            company_data[company_id.id][month]['so_kh_outsold_tc'] = c[0]
                # Lấy id bk thành công out sold:
                self.env.cr.execute(select_bk_id_tc_out_sold ,
                                    (start_time_month, end_time_month, company_id.id, tuple(list_partner)))
                result = self.env.cr.fetchall()
                for c in result:
                    list_cl.append(c[0])
                # Doanh số
                self.env.cr.execute(select_ds_tc_out_sold ,
                                    (start_time_month, end_time_month, company_id.id, tuple(list_cl)))
                result = self.env.cr.fetchall()
                for c in result:
                    if company_id.id in company_data:
                        if month in company_data[company_id.id]:
                            company_data[company_id.id][month]['ds_kh_outsold_tc'] = c[0] if c[0] else 0
                # Số lượng bk outsold
                self.env.cr.execute(select_partner_out_sold_bk,
                                    (start_time_month, end_time_month, company_id.id))
                result = self.env.cr.fetchall()
                for c in result:
                    if company_id.id in company_data:
                        if month in company_data[company_id.id]:
                            company_data[company_id.id][month]['kh_outsold'] = c[0]
        for month in range(start_month, end_month + 1):
            sum_kh_tm[month] = 0
            sum_kh_tum[month] = 0
            sum_kh_qtv[month] = 0
            sum_kh_dmd[month] = 0
            sum_kh_cnd[month] = 0
            sum_kh_khl[month] = 0
            tong_bk[month] = 0
            tong_ccgt[month] = 0
            today_last_day_in_start = monthrange(start_year, month)
            local_tz = timezone(self.env.user.tz or 'Etc/GMT+7')

            start_datetime = datetime(start_year, month, 1, 0, 0, 0)
            end_datetime = datetime(end_year, month, today_last_day_in_start[1], 23, 59, 59)
            start_datetime = local_tz.localize(start_datetime, is_dst=None)
            end_datetime = local_tz.localize(end_datetime, is_dst=None)

            start_time_month = start_datetime.astimezone(utc).replace(tzinfo=None)
            end_time_month = end_datetime.astimezone(utc).replace(tzinfo=None)
            tong_kh_tm = {}
            tong_kh_tum = {}
            tong_kh_qtv = {}
            tong_kh_dmd = {}
            tong_kh_cnd = {}
            tong_kh_khl = {}

            start_date_month = '%s-%s-01' % (start_year, month)
            end_date_month = '%s-%s-%s' % (end_year, month, today_last_day_in_start[1])
            survey_id = 'id_survey_%s' % brand
            id_survey = config.get_param(survey_id)
            ccd_ids = config.get_param('id_ccd')
            nps_9_10_ids = config.get_param('nps_9_10_ids')
            nps_0_1_ids = config.get_param('nps_0_1_ids')

            # Tổng số khách hàng không hài lòng
            self.env.cr.execute(select_so_luong_kh_khong_hai_long,
                                (start_time_month, end_time_month, list_khl, tuple(self.company_ids.ids)))
            result = self.env.cr.fetchall()
            for c in result:
                tong_kh_khl[c[0]] = c[1]
                sum_kh_khl[month] += c[1]
            # Tổng số khách hàng chấp nhận được
            self.env.cr.execute(select_so_luong_kh_chap_nhan_duoc,
                                (start_time_month, end_time_month, list_cnd, tuple(self.company_ids.ids)))
            result = self.env.cr.fetchall()
            for c in result:
                tong_kh_cnd[c[0]] = c[1]
                sum_kh_cnd[month] += c[1]
            # Tổng số khách hàng đạt mong đợi
            self.env.cr.execute(select_so_luong_kh_dat_mong_doi,
                                (start_time_month, end_time_month, list_dmd, tuple(self.company_ids.ids)))
            result = self.env.cr.fetchall()
            for c in result:
                tong_kh_dmd[c[0]] = c[1]
                sum_kh_dmd[month] += c[1]
            # Tổng số khách hàng thỏa mãn
            self.env.cr.execute(select_so_luong_kh_thoa_man,
                                (start_time_month, end_time_month, list_tm, tuple(self.company_ids.ids)))
            result = self.env.cr.fetchall()
            for c in result:
                tong_kh_tm[c[0]] = c[1]
                sum_kh_tm[month] += c[1]
            # Tổng số khách hàng trên ước muốn
            self.env.cr.execute(select_so_luong_kh_tren_uoc_muon,
                                (start_time_month, end_time_month, list_tum, tuple(self.company_ids.ids)))
            result = self.env.cr.fetchall()
            for c in result:
                tong_kh_tum[c[0]] = c[1]
                sum_kh_tum[month] += c[1]
            # Tổng số khách hàng quá tuyệt vời
            self.env.cr.execute(select_so_luong_kh_qua_tuyet_voi,
                                (start_time_month, end_time_month, list_qtv, tuple(self.company_ids.ids)))
            result = self.env.cr.fetchall()
            for c in result:
                tong_kh_qtv[c[0]] = c[1]
                sum_kh_qtv[month] += c[1]
            self.env.cr.execute(select_tong_khao_sat,
                                (start_time_month, end_time_month, tuple(self.company_ids.ids), int(id_survey)))
            result = self.env.cr.fetchall()
            for c in result:
                if c[0] in company_data:
                    if month in company_data[c[0]]:
                        company_data[c[0]][month]['so_luong_kh_khong_hai_long'] = tong_kh_khl[c[0]] / c[1] * 100 if c[
                                                                                                                        1] != 0 and \
                                                                                                                    c[
                                                                                                                        0] in tong_kh_khl and \
                                                                                                                    tong_kh_khl[
                                                                                                                        c[
                                                                                                                            0]] else 0
                        company_data[c[0]][month]['so_luong_kh_chap_nhan_duoc'] = tong_kh_cnd[c[0]] / c[1] * 100 if c[
                                                                                                                        1] != 0 and \
                                                                                                                    c[
                                                                                                                        0] in tong_kh_cnd and \
                                                                                                                    tong_kh_cnd[
                                                                                                                        c[
                                                                                                                            0]] else 0
                        company_data[c[0]][month]['so_luong_kh_dat_mong_doi'] = tong_kh_dmd[c[0]] / c[1] * 100 if c[
                                                                                                                      1] != 0 and \
                                                                                                                  c[
                                                                                                                      0] in tong_kh_dmd and \
                                                                                                                  tong_kh_dmd[
                                                                                                                      c[
                                                                                                                          0]] else 0
                        company_data[c[0]][month]['so_luong_kh_thoa_man_uoc_muon'] = tong_kh_tm[c[0]] / c[1] * 100 if c[
                                                                                                                          1] != 0 and \
                                                                                                                      c[
                                                                                                                          0] in tong_kh_tm and \
                                                                                                                      tong_kh_tm[
                                                                                                                          c[
                                                                                                                              0]] else 0
                        company_data[c[0]][month]['so_luong_kh_tren_ca_uoc_muon'] = tong_kh_tum[c[0]] / c[1] * 100 if c[
                                                                                                                          1] != 0 and \
                                                                                                                      c[
                                                                                                                          0] in tong_kh_tum and \
                                                                                                                      tong_kh_tum[
                                                                                                                          c[
                                                                                                                              0]] else 0
                        company_data[c[0]][month]['so_luong_kh_qua_tuyet_voi'] = tong_kh_qtv[c[0]] / c[1] * 100 if c[
                                                                                                                       1] != 0 and \
                                                                                                                   c[
                                                                                                                       0] in tong_kh_qtv and \
                                                                                                                   tong_kh_qtv[
                                                                                                                       c[
                                                                                                                           0]] else 0
                        company_data[c[0]][month]['tb_diem_tnnk'] = (company_data[c[0]][month][
                                                                         'so_luong_kh_khong_hai_long'] +
                                                                     company_data[c[0]][month][
                                                                         'so_luong_kh_chap_nhan_duoc'] * 2 +
                                                                     company_data[c[0]][month][
                                                                         'so_luong_kh_dat_mong_doi'] * 3 +
                                                                     company_data[c[0]][month][
                                                                         'so_luong_kh_thoa_man_uoc_muon'] * 4 +
                                                                     company_data[c[0]][month][
                                                                         'so_luong_kh_tren_ca_uoc_muon'] * 5 +
                                                                     company_data[c[0]][month][
                                                                         'so_luong_kh_qua_tuyet_voi'] * 6) / 100
                        company_data[c[0]][month]['space'] = '-'
            # Tổng Khảo sát
            self.env.cr.execute(select_tong_khao_sat,
                                (start_time_month, end_time_month, tuple(self.company_ids.ids), int(id_survey)))
            result = self.env.cr.fetchall()
            for c in result:
                if c[0] in company_data:
                    if month in company_data[c[0]]:
                        company_data[c[0]][month]['tong_khao_sat'] = c[1]

            # Số Booking outsold
            self.env.cr.execute(select_tong_bk_out_sold,
                                (start_time_month, end_time_month, tuple(self.company_ids.ids)))
            result = self.env.cr.fetchall()
            bk_os = {}
            for c in result:
                bk_os[c[0]] = c[1]
                tong_bk[month] += c[1]
            # Số phiếu khám thành công
            self.env.cr.execute(select_tong_phieu_kham,
                                (start_time_month, end_time_month, tuple(self.company_ids.ids)))
            result = self.env.cr.fetchall()
            pk_tc = {}
            for c in result:
                pk_tc[c[0]] = c[1]
                tong_bk[month] += c[1]
            # Số tái khám thành công
            self.env.cr.execute(select_tong_tai_kham,
                                (start_time_month, end_time_month, tuple(self.company_ids.ids)))
            result = self.env.cr.fetchall()
            tk_tc = {}
            for c in result:
                tk_tc[c[0]] = c[1]
                tong_bk[month] += c[1]

            # Tỉ lệ khách hàng chọn chắc chắn giới thiệu
            # self.env.cr.execute(select_ti_le_kh_ccgt,
            #                     (start_time_month, end_time_month, tuple(self.company_ids.ids),
            #                      tuple(map(int, ccd_ids.split(', ')))))
            # result = self.env.cr.fetchall()
            # for c in result:
            #     if c[0] in company_data:
            #         if month in company_data[c[0]]:
            #             company_data[c[0]][month]['ti_le_kh_chac_chan_gioi_thieu'] = c[1] / company_data[c[0]][month]['tong_khao_sat'] * 100 if company_data[c[0]][month]['tong_khao_sat'] != 0 else 0
            #             tong_ccgt[month] += c[1]

            # Tỉ lệ nps 9-10
            ti_le_9_10 = {}
            self.env.cr.execute(select_ti_le_nps,
                                (start_time_month, end_time_month, tuple(self.company_ids.ids),
                                 tuple(map(int, nps_9_10_ids.split(', ')))))
            result = self.env.cr.fetchall()
            for c in result:
                if c[0] in company_data:
                    if month in company_data[c[0]]:
                        ti_le_9_10[c[0]] = c[1]
                        tong_ccgt[month] += c[1]

            # Tỉ lệ 0-6
            ti_le_0_1 = {}
            self.env.cr.execute(select_ti_le_nps,
                                (start_time_month, end_time_month, tuple(self.company_ids.ids),
                                 tuple(map(int, nps_0_1_ids.split(', ')))))
            result = self.env.cr.fetchall()
            for c in result:
                if c[0] in company_data:
                    if month in company_data[c[0]]:
                        ti_le_0_1[c[0]] = c[1]
                        tong_ccgt[month] -= c[1]

            # Tỉ lệ thực hiện khảo sát
            self.env.cr.execute(select_tong_thuc_hien_khao_sat,
                                (start_time_month, end_time_month, tuple(self.company_ids.ids), int(id_survey)))
            result = self.env.cr.fetchall()
            sum_bk = 0
            for c in result:
                if c[0] in company_data:
                    if month in company_data[c[0]]:
                        booking = bk_os[c[0]] if c[0] in bk_os else 0
                        walkin = pk_tc[c[0]] if c[0] in pk_tc else 0
                        eva = tk_tc[c[0]] if c[0] in tk_tc else 0
                        sum_bk = booking + walkin + eva
                        company_data[c[0]][month]['ti_le_thuc_hien_ks'] = c[1] / sum_bk * 100 if sum_bk != 0 else 0
                        tl_9_10 = ti_le_9_10[c[0]] if c[0] in ti_le_9_10 else 0
                        tl_0_1 = ti_le_0_1[c[0]] if c[0] in ti_le_0_1 else 0
                        company_data[c[0]][month]['ti_le_kh_chac_chan_gioi_thieu'] = (tl_9_10 / c[1] - tl_0_1 / c[1]) * 100 if c[1] != 0 else 0

            # Số Khách hàng thành công
            self.env.cr.execute(select_tong_khach_book_thanh_cong,
                                (start_time_month, end_time_month, tuple(self.company_ids.ids)))
            result = self.env.cr.fetchall()
            for c in result:
                if c[0] in company_data:
                    if month in company_data[c[0]]:
                        company_data[c[0]][month]['so_kh_thanh_cong'] = c[1]

            # Số lượng khách hàng giới thiệu
            self.env.cr.execute(select, (start_time_month, end_time_month, tuple(self.company_ids.ids)))
            result = self.env.cr.fetchall()
            for c in result:
                if c[0] in company_data:
                    if month in company_data[c[0]]:
                        company_data[c[0]][month]['so_luong_kh_duoc_gioi_thieu'] = c[1]
                        company_data[c[0]][month]['ti_le_kh_duoc_gt_tren_dau_khach'] = c[1] / company_data[c[0]][month][
                            'so_kh_thanh_cong'] * 100 if company_data[c[0]][month]['so_kh_thanh_cong'] != 0 else 0

            # Doanh số do khách hàng giới thiệu mang lại
            self.env.cr.execute(select_kh_gioi_thieu, (start_date_month, end_date_month, tuple(self.company_ids.ids)))
            result = self.env.cr.fetchall()
            for c in result:
                if c[0] in company_data:
                    if month in company_data[c[0]]:
                        company_data[c[0]][month]['ds_do_kh_duoc_gioi_thieu_mang_lai'] = c[1]

            # Doanh số của chi nhánh:
            self.env.cr.execute(select_tong_cua_chi_nhanh,
                                (start_date_month, end_date_month, tuple(self.company_ids.ids)))
            result = self.env.cr.fetchall()
            for c in result:
                if c[0] in company_data:
                    if month in company_data[c[0]]:
                        # Tỉ lệ
                        if c[1]:
                            company_data[c[0]][month]['ti_le'] = company_data[c[0]][month][
                                                                     'ds_do_kh_duoc_gioi_thieu_mang_lai'] / c[1] * 100
                        else:
                            company_data[c[0]][month]['ti_le'] = 0
                        company_data[c[0]][month]['ds_tong_cua_chi_nhanh'] = c[1]

            self.env.cr.execute(select_so_kh_quay_lai, (start_time_month, end_time_month, tuple(self.company_ids.ids)))
            result = self.env.cr.fetchall()
            for c in result:
                if c[0] in company_data:
                    if month in company_data[c[0]]:
                        company_data[c[0]][month]['so_kh_quay_lai'] = c[1]
                        company_data[c[0]][month]['ti_le_quay_lai_theo_khach'] = company_data[c[0]][month][
                                                                                     'so_kh_quay_lai'] / \
                                                                                 company_data[c[0]][month][
                                                                                     'so_kh_thanh_cong'] * 100 if \
                            company_data[c[0]][month]['so_kh_thanh_cong'] else 0

            # Tong khach thang cong
            # self.env.cr.execute(select_tong_khach_book_thanh_cong,
            #                     (start_time_month, end_time_month, tuple(self.company_ids.ids)))
            # result = self.env.cr.fetchall()
            # for c in result:
            #     if c[0] in company_data:
            #         if month in company_data[c[0]]:

            # Tong dich vu thanh cong
            self.env.cr.execute(select_tong_dich_vu_thanh_cong,
                                (start_time_month, end_time_month, tuple(self.company_ids.ids)))
            result = self.env.cr.fetchall()
            for c in result:
                if c[0] in company_data:
                    if month in company_data[c[0]]:
                        company_data[c[0]][month]['tong_dv_thanh_cong'] = c[1]
                        company_data[c[0]][month]['so_dv_tb_tren_dau_khach'] = c[1] / company_data[c[0]][month][
                            'so_luong_kh_duoc_gioi_thieu'] if company_data[c[0]][month][
                                                                  'so_luong_kh_duoc_gioi_thieu'] != 0 else 0

            # Tổng số dịch vụ thành công của KH quay lại
            self.env.cr.execute(select_tong_dich_vu_thanh_cong_cua_kh_ql,
                                (start_time_month, end_time_month, tuple(self.company_ids.ids)))
            result = self.env.cr.fetchall()
            for c in result:
                if c[0] in company_data:
                    if month in company_data[c[0]]:
                        company_data[c[0]][month]['tong_so_dv_thanh_cong_kh_quay_lai'] = c[1]
                        company_data[c[0]][month]['ti_le_quay_lai_theo_dich_vu'] = c[1] / company_data[c[0]][month][
                            'so_kh_quay_lai'] if company_data[c[0]][month]['so_kh_quay_lai'] != 0 else 0

            # Doanh số Khách hàng quay lại
            self.env.cr.execute(select_ds_kh_quay_lai,
                                (start_date_month, end_date_month, tuple(self.company_ids.ids)))
            result = self.env.cr.fetchall()
            for c in result:
                if c[0] in company_data:
                    if month in company_data[c[0]]:
                        company_data[c[0]][month]['ds_kh_quay_lai'] = c[1]
                        company_data[c[0]][month]['ti_le_doanh_so'] = c[1] / company_data[c[0]][month][
                            'ds_tong_cua_chi_nhanh'] * 100 if company_data[c[0]][month][
                                                                  'ds_tong_cua_chi_nhanh'] != 0 else 0

            # Tổng book thành công trong tháng TODO
            self.env.cr.execute(select_tong_case,
                                (start_time_month, end_time_month, tuple(self.company_ids.ids)))
            result = self.env.cr.fetchall()
            tong_case = {}
            for c in result:
                company_data[c[0]][month]['tong_case'] = c[1]

            # # Case
            # self.env.cr.execute(select_tong_case_type, (start_time_month, end_time_month, tuple(self.company_ids.ids)))
            # result = self.env.cr.fetchall()
            # for c in result:
            #     # c la (company_id, type_case, number) ex (4, 'complain', 19)
            #     if c[0] in company_data:
            #         if month in company_data[c[0]]:
            #             if c[0] in tong_case and tong_case[c[0]]:
            #                 company_data[c[0]][month]['tong_case'] = tong_case[c[0]]
            #                 if c[1] == 'complain':
            #                     company_data[c[0]][month]['so_luong_case'] = c[2]
            #                     company_data[c[0]][month]['ti_le_case'] = c[2] / tong_case[c[0]] * 100 if tong_case[c[0]] != 0 else 0
            #                 if c[1] == 'gop_y':
            #                     company_data[c[0]][month]['so_luong_phan_anh'] = c[2]
            #                     company_data[c[0]][month]['ti_le_phan_anh'] = c[2] / tong_case[c[0]] * 100 if tong_case[c[0]] != 0 else 0
            # Case khiếu nại
            self.env.cr.execute(select_tong_case_complain,
                                (start_time_month, end_time_month, tuple(self.company_ids.ids)))
            result = self.env.cr.fetchall()
            for c in result:
                if c[0] in company_data:
                    if month in company_data[c[0]]:
                        company_data[c[0]][month]['so_luong_case'] = c[1]
                        company_data[c[0]][month]['ti_le_case'] = c[1] / company_data[c[0]][month][
                            'so_kh_thanh_cong'] * 100 if company_data[c[0]][month]['so_kh_thanh_cong'] != 0 else 0
            # Case góp ý - phản ánh
            self.env.cr.execute(select_tong_case_gop_y,
                                (start_time_month, end_time_month, tuple(self.company_ids.ids)))
            result = self.env.cr.fetchall()
            for c in result:
                if c[0] in company_data:
                    if month in company_data[c[0]]:
                        company_data[c[0]][month]['so_luong_phan_anh'] = c[1]
                        company_data[c[0]][month]['ti_le_phan_anh'] = c[1] / company_data[c[0]][month][
                            'so_kh_thanh_cong'] * 100 if company_data[c[0]][month]['so_kh_thanh_cong'] != 0 else 0

            # Case cảnh báo
            self.env.cr.execute(select_tong_case_warning,
                                (start_time_month, end_time_month, tuple(self.company_ids.ids)))
            result = self.env.cr.fetchall()
            for c in result:
                if c[0] in company_data:
                    if month in company_data[c[0]]:
                        company_data[c[0]][month]['so_luong_canh_bao'] = c[1]
                        company_data[c[0]][month]['ti_le_canh_bao'] = c[1] / company_data[c[0]][month][
                            'so_kh_thanh_cong'] * 100 if company_data[c[0]][month]['so_kh_thanh_cong'] != 0 else 0
        for month in range(start_month, end_month + 1):
            line = ['Tháng %s' % month]

            total_so_luong_kh_duoc_gioi_thieu = 0
            total_ds_do_kh_duoc_gioi_thieu_mang_lai = 0
            total_ds_tong_cua_chi_nhanh = 0
            total_so_kh_quay_lai = 0
            total_khl = sum_kh_khl[month]
            total_cnd = sum_kh_cnd[month]
            total_dmd = sum_kh_dmd[month]
            total_tm = sum_kh_tm[month]
            total_tum = sum_kh_tum[month]
            total_qtv = sum_kh_qtv[month]
            total_bk = tong_bk[month]
            total_ccgt = tong_ccgt[month]
            total_tong_ks = 0
            total_kh_thanh_cong = 0
            total_dv_thanh_cong = 0
            total_dv_thanh_cong_kh_ql = 0
            total_ds_kh_ql = 0
            tong_case = 0
            total_case = 0
            total_phan_anh = 0
            total_canh_bao = 0
            avage_diem_tnnk = 0
            avage_thuc_hien_ks = 0
            avage_ti_le_kh_ccd = 0
            total_kh_outsold = 0
            total_ds_outsold = 0
            total_bk_outsold_tc = 0
            for company_id in self.company_ids:
                data_line = company_data[company_id.id][month]
                for col in col_in_company:
                    line.append(data_line[col])

                    # Tính cột được cộng vào tổng
                    if col == 'so_luong_kh_duoc_gioi_thieu':
                        total_so_luong_kh_duoc_gioi_thieu += data_line[col]
                    if col == 'ds_do_kh_duoc_gioi_thieu_mang_lai':
                        total_ds_do_kh_duoc_gioi_thieu_mang_lai += data_line[col]
                    if col == 'ds_tong_cua_chi_nhanh':
                        total_ds_tong_cua_chi_nhanh += data_line[col]
                    if col == 'so_kh_quay_lai':
                        total_so_kh_quay_lai += data_line[col]
                    if col == 'tong_khao_sat':
                        total_tong_ks += data_line[col]
                    if col == 'tong_dv_thanh_cong':
                        total_dv_thanh_cong += data_line[col]
                    if col == 'so_kh_thanh_cong':
                        total_kh_thanh_cong += data_line[col]
                    if col == 'tong_so_dv_thanh_cong_kh_quay_lai':
                        total_dv_thanh_cong_kh_ql += data_line[col]
                    if col == 'ds_kh_quay_lai':
                        total_ds_kh_ql += data_line[col]
                    if col == 'tong_case':
                        tong_case += data_line[col]
                    if col == 'so_luong_case':
                        total_case += data_line[col]
                    if col == 'so_luong_phan_anh':
                        total_phan_anh += data_line[col]
                    if col == 'so_luong_canh_bao':
                        total_canh_bao += data_line[col]
                    if col == 'kh_outsold':
                        total_kh_outsold += data_line[col]
                    if col == 'so_kh_outsold_tc':
                        total_bk_outsold_tc += data_line[col]
                    if col == 'ds_kh_outsold_tc':
                        total_ds_outsold += data_line[col] / number_company
                    # if col == 'so_luong_kh_khong_hai_long':
                    #     avage_khl += data_line[col]/number_company
                    # if col == 'so_luong_kh_chap_nhan_duoc':
                    #     avage_cnd += data_line[col]/number_company
                    # if col == 'so_luong_kh_dat_mong_doi':
                    #     avage_dmd += data_line[col]/number_company
                    # if col == 'so_luong_kh_thoa_man_uoc_muon':
                    #     avage_tm += data_line[col]/number_company
                    # if col == 'so_luong_kh_tren_ca_uoc_muon':
                    #     avage_tum += data_line[col]/number_company
                    # if col == 'so_luong_kh_qua_tuyet_voi':
                    #     avage_qtv += data_line[col]/number_company
                    if col == 'tb_diem_tnnk':
                        avage_diem_tnnk += data_line[col] / number_company
                    if col == 'ti_le_thuc_hien_ks':
                        avage_thuc_hien_ks += data_line[col] / number_company
                    if col == 'ti_le_kh_chac_chan_gioi_thieu':
                        avage_ti_le_kh_ccd += data_line[col] / number_company

            # Tổng hợp các chi nhánh, những cột là tỉ lệ sẽ không tính tổng
            for col in col_in_company:
                if col == 'so_luong_kh_duoc_gioi_thieu':
                    line.append(total_so_luong_kh_duoc_gioi_thieu)
                elif col == 'ds_do_kh_duoc_gioi_thieu_mang_lai':
                    line.append(total_ds_do_kh_duoc_gioi_thieu_mang_lai)
                elif col == 'ds_tong_cua_chi_nhanh':
                    line.append(total_ds_tong_cua_chi_nhanh)
                elif col == 'so_kh_quay_lai':
                    line.append(total_so_kh_quay_lai)
                elif col == 'tong_khao_sat':
                    line.append(total_tong_ks)
                elif col == 'tong_dv_thanh_cong':
                    line.append(total_dv_thanh_cong)
                elif col == 'so_kh_thanh_cong':
                    line.append(total_kh_thanh_cong)
                elif col == 'tong_so_dv_thanh_cong_kh_quay_lai':
                    line.append(total_dv_thanh_cong_kh_ql)
                elif col == 'ds_kh_quay_lai':
                    line.append(total_ds_kh_ql)
                elif col == 'tong_case':
                    line.append(tong_case)
                elif col == 'so_luong_case':
                    line.append(total_case)
                elif col == 'so_luong_phan_anh':
                    line.append(total_phan_anh)
                elif col == 'so_luong_canh_bao':
                    line.append(total_canh_bao)
                elif col == 'kh_outsold':
                    line.append(total_kh_outsold)
                elif col == 'so_kh_outsold_tc':
                    line.append(total_bk_outsold_tc)
                elif col == 'ds_kh_outsold_tc':
                    line.append(total_ds_outsold)
                elif col == 'so_luong_kh_khong_hai_long':
                    avg = total_khl / total_tong_ks * 100 if total_tong_ks != 0 else 0
                    line.append(avg)
                elif col == 'so_luong_kh_chap_nhan_duoc':
                    avg = total_cnd / total_tong_ks * 100 if total_tong_ks != 0 else 0
                    line.append(avg)
                elif col == 'so_luong_kh_dat_mong_doi':
                    avg = total_dmd / total_tong_ks * 100 if total_tong_ks != 0 else 0
                    line.append(avg)
                elif col == 'so_luong_kh_thoa_man_uoc_muon':
                    avg = total_tm / total_tong_ks * 100 if total_tong_ks != 0 else 0
                    line.append(avg)
                elif col == 'so_luong_kh_tren_ca_uoc_muon':
                    avg = total_tum / total_tong_ks * 100 if total_tong_ks != 0 else 0
                    line.append(avg)
                elif col == 'so_luong_kh_qua_tuyet_voi':
                    avg = total_qtv / total_tong_ks * 100 if total_tong_ks != 0 else 0
                    line.append(avg)
                elif col == 'tb_diem_tnnk':
                    avg = (total_khl + total_cnd * 2 + total_dmd * 3 + total_tm * 4 + total_tum * 5 + total_qtv * 6) / total_tong_ks if total_tong_ks != 0 else 0
                    line.append(avg)
                elif col == 'ti_le_thuc_hien_ks':
                    avg = total_tong_ks / total_bk * 100 if total_bk != 0 else 0
                    line.append(avg)
                elif col == 'ti_le_kh_chac_chan_gioi_thieu':
                    avg = total_ccgt / total_tong_ks * 100 if total_tong_ks != 0 else 0
                    line.append(avg)
                elif col == 'so_dv_tb_tren_dau_khach':
                    avg = total_dv_thanh_cong / total_so_luong_kh_duoc_gioi_thieu if total_so_luong_kh_duoc_gioi_thieu != 0 else 0
                    line.append(avg)
                elif col == 'ti_le_kh_duoc_gt_tren_dau_khach':
                    avg = total_so_luong_kh_duoc_gioi_thieu / total_kh_thanh_cong * 100 if total_kh_thanh_cong != 0 else 0
                    line.append(avg)
                elif col == 'ti_le':
                    avg = total_ds_do_kh_duoc_gioi_thieu_mang_lai / total_ds_tong_cua_chi_nhanh * 100 if total_ds_tong_cua_chi_nhanh != 0 else 0
                    line.append(avg)
                elif col == 'ti_le_quay_lai_theo_dich_vu':
                    avg = total_dv_thanh_cong_kh_ql / total_so_kh_quay_lai if total_so_kh_quay_lai != 0 else 0
                    line.append(avg)
                elif col == 'ti_le_quay_lai_theo_khach':
                    avg = total_so_kh_quay_lai / total_kh_thanh_cong * 100 if total_kh_thanh_cong != 0 else 0
                    line.append(avg)
                elif col == 'ti_le_doanh_so':
                    avg = total_ds_kh_ql / total_ds_tong_cua_chi_nhanh * 100 if total_ds_tong_cua_chi_nhanh != 0 else 0
                    line.append(avg)
                elif col == 'ti_le_case':
                    avg = total_case / total_kh_thanh_cong * 100 if total_kh_thanh_cong != 0 else 0
                    line.append(avg)
                elif col == 'ti_le_phan_anh':
                    avg = total_phan_anh / total_kh_thanh_cong * 100 if total_kh_thanh_cong != 0 else 0
                    line.append(avg)
                elif col == 'so_luong_canh_bao':
                    line.append(total_canh_bao)
                elif col == 'ti_le_canh_bao':
                    avg = total_canh_bao / total_kh_thanh_cong * 100 if total_kh_thanh_cong != 0 else 0
                    line.append(avg)
                else:
                    line.append('-')

            data.append(line)

        # Tổng các tháng
        line = ['Tổng']
        first_row = 0
        for row in data:
            if first_row == 0:
                # Lấy dữ liệu dòng đầu tiên và chỉ những cột tính tổng hay số lượng
                col_index = 0
                col_in_c_index = 0
                for r in row:
                    if col_index > 0:
                        if col_in_c_index in [7, 10, 11, 12, 13, 16, 18, 19, 22, 24, 25, 26, 27, 28, 30, 32]:
                            line.append(r)
                        else:
                            line.append('-')

                        col_in_c_index += 1
                        if col_in_c_index == len_col_in_company:
                            col_in_c_index = 0
                    col_index += 1

            else:
                row_index = 0
                for r in row:
                    if row_index > 0:
                        if line[row_index] not in ['-', 'Tổng']:
                            line[row_index] += r
                    row_index += 1
            first_row += 1
        data.append(line)

        # # Trung bình các tháng
        # line_tb = ['Trung bình']
        # first_row = 0
        # number_row = 0
        # for row in data:
        #     number_row += 1
        # for row in data:
        #     if first_row == 0:
        #         # Lấy dữ liệu dòng đầu tiên và chỉ những cột tính tổng hay số lượng
        #         col_index = 0
        #         col_in_c_index = 0
        #         for r in row:
        #             if col_index > 0:
        #                 if col_in_c_index not in [7, 10, 11, 12, 13, 16, 18, 19, 22, 24, 25, 27, 29]:
        #                     line_tb.append(r/(number_row-1))
        #                 else:
        #                     line_tb.append('-')
        #
        #                 col_in_c_index += 1
        #                 if col_in_c_index == len_col_in_company:
        #                     col_in_c_index = 0
        #             col_index += 1
        #     else:
        #         row_index = 0
        #         for r in row:
        #             if row_index > 0:
        #                 if line_tb[row_index] not in ['-', 'Trung bình']:
        #                     line_tb[row_index] += r/(number_row-1) if r != '-' else 0
        #             row_index += 1
        #     first_row += 1
        # data.append(line_tb)

        # TODO xử lý gộp lại 1 dòng
        for company_id in self.company_ids:
            for col in col_in_company:
                cols.append('')

        # Tổng hợp các chi nhánh
        for col in col_in_company:
            cols.append('')

        # ----------- in dữ liệu
        template = self.env['ir.attachment'].browse(self.env.ref('crm_report.report_tnkh_attachment').id)
        decode = base64.b64decode(template.datas)
        wb = load_workbook(BytesIO(decode))
        ws = wb.active
        # if self.start_date == self.end_date:
        #     ws['A3'].value = self.end_date.strftime('Ngày %d tháng %m năm %Y')
        # else:
        #     ws['A3'].value = 'Từ ngày: %s đến ngày: %s' % (
        #         self.start_date.strftime('%d/%m/%Y'), self.end_date.strftime('%d/%m/%Y'))
        thin = borders.Side(style='thin')
        all_border_thin = borders.Border(left=thin, right=thin, top=thin, bottom=thin)
        line_font = Font(name='Times New Roman', size=13)

        # format_currency = '_(* #,##0_);_(* (#,##0);_(* "-"??_);_(@_)'

        # Xử lý header
        row = 1
        start_col = 2
        start_col_name = 2
        start_col_group = 2
        row_group = 2

        row_name = 3

        default_cell_header = None
        data_merge = []
        number_merge = 0

        for company_id in self.company_ids:
            cell = ws.cell(row, start_col)
            if default_cell_header is None:
                default_cell_header = cell
            cell.value = company_id.name
            cell.font = Font(name='Times New Roman', size=14, bold=True)
            start_col += len_col_in_company
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.fill = PatternFill(start_color='f9e5e1', end_color='f9e5e1', fill_type='solid')
            cell_merge = ws.cell(row, start_col - 1)
            data_merge.append("%s:%s" % (cell.coordinate, cell_merge.coordinate))

            # ws.merge_cells(start_row=row,start_column=start,end_row=row,end_column=start_col)
            for col_group in colgroup_in_company:
                cell = ws.cell(row_group, start_col_group)
                if col_group == 'Khảo sát sự hài lòng của khách hàng':
                    number_merge = 8
                    cell.fill = PatternFill(start_color='ff8b94', end_color='ff8b94', fill_type='solid')
                elif col_group == 'NPS':
                    number_merge = 0
                    cell.fill = PatternFill(start_color='FFA586', end_color='FFA586', fill_type='solid')
                elif col_group == 'Chi nhánh':
                    number_merge = 1
                    cell.fill = PatternFill(start_color='e6d7ff', end_color='e6d7ff', fill_type='solid')
                elif col_group == 'KH giới thiệu KH':
                    number_merge = 5
                    cell.fill = PatternFill(start_color='FFE286', end_color='FFE286', fill_type='solid')
                elif col_group == 'KH quay lại':
                    number_merge = 5
                    cell.fill = PatternFill(start_color='ffb3ba', end_color='ffb3ba', fill_type='solid')
                elif col_group == 'KH Outsold':
                    number_merge = 2
                    cell.fill = PatternFill(start_color='03fcec', end_color='03fcec', fill_type='solid')
                elif col_group == 'CASE':
                    number_merge = 0
                    cell.fill = PatternFill(start_color='ffffba', end_color='ffffba', fill_type='solid')
                elif col_group == 'Khiếu nại':
                    number_merge = 1
                    cell.fill = PatternFill(start_color='baffc9', end_color='baffc9', fill_type='solid')
                elif col_group == 'Phản ánh':
                    number_merge = 1
                    cell.fill = PatternFill(start_color='bae1ff', end_color='bae1ff', fill_type='solid')
                elif col_group == 'Cảnh báo':
                    number_merge = 1
                    cell.fill = PatternFill(start_color='dcedc1', end_color='dcedc1', fill_type='solid')
                elif col_group == '':
                    number_merge = 0

                cell.value = col_group
                cell.font = Font(name='Times New Roman', size=13, bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                ws.merge_cells(start_row=row_group, start_column=start_col_group, end_row=row_group,
                               end_column=start_col_group + number_merge)
                cell.border = all_border_thin
                start_col_group += number_merge + 1

            for col_name in colname_in_company:
                cell = ws.cell(row_name, start_col_name)
                cell.value = col_name
                # cell.fill = copy(default_cell_header.fill)
                cell.border = all_border_thin
                cell.font = Font(name='Times New Roman', size=13, bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                start_col_name += 1

        # Thêm tiêu đề tổng
        cell = ws.cell(row, start_col)
        cell.value = "TỔNG VÀ TRUNG BÌNH"
        cell.font = Font(name='Times New Roman', size=14, bold=True)
        cell.fill = PatternFill(start_color='f9e5e1', end_color='f9e5e1', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell_merge = ws.cell(row, start_col + len_col_in_company - 1)
        data_merge.append(("%s:%s" % (cell.coordinate, cell_merge.coordinate)))

        # Merge header tên của chi nhánh
        for merge in data_merge:
            ws.merge_cells(merge)
        for col_group in colgroup_in_company:
            cell = ws.cell(row_group, start_col_group)
            if col_group == 'Khảo sát sự hài lòng của khách hàng':
                number_merge = 8
                cell.fill = PatternFill(start_color='ff8b94', end_color='ff8b94', fill_type='solid')
            elif col_group == 'NPS':
                number_merge = 0
                cell.fill = PatternFill(start_color='FFA586', end_color='FFA586', fill_type='solid')
            elif col_group == 'Chi nhánh':
                number_merge = 1
                cell.fill = PatternFill(start_color='e6d7ff', end_color='e6d7ff', fill_type='solid')
            elif col_group == 'KH giới thiệu KH':
                number_merge = 5
                cell.fill = PatternFill(start_color='FFE286', end_color='FFE286', fill_type='solid')
            elif col_group == 'KH quay lại':
                number_merge = 5
                cell.fill = PatternFill(start_color='ffb3ba', end_color='ffb3ba', fill_type='solid')
            elif col_group == 'KH Outsold':
                number_merge = 2
                cell.fill = PatternFill(start_color='03fcec', end_color='03fcec', fill_type='solid')
            elif col_group == 'CASE':
                number_merge = 0
                cell.fill = PatternFill(start_color='ffffba', end_color='ffffba', fill_type='solid')
            elif col_group == 'Khiếu nại':
                number_merge = 1
                cell.fill = PatternFill(start_color='baffc9', end_color='baffc9', fill_type='solid')
            elif col_group == 'Phản ánh':
                number_merge = 1
                cell.fill = PatternFill(start_color='bae1ff', end_color='bae1ff', fill_type='solid')
            elif col_group == 'Cảnh báo':
                number_merge = 1
                cell.fill = PatternFill(start_color='dcedc1', end_color='dcedc1', fill_type='solid')
            elif col_group == '':
                number_merge = 0

            cell.value = col_group
            cell.font = Font(name='Times New Roman', size=13, bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = all_border_thin
            ws.merge_cells(start_row=row_group, start_column=start_col_group, end_row=row_group,
                           end_column=start_col_group + number_merge)
            start_col_group += number_merge + 1
        # Tên của header trong từng công ty
        for col_name in colname_in_company:
            cell = ws.cell(row_name, start_col_name)
            cell.value = col_name
            # cell.fill = copy(default_cell_header.fill)
            cell.border = all_border_thin
            cell.font = copy(default_cell_header.font)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            start_col_name += 1

        row = 4
        format_currency = '_(* #,##0_);_(* (#,##0);_(* "-"??_);_(@_)'
        format_decimal = '#,##0.00'

        for line in data:
            line_index = 0
            line_index_company = 0
            for col in range(0, len(cols) + 1):
                cell = ws.cell(row, col + 1)
                cell.font = line_font
                cell.border = all_border_thin
                cell.alignment = Alignment(horizontal='center', vertical='center')

                if line_index_company in [12, 17, 23]:
                    cell.value = line[col]
                    cell.number_format = format_currency
                elif line_index_company in [1, 2, 3, 4, 5, 6, 9, 10, 16, 18, 22, 24, 30, 32, 34]:
                    if line[col] != '-':
                        cell.value = "{:,.2f}".format(float(line[col])) + " %"
                    else:
                        cell.value = line[col]
                elif line_index_company in [7, 15, 21]:
                    cell.value = line[col]
                    cell.number_format = format_decimal
                else:
                    cell.value = line[col]
                    # cell.number_format = format_decimal

                line_index += 1
                line_index_company += 1
                if line_index_company == len_col_in_company:
                    line_index_company = 0

            row += 1
        row += 1

        fp = BytesIO()
        wb.save(fp)
        fp.seek(0)
        report = base64.encodebytes((fp.read()))
        fp.close()
        attachment = self.env['ir.attachment'].with_user(1).create({
            'name': 'bao_cao_trai_nghiem_khach_hang.xlsx',
            'datas': report,
            'res_model': 'temp.creation',
            'public': True,
        })
        url = "/web/content/?model=ir.attachment&id=%s&filename_field=name&field=datas&download=true" \
              % attachment.id

        return {
            'name': 'Báo cáo trải nghiệm khách hàng',
            'type': 'ir.actions.act_url',
            'url': url,
            'target': 'self',
        }
