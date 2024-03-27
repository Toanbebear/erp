from odoo import fields, api, models, _
from odoo.exceptions import ValidationError, UserError
from datetime import date, datetime, time
from calendar import monthrange
from openpyxl import load_workbook
from openpyxl.styles import Font, borders, Alignment
import base64
from io import BytesIO
from dateutil.relativedelta import relativedelta
from pytz import timezone, utc

thin = borders.Side(style='thin')
double = borders.Side(style='double')
all_border_thin = borders.Border(thin, thin, thin, thin)



class ReportServiceOverview(models.TransientModel):
    _name = 'report.service.overview'
    _description = 'Service Overview Report'

    start_date = fields.Date('Start date', default=date.today())
    end_date = fields.Date('End date', default=date.today())
    company_id = fields.Many2one('res.company', string='Company')

    # convert date to datetime for search domain, should be removed if using datetime directly
    start_datetime = fields.Datetime('Start datetime', compute='_compute_datetime')
    end_datetime = fields.Datetime('End datetime', compute='_compute_datetime')

    @api.depends('start_date', 'end_date')
    def _compute_datetime(self):
        self.start_datetime = False
        self.end_datetime = False
        if self.start_date and self.end_date:
            local_tz = timezone(self.env.user.tz or 'Etc/GMT+7')
            start_datetime = datetime(self.start_date.year, self.start_date.month, self.start_date.day, 0, 0, 0)
            end_datetime = datetime(self.end_date.year, self.end_date.month, self.end_date.day, 23, 59, 59)
            start_datetime = local_tz.localize(start_datetime, is_dst=None)
            end_datetime = local_tz.localize(end_datetime, is_dst=None)
            self.start_datetime = start_datetime.astimezone(utc).replace(tzinfo=None)
            self.end_datetime = end_datetime.astimezone(utc).replace(tzinfo=None)

    @api.onchange('start_date')
    def _onchange_start_date(self):
        if self.start_date:
            if self.start_date.month == fields.date.today().month:
                self.end_date = fields.date.today()
            else:
                self.end_date = date(self.start_date.year, self.start_date.month,
                                     monthrange(self.start_date.year, self.start_date.month)[1])

    @api.constrains('start_date', 'end_date')
    def check_dates(self):
        for record in self:
            start_date = fields.Date.from_string(record.start_date)
            end_date = fields.Date.from_string(record.end_date)
            if start_date > end_date:
                raise ValidationError(
                    _("Ngày kết thúc không thể ở trước ngày bắt đầu khi xuất báo cáo!!!"))

    def _get_data_report_service_overview(self):
        ret_data = []
        # start_date_search = self.start_date.strftime('%Y-%m-%d')
        # end_date_search = self.end_date.strftime('%Y-%m-%d')

        # print('aaaaa')
        # print(datetime.now())
        # a = len(self.env['crm.line'].search([('service_id.service_category', '=', 9), ('crm_id.type','=','lead')]).mapped('crm_id.id'))
        # print(datetime.now())
        # print(a)
        # print('bbbbbbb')
        # print(datetime.now())
        # b = "select count(distinct(cl.crm_id)) from crm_line cl left join sh_medical_health_center_service smhcs on cl.service_id = smhcs.id left join sh_medical_health_center_service_category smhcsc on smhcs.service_category = smhcsc.id left join crm_lead cl2 on cl.crm_id = cl2.id where cl2.type = 'lead' and smhcsc.id = 9"
        # self.env.cr.execute(b)
        # result_b = self.env.cr.dictfetchall()
        # print(datetime.now())
        # print(result_b)

        # lấy nhóm dịch vụ
        service_cat = self.env['sh.medical.health.center.service.category'].search([])
        for sc in service_cat:
            # ================================== LEAD =================================================================

            # data cột "data mới"
            query_new_data_field = "select count(distinct(cl.crm_id)) from crm_line cl left join sh_medical_health_center_service smhcs on cl.service_id = smhcs.id left join sh_medical_health_center_service_category smhcsc on smhcs.service_category = smhcsc.id left join crm_lead cl2 on cl.crm_id = cl2.id where cl2.type = 'lead' and cl2.type_data = 'new' and cl2.active = 'True' and smhcsc.id = %s and cl.company_id = %s and cl2.create_date between '%s' and '%s'" % \
                (sc.id, self.company_id.id, self.start_datetime, self.end_datetime)
            self.env.cr.execute(query_new_data_field)
            results_new_data_field = self.env.cr.dictfetchall()

            # data cột "tổng tương tác"
            query_total_interactive_field = "select count(distinct(cl.crm_id)) from crm_line cl left join sh_medical_health_center_service smhcs on cl.service_id = smhcs.id left join sh_medical_health_center_service_category smhcsc on smhcs.service_category = smhcsc.id left join crm_lead cl2 on cl.crm_id = cl2.id where cl2.type = 'lead' and cl2.active = 'True' and smhcsc.id = %s and cl.company_id = %s and cl2.stage_id not in (%s,%s) and cl2.create_date between '%s' and '%s'" % \
                (sc.id, self.company_id.id, self.env.ref('crm_base.crm_stage_lead_not_quality').id, self.env.ref('crm_base.crm_stage_test').id, self.start_datetime, self.end_datetime)
            self.env.cr.execute(query_total_interactive_field)
            results_total_interactive_field = self.env.cr.dictfetchall()

            # data cột "kết nối tư vấn"
            query_advisory_connect_field = "select count(distinct(cl.crm_id)) from crm_line cl left join sh_medical_health_center_service smhcs on cl.service_id = smhcs.id left join sh_medical_health_center_service_category smhcsc on smhcs.service_category = smhcsc.id left join crm_lead cl2 on cl.crm_id = cl2.id where cl2.type = 'lead' and cl2.active = 'True' and smhcsc.id = %s and cl.company_id = %s and cl2.stage_id in (%s,%s,%s) and cl2.create_date between '%s' and '%s'" % \
                                           (sc.id, self.company_id.id, self.env.ref('crm_base.crm_stage_refer').id, self.env.ref('crm_base.crm_stage_potential').id, self.env.ref('crm_base.crm_stage_booking').id, self.start_datetime, self.end_datetime)
            self.env.cr.execute(query_advisory_connect_field)
            results_advisory_connect_field = self.env.cr.dictfetchall()

            # data cột "% kết nối"
            if results_total_interactive_field[0]['count'] != 0:
                advisory_percent = results_advisory_connect_field[0]['count'] / results_total_interactive_field[0][
                    'count']
            else:
                advisory_percent = 0

            # dữ liệu cột "hẹn lịch"
            query_booking_lead_field = "select count(distinct(cl.crm_id)) from crm_line cl left join sh_medical_health_center_service smhcs on cl.service_id = smhcs.id left join sh_medical_health_center_service_category smhcsc on smhcs.service_category = smhcsc.id left join crm_lead cl2 on cl.crm_id = cl2.id where cl2.type = 'lead' and cl2.active = 'True' and smhcsc.id = %s and cl.company_id = %s and stage_id = %s and cl2.create_date between '%s' and '%s'" % \
                                       (sc.id, self.company_id.id, self.env.ref('crm_base.crm_stage_booking').id, self.start_datetime, self.end_datetime)
            self.env.cr.execute(query_booking_lead_field)
            results_booking_lead_field = self.env.cr.dictfetchall()

            # dữ liệu cột "% hẹn lịch"
            if results_advisory_connect_field[0]['count'] != 0:
                booking_percent = results_booking_lead_field[0]['count'] / results_advisory_connect_field[0][
                    'count']
            else:
                booking_percent = 0

            # ======================================== BOOKING =========================================================

            # dữ liệu cột "lịch hẹn"
            query_booking_field = "select count(distinct(cl.crm_id)) from crm_line cl left join sh_medical_health_center_service smhcs on cl.service_id = smhcs.id left join sh_medical_health_center_service_category smhcsc on smhcs.service_category = smhcsc.id left join crm_lead cl2 on cl.crm_id = cl2.id where cl2.type = 'opportunity' and cl2.type_data = 'new' and cl2.active = 'True' and smhcsc.id = %s and cl.company_id = %s and cl2.create_date between '%s' and '%s'" % \
                                  (sc.id, self.company_id.id, self.start_datetime, self.end_datetime)
            self.env.cr.execute(query_booking_field)
            results_booking_field = self.env.cr.dictfetchall()

            # dữ liệu cột "đến cửa"
            query_customer_come_field = "select count(distinct(cl.crm_id)) from crm_line cl left join sh_medical_health_center_service smhcs on cl.service_id = smhcs.id left join sh_medical_health_center_service_category smhcsc on smhcs.service_category = smhcsc.id left join crm_lead cl2 on cl.crm_id = cl2.id where cl2.type = 'opportunity' and cl2.customer_come = 'yes' and cl2.active = 'True' and smhcsc.id = %s and cl.company_id = %s and cl2.create_date between '%s' and '%s'" % \
                                  (sc.id, self.company_id.id, self.start_datetime, self.end_datetime)
            self.env.cr.execute(query_customer_come_field)
            results_customer_come_field = self.env.cr.dictfetchall()

            # dữ liệu cột "% đến cửa"
            if results_booking_field[0]['count'] != 0:
                percent_customer_come = results_customer_come_field[0]['count'] / results_booking_field[0]['count']
            else:
                percent_customer_come = 0

            # dữ liệu cột "thành công"
            query_booking_paid_field = "select count(distinct(cl.crm_id)) from crm_line cl left join sh_medical_health_center_service smhcs on cl.service_id = smhcs.id left join sh_medical_health_center_service_category smhcsc on smhcs.service_category = smhcsc.id left join crm_lead cl2 on cl.crm_id = cl2.id where cl2.type = 'opportunity' and cl2.active = 'True' and smhcsc.id = %s and cl.company_id = %s and cl2.stage_id = '%s' and cl2.create_date between '%s' and '%s'" % \
                                 (sc.id, self.company_id.id, self.env.ref('crm_base.crm_stage_paid').id, self.start_datetime, self.end_datetime)
            self.env.cr.execute(query_booking_paid_field)
            results_booking_paid_field = self.env.cr.dictfetchall()

            # dữ liệu cột "% thành công"
            if results_customer_come_field[0]['count'] != 0:
                percent_booking_paid = results_booking_paid_field[0]['count'] / results_customer_come_field[0]['count']
            else:
                percent_booking_paid = 0

            # lưu kết quả truy vấn theo dịch vụ vào ret_data
            ret_data.append({
                    'date': sc.name,
                    'new_data': results_new_data_field[0]['count'],
                    'interactive': results_total_interactive_field[0]['count'],
                    'advisory_connect': results_advisory_connect_field[0]['count'],
                    'advisory_percent': advisory_percent,
                    'booking_lead': results_booking_lead_field[0]['count'],
                    'booking_percent': booking_percent,
                    'booking': results_booking_field[0]['count'],
                    'customer_come': results_customer_come_field[0]['count'],
                    'percent_customer_come': percent_customer_come,
                    'booking_paid': results_booking_paid_field[0]['count'],
                    'percent_booking_paid': percent_booking_paid,
                })
        return ret_data

    def create_report_service_overview(self):
        data = self._get_data_report_service_overview()

        report_brand_overview_attachment = self.env['ir.attachment'].browse(
            self.env.ref('crm_report.report_service_overview_attachment').id)
        decode = base64.b64decode(report_brand_overview_attachment.datas)
        wb = load_workbook(BytesIO(decode))
        ws = wb.active
        line_font = Font(name='Times New Roman', size=14)
        ws['E3'].value = self.start_date.strftime('%d/%m/%Y')
        ws['G3'].value = self.end_date.strftime('%d/%m/%Y')
        ws['I3'].value = self.company_id.name
        key_col_list = list(range(1, 13))
        key_list = [
            'date',
            'new_data',
            'interactive',
            'advisory_connect',
            'advisory_percent',
            'booking_lead',
            'booking_percent',
            'booking',
            'customer_come',
            'percent_customer_come',
            'booking_paid',
            'percent_booking_paid',
        ]
        row = 7
        for line_data in data:
            for col, k in zip(key_col_list, key_list):
                cell = ws.cell(row, col)
                cell.value = line_data[k]
                cell.font = line_font
                cell.border = all_border_thin
                cell.alignment = Alignment(horizontal='left', vertical='center')
            row += 1

        fp = BytesIO()
        wb.save(fp)
        fp.seek(0)
        report = base64.encodebytes((fp.read()))
        fp.close()
        attachment = self.env['ir.attachment'].sudo().create({
            'name': 'bao_cao_tong_quan_theo_nhom_dich_vu.xlsx',
            'datas': report,
            'res_model': 'temp.creation',
            'public': True,
        })
        url = "/web/content/?model=ir.attachment&id=%s&filename_field=name&field=datas&download=true" \
              % attachment.id
        return {'name': 'BÁO CÁO TỔNG QUAN THEO NHÓM DỊCH VỤ',
                'type': 'ir.actions.act_url',
                'url': url,
                'target': 'self',
                }
