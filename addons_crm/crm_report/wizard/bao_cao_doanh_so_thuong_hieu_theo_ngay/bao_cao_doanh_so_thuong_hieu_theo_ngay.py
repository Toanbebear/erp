from odoo import fields, api, models, _
from odoo.exceptions import ValidationError, UserError
from datetime import date, datetime, time
from calendar import monthrange
from openpyxl import load_workbook
from openpyxl.styles import Font, borders, Alignment
import base64
from io import BytesIO
from dateutil.relativedelta import relativedelta
from pytz import timezone, utc

thin = borders.Side(style='thin')
double = borders.Side(style='double')
all_border_thin = borders.Border(thin, thin, thin, thin)


class BrandSalesReport(models.TransientModel):
    _name = 'report.brand.sale'
    _description = 'Brand Sale Report'

    start_date = fields.Date('Start date', default=date.today())
    end_date = fields.Date('End date', default=date.today())

    # convert date to datetime for search domain, should be removed if using datetime directly
    start_datetime = fields.Datetime('Start datetime', compute='_compute_datetime')
    end_datetime = fields.Datetime('End datetime', compute='_compute_datetime')

    @api.depends('start_date', 'end_date')
    def _compute_datetime(self):
        self.start_datetime = False
        self.end_datetime = False
        if self.start_date and self.end_date:
            local_tz = timezone(self.env.user.tz or 'Etc/GMT+7')
            start_datetime = datetime(self.start_date.year, self.start_date.month, self.start_date.day, 0, 0, 0)
            end_datetime = datetime(self.end_date.year, self.end_date.month, self.end_date.day, 23, 59, 59)
            start_datetime = local_tz.localize(start_datetime, is_dst=None)
            end_datetime = local_tz.localize(end_datetime, is_dst=None)
            self.start_datetime = start_datetime.astimezone(utc).replace(tzinfo=None)
            self.end_datetime = end_datetime.astimezone(utc).replace(tzinfo=None)

    @api.onchange('start_date')
    def _onchange_start_date(self):
        if self.start_date:
            if self.start_date.month == fields.date.today().month:
                self.end_date = fields.date.today()
            else:
                self.end_date = date(self.start_date.year, self.start_date.month,
                                     monthrange(self.start_date.year, self.start_date.month)[1])

    @api.constrains('start_date', 'end_date')
    def check_dates(self):
        for record in self:
            start_date = fields.Date.from_string(record.start_date)
            end_date = fields.Date.from_string(record.end_date)
            if start_date > end_date:
                raise ValidationError(
                    _("Ngày kết thúc không thể ở trước ngày bắt đầu khi xuất báo cáo!!!"))
            if start_date.month != end_date.month:
                raise ValidationError(
                    _("Ngày bắt dầu và ngày kết thúc khi xuất báo cáo phải nằm trong cùng 1 tháng!!!"))

    def _get_data_report_brand_sale(self):
        local_tz = timezone(self.env.user.tz or 'Etc/GMT+7')

        # lấy số ngày theo ngày bắt đầu và ngày kết thúc đã chọn ( trong cùng 1 tháng)
        days = self.end_date - self.start_date

        # lấy tháng và năm
        month = self.end_date.month
        year = self.end_date.year

        # khởi tạo biến
        ret_data = []  # in báo cáo
        data_total = {'company': 'TOTAL', 'amount_total': 0}  # ghi nhận giá trị total - dòng cuối của báo cáo
        amount_total_company = 0  # biến total tổng doanh số của tất cả các thương hiệu- dùng để format định dạng khi in

        # tạo các key thuộc data_total
        if self.start_date.day == 1:
            days_loop = days.days + 2
            for i in range(1, days_loop):
                # khởi tạo  ngày tìm
                date_search = date(year, month, i)
                data_total['%s' % date_search] = 0
        else:
            days_loop = days.days + 1
            for i in range(0, days_loop):
                date_int = self.start_date.day + i

                # khởi tạo ngày tìm
                date_search = date(year, month, date_int)
                data_total['%s' % date_search] = 0

        # lấy danh sách cty
        companies = self.env['res.company'].search([])
        for company in companies:
            item_data = {'company': company.name}
            # truy vấn tổng doanh số theo start_date và end_date
            query = "select sum(cl.amount_total) from crm_lead cl where cl.type = 'opportunity' and active = 'True' and company_id = '%s' and stage_id = '%s' and create_date between '%s' and '%s'" % (
                company.id, self.env.ref('crm_base.crm_stage_confirm').id, self.start_datetime, self.end_datetime)
            self.env.cr.execute(query)
            results_query = self.env.cr.dictfetchall()
            item_data['amount_total'] = format(int(results_query[0]['sum']), ',d').replace(',', '.') if \
            results_query[0]['sum'] is not None else '-'
            ret_data.append(item_data)

            # cập nhật kết quả vào data_total
            if results_query[0]['sum']:
                amount_total_company += results_query[0]['sum']
                data_total['amount_total'] = format(int(amount_total_company), ',d').replace(',', '.')

            # nếu ngày bắt đầu là đầu tháng thì khởi tạo date_search từ ngày mùng 01 của tháng đó
            if self.start_date.day == 1:
                days_loop = days.days + 2
                for i in range(1, days_loop):
                    # amount_total_day = 0
                    # khởi tạo ngày tìm
                    date_search = date(year, month, i)

                    # ngày giờ để search trên server
                    start_date_search = datetime(year, month, i, 0, 0, 0)
                    end_date_search = datetime(year, month, i, 23, 59, 59)
                    start_date_search = local_tz.localize(start_date_search, is_dst=None)
                    end_date_search = local_tz.localize(end_date_search, is_dst=None)
                    start_datetime_search = start_date_search.astimezone(utc).replace(tzinfo=None)
                    end_datetime_search = end_date_search.astimezone(utc).replace(tzinfo=None)

                    # truy vấn doanh số theo ngày
                    query = "select sum(cl.amount_total) from crm_lead cl where cl.type = 'opportunity' and active = 'True' and company_id = '%s' and stage_id = '%s' and create_date between '%s' and '%s'" % (
                        company.id, self.env.ref('crm_base.crm_stage_confirm').id, start_datetime_search, end_datetime_search)
                    self.env.cr.execute(query)
                    results_query = self.env.cr.dictfetchall()
                    item_data['%s' % date_search] = format(int(results_query[0]['sum']), ',d').replace(',', '.') if \
                    results_query[0]['sum'] is not None else '-'

                    # cập nhật kết quả vào data_total
                    if results_query[0]['sum']:
                        # amount_total_day += results_query[0]['sum']
                        data_total['%s' % date_search] += results_query[0]['sum']

            # nếu ngày bắt đầu không phải đầu tháng thì khởi tạo date_search = start_date
            else:
                days_loop = days.days + 1
                for i in range(0, days_loop):
                    # amount_total_day = 0
                    date_int = self.start_date.day + i

                    # khởi tạo ngày tìm
                    date_search = date(year, month, date_int)

                    # ngày giờ để search trên server
                    start_date_search = datetime(year, month, date_int, 0, 0, 0)
                    end_date_search = datetime(year, month, date_int, 23, 59, 59)
                    start_date_search = local_tz.localize(start_date_search, is_dst=None)
                    end_date_search = local_tz.localize(end_date_search, is_dst=None)
                    start_datetime_search = start_date_search.astimezone(utc).replace(tzinfo=None)
                    end_datetime_search = end_date_search.astimezone(utc).replace(tzinfo=None)

                    # truy vấn doanh số theo ngày
                    query = "select sum(cl.amount_total) from crm_lead cl where cl.type = 'opportunity' and active = 'True' and company_id = '%s' and stage_id = '%s' and create_date between '%s' and '%s'" % (
                        company.id, self.env.ref('crm_base.crm_stage_confirm').id, start_datetime_search, end_datetime_search)
                    self.env.cr.execute(query)
                    results_query = self.env.cr.dictfetchall()
                    item_data['%s' % date_search] = format(int(results_query[0]['sum']), ',d').replace(',', '.') if \
                    results_query[0]['sum'] is not None else '-'

                    # cập nhật kết quả vào data_total
                    if results_query[0]['sum']:
                        # amount_total_day += results_query[0]['sum']
                        data_total['%s' % date_search] += results_query[0]['sum']
        # thêm data_total vào cuối ret_data in báo cáo
        ret_data.append(data_total)

        return ret_data

    def create_report_brand_sale(self):
        # lấy số ngày theo ngày bắt đầu và ngày kết thúc đã chọn ( trong cùng 1 tháng)
        days = self.end_date - self.start_date

        # lấy tháng và năm
        month = self.end_date.month
        year = self.end_date.year

        # lấy data
        data = self._get_data_report_brand_sale()

        report_brand_sales_attachment = self.env['ir.attachment'].browse(
            self.env.ref('crm_report.report_brand_sale_attachment').id)
        decode = base64.b64decode(report_brand_sales_attachment.datas)
        wb = load_workbook(BytesIO(decode))
        ws = wb.active
        line_font = Font(name='Times New Roman', size=14)

        if self.start_date.day == 1:
            days_loop = days.days + 2

            # tạo key để in tiêu đề cột
            key_col_list_structure = list(range(4, days_loop + 3))
            key_list_structure = []

            # tạo key để in dữ liệu báo cáo
            key_col_list = list(range(2, days_loop + 3))
            key_list = ['company', 'amount_total']
            for i in range(1, days_loop):
                date_search = date(year, month, i)
                key_list_structure.append('%s' % str(date_search))
                key_list.append('%s' % str(date_search))
        else:
            days_loop = days.days + 1

            # tạo key để in tiêu đề cột
            key_col_list_structure = list(range(4, days_loop + 4))
            key_list_structure = []

            # tạo key để in dữ liệu báo cáo
            key_col_list = list(range(2, days_loop + 4))
            key_list = ['company', 'amount_total']
            for i in range(0, days_loop):
                date_int = self.start_date.day + i

                # khởi tạo ngày tìm
                date_search = date(year, month, date_int)

                key_list_structure.append('%s' % str(date_search))
                key_list.append('%s' % str(date_search))

        # in tiêu đề cột
        row = 4
        for col, k in zip(key_col_list_structure, key_list_structure):
            cell = ws.cell(row, col)
            cell.value = k
            cell.font = line_font
            cell.border = all_border_thin
            cell.alignment = Alignment(horizontal='left', vertical='center')

        # in báo cáo
        ws['G2'].value = self.start_date.strftime('%d/%m/%Y')
        ws['I2'].value = self.end_date.strftime('%d/%m/%Y')
        row = 5
        for line_data in data:
            for col, k in zip(key_col_list, key_list):
                cell = ws.cell(row, col)
                cell.value = line_data[k]
                cell.font = line_font
                cell.border = all_border_thin
                cell.alignment = Alignment(horizontal='left', vertical='center')
            row += 1

        fp = BytesIO()
        wb.save(fp)
        fp.seek(0)
        report = base64.encodebytes((fp.read()))
        fp.close()
        attachment = self.env['ir.attachment'].sudo().create({
            'name': 'bao_cao_doanh_thu_thuong_hieu_theo_ngay.xlsx',
            'datas': report,
            'res_model': 'temp.creation',
            'public': True,
        })
        url = "/web/content/?model=ir.attachment&id=%s&filename_field=name&field=datas&download=true" \
              % attachment.id

        return {'name': 'BÁO CÁO DOANH THU THƯƠNG HIỆU THEO NGÀY',
                'type': 'ir.actions.act_url',
                'url': url,
                'target': 'self',
                }
