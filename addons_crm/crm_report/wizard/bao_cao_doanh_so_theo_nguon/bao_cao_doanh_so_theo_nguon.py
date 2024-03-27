from odoo import fields, api, models, _
from odoo.exceptions import ValidationError, UserError
from datetime import date, datetime, time
from calendar import monthrange
from openpyxl import load_workbook
from openpyxl.styles import Font, borders, Alignment
import base64
from io import BytesIO
from dateutil.relativedelta import relativedelta
from pytz import timezone, utc

thin = borders.Side(style='thin')
double = borders.Side(style='double')
all_border_thin = borders.Border(thin, thin, thin, thin)


class SaleReportBySource(models.TransientModel):
    _name = 'report.sale.source'
    _description = 'Sales report by source'

    start_date = fields.Date('Start date', default=date.today())
    end_date = fields.Date('End date', default=date.today())
    company_id = fields.Many2one('res.company', string='Company')

    # convert date to datetime for search domain, should be removed if using datetime directly
    start_datetime = fields.Datetime('Start datetime', compute='_compute_datetime')
    end_datetime = fields.Datetime('End datetime', compute='_compute_datetime')

    @api.depends('start_date', 'end_date')
    def _compute_datetime(self):
        self.start_datetime = False
        self.end_datetime = False
        if self.start_date and self.end_date:
            local_tz = timezone(self.env.user.tz or 'Etc/GMT+7')
            start_datetime = datetime(self.start_date.year, self.start_date.month, self.start_date.day, 0, 0, 0)
            end_datetime = datetime(self.end_date.year, self.end_date.month, self.end_date.day, 23, 59, 59)
            start_datetime = local_tz.localize(start_datetime, is_dst=None)
            end_datetime = local_tz.localize(end_datetime, is_dst=None)
            self.start_datetime = start_datetime.astimezone(utc).replace(tzinfo=None)
            self.end_datetime = end_datetime.astimezone(utc).replace(tzinfo=None)

    @api.onchange('start_date')
    def _onchange_start_date(self):
        if self.start_date:
            if self.start_date.month == fields.date.today().month:
                self.end_date = fields.date.today()
            else:
                self.end_date = date(self.start_date.year, self.start_date.month,
                                     monthrange(self.start_date.year, self.start_date.month)[1])

    @api.constrains('start_date', 'end_date')
    def check_dates(self):
        for record in self:
            start_date = fields.Date.from_string(record.start_date)
            end_date = fields.Date.from_string(record.end_date)
            if start_date > end_date:
                raise ValidationError(
                    _("Ngày kết thúc không thể ở trước ngày bắt đầu khi xuất báo cáo!!!"))

    def _get_data_sales_source_report(self):
        ret_data = []
        total_amount_sales_source = 0
        sources = self.env['utm.source'].search([])


        # for source in sources:
        #     amount_sales_source = 0
        #     bookings = self.env['crm.lead'].search([('create_date', '>=', self.start_date), ('create_date', '<=', self.end_date), ('type', '=', 'opportunity'), ('company_id', '=', self.company_id.id),('source_id', '=', source.id), ('stage_id', '=', self.env.ref('crm_base.crm_stage_confirm').id)])
        #     if bookings:
        #         for bk in bookings:
        #             amount_sales_source += bk.amount_total
        #     ret_data.append({
        #         'source': source.name,
        #         'amout_source': format(int(amount_sales_source), ',d').replace(',', '.'),
        #     })
        #     total_amount_sales_source += amount_sales_source
        for source in sources:
            amount_sales_source = 0
            query = "select sum(cl.amount_total) from crm_lead cl where active = 'True' and cl.type = 'opportunity' and cl.create_date >= '%s' and cl.create_date <= '%s' and cl.company_id = '%s' and cl.source_id = '%s' and cl.stage_id = '%s'" % (
            self.start_datetime, self.end_datetime, self.company_id.id, source.id,
            self.env.ref('crm_base.crm_stage_confirm').id)
            self.env.cr.execute(query)
            result_query = self.env.cr.dictfetchall()
            if result_query[0]['sum']:
                amount_sales_source += int(result_query[0]['sum'])
            ret_data.append({
                'source': source.name,
                'amount_source': amount_sales_source,
            })
            total_amount_sales_source += amount_sales_source
        # thêm trường tỷ trọng vào từng item (do trường tỷ trọng chỉ có thể tính được sau khi tính tổng tất cả doanh thu của các nguồn)
        for item in ret_data:
            if total_amount_sales_source != 0:
                item['proportion'] = int(item['amount_source']) / int(total_amount_sales_source)
                item['amount_source'] = format(int(item['amount_source']), ',d').replace(',', '.')
            else:
                item['proportion'] = '-'
        # thêm item total vào cuối ret_data
        ret_data.append({
            'source': 'TOTAL',
            'amount_source': format(int(total_amount_sales_source), ',d').replace(',', '.'),
            'proportion': '-',
        })
        return ret_data

    def create_sales_source_report(self):
        # lấy data
        data = self._get_data_sales_source_report()

        # in dữ liệu
        report_brand_overview_attachment = self.env['ir.attachment'].browse(
            self.env.ref('crm_report.report_sales_source_attachment').id)
        decode = base64.b64decode(report_brand_overview_attachment.datas)
        wb = load_workbook(BytesIO(decode))
        ws = wb.active
        line_font = Font(name='Times New Roman', size=14)
        ws['B3'].value = 'Từ %s đến %s' % (self.start_date.strftime('%d/%m/%Y'), self.end_date.strftime('%d/%m/%Y'))
        ws['C2'].value = self.company_id.name
        key_col_list = (2,3,4)
        key_list = [
            'source',
            'amount_source',
            'proportion',
        ]
        row = 6
        for line_data in data:
            for col, k in zip(key_col_list, key_list):
                cell = ws.cell(row, col)
                cell.value = line_data[k]
                cell.font = line_font
                cell.border = all_border_thin
                cell.alignment = Alignment(horizontal='left', vertical='center')
            row += 1
        fp = BytesIO()
        wb.save(fp)
        fp.seek(0)
        report = base64.encodebytes((fp.read()))
        fp.close()
        attachment = self.env['ir.attachment'].sudo().create({
            'name': 'bao_cao_doanh_so_theo_nguon.xlsx',
            'datas': report,
            'res_model': 'temp.creation',
            'public': True,
        })
        url = "/web/content/?model=ir.attachment&id=%s&filename_field=name&field=datas&download=true" \
              % attachment.id
        return {'name': 'BÁO CÁO DOANH SỐ THEO NGUỒN',
                'type': 'ir.actions.act_url',
                'url': url,
                'target': 'self',
                }


