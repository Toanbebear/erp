from odoo import fields, api, models, _
from odoo.exceptions import ValidationError, UserError
from datetime import date, datetime, time
from calendar import monthrange
from openpyxl import load_workbook
from openpyxl.styles import Font, borders, Alignment
import base64
from io import BytesIO
from dateutil.relativedelta import relativedelta
from pytz import timezone, utc

thin = borders.Side(style='thin')
double = borders.Side(style='double')
all_border_thin = borders.Border(thin, thin, thin, thin)


class BrandOverviewReport(models.TransientModel):
    _name = 'report.brand.overview'
    _description = 'Brand Overview Report'

    start_date = fields.Date('Start date', default=date.today())
    end_date = fields.Date('End date', default=date.today())
    # convert date to datetime for search domain, should be removed if using datetime directly
    start_datetime = fields.Datetime('Start datetime', compute='_compute_datetime')
    end_datetime = fields.Datetime('End datetime', compute='_compute_datetime')

    company_id = fields.Many2one('res.company', string='Company')
    year = fields.Selection([(str(num), str(num)) for num in range(2019, datetime.now().year + 1)], string='Year')
    type_display = fields.Selection([('date', 'Hiện theo ngày'), ('month', 'Hiện theo tháng')], string='Type Display',
                                    default='date')

    @api.depends('start_date', 'end_date')
    def _compute_datetime(self):
        self.start_datetime = False
        self.end_datetime = False
        if self.start_date and self.end_date:
            local_tz = timezone(self.env.user.tz or 'Etc/GMT+7')
            start_datetime = datetime(self.start_date.year, self.start_date.month, self.start_date.day, 0, 0, 0)
            end_datetime = datetime(self.end_date.year, self.end_date.month, self.end_date.day, 23, 59, 59)
            start_datetime = local_tz.localize(start_datetime, is_dst=None)
            end_datetime = local_tz.localize(end_datetime, is_dst=None)
            self.start_datetime = start_datetime.astimezone(utc).replace(tzinfo=None)
            self.end_datetime = end_datetime.astimezone(utc).replace(tzinfo=None)

    @api.onchange('start_date')
    def _onchange_start_date(self):
        if self.start_date:
            if self.start_date.month == fields.date.today().month:
                self.end_date = fields.date.today()
            else:
                self.end_date = date(self.start_date.year, self.start_date.month,
                                     monthrange(self.start_date.year, self.start_date.month)[1])

    @api.constrains('start_date', 'end_date')
    def check_dates(self):
        for record in self:
            start_date = fields.Date.from_string(record.start_date)
            end_date = fields.Date.from_string(record.end_date)
            if start_date > end_date:
                raise ValidationError(
                    _("Ngày kết thúc không thể ở trước ngày bắt đầu khi xuất báo cáo!!!"))
            if start_date.month != end_date.month or start_date.year != end_date.year:
                raise ValidationError(
                    _("Ngày bắt dầu và ngày kết thúc khi xuất báo cáo phải nằm trong cùng 1 tháng!!!"))

    def _get_data_report_brand_overview_per_day(self):
        local_tz = timezone(self.env.user.tz or 'Etc/GMT+7')

        # lấy số ngày theo ngày bắt đầu và ngày kết thúc đã chọn ( trong cùng 1 tháng)
        # days = self.end_date - self.start_date
        days = self.end_datetime - self.start_datetime

        # lấy tháng và năm
        month = self.end_date.month
        year = self.end_date.year

        # khởi tạo biến total
        total_new_data = 0
        total_interactive = 0
        total_advisory_connect = 0
        total_advisory_percent = 0
        total_booking_lead = 0
        total_booking_percent = 0
        total_booking = 0
        total_customer_come = 0
        total_percent_customer_come = 0
        total_booking_paid = 0
        total_percent_booking_paid = 0
        ret_data = []
        # nếu ngày bắt đầu là đầu tháng thì khởi tạo date_search từ ngày mùng 01 của tháng đó
        if self.start_date.day == 1:
            days_loop = days.days + 2
            for i in range(1, days_loop):
                # ngày để in ra báo cáo
                date_search = date(year, month, i)

                # ngày giờ để search trên server
                start_date_search = datetime(year, month, i, 0, 0, 0)
                end_date_search = datetime(year, month, i, 23, 59, 59)
                start_date_search = local_tz.localize(start_date_search, is_dst=None)
                end_date_search = local_tz.localize(end_date_search, is_dst=None)
                start_datetime_search = start_date_search.astimezone(utc).replace(tzinfo=None)
                end_datetime_search = end_date_search.astimezone(utc).replace(tzinfo=None)

                # print(date_search)
                # print(start_datetime_search)
                # print(end_datetime_search)

                # ================================== lead =============================================================
                # dữ liệu cột "data mới"
                query_new_data_field = "select count(*) from crm_lead where stage_id not in (%s,%s) and type = 'lead' and company_id = '%s' and " \
                                       "type_data = 'new' and active = 'True' and " \
                                       "create_date between " \
                                       "'%s' and '%s'" % (
                                       self.env.ref('crm_base.crm_stage_lead_not_quality').id,
                                       self.env.ref('crm_base.crm_stage_test').id,
                                       self.company_id.id, start_datetime_search, end_datetime_search)
                self.env.cr.execute(query_new_data_field)
                results_new_data = self.env.cr.dictfetchall()

                # dữ liệu cột "tổng tương tác"
                query_total_interactive_field = "select count(*) from crm_lead where stage_id not in (%s,%s) and " \
                                                "active = 'True' and type = 'lead' and company_id = " \
                                                "'%s'and create_date between '%s' " \
                                                "and '%s'" % (
                                                    self.env.ref('crm_base.crm_stage_lead_not_quality').id,
                                                    self.env.ref('crm_base.crm_stage_test').id,
                                                    self.company_id.id,
                                                    start_datetime_search, end_datetime_search)
                self.env.cr.execute(query_total_interactive_field)
                results_interactive_field = self.env.cr.dictfetchall()

                # dữ liệu cột "kết nối tư vấn"
                query_advisory_connect_field = "select count(*) from crm_lead where stage_id in (%s,%s,%s) and active " \
                                               "= 'True' and type = 'lead' and company_id = " \
                                               "'%s' and create_date between '%s' " \
                                               "and '%s'" % (
                                                   self.env.ref('crm_base.crm_stage_refer').id,
                                                   self.env.ref('crm_base.crm_stage_potential').id,
                                                   self.env.ref('crm_base.crm_stage_booking').id,
                                                   self.company_id.id,
                                                   start_datetime_search, end_datetime_search)
                self.env.cr.execute(query_advisory_connect_field)
                results_advisory_connect_field = self.env.cr.dictfetchall()

                # dữ liệu cột "% kết nối"
                if results_interactive_field[0]['count'] != 0:
                    advisory_percent = results_advisory_connect_field[0]['count'] / results_interactive_field[0][
                        'count']
                else:
                    advisory_percent = 0

                # dữ liệu cột "hẹn lịch"
                query_booking_lead_field = "select count(*) from crm_lead where stage_id = %s and active = 'True' and " \
                                           "type = 'lead' and company_id = " \
                                           "'%s' and create_date between '%s' " \
                                           "and '%s'" % (
                                               self.env.ref('crm_base.crm_stage_booking').id,
                                               self.company_id.id,
                                               start_datetime_search, end_datetime_search)
                self.env.cr.execute(query_booking_lead_field)
                results_booking_lead_field = self.env.cr.dictfetchall()

                # dữ liệu cột "% hẹn lịch"
                if results_advisory_connect_field[0]['count'] != 0:
                    booking_percent = results_booking_lead_field[0]['count'] / results_advisory_connect_field[0][
                        'count']
                else:
                    booking_percent = 0

                # ================================ booking =============================================================
                # dữ liệu cột "lịch hẹn"
                query_booking_field = "select count(*) from crm_lead cl join product_pricelist pp on cl.price_list_id " \
                                      "= pp.id where cl.type = 'opportunity' and pp.type = " \
                                      "'service' and cl.active = 'True' and cl.company_id = '%s' and cl.create_date " \
                                      "between '%s' and '%s' " % (
                                          self.company_id.id, start_datetime_search, end_datetime_search)
                self.env.cr.execute(query_booking_field)
                results_booking_field = self.env.cr.dictfetchall()

                # dữ liệu cột "đến cửa"
                query_customer_come = "select count(*) from crm_lead cl join product_pricelist pp on cl.price_list_id " \
                                      "= pp.id where cl.type = 'opportunity' and pp.type = 'service' and cl.active = " \
                                      "'True' and customer_come = 'yes' and cl.company_id = '%s' and cl.create_date between " \
                                      "'%s' and '%s'" % (self.company_id.id, start_datetime_search,
                                                                           end_datetime_search)
                self.env.cr.execute(query_customer_come)
                results_customer_come_field = self.env.cr.dictfetchall()

                # dữ liệu cột "% đến cửa"
                if results_booking_field[0]['count'] != 0:
                    percent_customer_come = results_customer_come_field[0]['count'] / results_booking_field[0]['count']
                else:
                    percent_customer_come = 0

                # dữ liệu cột "thành công"
                query_booking_paid = "select count(*) from crm_lead where active = 'True' and type = 'opportunity' " \
                                     "and stage_id = '%s' and company_id = '%s' and create_date between '%s'" \
                                     " and '%s'" % (self.env.ref('crm_base.crm_stage_paid').id,
                                                             self.company_id.id,
                                                             start_datetime_search, end_datetime_search)
                self.env.cr.execute(query_booking_paid)
                results_booking_paid_field = self.env.cr.dictfetchall()

                # dữ liệu cột "% thành công"
                if results_customer_come_field[0]['count'] != 0:
                    percent_booking_paid = results_booking_paid_field[0]['count'] / results_customer_come_field[0][
                        'count']
                else:
                    percent_booking_paid = 0

                # lưu kết quả truy vấn theo ngày vào ret_data
                ret_data.append({
                    'date': date_search,
                    'new_data': results_new_data[0]['count'],
                    'interactive': results_interactive_field[0]['count'],
                    'advisory_connect': results_advisory_connect_field[0]['count'],
                    'advisory_percent': advisory_percent,
                    'booking_lead': results_booking_lead_field[0]['count'],
                    'booking_percent': booking_percent,
                    'booking': results_booking_field[0]['count'],
                    'customer_come': results_customer_come_field[0]['count'],
                    'percent_customer_come': percent_customer_come,
                    'booking_paid': results_booking_paid_field[0]['count'],
                    'percent_booking_paid': percent_booking_paid,
                })
                # tính lại các biến total
                total_new_data += results_new_data[0]['count']
                total_interactive += results_interactive_field[0]['count']
                total_advisory_connect += results_advisory_connect_field[0]['count']
                total_advisory_percent += advisory_percent
                total_booking_lead += results_booking_lead_field[0]['count']
                total_booking_percent += booking_percent
                total_booking += results_booking_field[0]['count']
                total_customer_come += results_customer_come_field[0]['count']
                total_percent_customer_come += percent_customer_come
                total_booking_paid += results_booking_paid_field[0]['count']
                total_percent_booking_paid += percent_booking_paid
        # nếu ngày bắt đầu không phải đầu tháng thì khởi tạo date_search = start_date
        else:
            days_loop = days.days + 1
            for i in range(0, days_loop):
                date_int = self.start_date.day + i

                # ngày để in ra báo cáo
                date_search = date(year, month, date_int)

                # ngày giờ để search trên server
                start_date_search = datetime(year, month, date_search.day, 0, 0, 0)
                end_date_search = datetime(year, month, date_search.day, 23, 59, 59)
                start_date_search = local_tz.localize(start_date_search, is_dst=None)
                end_date_search = local_tz.localize(end_date_search, is_dst=None)
                start_datetime_search = start_date_search.astimezone(utc).replace(tzinfo=None)
                end_datetime_search = end_date_search.astimezone(utc).replace(tzinfo=None)

                # print(date_search)
                # print(start_datetime_search)
                # print(end_datetime_search)
                # ================================== lead =============================================================
                # dữ liệu cột "data mới"
                query_new_data_field = "select count(*) from crm_lead where type = 'lead' and company_id = '%s' and " \
                                       "type_data = 'new' and active = 'True' and " \
                                       "create_date between " \
                                       "'%s' and '%s'" % (
                                           self.company_id.id, start_datetime_search, end_datetime_search)
                self.env.cr.execute(query_new_data_field)
                results_new_data = self.env.cr.dictfetchall()

                # dữ liệu cột "tổng tương tác"
                query_total_interactive_field = "select count(*) from crm_lead where stage_id not in (%s,%s) and " \
                                                "active = 'True' and type = 'lead' and company_id = " \
                                                "'%s'and create_date between '%s' " \
                                                "and '%s'" % (
                                                    self.env.ref('crm_base.crm_stage_lead_not_quality').id,
                                                    self.env.ref('crm_base.crm_stage_test').id,
                                                    self.company_id.id,
                                                    start_datetime_search, end_datetime_search)
                self.env.cr.execute(query_total_interactive_field)
                results_interactive_field = self.env.cr.dictfetchall()

                # dữ liệu cột "kết nối tư vấn"
                query_advisory_connect_field = "select count(*) from crm_lead where stage_id in (%s,%s,%s) and active " \
                                               "= 'True' and type = 'lead' and company_id = " \
                                               "'%s' and create_date between '%s' " \
                                               "and '%s'" % (
                                                   self.env.ref('crm_base.crm_stage_refer').id,
                                                   self.env.ref('crm_base.crm_stage_potential').id,
                                                   self.env.ref('crm_base.crm_stage_booking').id,
                                                   self.company_id.id,
                                                   start_datetime_search, end_datetime_search)
                self.env.cr.execute(query_advisory_connect_field)
                results_advisory_connect_field = self.env.cr.dictfetchall()

                # dữ liệu cột "% kết nối"
                if results_interactive_field[0]['count'] != 0:
                    advisory_percent = results_advisory_connect_field[0]['count'] / results_interactive_field[0][
                        'count']
                else:
                    advisory_percent = 0

                # dữ liệu cột "hẹn lịch"
                query_booking_lead_field = "select count(*) from crm_lead where stage_id = %s and active = 'True' and type = 'lead' and company_id = " \
                                           "'%s' and create_date between '%s' " \
                                           "and '%s'" % (
                                               self.env.ref('crm_base.crm_stage_booking').id,
                                               self.company_id.id,
                                               start_datetime_search, end_datetime_search)
                self.env.cr.execute(query_booking_lead_field)
                results_booking_lead_field = self.env.cr.dictfetchall()

                # dữ liệu cột "% hẹn lịch"
                if results_advisory_connect_field[0]['count'] != 0:
                    booking_percent = results_booking_lead_field[0]['count'] / results_advisory_connect_field[0][
                        'count']
                else:
                    booking_percent = 0

                # ================================ booking =============================================================
                # dữ liệu cột "lịch hẹn"
                query_booking_field = "select count(*) from crm_lead cl join product_pricelist pp on cl.price_list_id " \
                                      "= pp.id where cl.type_data = 'new' and cl.type = 'opportunity' and pp.type = " \
                                      "'service' and cl.active = 'True' and cl.company_id = '%s' and cl.create_date " \
                                      "between '%s' and '%s' " % (
                                          self.company_id.id, start_datetime_search, end_datetime_search)
                self.env.cr.execute(query_booking_field)
                results_booking_field = self.env.cr.dictfetchall()

                # dữ liệu cột "đến cửa"
                query_customer_come = "select count(*) from crm_lead where active = 'True' and type = 'opportunity' " \
                                      "and customer_come = 'yes' and company_id = '%s' and create_date between '%s' " \
                                      "and '%s'" % (self.company_id.id, start_datetime_search, end_datetime_search)
                self.env.cr.execute(query_customer_come)
                results_customer_come_field = self.env.cr.dictfetchall()

                # dữ liệu cột "% đến cửa"
                if results_booking_field[0]['count'] != 0:
                    percent_customer_come = results_customer_come_field[0]['count'] / results_booking_field[0]['count']
                else:
                    percent_customer_come = 0

                # dữ liệu cột "thành công"
                query_booking_paid = "select count(*) from crm_lead where active = 'True' and type = 'opportunity' " \
                                     "and stage_id = '%s' and company_id = '%s' and create_date between '%s'" \
                                     " and '%s'" % (self.env.ref('crm_base.crm_stage_paid').id,
                                                             self.company_id.id,
                                                             start_datetime_search, end_datetime_search)
                self.env.cr.execute(query_booking_paid)
                results_booking_paid_field = self.env.cr.dictfetchall()

                # dữ liệu cột "% thành công"
                if results_customer_come_field[0]['count'] != 0:
                    percent_booking_paid = results_booking_paid_field[0]['count'] / results_customer_come_field[0][
                        'count']
                else:
                    percent_booking_paid = 0

                # lưu kết quả truy vấn theo ngày vào ret_data
                ret_data.append({
                    'date': date_search,
                    'new_data': results_new_data[0]['count'],
                    'interactive': results_interactive_field[0]['count'],
                    'advisory_connect': results_advisory_connect_field[0]['count'],
                    'advisory_percent': advisory_percent,
                    'booking_lead': results_booking_lead_field[0]['count'],
                    'booking_percent': booking_percent,
                    'booking': results_booking_field[0]['count'],
                    'customer_come': results_customer_come_field[0]['count'],
                    'percent_customer_come': percent_customer_come,
                    'booking_paid': results_booking_paid_field[0]['count'],
                    'percent_booking_paid': percent_booking_paid,
                })

                # tính lại các biến total
                total_new_data += results_new_data[0]['count']
                total_interactive += results_interactive_field[0]['count']
                total_advisory_connect += results_advisory_connect_field[0]['count']
                total_booking_lead += results_booking_lead_field[0]['count']
                total_booking += results_booking_field[0]['count']
                total_customer_come += results_customer_come_field[0]['count']
                total_booking_paid += results_booking_paid_field[0]['count']

        # thêm kết quả total vào cuối danh sách data
        ret_data.append({
            'date': 'TOTAL',
            'new_data': total_new_data,
            'interactive': total_interactive,
            'advisory_connect': total_advisory_connect,
            'advisory_percent': total_advisory_connect / total_interactive if total_interactive != 0 else '-',
            'booking_lead': total_booking_lead,
            'booking_percent': total_booking_lead / total_advisory_connect if total_advisory_connect != 0 else '-',
            'booking': total_booking,
            'customer_come': total_customer_come,
            'percent_customer_come': total_customer_come / total_booking if total_booking != 0 else '-',
            'booking_paid': total_booking_paid,
            'percent_booking_paid': total_booking_paid / total_customer_come if total_customer_come != 0 else '-',
        })
        return ret_data

    def _get_data_report_brand_overview_per_month(self):
        # khởi tạo biến total
        total_new_data = 0
        total_interactive = 0
        total_advisory_connect = 0
        total_advisory_percent = 0
        total_booking_lead = 0
        total_booking_percent = 0
        total_booking = 0
        total_customer_come = 0
        total_percent_customer_come = 0
        total_booking_paid = 0
        total_percent_booking_paid = 0
        ret_data = []
        year_search = self.year
        for month in range(1, 13):
            # ================================== lead ==================================================================
            # dữ liệu cột "data mới"
            query_new_data_field = "select count(*) from crm_lead cl where extract(month from cl.create_date) = %s " \
                                   "and extract(year from cl.create_date) = %s and type = 'lead' and cl.type_data = 'new' and cl.active " \
                                   "= 'True' and stage_id not in (%s,%s) and cl.company_id = %s" % (month, year_search,
                                                                                                    self.env.ref(
                                                                                                        'crm_base.crm_stage_lead_not_quality').id,
                                                                                                    self.env.ref(
                                                                                                        'crm_base.crm_stage_test').id,
                                                                                                    self.company_id.id)
            self.env.cr.execute(query_new_data_field)
            results_new_data = self.env.cr.dictfetchall()

            # dữ liệu cột "tổng tương tác"
            query_total_interactive_field = "select count(*) from crm_lead cl where extract(month from " \
                                            "cl.create_date) = %s and extract(year from cl.create_date) = %s and " \
                                            "cl.active = 'True' and type = 'lead' and cl.company_id = %s and stage_id " \
                                            "not in (%s,%s)" % (month, year_search, self.company_id.id, self.env.ref(
                'crm_base.crm_stage_lead_not_quality').id, self.env.ref('crm_base.crm_stage_test').id)
            self.env.cr.execute(query_total_interactive_field)
            results_interactive_field = self.env.cr.dictfetchall()

            # dữ liệu cột "kết nối tư vấn"
            query_advisory_connect_field = "select count(*) from crm_lead cl where extract(month from cl.create_date) " \
                                           "= %s and extract(year from cl.create_date) = %s and cl.active = 'True' " \
                                           "and company_id = %s and stage_id in (%s,%s,%s)" % (month, year_search,
                                                                                               self.company_id.id,
                                                                                               self.env.ref(
                                                                                                   'crm_base.crm_stage_refer').id,
                                                                                               self.env.ref(
                                                                                                   'crm_base.crm_stage_potential').id,
                                                                                               self.env.ref(
                                                                                                   'crm_base.crm_stage_booking').id)
            self.env.cr.execute(query_advisory_connect_field)
            results_advisory_connect_field = self.env.cr.dictfetchall()

            # dữ liệu cột "% kết nối"
            if results_interactive_field[0]['count'] != 0:
                advisory_percent = results_advisory_connect_field[0]['count'] / results_interactive_field[0]['count']
            else:
                advisory_percent = 0

            # dữ liệu cột "hẹn lịch"
            query_booking_lead_field = "select count(*) from crm_lead cl where extract(month from cl.create_date) = " \
                                       "%s and extract(year from cl.create_date) = %s and type = 'lead' and " \
                                       "company_id = %s and cl.active = 'True' and stage_id = %s" % (month, year_search,
                                                                                                 self.company_id.id,
                                                                                                 self.env.ref(
                                                                                                     'crm_base.crm_stage_booking').id)
            self.env.cr.execute(query_booking_lead_field)
            results_booking_lead_field = self.env.cr.dictfetchall()

            # dữ liệu cột "% hẹn lịch"
            if results_advisory_connect_field[0]['count'] != 0:
                booking_percent = results_booking_lead_field[0]['count'] / results_advisory_connect_field[0]['count']
            else:
                booking_percent = 0

            # ================================ booking ===========================================================
            # dữ liệu cột "lịch hẹn"
            # a = "select count(*) from crm_lead cl join product_pricelist pp on cl.price_list_id = pp.id where cl.type = 'opportunity' and pp.type = 'service' and cl.active = 'True' and extract(month from cl.create_date) = %s and extract(year from cl.create_date) = %s and cl.company_id = %s"
            query_booking_field = "select count(*) from crm_lead cl join product_pricelist pp on cl.price_list_id = " \
                                  "pp.id where cl.type = 'opportunity' and pp.type = 'service' and cl.active = 'True' " \
                                  "and extract(month from cl.create_date) = %s and extract(year from cl.create_date) " \
                                  "= %s and cl.company_id = %s" % (month, year_search, self.company_id.id)
            self.env.cr.execute(query_booking_field)
            results_booking_field = self.env.cr.dictfetchall()

            # dữ liệu cột "đến cửa"
            query_customer_come = "select count(*) from crm_lead cl join product_pricelist pp on cl.price_list_id = " \
                                  "pp.id where cl.type = 'opportunity' and pp.type = 'service' and cl.active = 'True' " \
                                  "and extract(month from cl.create_date) = %s and extract(year from cl.create_date) " \
                                  "= %s and cl.company_id = %s and cl.customer_come = 'yes'" % (
                                  month, year_search, self.company_id.id)
            self.env.cr.execute(query_customer_come)
            results_customer_come_field = self.env.cr.dictfetchall()

            # dữ liệu cột "% đến cửa"
            if results_booking_field[0]['count'] != 0:
                percent_customer_come = results_customer_come_field[0]['count'] / results_booking_field[0]['count']
            else:
                percent_customer_come = 0

            # dữ liệu cột "thành công"
            query_booking_paid = "select count(*) from crm_lead cl where extract(month from cl.create_date) = %s and " \
                                 "extract(year from cl.create_date) = %s and active = 'True' and type = 'opportunity' " \
                                 "and company_id = '%s' and stage_id = '%s'" % (month, year_search,
                                                                                self.company_id.id, self.env.ref(
                'crm_base.crm_stage_paid').id)
            self.env.cr.execute(query_booking_paid)
            results_booking_paid_field = self.env.cr.dictfetchall()

            # dữ liệu cột "% thành công"
            if results_customer_come_field[0]['count'] != 0:
                percent_booking_paid = results_booking_paid_field[0]['count'] / results_customer_come_field[0]['count']
            else:
                percent_booking_paid = 0

            ret_data.append({
                'date': 'Tháng ' + str(month),
                'new_data': results_new_data[0]['count'],
                'interactive': results_interactive_field[0]['count'],
                'advisory_connect': results_advisory_connect_field[0]['count'],
                'advisory_percent': advisory_percent,
                'booking_lead': results_booking_lead_field[0]['count'],
                'booking_percent': booking_percent,
                'booking': results_booking_field[0]['count'],
                'customer_come': results_customer_come_field[0]['count'],
                'percent_customer_come': percent_customer_come,
                'booking_paid': results_booking_paid_field[0]['count'],
                'percent_booking_paid': percent_booking_paid,
            })
            # tính lại các biến total
            total_new_data += results_new_data[0]['count']
            total_interactive += results_interactive_field[0]['count']
            total_advisory_connect += results_advisory_connect_field[0]['count']
            total_booking_lead += results_booking_lead_field[0]['count']
            total_booking += results_booking_field[0]['count']
            total_customer_come += results_customer_come_field[0]['count']
            total_booking_paid += results_booking_paid_field[0]['count']

        # thêm kết quả total vào cuối danh sách data
        ret_data.append({
            'date': 'TOTAL',
            'new_data': total_new_data,
            'interactive': total_interactive,
            'advisory_connect': total_advisory_connect,
            'advisory_percent': total_advisory_connect / total_interactive if total_interactive != 0 else '-',
            'booking_lead': total_booking_lead,
            'booking_percent': total_booking_lead / total_advisory_connect if total_advisory_connect != 0 else '-',
            'booking': total_booking,
            'customer_come': total_customer_come,
            'percent_customer_come': total_customer_come / total_booking if total_booking != 0 else '-',
            'booking_paid': total_booking_paid,
            'percent_booking_paid': total_booking_paid / total_customer_come if total_customer_come != 0 else '-',
        })

        return ret_data

    def create_report_brand_overview(self):
        # lấy data
        if self.type_display == 'date':
            # hiện theo ngày
            data = self._get_data_report_brand_overview_per_day()
        else:
            # hiện theo tháng
            data = self._get_data_report_brand_overview_per_month()

        # in báo cáo
        report_brand_overview_attachment = self.env['ir.attachment'].browse(
            self.env.ref('crm_report.report_brand_overview_customer_attachment').id)
        decode = base64.b64decode(report_brand_overview_attachment.datas)
        wb = load_workbook(BytesIO(decode))
        ws = wb.active
        line_font = Font(name='Times New Roman', size=14)
        if self.type_display == 'date':
            ws['E3'].value = self.start_date.strftime('%d/%m/%Y')
            ws['G3'].value = self.end_date.strftime('%d/%m/%Y')
        else:
            ws['E3'].value = '01/01/%s' % self.year
            ws['G3'].value = '31/12/%s' % self.year
        ws['I3'].value = self.company_id.name
        key_col_list = list(range(1, 13))
        key_list = [
            'date',
            'new_data',
            'interactive',
            'advisory_connect',
            'advisory_percent',
            'booking_lead',
            'booking_percent',
            'booking',
            'customer_come',
            'percent_customer_come',
            'booking_paid',
            'percent_booking_paid',
        ]
        row = 7
        for line_data in data:
            for col, k in zip(key_col_list, key_list):
                cell = ws.cell(row, col)
                cell.value = line_data[k]
                cell.font = line_font
                cell.border = all_border_thin
                cell.alignment = Alignment(horizontal='left', vertical='center')
            row += 1

        fp = BytesIO()
        wb.save(fp)
        fp.seek(0)
        report = base64.encodebytes((fp.read()))
        fp.close()
        attachment = self.env['ir.attachment'].sudo().create({
            'name': 'bao_cao_tong_quan_theo_khach.xlsx',
            'datas': report,
            'res_model': 'temp.creation',
            'public': True,
        })
        url = "/web/content/?model=ir.attachment&id=%s&filename_field=name&field=datas&download=true" \
              % attachment.id
        return {'name': 'BÁO CÁO TỔNG QUAN THEO ĐẦU KHÁCH',
                'type': 'ir.actions.act_url',
                'url': url,
                'target': 'self',
                }
