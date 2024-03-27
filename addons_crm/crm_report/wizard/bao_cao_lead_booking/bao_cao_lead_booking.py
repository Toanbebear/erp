from odoo import fields, api, models, _
from odoo.exceptions import ValidationError, UserError
from datetime import date, datetime, time
from calendar import monthrange
from openpyxl import load_workbook
from openpyxl.styles import Font, borders, Alignment
import base64
from io import BytesIO
from dateutil.relativedelta import relativedelta
from pytz import timezone, utc

thin = borders.Side(style='thin')
double = borders.Side(style='double')
all_border_thin = borders.Border(thin, thin, thin, thin)
TEMPLATE_BOOKING = [
    'khach_hang_den_cua',
    'trang_thai_booking',
    'ma_booking',
    'phan_loai_khach_hang',
    'dien_thoai',
    'ten_lien_he',
    'thuong_hieu',
    'chi_nhanh',
    'ngay_hen_lich',
    'ngay_tao',
    'nguoi_tao',
    'phong_ban_tao',
    'bang_gia',
    'nguon',
    'trang_thai_hieu_luc',
    'loai_ban_ghi',
    'ngay_den_cua',
    'chien_dich'
]

TEMPLATE_LEAD = [
    'ten_lien_he',
    'quan_huyen',
    'tinh_thanh_pho',
    'thuong_hieu',
    'chi_nhanh',
    'nguon',
    'trang_thai',
    'dong_dich_vu',
    'nguoi_tao',
    'phong_ban_nguoi_tao',
    'ngay_tao',
    'dien_thoai',
    'nhom_dich_vu',
    'kieu_du_lieu',
    'nguon_mo_rong'
]


class BookingExportReport(models.TransientModel):
    _name = 'report.export.booking'
    _description = 'Export Booking Report'

    start_date = fields.Date('Start date', default=date.today())
    end_date = fields.Date('End date', default=date.today())
    # convert date to datetime for search domain, should be removed if using datetime directly
    start_datetime = fields.Datetime('Start datetime', compute='_compute_datetime')
    end_datetime = fields.Datetime('End datetime', compute='_compute_datetime')

    type_date_search = fields.Selection([('cd', 'Ngày tạo'), ('bd', 'Ngày đặt lịch')], string='Type Record',
                                        default='cd')
    # company_id = fields.Many2one('res.company', string='Company', default=lambda self: self.env.company, domain=lambda self: [('id', 'in', self.env.companies.ids)])
    company_id = fields.Many2many(string='Chi nhánh', comodel_name='res.company', domain=lambda self: [('id', 'in', self.env.user.company_ids.ids)])
    product_category = fields.Many2one('sh.medical.health.center.service.category', string='Product Caterory')
    type_crm = fields.Selection([('lead', 'Lead'), ('opportunity', 'Booking')], string='Type Record',
                                default='lead')

    @api.depends('start_date', 'end_date')
    def _compute_datetime(self):
        self.start_datetime = False
        self.end_datetime = False
        if self.start_date and self.end_date:
            local_tz = timezone(self.env.user.tz or 'Etc/GMT+7')
            start_datetime = datetime(self.start_date.year, self.start_date.month, self.start_date.day, 0, 0, 0)
            end_datetime = datetime(self.end_date.year, self.end_date.month, self.end_date.day, 23, 59, 59)
            start_datetime = local_tz.localize(start_datetime, is_dst=None)
            end_datetime = local_tz.localize(end_datetime, is_dst=None)
            self.start_datetime = start_datetime.astimezone(utc).replace(tzinfo=None)
            self.end_datetime = end_datetime.astimezone(utc).replace(tzinfo=None)

    @api.onchange('start_date')
    def _onchange_start_date(self):
        if self.start_date:
            if self.start_date.month == fields.date.today().month:
                self.end_date = fields.date.today()
            else:
                self.end_date = date(self.start_date.year, self.start_date.month,
                                     monthrange(self.start_date.year, self.start_date.month)[1])

    @api.constrains('start_date', 'end_date')
    def check_dates(self):
        for record in self:
            start_date = fields.Date.from_string(record.start_date)
            end_date = fields.Date.from_string(record.end_date)
            if start_date > end_date:
                raise ValidationError(
                    _("Ngày kết thúc không thể ở trước ngày bắt đầu khi xuất báo cáo!!!"))

    def _get_data_booking(self, booking, state_booking):
        results_data = []

        lang_code = self.env.context.get('lang') or 'en_US'
        date_format = self.env['res.lang']._lang_get(lang_code).date_format
        time_format = self.env['res.lang']._lang_get(lang_code).time_format

        if booking.arrival_date:
            arrival_date = fields.Datetime.context_timestamp(self.with_context(tz=self.env.user.tz), booking.arrival_date if booking.arrival_date else None).strftime(
                ('%s %s') % (date_format, time_format)),
        else:
            arrival_date = None

        results_data.append({
            'khach_hang_den_cua': booking.customer_come if booking.customer_come else None,
            'trang_thai_booking': state_booking,
            'ma_booking': booking.name,
            'phan_loai_khach_hang': None,
            'dien_thoai': None,
            'ten_lien_he': booking.contact_name,
            'thuong_hieu': booking.brand_id.name,
            'chi_nhanh': booking.company_id.name,
            'ngay_hen_lich': '%s' % fields.Datetime.context_timestamp(self.with_context(tz=self.env.user.tz), booking.booking_date).strftime(
                ('%s %s') % (date_format, time_format)) if self.type_crm == 'opportunity' else None,
            'ngay_tao': '%s' % fields.Datetime.context_timestamp(self.with_context(tz=self.env.user.tz), booking.create_on).strftime(('%s %s') % (date_format, time_format)),
            'nguoi_tao': booking.create_by.name,
            'phong_ban_tao': booking.department_id.name,
            'bang_gia': booking.price_list_id.name,
            'nguon': booking.source_id.name,
            'trang_thai_hieu_luc': booking.effect,
            'loai_ban_ghi': booking.type_crm_id.name or None,
            'ngay_den_cua': '%s' % arrival_date,
            'chien_dich': booking.campaign_id.name or None
        })
        # if self.product_category:
        #     results_data = [element for element in results_data if element['service_category'] == self.product_category]

        return results_data

    def _get_data_lead(self, lead):
        results_data = []

        lang_code = self.env.context.get('lang') or 'en_US'
        date_format = self.env['res.lang']._lang_get(lang_code).date_format
        time_format = self.env['res.lang']._lang_get(lang_code).time_format

        if lead.crm_line_ids:
            for line in lead.crm_line_ids:
                val = {
                    'ten_lien_he': lead.contact_name,
                    'quan_huyen': lead.district_id.name or None,
                    'tinh_thanh_pho': lead.state_id.name or None,
                    'thuong_hieu': lead.brand_id.name,
                    'chi_nhanh': lead.company_id.name,
                    'nguon': lead.original_source_id.name,
                    'trang_thai': lead.stage_id.name,
                    'dong_dich_vu': line.service_id.name,
                    'nguoi_tao': lead.create_by.name,
                    'phong_ban_nguoi_tao': lead.department_id.name or None,
                    'ngay_tao': '%s' % fields.Datetime.context_timestamp(self.with_context(tz=self.env.user.tz), lead.create_on).strftime(('%s %s') % (date_format, time_format)),
                    'dien_thoai': None,
                    'nhom_dich_vu': line.service_id.service_category.name,
                    'kieu_du_lieu': lead.type_data,
                    'nguon_mo_rong': line.source_extend_id.name
                }
                results_data.append(val)
        else:
            val = {
                'ten_lien_he': lead.contact_name,
                'quan_huyen': lead.district_id.name or None,
                'tinh_thanh_pho': lead.state_id.name or None,
                'thuong_hieu': lead.brand_id.name,
                'chi_nhanh': lead.company_id.name,
                'nguon': lead.original_source_id.name,
                'trang_thai': lead.stage_id.name,
                'dong_dich_vu': None,
                'nguoi_tao': lead.create_by.name,
                'phong_ban_nguoi_tao': lead.department_id.name or None,
                'ngay_tao': '%s' % fields.Datetime.context_timestamp(self.with_context(tz=self.env.user.tz), lead.create_on).strftime(('%s %s') % (date_format, time_format)),
                'dien_thoai': None,
                'nhom_dich_vu': None,
                'kieu_du_lieu': lead.type_data,
                'nguon_mo_rong': None
            }
            results_data.append(val)
        return results_data

    def create_report_export_bk(self):
        datas = []
        # Tạo domain từ đầu vào
        if self.type_date_search == "cd":  # Theo ngày tạo
            domain = [('create_date', '>=', self.start_datetime), ('create_date', '<=', self.end_datetime),
                      ('type', '=', self.type_crm), ('company_id', 'in', self.company_id.ids)]
        else:  # Theo ngày đặt lịch
            domain = [('booking_date', '>=', self.start_datetime),
                      ('booking_date', '<=', self.end_datetime),
                      ('type', '=', self.type_crm), ('company_id', 'in', self.company_id.ids)]

        bookings = self.env['crm.lead'].sudo().search(domain)

        if self.type_crm == 'opportunity':
            for booking in bookings:
                state_booking = booking.stage_id.name
                # Kiểm tra dịch vụ và trạng thái của line dịch vụ
                crm_line = booking.crm_line_ids

                crm_line_covid = crm_line.filtered(
                    lambda line: line.crm_id.price_list_id.item_ids.filtered(lambda product: product.product_id.id == line.product_id.id and product.target_sale_marketing == True))

                crm_line_not_covid = crm_line.filtered(
                    lambda line: line.crm_id.price_list_id.item_ids.filtered(lambda product: product.product_id.id == line.product_id.id and product.target_sale_marketing == False))

                if crm_line_covid:
                    if all([line.stage != 'done' for line in crm_line_not_covid]) and any([line.stage == 'done' for line in crm_line_covid]):
                        state_booking = None
                # get data
                result = self._get_data_booking(booking, state_booking)
                for e in result:
                    datas.append(e)
                report_brand_overview_attachment = self.env['ir.attachment'].browse(
                    self.env.ref('crm_report.xuat_booking_bk_attachment').id)
        elif self.type_crm == 'lead':
            for booking in bookings:
                result = self._get_data_lead(booking)
                for e in result:
                    datas.append(e)
                report_brand_overview_attachment = self.env['ir.attachment'].browse(
                    self.env.ref('crm_report.xuat_lead_bk_attachment').id)

        # in báo cáo
        decode = base64.b64decode(report_brand_overview_attachment.datas)
        wb = load_workbook(BytesIO(decode))
        ws = wb.active
        line_font = Font(name='Times New Roman', size=14)
        if self.type_crm == 'lead':
            name = 'LEAD'
        else:
            name = 'BOOKING'

        ws['A2'].value = 'DỮ LIỆU %s' % name
        ws['A2'].font = Font(name='Times New Roman', size=20)
        if self.type_crm == 'lead':
            ws['E3'].value = 'Ngày tạo từ: '
            ws['E3'].font = line_font
            ws['F3'].value = self.start_date.strftime('%d/%m/%Y')
            ws['F3'].font = line_font
            ws['G3'].value = 'Đến: '
            ws['G3'].font = line_font
            ws['H3'].value = self.end_date.strftime('%d/%m/%Y')
            ws['H3'].font = line_font
        else:
            if self.type_date_search == 'cd':
                ws['E3'].value = 'Ngày tạo từ: '
                ws['E3'].font = line_font
                ws['F3'].value = self.start_date.strftime('%d/%m/%Y')
                ws['F3'].font = line_font
                ws['G3'].value = 'Đến: '
                ws['G3'].font = line_font
                ws['H3'].value = self.end_date.strftime('%d/%m/%Y')
                ws['H3'].font = line_font
            else:
                ws['E3'].value = 'Ngày đặt lịch từ: '
                ws['E3'].font = line_font
                ws['F3'].value = self.start_date.strftime('%d/%m/%Y')
                ws['F3'].font = line_font
                ws['G3'].value = 'Đến: '
                ws['G3'].font = line_font
                ws['H3'].value = self.end_date.strftime('%d/%m/%Y')
                ws['H3'].font = line_font
        ws['E4'].value = 'Chi nhánh: '
        ws['E4'].font = line_font
        ws['F4'].value = ', '.join(rec.name for rec in self.company_id)
        ws['F4'].font = line_font

        if self.product_category:
            name = name + '_' + 'Nhóm_dịch_vụ: ' + self.product_category.name
            ws['E5'].value = 'Nhóm dịch vụ: '
            ws['E5'].font = line_font
            ws['F5'].value = name
            ws['F5'].font = line_font
        key_col_list_booking = list(range(1, len(TEMPLATE_BOOKING) + 1))
        key_col_list_lead = list(range(1, len(TEMPLATE_LEAD) + 1))

        row = 8
        if self.type_crm == 'opportunity':
            for line_data in datas:
                for col, k in zip(key_col_list_booking, TEMPLATE_BOOKING):
                    cell = ws.cell(row, col)
                    cell.value = line_data[k]
                    cell.font = line_font
                    cell.border = all_border_thin
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                row += 1
        elif self.type_crm == 'lead':
            for line_data in datas:
                for col, k in zip(key_col_list_lead, TEMPLATE_LEAD):
                    cell = ws.cell(row, col)
                    cell.value = line_data[k]
                    cell.font = line_font
                    cell.border = all_border_thin
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                row += 1

        fp = BytesIO()
        wb.save(fp)
        fp.seek(0)
        report = base64.encodebytes((fp.read()))
        fp.close()

        attachment = self.env['ir.attachment'].sudo().create({
            'name': 'xuat_du_lieu_%s.xlsx' % name,
            'datas': report,
            'res_model': 'temp.creation',
            'public': True,
        })
        url = "/web/content/?model=ir.attachment&id=%s&filename_field=name&field=datas&download=true" \
              % attachment.id
        return {'name': 'XUẤT DỮ LIỆU LEAD/ BOOKING',
                'type': 'ir.actions.act_url',
                'url': url,
                'target': 'self',
                }
