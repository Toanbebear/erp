from odoo import fields, api, models, _
from odoo.exceptions import ValidationError, UserError
from datetime import date, datetime, time
from calendar import monthrange
from openpyxl import load_workbook
from openpyxl.styles import Font, borders, Alignment, PatternFill
import base64
from io import BytesIO
from dateutil.relativedelta import relativedelta
from pytz import timezone, utc

import logging

thin = borders.Side(style='thin')
double = borders.Side(style='double')
all_border_thin = borders.Border(thin, thin, thin, thin)


class DailySalesReport(models.TransientModel):
    _name = 'erp.daily.sales.report'
    _description = 'Sales report'

    start_date = fields.Date('Start date', default=date.today())
    end_date = fields.Date('End date', default=date.today())
    # convert date to datetime for search domain, should be removed if using datetime directly
    # start_datetime = fields.Datetime('Start datetime', compute='_compute_datetime')
    # end_datetime = fields.Datetime('End datetime', compute='_compute_datetime')

    # @api.depends('start_date', 'end_date')
    # def _compute_datetime(self):
    #     self.start_datetime = False
    #     self.end_datetime = False
    #     if self.start_date and self.end_date:
    #         local_tz = timezone(self.env.user.tz or 'Etc/GMT+7')
    #         start_datetime = datetime(self.start_date.year, self.start_date.month, self.start_date.day, 0, 0, 0)
    #         end_datetime = datetime(self.end_date.year, self.end_date.month, self.end_date.day, 23, 59, 59)
    #         start_datetime = local_tz.localize(start_datetime, is_dst=None)
    #         end_datetime = local_tz.localize(end_datetime, is_dst=None)
    #         self.start_datetime = start_datetime.astimezone(utc).replace(tzinfo=None)
    #         self.end_datetime = end_datetime.astimezone(utc).replace(tzinfo=None)

    @api.onchange('start_date')
    def _onchange_start_date(self):
        if self.start_date:
            if self.start_date.month == fields.date.today().month:
                self.end_date = fields.date.today()
            else:
                self.end_date = date(self.start_date.year, self.start_date.month,
                                     monthrange(self.start_date.year, self.start_date.month)[1])

    @api.constrains('start_date', 'end_date')
    def check_dates(self):
        for record in self:
            start_date = fields.Date.from_string(record.start_date)
            end_date = fields.Date.from_string(record.end_date)

            if start_date > end_date:
                raise ValidationError(
                    _("Ngày kết thúc phải sau ngày bắt đầu."))
            elif (end_date - start_date).days > 31:
                raise ValidationError(
                    _("Chỉ tải được dữ liệu trong một tháng"))

    def report_daily_sales(self):
        ret_data = []
        # tổng tiền booking
        total_amount_total_booking_inbound = 0
        total_amount_total_booking_outbound = 0

        # tổng tiền thu (VND)
        total_amount_vnd_inbound = 0
        total_amount_vnd_outbound = 0

        # tổng tiền đã nộp trước
        total_amount_paid_inbound = 0
        total_amount_paid_outbound = 0

        # tổng tiền còn lại
        total_amount_remain_inbound = 0
        total_amount_remain_outbound = 0

        # tổng số tiền đã sử dụng
        total_amount_used_inbound = 0
        total_amount_used_outbound = 0

        # lấy data từ payment
        domain = [('payment_date', '>=', self.start_date), ('payment_date', '<=', self.end_date),
                  ('state', '!=', 'draft'), ('state', '!=', 'cancelled'), ('payment_type', '!=', 'internal')]
        list_payment = self.env['account.payment'].search(domain)

        # lấy data đơn vị tiền sử dụng
        list_currency_used = self.env['res.currency'].search([])
        total_amount_foreign_currency_inbound = {cur.name: 0 for cur in list_currency_used}
        total_amount_foreign_currency_outbound = {cur.name: 0 for cur in list_currency_used}

        # tổng số tiền theo hình thức thanh toán phiếu thu
        total_amount_tm_inbound = 0
        total_amount_ck_inbound = 0
        total_amount_nb_inbound = 0
        total_amount_pos_inbound = 0
        total_amount_vdt_inbound = 0

        # tổng số tiền theo hình thức thanh toán phiếu thu
        total_amount_tm_outbound = 0
        total_amount_ck_outbound = 0
        total_amount_nb_outbound = 0
        total_amount_pos_outbound = 0
        total_amount_vdt_outbound = 0

        list_booking_total = []
        list_booking_paid = []
        list_booking_remain = []
        list_booking_used = []

        list_booking_total_outbound = []
        list_booking_paid_outbound = []
        list_booking_remain_outbound = []
        list_booking_used_outbound = []

        for rec in list_payment:
            rec = rec.sudo()
            amount_foreign_currency = 0
            amount = 0
            payment_type = None
            # ---------------- ghi nhận kiểu thanh toán & tính tổng tiền
            if rec.payment_type:
                if rec.payment_type == 'outbound':
                    payment_type = 'Hoàn tiền'

                    # tổng tiền booking theo phiếu hoàn tiền
                    if rec.crm_id.id not in list_booking_total_outbound:
                        if rec.crm_id.amount_total:
                            list_booking_total_outbound.append(rec.crm_id.id)
                            total_amount_total_booking_outbound += rec.crm_id.amount_total

                    # tổng tiền thu(nguyên tệ) theo phiếu hoàn tiền
                    for i in list_currency_used:
                        if rec.currency_id.name == i.name:
                            total_amount_foreign_currency_outbound[i.name] += rec.amount

                    # tổng tiền đã nộp(theo booking) theo phiếu hoàn tiền
                    if rec.crm_id.id not in list_booking_paid_outbound:
                        if rec.crm_id.amount_paid:
                            list_booking_paid_outbound.append(rec.crm_id.id)
                            total_amount_paid_outbound += rec.crm_id.amount_paid

                    # tổng tiền còn lại(theo booking) theo phiếu hoàn tiền
                    if rec.crm_id.id not in list_booking_remain_outbound:
                        if rec.crm_id.amount_remain:
                            list_booking_remain_outbound.append(rec.crm_id.id)
                            total_amount_remain_outbound += rec.crm_id.amount_remain

                    # tổng số tiền đã sử dụng(theo booking) theo phiếu hoàn tiền
                    if rec.crm_id.id not in list_booking_used_outbound:
                        if rec.crm_id.amount_used:
                            list_booking_used_outbound.append(rec.crm_id.id)
                            total_amount_used_outbound += rec.crm_id.amount_remain

                elif rec.payment_type == 'inbound':
                    payment_type = 'Thu tiền'

                    # tổng tiền booking theo phiếu thu
                    if rec.crm_id.id not in list_booking_total:
                        if rec.crm_id.amount_total:
                            list_booking_total.append(rec.crm_id.id)
                            total_amount_total_booking_inbound += rec.crm_id.amount_total

                    # tổng tiền thu(nguyên tệ) theo phiếu thu tiền
                    for i in list_currency_used:
                        if rec.currency_id.name == i.name:
                            total_amount_foreign_currency_inbound[i.name] += rec.amount

                    # tổng tiền đã nộp(lấy theo booking) theo phiếu thu tiền:
                    if rec.crm_id.id not in list_booking_paid:
                        if rec.crm_id.amount_paid:
                            list_booking_paid.append(rec.crm_id.id)
                            total_amount_paid_inbound += rec.crm_id.amount_paid

                    # tổng tiền còn lại (theo booking) theo phiếu thu tiền
                    if rec.crm_id.id not in list_booking_remain:
                        if rec.crm_id.amount_remain:
                            list_booking_remain.append(rec.crm_id.id)
                            total_amount_remain_inbound += rec.crm_id.amount_remain

                    # tổng số tiền đã sử dụng(theo booking) theo phiếu hoàn tiền
                    if rec.crm_id.id not in list_booking_used:
                        if rec.crm_id.amount_used:
                            list_booking_used.append(rec.crm_id.id)
                            total_amount_used_inbound += rec.crm_id.amount_remain

            # ------------------ kiểm tra ngoại tệ ----------------------------------------
            if rec.currency_rate_id:
                amount_foreign_currency = format(int(rec.amount), ',d').replace(',', '.')
                amount_foreign_currency = str(amount_foreign_currency) + ' ' + rec.currency_id.name
                amount = None
            if not rec.currency_rate_id:
                # amount = format(int(rec.amount), ',d').replace(',', '.')
                # amount = str(amount) + ' ' + rec.currency_id.name
                amount = rec.amount
                amount_foreign_currency = None
                if rec.payment_type == 'outbound':
                    total_amount_vnd_outbound += rec.amount
                if rec.payment_type == 'inbound':
                    total_amount_vnd_inbound += rec.amount
            # ------------------------- lưu dữ liệu --------------------------------------
            # get payment_method
            if rec.payment_method == 'tm':
                payment_method = 'Tiền mặt'
                if rec.payment_type == 'outbound':
                    total_amount_tm_outbound += rec.amount
                if rec.payment_type == 'inbound':
                    total_amount_tm_inbound += rec.amount
            elif rec.payment_method == 'ck':
                payment_method = 'Chuyển khoản'
                if rec.payment_type == 'outbound':
                    total_amount_ck_outbound += rec.amount
                if rec.payment_type == 'inbound':
                    total_amount_ck_inbound += rec.amount
            elif rec.payment_method == 'nb':
                payment_method = 'Thanh toán nội bộ'
                if rec.payment_type == 'outbound':
                    total_amount_nb_outbound += rec.amount
                if rec.payment_type == 'inbound':
                    total_amount_nb_inbound += rec.amount
            elif rec.payment_method == 'pos':
                payment_method = 'Quẹt thẻ qua POS'
                if rec.payment_type == 'outbound':
                    total_amount_pos_outbound += rec.amount
                if rec.payment_type == 'inbound':
                    total_amount_pos_inbound += rec.amount
            else:
                payment_method = 'Thanh toán qua thẻ điện tử'
                if rec.payment_type == 'outbound':
                    total_amount_vdt_outbound += rec.amount
                if rec.payment_type == 'inbound':
                    total_amount_vdt_inbound += rec.amount

            ret_data.append({
                'payment_date': rec.payment_date.strftime("%d/%m/%Y"),
                'name': rec.name,
                'payment_type': payment_type,
                'payment_method': payment_method,
                'journal_id': rec.journal_id.name,
                'communication': rec.communication,
                'code_customer': rec.partner_id.code_customer,
                'name_customer': rec.partner_id.name,
                'address_customer': rec.partner_id.street or '-',
                'code_booking': rec.crm_id.name or '-',
                'amount_total': rec.crm_id.amount_total,
                # 'amount_total': format(int(rec.crm_id.amount_total), ',d').replace(',', '.') or '-',
                'amount': rec.amount,
                'currency': rec.currency_id.name,
                'amount_vnd': rec.amount_vnd,
                # 'amount_used': format(int(rec.crm_id.amount_used), ',d').replace(',', '.') or '-',
                # 'amount_paid': format(int(rec.crm_id.amount_paid), ',d').replace(',', '.') or '-',
                # 'amount_remain': format(int(rec.crm_id.amount_remain), ',d').replace(',', '.') or '-',
                'amount_used': rec.crm_id.amount_used or '-',
                'amount_paid': rec.crm_id.amount_paid or '-',
                'amount_remain': rec.crm_id.amount_remain or '-',
                'company': rec.company_id.name,
                'source': rec.crm_id.source_id.name or '-',
                'user': rec.user.name,
                'note': None,
            })
        # ----------- in dữ liệu
        daily_sales_attachment = self.env['ir.attachment'].browse(
            self.env.ref('crm_report.erp_daily_sales_report_attachment').id)
        decode = base64.b64decode(daily_sales_attachment.datas)
        wb = load_workbook(BytesIO(decode))
        ws = wb.active
        if self.start_date == self.end_date:
            ws['A3'].value = self.end_date.strftime('Ngày %d tháng %m năm %Y')
        else:
            ws['A3'].value = 'Từ ngày: %s đến ngày: %s' % (
                self.start_date.strftime('%d/%m/%Y'), self.end_date.strftime('%d/%m/%Y'))
        thin = borders.Side(style='thin')
        all_border_thin = borders.Border(left=thin, right=thin, top=thin, bottom=thin)
        line_font = Font(name='Times New Roman', size=12)
        key_col_list = list(range(2, 23))
        key_list = [
            'payment_date',
            'name',
            'payment_type',
            'payment_method',
            'journal_id',
            'communication',
            'code_customer',
            'name_customer',
            'address_customer',
            'code_booking',
            'amount_total',
            'amount',
            'currency',
            'amount_vnd',
            'amount_used',
            'amount_paid',
            'amount_remain',
            'company',
            'source',
            'user',
            'note',
        ]
        format_currency = '_(* #,##0_);_(* (#,##0);_(* "-"??_);_(@_)'
        row = 6
        for line_data in ret_data:
            ws.cell(row, 1).border = all_border_thin
            ws.cell(row, 1).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(row, 1).value = row - 5
            for col, k in zip(key_col_list, key_list):
                cell = ws.cell(row, col)
                cell.value = line_data[k]
                cell.font = line_font
                cell.border = all_border_thin
                cell.alignment = Alignment(horizontal='left', vertical='center')
                if 11 < col < 19:
                    cell.number_format = format_currency
            row += 1
        row += 1

        ### CÁC LOẠI TỔNG TIỀN THEO BOOKING
        title_in_outbound_fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')

        # TIÊU ĐỀ BẢNG
        ws.cell(row, 9).border, ws.cell(row, 9).value, ws.cell(row, 9).fill, ws.cell(row, 9).font, ws.cell(row, 9).alignment = all_border_thin, 'Số tiền theo booking', title_in_outbound_fill, Font(name='Times New Roman', size=12, bold=True), Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.cell(row, 10).border, ws.cell(row, 10).value, ws.cell(row, 10).fill, ws.cell(row, 10).font, ws.cell(row, 10).alignment = all_border_thin, 'Số tiền thu (nguyên tệ)', title_in_outbound_fill, Font(name='Times New Roman', size=12, bold=True), Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.cell(row, 11).border, ws.cell(row, 11).value, ws.cell(row, 11).fill, ws.cell(row, 11).font, ws.cell(row, 11).alignment = all_border_thin, 'Tổng số tiền quy đổi VND', title_in_outbound_fill, Font(name='Times New Roman', size=12, bold=True), Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.cell(row, 12).border, ws.cell(row, 12).value, ws.cell(row, 12).fill, ws.cell(row, 12).font, ws.cell(row, 12).alignment = all_border_thin, 'Tổng tiền khách đã sử dụng', title_in_outbound_fill, Font(name='Times New Roman', size=12, bold=True), Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.cell(row, 13).border, ws.cell(row, 13).value, ws.cell(row, 13).fill, ws.cell(row, 13).font, ws.cell(row, 13).alignment = all_border_thin, 'Tổng tiền khách đã trả', title_in_outbound_fill, Font(name='Times New Roman', size=12, bold=True), Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.cell(row, 14).border, ws.cell(row, 14).value, ws.cell(row, 14).fill, ws.cell(row, 14).font, ws.cell(row, 14).alignment = all_border_thin, 'Tổng tiền khách còn lại', title_in_outbound_fill, Font(name='Times New Roman', size=12, bold=True), Alignment(horizontal='center', vertical='center', wrap_text=True)

        # #GIÁ TRỊ BẢNG
        ws.cell(row + 1, 8).border, ws.cell(row + 1, 8).value, ws.cell(row + 1, 8).font, ws.cell(row + 1, 8).number_format = all_border_thin, 'TỔNG THU', Font(name='Times New Roman', size=14, bold=True),format_currency
        ws.cell(row + 2, 8).border, ws.cell(row + 2, 8).value, ws.cell(row + 2, 8).font, ws.cell(row + 2, 8).number_format= all_border_thin, 'TỔNG HOÀN', Font(name='Times New Roman', size=14, bold=True),format_currency

        ws.cell(row + 1, 9).border, ws.cell(row + 1, 9).value, ws.cell(row + 1, 9).number_format = all_border_thin, total_amount_total_booking_inbound,format_currency
        ws.cell(row + 2, 9).border, ws.cell(row + 2, 9).value, ws.cell(row + 2, 9).number_format = all_border_thin, total_amount_total_booking_outbound,format_currency

        # in tổng số tiền thu(nguyên tệ)
        list_test_total_amount_foreign_currency_inbound = [
            str(key) + ': ' + str(format(int(value), ',d').replace(',', '.')) for key, value in
            total_amount_foreign_currency_inbound.items()]
        list_test_total_amount_foreign_currency_inbound = '\n'.join(list_test_total_amount_foreign_currency_inbound)

        list_test_total_amount_foreign_currency_outbound = [
            str(key) + ': ' + str(format(int(value), ',d').replace(',', '.')) for key, value in
            total_amount_foreign_currency_outbound.items()]
        list_test_total_amount_foreign_currency_outbound = '\n'.join(list_test_total_amount_foreign_currency_outbound)

        ws.cell(row + 1, 10).border, ws.cell(row + 1, 10).value, ws.cell(row + 1, 10).number_format, ws.cell(row + 1, 10).alignment = all_border_thin, list_test_total_amount_foreign_currency_inbound,format_currency, Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.cell(row + 2, 10).border, ws.cell(row + 2, 10).value, ws.cell(row + 2, 10).number_format, ws.cell(row + 2, 10).alignment = all_border_thin, list_test_total_amount_foreign_currency_outbound,format_currency, Alignment(horizontal='center', vertical='center', wrap_text=True)

        # in tổng số tiền quy đổi (VND)
        ws.cell(row + 1, 11).border, ws.cell(row + 1, 11).value, ws.cell(row + 1,
                                                                         11).number_format = all_border_thin, total_amount_vnd_inbound, format_currency
        ws.cell(row + 2, 11).border, ws.cell(row + 2, 11).value, ws.cell(row + 2,
                                                                         11).number_format = all_border_thin, total_amount_vnd_outbound, format_currency

        # in tổng số tiền đã sử dụng(VND)
        ws.cell(row+1, 12).value, ws.cell(row+1, 12).border, ws.cell(row + 1, 12).number_format = total_amount_used_inbound, all_border_thin,format_currency
        ws.cell(row+2, 12).value, ws.cell(row+2, 12).border, ws.cell(row + 2, 12).number_format = total_amount_used_outbound, all_border_thin,format_currency

        # in tổng tiền khách đã trả
        ws.cell(row + 1, 13).value, ws.cell(row + 1, 13).border, ws.cell(row + 1, 13).number_format = total_amount_paid_inbound, all_border_thin,format_currency
        ws.cell(row + 2, 13).value, ws.cell(row + 2, 13).border, ws.cell(row + 2, 13).number_format = total_amount_paid_outbound, all_border_thin,format_currency

        # in tổng tiền còn lại
        ws.cell(row + 1, 14).value, ws.cell(row + 1, 14).border, ws.cell(row + 1, 14).number_format = total_amount_remain_inbound, all_border_thin,format_currency
        ws.cell(row + 2, 14).value, ws.cell(row + 2, 14).border, ws.cell(row + 2, 14).number_format = total_amount_remain_outbound, all_border_thin,format_currency

        # in tổng tiền theo hình thức thanh toán phiếu thu
        title_inbound_fill = PatternFill(start_color='9BC2E6', end_color='9BC2E6', fill_type='solid')
        title_outbound_fill = PatternFill(start_color='F8CBAD', end_color='F8CBAD', fill_type='solid')

        ws.cell(row, 2).value  = 'Hình thức'
        ws.cell(row, 3).value = 'Tổng thu'
        ws.cell(row, 2).fill, ws.cell(row, 2).alignment, ws.cell(row, 2).font = title_inbound_fill,Alignment(horizontal='center', vertical='center'),Font(name='Times New Roman', size=14, bold=True)
        ws.cell(row, 3).fill, ws.cell(row, 3).alignment, ws.cell(row, 3).font = title_inbound_fill,Alignment(horizontal='center', vertical='center'),Font(name='Times New Roman', size=14, bold=True)
        # ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)

        for i in range(2, 4):
            ws.cell(row, i).border = all_border_thin
            # tiêu đề
        ws.cell(row + 1, 2).value = 'Tiền mặt'
        ws.cell(row + 1, 2).border = all_border_thin
        ws.cell(row + 2, 2).value = 'Chuyển khoản'
        ws.cell(row + 2, 2).border = all_border_thin
        ws.cell(row + 3, 2).value = 'Thanh toán nội bộ'
        ws.cell(row + 3, 2).border = all_border_thin
        ws.cell(row + 4, 2).value = 'Quẹt thẻ qua POS'
        ws.cell(row + 4, 2).border = all_border_thin
        ws.cell(row + 5, 2).value = 'Ví điện tử'
        ws.cell(row + 5, 2).border = all_border_thin

        # TỔNG DOANH THU
        tong_doanh_so = total_amount_tm_inbound + total_amount_ck_inbound + total_amount_nb_inbound + total_amount_pos_inbound + total_amount_vdt_inbound - total_amount_tm_outbound - total_amount_ck_outbound - total_amount_nb_outbound
        ws.cell(row + 7, 2).value, ws.cell(row + 7, 2).border, ws.cell(row + 7, 2).font = 'TỔNG DOANH THU',all_border_thin,Font(name='Times New Roman', size=12, bold=True)
        ws.cell(row + 7, 3).value, ws.cell(row + 7, 3).border, ws.cell(row + 7, 3).font, ws.cell(row + 7, 3).number_format = tong_doanh_so ,all_border_thin,Font(name='Times New Roman', size=16, bold=True, color='FF0000'),format_currency

        # DIỄN GIẢI CHI
        title_out_fill = PatternFill(start_color='DBDBDB', end_color='DBDBDB', fill_type='solid')
        ws.cell(row + 7, 5).value, ws.cell(row + 7, 5).border, ws.cell(row + 7, 5).font, ws.cell(row + 7,
                                                                                                 5).fill, ws.cell(
            row + 7, 5).alignment = 'STT', all_border_thin, Font(name='Times New Roman', size=12,
                                                                 bold=True), title_out_fill, Alignment(
            horizontal='center', vertical='center')
        # ws.merge_cells(start_row=row + 7, start_column=6, end_row=row+7, end_column=7)
        ws.cell(row + 7, 6).value, ws.cell(row + 7, 6).border, ws.cell(row + 7, 6).font, ws.cell(row + 7,
                                                                                                 6).fill, ws.cell(
            row + 7, 6).alignment = 'Chi phí', all_border_thin, Font(name='Times New Roman', size=12,
                                                                     bold=True), title_out_fill, Alignment(
            horizontal='center', vertical='center')
        ws.cell(row + 7, 7).value, ws.cell(row + 7, 7).border, ws.cell(row + 7, 7).font, ws.cell(row + 7,
                                                                                                 7).fill, ws.cell(
            row + 7, 7).alignment = 'Diễn giải', all_border_thin, Font(name='Times New Roman', size=12,
                                                                       bold=True), title_out_fill, Alignment(
            horizontal='center', vertical='center')

        for col_chi in range(5, 8):
            if col_chi == 6:
                ws.cell(row + 8, col_chi).number_format = format_currency
                ws.cell(row + 9, col_chi).number_format = format_currency
                ws.cell(row + 10, col_chi).number_format = format_currency
                ws.cell(row + 11, col_chi).number_format = format_currency
                ws.cell(row + 12, col_chi).number_format = format_currency
                ws.cell(row + 13, col_chi).number_format = format_currency
                ws.cell(row + 14, col_chi).number_format = format_currency
                ws.cell(row + 15, col_chi).number_format = format_currency
                ws.cell(row + 16, col_chi).number_format = format_currency
                ws.cell(row + 17, col_chi).number_format = format_currency
                ws.cell(row + 18, col_chi).number_format = format_currency

            ws.cell(row + 8, col_chi).border, ws.cell(row + 8, col_chi).fill = all_border_thin, title_out_fill
            ws.cell(row + 9, col_chi).border,ws.cell(row + 9, col_chi).fill = all_border_thin, title_out_fill
            ws.cell(row + 10, col_chi).border,ws.cell(row + 10, col_chi).fill = all_border_thin, title_out_fill
            ws.cell(row + 11, col_chi).border,ws.cell(row + 11, col_chi).fill = all_border_thin, title_out_fill
            ws.cell(row + 12, col_chi).border,ws.cell(row + 12, col_chi).fill = all_border_thin, title_out_fill
            ws.cell(row + 13, col_chi).border,ws.cell(row + 13, col_chi).fill = all_border_thin, title_out_fill
            ws.cell(row + 14, col_chi).border,ws.cell(row + 14, col_chi).fill = all_border_thin, title_out_fill
            ws.cell(row + 15, col_chi).border,ws.cell(row + 15, col_chi).fill = all_border_thin, title_out_fill
            ws.cell(row + 16, col_chi).border,ws.cell(row + 16, col_chi).fill = all_border_thin, title_out_fill
            ws.cell(row + 17, col_chi).border,ws.cell(row + 17, col_chi).fill = all_border_thin, title_out_fill
            ws.cell(row + 18, col_chi).border,ws.cell(row + 18, col_chi).fill,ws.cell(row + 18, col_chi).font = all_border_thin, title_out_fill,Font(name='Times New Roman', size=13,
                                                                       bold=True)
        ws.cell(row + 18, 5).value = "Tổng chi phí"


        # giá trị
        ws.cell(row + 1, 3).value = total_amount_tm_inbound
        ws.cell(row + 1, 3).border = all_border_thin
        ws.cell(row + 1, 3).number_format = format_currency
        ws.cell(row + 2, 3).value = total_amount_ck_inbound
        ws.cell(row + 2, 3).border = all_border_thin
        ws.cell(row + 2, 3).number_format = format_currency
        ws.cell(row + 3, 3).value = total_amount_nb_inbound
        ws.cell(row + 3, 3).border = all_border_thin
        ws.cell(row + 3, 3).number_format = format_currency
        ws.cell(row + 4, 3).value = total_amount_pos_inbound
        ws.cell(row + 4, 3).border = all_border_thin
        ws.cell(row + 4, 3).number_format = format_currency
        ws.cell(row + 5, 3).value = total_amount_vdt_inbound
        ws.cell(row + 5, 3).border = all_border_thin
        ws.cell(row + 5, 3).number_format = format_currency

        # in tổng tiền theo hình thức thanh toán phiếu thu
        ws.cell(row, 5).value = 'Hình thức'
        ws.cell(row, 6).value = 'Tổng hoàn'

        ws.cell(row, 5).fill, ws.cell(row, 5).alignment, ws.cell(row, 5).font = title_outbound_fill, Alignment(
            horizontal='center', vertical='center'), Font(name='Times New Roman', size=14, bold=True)
        ws.cell(row, 6).fill, ws.cell(row, 6).alignment, ws.cell(row, 6).font = title_outbound_fill, Alignment(
            horizontal='center', vertical='center'), Font(name='Times New Roman', size=14, bold=True)
        # ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=6)
        for i in range(5, 7):
            ws.cell(row, i).border = all_border_thin
            # tiêu đề
        ws.cell(row + 1, 5).value = 'Tiền mặt'
        ws.cell(row + 1, 5).border = all_border_thin
        ws.cell(row + 2, 5).value = 'Chuyển khoản'
        ws.cell(row + 2, 5).border = all_border_thin
        ws.cell(row + 3, 5).value = 'Thanh toán nội bộ'
        ws.cell(row + 3, 5).border = all_border_thin
            # giá trị
        ws.cell(row + 1, 6).value = total_amount_tm_outbound
        ws.cell(row + 1, 6).border = all_border_thin
        ws.cell(row + 1, 6).number_format = format_currency
        ws.cell(row + 2, 6).value = total_amount_ck_outbound
        ws.cell(row + 2, 6).border = all_border_thin
        ws.cell(row + 2, 6).number_format = format_currency
        ws.cell(row + 3, 6).value = total_amount_nb_outbound
        ws.cell(row + 3, 6).border = all_border_thin
        ws.cell(row + 3, 6).number_format = format_currency

        fp = BytesIO()
        wb.save(fp)
        fp.seek(0)
        report = base64.encodebytes((fp.read()))
        fp.close()
        attachment = self.env['ir.attachment'].with_user(1).create({
            'name': 'bao_cao_doanh_so_ngay.xlsx',
            'datas': report,
            'res_model': 'temp.creation',
            'public': True,
        })
        url = "/web/content/?model=ir.attachment&id=%s&filename_field=name&field=datas&download=true" \
              % attachment.id
        # return {
        #     'name': 'Báo cáo doanh số ngày',
        #     'type': 'ir.actions.act_window',
        #     'res_model': 'temp.wizard',
        #     'view_mode': 'form',
        #     'view_type': 'form',
        #     'target': 'inline',
        #     'view_id': self.env.ref('ms_templates.report_wizard').id,
        #     'context': {'attachment_id': attachment.id}
        # }
        return {
            'name': 'Báo cáo doanh số ngày',
            'type': 'ir.actions.act_url',
            'url': url,
            'target': 'self',
        }
