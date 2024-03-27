from odoo import fields, api, models, _
from odoo.exceptions import ValidationError, UserError
from datetime import date, datetime, time
from calendar import monthrange
import openpyxl
from openpyxl.drawing.image import Image
from openpyxl import load_workbook
from openpyxl.styles import Font, borders, Alignment, PatternFill
from odoo.addons.report_sale.wizard.theme_report import ThemeReport
import base64
from io import BytesIO
from dateutil.relativedelta import relativedelta
from pytz import timezone, utc
from odoo.modules.module import get_module_resource


class SpecialTreatmentReport(models.TransientModel):
    _name = 'report.revenue.sale'
    _description = 'Báo cáo doanh thu bán sản phẩm'

    start_date = fields.Date('Start date', default=date.today())
    end_date = fields.Date('End date', default=date.today())

    # convert date to datetime for search domain, should be removed if using datetime directly
    start_datetime = fields.Datetime('Start datetime', compute='_compute_datetime')
    end_datetime = fields.Datetime('End datetime', compute='_compute_datetime')

    brand_id = fields.Many2one('res.brand', string='Thương hiệu')
    company_id = fields.Many2one(string='Chi nhánh', comodel_name='res.company',
                                   domain="[('brand_id', '=', brand_id),('name', 'not like', 'KHÔNG XÁC ĐỊNH')]")

    @api.depends('start_date', 'end_date')
    def _compute_datetime(self):
        self.start_datetime = False
        self.end_datetime = False
        if self.start_date and self.end_date:
            local_tz = timezone(self.env.user.tz or 'Etc/GMT+7')
            start_datetime = datetime(self.start_date.year, self.start_date.month, self.start_date.day, 0, 0, 0)
            end_datetime = datetime(self.end_date.year, self.end_date.month, self.end_date.day, 23, 59, 59)
            start_datetime = local_tz.localize(start_datetime, is_dst=None)
            end_datetime = local_tz.localize(end_datetime, is_dst=None)
            self.start_datetime = start_datetime.astimezone(utc).replace(tzinfo=None)
            self.end_datetime = end_datetime.astimezone(utc).replace(tzinfo=None)

    @api.onchange('start_date')
    def _onchange_start_date(self):
        if self.start_date:
            if self.start_date.month == fields.date.today().month:
                self.end_date = fields.date.today()
            else:
                self.end_date = date(self.start_date.year, self.start_date.month,
                                     monthrange(self.start_date.year, self.start_date.month)[1])

    @api.constrains('start_date', 'end_date')
    def check_dates(self):
        for record in self:
            start_date = fields.Date.from_string(record.start_date)
            end_date = fields.Date.from_string(record.end_date)
            if start_date > end_date:
                raise ValidationError(
                    _("Ngày kết thúc không thể ở trước ngày bắt đầu khi xuất báo cáo!!!"))

    def report_sale(self):
        template = self.env['ir.attachment'].browse(
            self.env.ref('crm_report.report_revenue_sale_attachment').id)
        decode = base64.b64decode(template.datas)
        wb = load_workbook(BytesIO(decode))
        ws = wb.active
        thin = borders.Side(style='thin')
        all_border_thin = borders.Border(left=thin, right=thin, top=thin, bottom=thin)
        big_data = {}
        # in tiêu đề và giá trị
        key_list = [
            'index',
            'code_staff',
            'name_staff',
            'department',
            'role_advise',
            'name_customer',
            'code_booking',
            'category_source_1',
            'category_source_2',
            'service',
            'quantity',
            'total',
            'type_customer',
            'service_point',
            'source_begin',
            'source_extend',
            'type_extend'
        ]
        key_list_title = [
            'STT',
            'Mã nhân viên',
            'Nhân viên tư vấn',
            'Bộ phận',
            'Vai trò tư vấn',
            'Tên khách hàng',
            'Mã booking',
            'Nhóm dịch vụ cấp 1',
            'Nhóm dịch vụ cấp 2',
            'Tên dịch vụ',
            'Số lượng đăng kí',
            'Số tiền',
            'Phân loại khách hàng',
            'Điểm dịch vụ',
            'Nguồn khách ban đầu',
            'Nguồn mở rộng',
            'Phân loại mở rộng'
        ]

        select_crm_sale_payment = """
        select cln.id, hr.employee_code, hr.name, hj.name, cln.consulting_role_1, rp.name, cl.name, hcs.his_service_type,
        pc.name, pt.name, cln.quantity, cln.total, cl.type_data_partner, hcs.kpi_point, us1.name, us2.name, cln.extensive_source_classification,
        csp.amount_proceeds, cln.number_used, cln.total_received
        from crm_sale_payment csp
        left join sh_medical_health_center_service hcs on hcs.id = csp.service_id 
        left join crm_lead cl on cl.id = csp.booking_id
        left join res_partner rp on rp.id = csp.partner_id
        left join sh_medical_health_center_service_category csc on csc.id = hcs.service_category
        left join product_category pc on pc.id = csc.product_cat_id
        left join utm_source us1 on us1.id = cl.original_source_id
        left join crm_line cln on cln.id = csp.crm_line_id
        left join res_users ru on ru.id = cln.consultants_1
        left join hr_employee hr on hr.id = ru.employee_id_store
        left join hr_job hj on hj.id = hr.job_id
        left join hr_group_job hgj on hgj.id = hj.group_job
        left join hr_department hd on hd.id = hr.department_id
        left join utm_source us2 on us2.id = cln.source_extend_id
        left join product_product pp on pp.id = hcs.product_id
        left join product_template pt on pt.id = pp.product_tmpl_id
        where csp.create_date >= %s and csp.create_date <= %s and csp.company_id = %s and csp.service_id = cln.service_id
        and cln.stage != 'cancel'
        """
        self.env.cr.execute(select_crm_sale_payment,
                            [self.start_datetime.strftime('%Y-%m-%d %H:%M:%S'),
                             self.end_datetime.strftime('%Y-%m-%d %H:%M:%S'),
                             self.company_id.id
                             ])

        datas = self.env.cr.fetchall()
        index = 1
        sum_csp = {}
        for data in datas:
            big_data[int(data[0])] = {}
            big_data[int(data[0])]['name_customer'] = data[5]
            big_data[int(data[0])]['code_booking'] = data[6]
            if data[7].lower() == 'spa':
                big_data[int(data[0])]['category_source_1'] = 'Spa'
            elif data[7].lower() == 'laser':
                big_data[int(data[0])]['category_source_1'] = 'Laser'
            elif data[7].lower() == 'odontology':
                big_data[int(data[0])]['category_source_1'] = 'Nha'
            elif data[7].lower() == 'surgery':
                big_data[int(data[0])]['category_source_1'] = 'Phẫu thuật'
            elif data[7].lower() == 'chiphi':
                big_data[int(data[0])]['category_source_1'] = 'Chi phí khác'
            big_data[int(data[0])]['category_source_2'] = data[8]
            if data[12] == 'old':
                big_data[int(data[0])]['type_customer'] = 'Khách hàng cũ'
            elif data[12] == 'new':
                big_data[int(data[0])]['type_customer'] = 'Khách hàng mới'
            else:
                big_data[int(data[0])]['type_customer'] = data[12]
            big_data[int(data[0])]['service_point'] = data[13]
            big_data[int(data[0])]['source_begin'] = data[14]
            big_data[int(data[0])]['index'] = index
            big_data[int(data[0])]['code_staff'] = data[1]
            big_data[int(data[0])]['name_staff'] = data[2]
            big_data[int(data[0])]['department'] = data[3]
            if data[4] == '1':
                big_data[int(data[0])]['role_advise'] = 'Tư vấn độc lập'
            elif data[4] == '2':
                big_data[int(data[0])]['role_advise'] = 'Tư vấn đồng thời'
            elif data[4] == '3':
                big_data[int(data[0])]['role_advise'] = 'Lễ tân - CVTV cùng tư vấn'
            elif data[4] == '4':
                big_data[int(data[0])]['role_advise'] = 'BS da liễu - KTV cùng tư vấn'
            elif data[4] == '5':
                big_data[int(data[0])]['role_advise'] = 'Tư vấn chính'
            elif data[4] == '6':
                big_data[int(data[0])]['role_advise'] = 'Tư vấn phụ'
            else:
                big_data[int(data[0])]['role_advise'] = data[4]
            big_data[int(data[0])]['service'] = data[9]
            big_data[int(data[0])]['quantity'] = data[10]
            if data[12] == 'new':
                big_data[int(data[0])]['type_customer'] = 'Khách hàng mới'
            elif data[12] == 'old':
                big_data[int(data[0])]['type_customer'] = 'Khách hàng cũ'
            big_data[int(data[0])]['source_extend'] = data[15]
            if data[16] == 'ext01':
                big_data[int(data[0])]['type_extend'] = 'Mở rộng_Dịch vụ trong Phòng/Bộ phận'
            elif data[16] == 'ext02':
                big_data[int(data[0])]['type_extend'] = 'Bán chéo_Dịch vụ ngoài Phòng/Bộ phận'
            elif data[16] == 'ext03':
                big_data[int(data[0])]['type_extend'] = 'Upsale_Thay đổi dịch vụ'
            elif data[16] == 'ext04':
                big_data[int(data[0])]['type_extend'] = 'Mở rộng_Từ KH tái khám'
            elif data[16] == 'ext05':
                big_data[int(data[0])]['type_extend'] = 'Mở rộng_Từ KH khiếu nại'
            elif data[16] == 'ext06':
                big_data[int(data[0])]['type_extend'] = 'Mở rộng_Từ KH bảo hành'
            else:
                big_data[int(data[0])]['type_extend'] = data[16]
            if int(data[0]) not in sum_csp:
                sum_csp[int(data[0])] = data[17]
            else:
                sum_csp[int(data[0])] += data[17]

            if data[11] == sum_csp[int(data[0])]:
                big_data[int(data[0])]['total'] = data[11]
            elif sum_csp[int(data[0])] < data[11] and data[18] >= 1 and data[19] < data[11]:
                big_data[int(data[0])]['total'] = data[11] / data[10]
            elif data[19] == data[11] > sum_csp[int(data[0])]:
                big_data[int(data[0])]['total'] = data[11] - data[11] / data[10]
            else:
                big_data[int(data[0])]['total'] = 0
            index += 1
        format_decimal = '#,##0.00'
        row = 3
        for data in big_data:
            col = 1
            for line in key_list:
                cell = ws.cell(row, col)
                cell.value = big_data[data][line]
                cell.font = Font(name='Times New Roman', size=13)
                cell.border = all_border_thin
                if line == 'total':
                    cell.number_format = format_decimal
                elif line in ['index', 'code_staff', 'code_booking','category_source_1','quantity','service_point']:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                col += 1
            row += 1

        fp = BytesIO()
        wb.save(fp)
        fp.seek(0)
        report = base64.encodebytes((fp.read()))
        fp.close()
        attachment = self.env['ir.attachment'].with_user(1).create({
            'name': 'Bao_cao_doanh_thu_ban_san_pham.xlsx',
            'datas': report,
            'res_model': 'temp.creation',
            'public': True,
        })
        url = "/web/content/?model=ir.attachment&id=%s&filename_field=name&field=datas&download=true" \
              % attachment.id

        return {
            'name': 'Báo cáo doanh thu bán sản phẩm',
            'type': 'ir.actions.act_url',
            'url': url,
            'target': 'self',
        }