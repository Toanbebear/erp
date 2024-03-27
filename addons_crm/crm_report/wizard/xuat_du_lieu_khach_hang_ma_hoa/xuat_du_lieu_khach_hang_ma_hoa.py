import hashlib

from odoo import fields, api, models, _
from odoo.exceptions import ValidationError, UserError
from datetime import date, datetime, time
from calendar import monthrange
from openpyxl import load_workbook
from openpyxl.styles import Font, borders, Alignment, PatternFill

import base64
from io import BytesIO
from dateutil.relativedelta import relativedelta
from pytz import timezone, utc
from copy import copy
import requests


class ExportCustomerEncode(models.TransientModel):
    _name = 'export.customer.encode'
    _description = 'Xuất dữ liệu '

    def domain_stage(self):
        crm_type = self.env['crm.type'].sudo().browse(1)
        return [('id', 'in', crm_type.stage_id.ids)]

    start_date = fields.Date('Start date', default=date.today())
    end_date = fields.Date('End date', default=date.today())
    start_datetime = fields.Datetime('Start datetime', compute='_compute_datetime')
    end_datetime = fields.Datetime('End datetime', compute='_compute_datetime')
    brand_id = fields.Many2one('res.brand', string='Thương hiệu')
    stage_id = fields.Many2many('crm.stage', string='Trạng thái', domain=domain_stage)

    @api.depends('start_date', 'end_date')
    def _compute_datetime(self):
        self.start_datetime = False
        self.end_datetime = False
        if self.start_date and self.end_date:
            local_tz = timezone(self.env.user.tz or 'Etc/GMT+7')
            start_datetime = datetime(self.start_date.year, self.start_date.month, self.start_date.day, 0, 0, 0)
            end_datetime = datetime(self.end_date.year, self.end_date.month, self.end_date.day, 23, 59, 59)
            start_datetime = local_tz.localize(start_datetime, is_dst=None)
            end_datetime = local_tz.localize(end_datetime, is_dst=None)
            self.start_datetime = start_datetime.astimezone(utc).replace(tzinfo=None)
            self.end_datetime = end_datetime.astimezone(utc).replace(tzinfo=None)

    @api.constrains('start_date', 'end_date')
    def check_dates(self):
        for record in self:
            start_date = fields.Date.from_string(record.start_date)
            end_date = fields.Date.from_string(record.end_date)
            if start_date > end_date:
                raise ValidationError(
                    _("End Date cannot be set before Start Date."))

    def report(self):
        template = self.env['ir.attachment'].browse(self.env.ref('crm_report.export_customer_encode_attachment').id)
        decode = base64.b64decode(template.datas)
        wb = load_workbook(BytesIO(decode))
        ws = wb.active
        thin = borders.Side(style='thin')
        all_border_thin = borders.Border(left=thin, right=thin, top=thin, bottom=thin)
        line_font = Font(name='Times New Roman', size=13)
        datas = []

        select = """
        select cld.id, cld.email_from, cld.phone, cs.name, cld.create_date, cld.id, cld.id
        from crm_lead cld
        left join crm_stage cs on cs.id = cld.stage_id
        where cld.brand_id = %s and cs.id = %s and cld.create_date >= %s and cld.create_date <= %s and cld.type='lead'
        """
        self.env.cr.execute(select,
                            (self.brand_id.id, tuple(self.stage_id.ids), self.start_datetime, self.end_datetime))
        result = self.env.cr.fetchall()
        for c in result:
            datas.append(c)
        index = 1
        row = 2
        for data in datas:
            col = 1
            for value in data:
                cell = ws.cell(row, col)
                if col == 1:
                    cell.value = index
                elif col == 3:
                    phone = '+84' + value[1:] if value else ''
                    hash_object = hashlib.sha256()
                    hash_object.update(phone.encode('utf-8'))
                    cell.value = hash_object.hexdigest()
                elif col == 5:
                    cell.value = fields.Datetime.context_timestamp(self.with_context(tz=self.env.user.tz),value).strftime('%d/%m/%Y %H:%M:%S')
                elif col in (6, 7):
                    cell.value = ''
                else:
                    cell.value = value
                cell.font = line_font
                cell.border = all_border_thin
                if col == 1:
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                else:
                    cell.alignment = Alignment(wrap_text=True)
                col += 1
            index += 1
            row += 1
        fp = BytesIO()
        wb.save(fp)
        fp.seek(0)
        report = base64.encodebytes((fp.read()))
        fp.close()
        attachment = self.env['ir.attachment'].sudo().create({
            'name': 'Xuất dữ liệu khách hàng mã hóa.xlsx',
            'datas': report,
            'res_model': 'temp.creation',
            'public': True,
        })
        url = "/web/content/?model=ir.attachment&id=%s&filename_field=name&field=datas&download=true" \
              % attachment.id
        return {'name': 'Xuất dữ liệu khách hàng mã hóa.xlsx',
                'type': 'ir.actions.act_url',
                'url': url,
                'target': 'self',
                }