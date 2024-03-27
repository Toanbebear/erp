import base64
from calendar import monthrange
from datetime import date, datetime
from io import BytesIO

from openpyxl import load_workbook
from openpyxl.styles import Font, borders, Alignment
from pytz import timezone, utc

from odoo import fields, api, models, _
from odoo.exceptions import ValidationError

thin = borders.Side(style='thin')
double = borders.Side(style='double')
all_border_thin = borders.Border(thin, thin, thin, thin)


class CtvReport(models.TransientModel):
    _name = 'crm.ctv.report'
    _description = 'Báo cáo chi tiết doanh số'

    type_report = fields.Selection(
        [('type_1', 'Báo cáo tổng hợp'), ('type_2', 'Báo cáo chi tiết')], string='Loại báo cáo', default="type_1"
    )
    # type_report = fields.Selection([('type_1', 'Báo cáo chi tiết doanh số')], string='Loại báo cáo')
    company_id = fields.Many2one('res.company', string='Công ty', default=lambda self: self.env.company)
    collaborator_id = fields.Many2one('collaborator.collaborator', string='Cộng tác viên')
    start_date = fields.Date('Ngày bắt đầu', default=date.today().replace(day=1))
    end_date = fields.Date('Ngày kết thúc')
    # convert date to datetime for search domain, should be removed if using datetime directly
    start_datetime = fields.Datetime('Start datetime', compute='_compute_datetime')
    end_datetime = fields.Datetime('End datetime', compute='_compute_datetime')

    # check ngày
    @api.depends('start_date', 'end_date')
    def _compute_datetime(self):
        self.start_datetime = False
        self.end_datetime = False
        if self.start_date and self.end_date:
            local_tz = timezone(self.env.user.tz or 'Etc/GMT+7')
            start_datetime = datetime(self.start_date.year, self.start_date.month, self.start_date.day, 0, 0, 0)
            end_datetime = datetime(self.end_date.year, self.end_date.month, self.end_date.day, 23, 59, 59)
            start_datetime = local_tz.localize(start_datetime, is_dst=None)
            end_datetime = local_tz.localize(end_datetime, is_dst=None)
            self.start_datetime = start_datetime.astimezone(utc).replace(tzinfo=None)
            self.end_datetime = end_datetime.astimezone(utc).replace(tzinfo=None)

    @api.onchange('start_date')
    def _onchange_start_date(self):
        if self.start_date:
            if self.start_date.month == fields.date.today().month:
                self.end_date = fields.date.today()
            else:
                self.end_date = date(self.start_date.year, self.start_date.month,
                                     monthrange(self.start_date.year, self.start_date.month)[1])

    @api.constrains('start_date', 'end_date')
    def check_dates(self):
        for record in self:
            start_date = fields.Date.from_string(record.start_date)
            end_date = fields.Date.from_string(record.end_date)
            if start_date > end_date:
                raise ValidationError(
                    _("Ngày kết thúc không thể ở trước ngày bắt đầu khi xuất báo cáo!!!"))

    @api.onchange('type_report')
    def check_type_report(self):
        if self.type_report == 'type_1':
            self.collaborator_id = False

    # # # #lấy dữ liệu
    def _get_data_report(self):
        # Phan loai kết quả
        dict_stage = {
            'draft': 'Nháp',
            'open': 'Mở lại',
            'new': 'Có hiệu lực',
            'done': 'Hết hiệu lực',
            'cancel': 'Đã hủy',
        }

        if self.type_report == 'type_2':
            if self.collaborator_id:
                ret_data = []
                domain = [('service_date', '>=', self.start_date), ('service_date', '<=', self.end_date),
                          ('company_id', '=', self.company_id.id), ('collaborator_id', '=', self.collaborator_id.id)]
                detail = self.env['crm.detail.sale'].search(domain)

                stt = 1
                for record in detail:
                    ret_data.append({
                        "stt": stt,
                        "thuong_hieu": record.brand_id.name,
                        "chi_nhanh": record.company_id.name,
                        "ma_ctv": record.collaborator_id.code,
                        "ten_ctv": record.collaborator_id.name,
                        "ma_hop_dong": record.contract_id.default_code,
                        "ngay_bat_dau": record.contract_id.start_date.strftime('%d/%m/%Y'),
                        "ngay_ket_thuc": record.contract_id.end_date.strftime('%d/%m/%Y'),
                        "trang_thai_hop_dong": dict_stage['%s' % record.contract_id.stage],
                        "khach_hang": record.sale_order.partner_id.name,
                        "booking": record.booking_id.name,
                        "phieu_kham": record.walkin_id.name,
                        "so": record.sale_order.name,
                        "dich_vu_su_dung": record.service_id.name,
                        "ngay_thuc_hien": record.walkin_id.service_date_start.strftime('%d/%m/%Y'),
                        "ngay_hoan_thanh": record.service_date.strftime('%d/%m/%Y'),
                        "don_gia": record.amount_total,
                        "%_hoa_hong": record.discount_percent,
                        "tien_hoa_hong": record.amount_used,
                        # "da_chi": record.amount_paid,
                        # "con_lai": record.amount_remain,
                        # "hinh_thuc_thanh_toan":
                        "ngan_hang": record.collaborator_id.bank if record.collaborator_id.bank else None,
                        "so_tai_khoan": record.collaborator_id.card_number if record.collaborator_id.card_number else None,
                        "ten_chu_tai_khoan": record.collaborator_id.user_bank if record.collaborator_id.user_bank else None,
                        "ghi_chu": None,
                    })
                    stt += 1
                sum_amount_don_gia = sum([element['don_gia'] for element in ret_data])
                sum_amount_tien_hoa_hong = sum([element['tien_hoa_hong'] for element in ret_data])
                # sum_amount_da_chi = sum([element['da_chi'] for element in ret_data])
                # sum_amount_con_lai = sum([element['con_lai'] for element in ret_data])
                val = {
                    "stt": None,
                    "thuong_hieu": None,
                    "chi_nhanh": None,
                    "ma_ctv": None,
                    "ten_ctv": None,
                    "ma_hop_dong": None,
                    "ngay_bat_dau": None,
                    "ngay_ket_thuc": None,
                    "trang_thai_hop_dong": None,
                    "khach_hang": None,
                    "booking": None,
                    "phieu_kham": None,
                    "so": None,
                    "dich_vu_su_dung": None,
                    "ngay_thuc_hien": None,
                    "ngay_hoan_thanh": 'TỔNG CỘNG:',
                    "don_gia": sum_amount_don_gia,
                    "%_hoa_hong": None,
                    "tien_hoa_hong": sum_amount_tien_hoa_hong,
                    # "da_chi": sum_amount_da_chi,
                    # "con_lai": sum_amount_con_lai,
                    # "hinh_thuc_thanh_toan": None,
                    "ngan_hang": None,
                    "so_tai_khoan": None,
                    "ten_chu_tai_khoan": None,
                    "ghi_chu": None,
                }
                ret_data.append(val)
                return ret_data
        else:
            ret_data = []
            domain = [('service_date', '>=', self.start_date), ('service_date', '<=', self.end_date),
                      ('company_id', '=', self.company_id.id)]
            detail = self.env['crm.detail.sale'].search(domain)

            stt = 1
            for record in detail:
                ret_data.append({
                    "stt": stt,
                    "thuong_hieu": record.brand_id.name,
                    "chi_nhanh": record.company_id.name,
                    "ma_ctv": record.collaborator_id.code,
                    "ten_ctv": record.collaborator_id.name,
                    "ma_hop_dong": record.contract_id.default_code,
                    "ngay_bat_dau": record.contract_id.start_date.strftime('%d/%m/%Y'),
                    "ngay_ket_thuc": record.contract_id.end_date.strftime('%d/%m/%Y'),
                    "trang_thai_hop_dong": dict_stage['%s' % record.contract_id.stage],
                    "khach_hang": record.sale_order.partner_id.name,
                    "booking": record.booking_id.name,
                    "phieu_kham": record.walkin_id.name,
                    "so": record.sale_order.name,
                    "dich_vu_su_dung": record.service_id.name if record.service_id else None,
                    "ngay_thuc_hien": record.walkin_id.service_date_start.strftime(
                        '%d/%m/%Y') if record.walkin_id.service_date_start else None,
                    "ngay_hoan_thanh": record.service_date.strftime('%d/%m/%Y') if record.service_date else None,
                    "don_gia": record.amount_total,
                    "%_hoa_hong": record.discount_percent,
                    "tien_hoa_hong": record.amount_used,
                    # "da_chi": record.amount_paid,
                    # "con_lai": record.amount_remain,
                    # "hinh_thuc_thanh_toan":
                    "ngan_hang": record.collaborator_id.bank if record.collaborator_id.bank else None,
                    "so_tai_khoan": record.collaborator_id.card_number if record.collaborator_id.card_number else None,
                    "ten_chu_tai_khoan": record.collaborator_id.user_bank if record.collaborator_id.user_bank else None,
                    "ghi_chu": None,
                })
                stt += 1
            sum_amount_don_gia = sum([element['don_gia'] for element in ret_data])
            sum_amount_tien_hoa_hong = sum([element['tien_hoa_hong'] for element in ret_data])
            # sum_amount_da_chi = sum([element['da_chi'] for element in ret_data])
            # sum_amount_con_lai = sum([element['con_lai'] for element in ret_data])
            val = {
                "stt": None,
                "thuong_hieu": None,
                "chi_nhanh": None,
                "ma_ctv": None,
                "ten_ctv": None,
                "ma_hop_dong": None,
                "ngay_bat_dau": None,
                "ngay_ket_thuc": None,
                "trang_thai_hop_dong": None,
                "khach_hang": None,
                "booking": None,
                "phieu_kham": None,
                "so": None,
                "dich_vu_su_dung": None,
                "ngay_thuc_hien": None,
                "ngay_hoan_thanh": 'TỔNG CỘNG:',
                "don_gia": sum_amount_don_gia,
                "%_hoa_hong": None,
                "tien_hoa_hong": sum_amount_tien_hoa_hong,
                # "da_chi": sum_amount_da_chi,
                # "con_lai": sum_amount_con_lai,
                # "hinh_thuc_thanh_toan": None,
                "ngan_hang": None,
                "so_tai_khoan": None,
                "ten_chu_tai_khoan": None,
                "ghi_chu": None,
            }
            ret_data.append(val)
            return ret_data
        # else:
        #     # Phan loai kết quả
        #     dict_stage = {
        #         'draft': 'Nháp',
        #         'open': 'Mở lại',
        #         'new': 'Có hiệu lực',
        #         'done': 'Hết hiệu lực',
        #         'cancel': 'Đã hủy',
        #     }
        #
        #     ret_data = []
        #     """
        #            Đứng ở SO lọc ra tất cả bản ghi thảo mãn trong khoảng thời gian, và có trạng thái là sale
        #            Lọc ra các SO có cộng tác viên trong Booking bằng với công tác viên cần tìm kiếm
        #            :return:
        #            """
        #     domain = [('date_order', '>=', self.start_datetime), ('date_order', '<=', self.end_datetime),
        #               ('state', 'in', ('sale', 'done'))]
        #     SO = self.env['sale.order'].search(domain)
        #
        #     # ctv = self.env['collaborator.collaborator'].search([('company_id', '=', self.company_id.id)])
        #     contract = self.env['collaborator.contract'].search(
        #         [('company_id', '=', self.company_id.id), ('stage', 'in', ('new', 'open', 'done'))])
        #
        #     stt = 1
        #     for hd in contract:
        #         # lấy ra loại hợp đồng
        #         services = hd.contract_type_id.filtered(lambda rec: rec.state in ('new', 'open'))
        #
        #         for ctv in hd.collaborator_id:
        #             # Lấy ra SO
        #             list_sale_order = SO.filtered(lambda rec: rec.booking_id.collaborator_id.id == ctv.id)
        #
        #             # duyệt qua dịch vụ đã kí
        #             for so in list_sale_order:
        #                 sols = [(s.product_id.default_code, s) for s in so.order_line]
        #                 for line in services:
        #                     list_service = list()
        #                     "duyệt qua dịch vụ trong hợp đồng"
        #                     for ser in line.service_id:
        #                         list_service.append(ser.default_code)
        #                     "duyệt qua dịch vụ trong S0"
        #                     for sol in sols:
        #                         if sol[0] in list_service:
        #                             ret_data.append({
        #                                 "stt": stt,
        #                                 "ma_ctv": ctv.code,
        #                                 "ten_ctv": ctv.collaborator,
        #                                 "ma_hop_dong": hd.default_code,
        #                                 "ngay_bat_dau": hd.start_date.strftime('%d/%m/%Y'),
        #                                 "ngay_ket_thuc": hd.end_date.strftime('%d/%m/%Y'),
        #                                 "trang_thai": dict_stage['%s' % hd.stage],
        #                                 "khach_hang": so.booking_id.contact_name,
        #                                 "booking": so.booking_id.name,
        #                                 "dich_vu": sol[1].name,
        #                                 "don_ban_hang": so.name,
        #                                 "ngay_hoan_thanh": so.date_order.strftime('%d/%m/%Y'),
        #                                 "don_gia": sol[1].price_unit,
        #                                 "hoa_hong": line.discount_percent,
        #                                 "tien_hoa_hong": sol[1].price_unit * line.discount_percent / 100,
        #                                 "ngan_hang": ctv.bank or None,
        #                                 "so_tai_khoan": ctv.card_number or None,
        #                                 "ghi_chu": hd.description or None,
        #                             })
        #                             stt += 1
        #     sum_amount = sum([element['tien_hoa_hong'] for element in ret_data])
        #     val = {
        #         "stt": None,
        #         "ma_ctv": None,
        #         "ten_ctv": None,
        #         "ma_hop_dong": None,
        #         "ngay_bat_dau": None,
        #         "ngay_ket_thuc": None,
        #         "trang_thai": None,
        #         "khach_hang": None,
        #         "booking": None,
        #         "dich_vu": None,
        #         "don_ban_hang": None,
        #         "ngay_hoan_thanh": None,
        #         "don_gia": None,
        #         "hoa_hong": 'Tổng Tiền hoa hồng:',
        #         "tien_hoa_hong": sum_amount,
        #         "ngan_hang": None,
        #         "so_tai_khoan": None,
        #         "ghi_chu": None,
        #     }
        #     ret_data.append(val)
        #     return ret_data

    # báo cáo chi tiết doanh số
    # def create_report(self):
    #     datas = self._get_data_report()
    #     # in du lieu
    #     report_brand_overview_attachment = self.env['ir.attachment'].browse(self.env.ref('collaborator.bao_cao_doanh_thu_hoa_hong_attachment').id)
    #     decode = base64.b64decode(report_brand_overview_attachment.datas)
    #     wb = load_workbook(BytesIO(decode))
    #     ws = wb.active
    #     line_font = Font(name='Times New Roman', size=12)
    #
    #     ws['A1'].value = self.company_id.name
    #     ws['E3'].value += self.start_date.strftime('%d/%m/%Y')
    #     ws['G3'].value += self.end_datetime.strftime('%d/%m/%Y')
    #
    #     key_col_list = list(range(1, 19))
    #     key_list = [
    #         "stt",
    #         "ma_ctv",
    #         "ten_ctv",
    #         "ma_hop_dong",
    #         "ngay_bat_dau",
    #         "ngay_ket_thuc",
    #         "trang_thai",
    #         "khach_hang",
    #         "booking",
    #         "dich_vu",
    #         "don_ban_hang",
    #         "ngay_hoan_thanh",
    #         "don_gia",
    #         "hoa_hong",
    #         "tien_hoa_hong",
    #         "ngan_hang",
    #         "so_tai_khoan",
    #         "ghi_chu",
    #     ]
    #     row = 6
    #     # tong = 0
    #     for data in datas:
    #         for col, k in zip(key_col_list, key_list):
    #             beforeCell = ws.cell(4, col)
    #             beforeCell.font = Font(name='Times New Roman', size=12, color='FFFFFF')
    #             cell = ws.cell(row, col)
    #             cell.value = data[k]
    #             cell.font = line_font
    #             cell.border = all_border_thin
    #             # if col == 9:
    #             #     tong += data['tien_hoa_hong']
    #
    #             cell.alignment = Alignment(horizontal='center', vertical='center')
    #             ws['M' + str(row)].number_format = '#,##0" đ"'
    #             ws['O' + str(row)].number_format = '#,##0" đ"'
    #             ws['N' + str(row)].number_format = '0.0"%"'
    #
    #         row += 1
    #
    #     fp = BytesIO()
    #     wb.save(fp)
    #     fp.seek(0)
    #     report = base64.encodebytes((fp.read()))
    #     fp.close()
    #     attachment = self.env['ir.attachment'].sudo().create({
    #         'name': 'bao_cao_doanh_so_ctv',
    #         'datas': report,
    #         'res_model': 'temp.creation',
    #         'public': True,
    #     })
    #     url = "/web/content/?model=ir.attachment&id=%s&filename_field=name&field=datas&download=true" \
    #           % attachment.id
    #     return {'name': 'BÁO CÁO DOANH SỐ CTV',
    #             'type': 'ir.actions.act_url',
    #             'url': url,
    #             'target': 'self',
    #             }

    # báo cáo chi tiết giao dịch
    def create_report(self):
        datas = self._get_data_report()
        # in du lieu
        report_brand_overview_attachment = self.env['ir.attachment'].browse(
            self.env.ref('collaborator.bao_cao_chi_tiet_doanh_so_attachment').id)
        decode = base64.b64decode(report_brand_overview_attachment.datas)
        wb = load_workbook(BytesIO(decode))
        ws = wb.active
        line_font = Font(name='Times New Roman', size=12)

        ws['A1'].value = self.company_id.name
        ws['E3'].value += self.start_date.strftime('%d/%m/%Y')
        ws['F3'].value += self.end_datetime.strftime('%d/%m/%Y')

        key_col_list = list(range(1, 24))
        key_list = [
            "stt",
            "thuong_hieu",
            "chi_nhanh",
            "ma_ctv",
            "ten_ctv",
            "ma_hop_dong",
            "ngay_bat_dau",
            "ngay_ket_thuc",
            "trang_thai_hop_dong",
            "khach_hang",
            "booking",
            "phieu_kham",
            "so",
            "dich_vu_su_dung",
            "ngay_thuc_hien",
            "ngay_hoan_thanh",
            "don_gia",
            "%_hoa_hong",
            "tien_hoa_hong",
            "ngan_hang",
            "so_tai_khoan",
            "ten_chu_tai_khoan",
            "ghi_chu",
        ]
        row = 6
        # tong = 0
        for data in datas:
            for col, k in zip(key_col_list, key_list):
                beforeCell = ws.cell(4, col)
                beforeCell.font = Font(name='Times New Roman', size=12, color='FFFFFF')
                cell = ws.cell(row, col)
                cell.value = data[k]
                cell.font = line_font
                cell.border = all_border_thin
                # if col == 9:
                #     tong += data['tien_hoa_hong']

                cell.alignment = Alignment(horizontal='center', vertical='center')
                ws['Q' + str(row)].number_format = '#,##0" đ"'
                ws['S' + str(row)].number_format = '#,##0" đ"'
            row += 1

        fp = BytesIO()
        wb.save(fp)
        fp.seek(0)
        report = base64.encodebytes((fp.read()))
        fp.close()
        attachment = self.env['ir.attachment'].sudo().create({
            'name': 'BÁO CÁO CHI TIẾT DOANH SỐ HOA HỒNG CTV',
            'datas': report,
            'res_model': 'temp.creation',
            'public': True,
        })
        url = "/web/content/?model=ir.attachment&id=%s&filename_field=name&field=datas&download=true" \
              % attachment.id
        return {'name': 'BÁO CÁO CHI TIẾT DOANH SỐ HOA HỒNG CTV',
                'type': 'ir.actions.act_url',
                'url': url,
                'target': 'self',
                }
