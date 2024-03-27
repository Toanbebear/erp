import base64
from calendar import monthrange
from datetime import date, datetime
from io import BytesIO

from openpyxl import load_workbook
from openpyxl.styles import Font, borders, Alignment
from pytz import timezone, utc

from odoo import fields, api, models, _
from odoo.exceptions import ValidationError

thin = borders.Side(style='thin')
double = borders.Side(style='double')
all_border_thin = borders.Border(thin, thin, thin, thin)


class CollaboratorReportPayment(models.TransientModel):
    _name = 'collaborator.collaborator.payment.report'
    _description = 'Báo cáo hoa hồng cộng tác viên'

    type_report = fields.Selection(
        [('type_1', 'Tổng hợp'), ('type_2', 'Báo cáo chi tiết')], string='Loại báo cáo', default="type_1")
    company_id = fields.Many2one('res.company', string='Công ty', default=lambda self: self.env.company)
    collaborator_id = fields.Many2one('collaborator.collaborator', string='Cộng tác viên')
    start_date = fields.Date('Ngày bắt đầu', default=date.today().replace(day=1))
    end_date = fields.Date('Ngày kết thúc')
    # convert date to datetime for search domain, should be removed if using datetime directly
    start_datetime = fields.Datetime('Start datetime', compute='_compute_datetime')
    end_datetime = fields.Datetime('End datetime', compute='_compute_datetime')

    # check ngày
    @api.depends('start_date', 'end_date')
    def _compute_datetime(self):
        self.start_datetime = False
        self.end_datetime = False
        if self.start_date and self.end_date:
            local_tz = timezone(self.env.user.tz or 'Etc/GMT+7')
            start_datetime = datetime(self.start_date.year, self.start_date.month, self.start_date.day, 0, 0, 0)
            end_datetime = datetime(self.end_date.year, self.end_date.month, self.end_date.day, 23, 59, 59)
            start_datetime = local_tz.localize(start_datetime, is_dst=None)
            end_datetime = local_tz.localize(end_datetime, is_dst=None)
            self.start_datetime = start_datetime.astimezone(utc).replace(tzinfo=None)
            self.end_datetime = end_datetime.astimezone(utc).replace(tzinfo=None)

    @api.onchange('start_date')
    def _onchange_start_date(self):
        if self.start_date:
            if self.start_date.month == fields.date.today().month:
                self.end_date = fields.date.today()
            else:
                self.end_date = date(self.start_date.year, self.start_date.month,
                                     monthrange(self.start_date.year, self.start_date.month)[1])

    @api.constrains('start_date', 'end_date')
    def check_dates(self):
        for record in self:
            start_date = fields.Date.from_string(record.start_date)
            end_date = fields.Date.from_string(record.end_date)
            if start_date > end_date:
                raise ValidationError(
                    _("Ngày kết thúc không thể ở trước ngày bắt đầu khi xuất báo cáo!!!"))

    @api.onchange('type_report')
    def check_type_report(self):
        if self.type_report == 'type_1':
            self.collaborator_id = False

    # # # #lấy dữ liệu
    def _get_data_report(self):
        # Phan loai kết quả
        dict_stage = {
            'draft': 'Nháp',
            'open': 'Mở lại',
            'new': 'Có hiệu lực',
            'done': 'Hết hiệu lực',
            'cancel': 'Đã hủy',
        }

        dict_method = {
            'tm': 'Tiền mặt',
            'ck': 'Chuyển khoản',
        }

        if self.type_report == 'type_2':
            if self.collaborator_id:
                ret_data = []
                domain = [('ctv_payment_date', '>=', self.start_date), ('ctv_payment_date', '<=', self.end_date),
                          ('company_id', '=', self.company_id.id),
                          ('collaborator_id', '=', self.collaborator_id.id), ('state', '=', 'posted')]
                ctv = self.env['account.payment.ctv'].search(domain)
                stt = 1
                for record in ctv:
                    ret_data.append({
                        "stt": stt,
                        "ma_phieu_chi": record.name,
                        "ngay_chi": record.ctv_payment_date.strftime('%d/%m/%Y'),
                        "hinh_thuc_chi": dict_method['%s' % record.ctv_payment_method],
                        "thuong_hieu": record.brand_id.name,
                        "chi_nhanh": record.company_id.name,
                        "ma_ctv": record.collaborator_id.code,
                        "ten_ctv": record.collaborator_id.name,
                        "ma_hop_dong": record.contract_id.default_code if record.contract_id else None,
                        "ngay_bat_dau": record.contract_id.start_date.strftime(
                            '%d/%m/%Y') if record.contract_id else None,
                        "ngay_ket_thuc": record.contract_id.end_date.strftime(
                            '%d/%m/%Y') if record.contract_id else None,
                        "trang_thai_hop_dong": dict_stage[
                            '%s' % record.contract_id.stage] if record.contract_id else None,
                        "da_chi": record.amount,
                        "con_lai": None,
                        "ngan_hang": record.collaborator_id.bank if record.collaborator_id.bank else None,
                        "so_tai_khoan": record.collaborator_id.card_number if record.collaborator_id.card_number else None,
                        "ten_chu_tai_khoan": record.collaborator_id.user_bank if record.collaborator_id.user_bank else None,
                        "ghi_chu": None,
                    })
                    stt += 1
                sum_amount_da_chi = sum([element['da_chi'] for element in ret_data])
                val = {
                    "stt": None,
                    "ma_phieu_chi": None,
                    "ngay_chi": None,
                    "hinh_thuc_chi": None,
                    "thuong_hieu": None,
                    "chi_nhanh": None,
                    "ma_ctv": None,
                    "ten_ctv": None,
                    "ma_hop_dong": None,
                    "ngay_bat_dau": None,
                    "ngay_ket_thuc": None,
                    "trang_thai_hop_dong": 'Tổng:',
                    "da_chi": sum_amount_da_chi,
                    "con_lai": None,
                    "ngan_hang": None,
                    "so_tai_khoan": None,
                    "ten_chu_tai_khoan": None,
                    "ghi_chu": None,
                }
                ret_data.append(val)
                return ret_data
        else:
            ret_data = []
            domain = [('ctv_payment_date', '>=', self.start_date), ('ctv_payment_date', '<=', self.end_date),
                      ('company_id', '=', self.company_id.id), ('state', '=', 'posted')]
            ctv = self.env['account.payment.ctv'].search(domain)
            stt = 1
            for record in ctv:
                ret_data.append({
                    "stt": stt,
                    "ma_phieu_chi": record.name,
                    "ngay_chi": record.ctv_payment_date.strftime('%d/%m/%Y'),
                    "hinh_thuc_chi": dict_method['%s' % record.ctv_payment_method],
                    "thuong_hieu": record.brand_id.name,
                    "chi_nhanh": record.company_id.name,
                    "ma_ctv": record.collaborator_id.code,
                    "ten_ctv": record.collaborator_id.name,
                    "ma_hop_dong": record.contract_id.default_code if record.contract_id else None,
                    "ngay_bat_dau": record.contract_id.start_date.strftime('%d/%m/%Y') if record.contract_id else None,
                    "ngay_ket_thuc": record.contract_id.end_date.strftime('%d/%m/%Y') if record.contract_id else None,
                    "trang_thai_hop_dong": dict_stage['%s' % record.contract_id.stage] if record.contract_id else None,
                    "da_chi": record.amount,
                    "con_lai": None,
                    "ngan_hang": record.collaborator_id.bank if record.collaborator_id.bank else None,
                    "so_tai_khoan": record.collaborator_id.card_number if record.collaborator_id.card_number else None,
                    "ten_chu_tai_khoan": record.collaborator_id.user_bank if record.collaborator_id.user_bank else None,
                    "ghi_chu": None,
                })
                stt += 1
            sum_amount_da_chi = sum([element['da_chi'] for element in ret_data])
            val = {
                "stt": None,
                "ma_phieu_chi": None,
                "ngay_chi": None,
                "hinh_thuc_chi": None,
                "thuong_hieu": None,
                "chi_nhanh": None,
                "ma_ctv": None,
                "ten_ctv": None,
                "ma_hop_dong": None,
                "ngay_bat_dau": None,
                "ngay_ket_thuc": None,
                "trang_thai_hop_dong": 'Tổng:',
                "da_chi": sum_amount_da_chi,
                "con_lai": None,
                "ngan_hang": None,
                "so_tai_khoan": None,
                "ten_chu_tai_khoan": None,
                "ghi_chu": None,
            }
            ret_data.append(val)
            return ret_data

    def create_report(self):
        datas = self._get_data_report()
        # in du lieu
        report_brand_overview_attachment = self.env['ir.attachment'].browse(
            self.env.ref('collaborator.bao_cao_hoa_hong_cong_tac_vien_attachment').id)
        decode = base64.b64decode(report_brand_overview_attachment.datas)
        wb = load_workbook(BytesIO(decode))
        ws = wb.active
        line_font = Font(name='Times New Roman', size=12)

        ws['A1'].value = self.company_id.name
        ws['F3'].value += self.start_date.strftime('%d/%m/%Y')
        ws['G3'].value += self.end_datetime.strftime('%d/%m/%Y')

        key_col_list = list(range(1, 19))
        key_list = [
            "stt",
            "ma_phieu_chi",
            "ngay_chi",
            "hinh_thuc_chi",
            "thuong_hieu",
            "chi_nhanh",
            "ma_ctv",
            "ten_ctv",
            "ma_hop_dong",
            "ngay_bat_dau",
            "ngay_ket_thuc",
            "trang_thai_hop_dong",
            "da_chi",
            "con_lai",
            "ngan_hang",
            "so_tai_khoan",
            "ten_chu_tai_khoan",
            "ghi_chu",
        ]
        row = 6
        # tong = 0
        for data in datas:
            for col, k in zip(key_col_list, key_list):
                beforeCell = ws.cell(4, col)
                beforeCell.font = Font(name='Times New Roman', size=12, color='FFFFFF')
                cell = ws.cell(row, col)
                cell.value = data[k]
                cell.font = line_font
                cell.border = all_border_thin
                cell.alignment = Alignment(horizontal='center', vertical='center')
                ws['M' + str(row)].number_format = '#,##0" đ"'
            row += 1

        fp = BytesIO()
        wb.save(fp)
        fp.seek(0)
        report = base64.encodebytes((fp.read()))
        fp.close()
        attachment = self.env['ir.attachment'].sudo().create({
            'name': 'BÁO CÁO HOA HỒNG HOA HỒNG CTV',
            'datas': report,
            'res_model': 'temp.creation',
            'public': True,
        })
        url = "/web/content/?model=ir.attachment&id=%s&filename_field=name&field=datas&download=true" \
              % attachment.id
        return {'name': 'BÁO CÁO HOA HỒNG HOA HỒNG CTV',
                'type': 'ir.actions.act_url',
                'url': url,
                'target': 'self',
                }
