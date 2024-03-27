import openpyxl

from odoo import fields, api, models, _
from odoo.exceptions import ValidationError, UserError
from datetime import date, datetime, time
from calendar import monthrange
from openpyxl import load_workbook
from openpyxl.styles import Font, borders, Alignment
import base64
from io import BytesIO
from dateutil.relativedelta import relativedelta
from pytz import timezone, utc

from odoo.modules.module import get_module_resource
from odoo.addons.report_sale.wizard.theme_report import ThemeReport

thin = borders.Side(style='thin')
double = borders.Side(style='double')
all_border_thin = borders.Border(thin, thin, thin, thin)
SERVICE_HIS_TYPE = [('Spa', 'Spa'), ('Laser', 'Laser'), ('Odontology', 'Nha'), ('Surgery', 'Phẫu thuật'), ('ChiPhi', 'Chi phí khác')]
dict_type = dict((key, value) for key, value in SERVICE_HIS_TYPE)

class ServiceSaleReport(models.TransientModel):
    _name = 'kpi.service.sale.report'
    _description = 'Bao cao doanh thu ban dich vu'

    start_date = fields.Date('Start date', default=date.today())
    end_date = fields.Date('End date', default=date.today())
    company_id = fields.Many2one('res.company', string='Company', default=lambda self: self.env.company)
    # domain="[('brand_id.code', '=', 'KN')]",
    # convert date to datetime for search domain, should be removed if using datetime directly
    start_datetime = fields.Datetime('Start datetime', compute='_compute_datetime')
    end_datetime = fields.Datetime('End datetime', compute='_compute_datetime')

    @api.depends('start_date', 'end_date')
    def _compute_datetime(self):
        self.start_datetime = False
        self.end_datetime = False
        if self.start_date and self.end_date:
            local_tz = timezone(self.env.user.tz or 'Etc/GMT+7')
            start_datetime = datetime(self.start_date.year, self.start_date.month, self.start_date.day, 0, 0, 0)
            end_datetime = datetime(self.end_date.year, self.end_date.month, self.end_date.day, 23, 59, 59)
            start_datetime = local_tz.localize(start_datetime, is_dst=None)
            end_datetime = local_tz.localize(end_datetime, is_dst=None)
            self.start_datetime = start_datetime.astimezone(utc).replace(tzinfo=None)
            self.end_datetime = end_datetime.astimezone(utc).replace(tzinfo=None)

    @api.onchange('start_date')
    def _onchange_start_date(self):
        if self.start_date:
            if self.start_date.month == fields.date.today().month:
                self.end_date = fields.date.today()
            else:
                self.end_date = date(self.start_date.year, self.start_date.month,
                                     monthrange(self.start_date.year, self.start_date.month)[1])

    def walkin_num(self, walkin, service):
        list_done_walkin = walkin.booking_id.walkin_ids.filtered(
            lambda w: service.id in w.service.ids and w.state == 'Completed').sorted('service_date').mapped(
            'service_date')
        index = [i for i, x in enumerate(list_done_walkin) if x == walkin.service_date]

        if len(list_done_walkin) > 0 and len(index) > 0:
            walkin_num = index[0] + 1
            # print("Đã có phiếu khám hoàn thành")
        else:
            # print("Chưa có phiếu khám hoàn thành")
            walkin_num = len(list_done_walkin) + 1
        return walkin_num

    @api.constrains('start_date', 'end_date')
    def check_dates(self):
        for record in self:
            start_date = fields.Date.from_string(record.start_date)
            end_date = fields.Date.from_string(record.end_date)
            if start_date > end_date:
                raise ValidationError(
                    _("Ngày kết thúc không thể ở trước ngày bắt đầu khi xuất báo cáo!!!"))

    def _get_data_kpi_service_sale_report(self):
        ret_data = []

        domain = [('order_id.state', '=', 'sale'), ('order_id.company_id', '=', self.company_id.id),
                  ('order_id.date_order', '>=', self.start_datetime),
                  ('order_id.date_order', '<=', self.end_datetime), ('order_id.pricelist_id.type', '=', 'service')]

        # if self.env.user.has_group('crm_base.receptionist_crm') or self.env.user.has_group('shealth_all_in_one.group_sh_medical_receptionist'):
        #     domain.append(('create_uid', '=', self.env.user.id))

        sale_order_lines = self.env['sale.order.line'].sudo().search(domain)

        total_revenue = 0
        for sol in sale_order_lines:
            # get so
            sale_order = sol.order_id

            # get booking
            booking = sale_order.booking_id

            crm_line = sol.crm_line_id

            # get consultants
            user_1 = None
            user_name_1 = None
            code_user_1 = None
            consulting_role_1 = None
            department_1 = None
            user_2 = None
            user_name_2 = None
            code_user_2 = None
            consulting_role_2 = None
            department_2 = None

            if crm_line.consultants_1:
                user_1 = crm_line.consultants_1
                user_name_1 = user_1.name
                consulting_role_1 = crm_line.consulting_role_1
                if user_1.employee_ids:
                    employ = user_1.employee_ids[0]
                    if employ.employee_code:
                        code_user_1 = employ.employee_code
                    else:
                        code_user_1 = employ.employee_id
                    if employ.department_id:
                        department_1 = employ.department_id.name
            if crm_line.consultants_2:
                user_2 = crm_line.consultants_2
                user_name_2 = user_2.name
                consulting_role_2 = crm_line.consulting_role_2
                if user_2.employee_ids:
                    employ = user_2.employee_ids[0]
                    if employ.employee_code:
                        code_user_2 = employ.employee_code
                    else:
                        code_user_2 = employ.employee_id
                    if employ.department_id:
                        department_2 = employ.department_id.name

            if consulting_role_1 == '1':
                consulting_role_1 = 'Tư vấn độc lập'
            elif consulting_role_1 == '2':
                consulting_role_1 = 'Tư vấn đồng thời'
            elif consulting_role_1 == '3':
                consulting_role_1 = 'Lễ tân - CVTV cùng tư vấn'
            elif consulting_role_1 == '4':
                consulting_role_1 = 'BS da liễu - KTV cùng tư vấn'
            elif consulting_role_1 == '5':
                consulting_role_1 = 'Tư vấn chính'
            elif consulting_role_1 == '6':
                consulting_role_1 = 'Tư vấn phụ'
            else:
                consulting_role_1 = ''

            if consulting_role_2 == '1':
                consulting_role_2 = 'Tư vấn độc lập'
            elif consulting_role_2 == '2':
                consulting_role_2 = 'Tư vấn đồng thời'
            elif consulting_role_2 == '3':
                consulting_role_2 = 'Lễ tân - CVTV cùng tư vấn'
            elif consulting_role_2 == '4':
                consulting_role_2 = 'BS da liễu - KTV cùng tư vấn'
            elif consulting_role_2 == '5':
                consulting_role_2 = 'Tư vấn chính'
            elif consulting_role_2 == '6':
                consulting_role_2 = 'Tư vấn phụ'
            else:
                consulting_role_2 = ''

            # get service his
            service = self.env['sh.medical.health.center.service'].search([('product_id', '=', sol.product_id.id)])

            # get walkin
            payment = ''
            walkin = self.env['sh.medical.appointment.register.walkin'].search([('sale_order_id', '=', sale_order.id)])
            walkin_num = self.walkin_num(walkin, service)
            if walkin.payment_ids:
                for rec in walkin.payment_ids:
                    if rec.state == 'posted':
                        payment += rec.name
                        payment += ","
            # trường số lượng và đơn vị xử lý các thương hiệu sử dụng khách nhau
                # thương hiệu KN: các dvu spa, laser, nha lấy từ bk còn phẫu thuật lấy từ sol
                # thương hiệu DA, PR: toàn bộ lấy từ bk
            # check thương hiệu
            if self.company_id.brand_id.code.lower() == 'da' or self.company_id.brand_id.code.lower() == 'pr':
                ret_data.append({
                    'date': sale_order.date_order.strftime('%d/%m/%Y'),
                    'code_user_1': code_user_1,
                    'name_user_1': user_name_1,
                    'department_1': department_1,
                    'consulting_role_1': consulting_role_1,
                    'code_user_2': code_user_2,
                    'name_user_2': user_name_2,
                    'department_2': department_2,
                    'consulting_role_2': consulting_role_2,
                    'code_booking': booking.name,
                    'source_booking': booking.source_id.name,
                    'code_customer': sale_order.code_customer,
                    'name_customer': sale_order.partner_id.name,
                    'category_service': service.service_category.name,
                    'service': sol.product_id.name,
                    'service_type': dict_type.get(service.his_service_type),
                    'walkin_num': walkin_num,
                    'uom_price': crm_line.uom_price,
                    'qty': crm_line.quantity,
                    'revenue': sol.price_subtotal,
                    'payment': payment,
                    'kpi_point': service.kpi_point if service else None,
                    'source_extend': crm_line.source_extend_id.name if crm_line.source_extend_id else None,
                    'create_user': sol.create_uid.name,
                    'department_user': sol.create_uid.employee_ids[
                        0].department_id.name if sol.create_uid.employee_ids else None,
                })
            # check thương hiệu
            if self.company_id.brand_id.code.lower() == 'kn':
                if service.his_service_type in ['Spa', 'Laser', 'Odontology']:
                    ret_data.append({
                        'date': sale_order.date_order.strftime('%d/%m/%Y'),
                        'code_user_1': code_user_1,
                        'name_user_1': user_name_1,
                        'department_1': department_1,
                        'consulting_role_1': consulting_role_1,
                        'code_user_2': code_user_2,
                        'name_user_2': user_name_2,
                        'department_2': department_2,
                        'consulting_role_2': consulting_role_2,
                        'code_booking': booking.name,
                        'source_booking': booking.source_id.name,
                        'code_customer': sale_order.code_customer,
                        'name_customer': sale_order.partner_id.name,
                        'category_service': service.service_category.name,
                        'service': sol.product_id.name,
                        'service_type': dict_type.get(service.his_service_type),
                        'walkin_num': walkin_num,
                        'uom_price': crm_line.uom_price,
                        'qty': crm_line.quantity,
                        'revenue': sol.price_subtotal,
                        'payment': payment,
                        'kpi_point': service.kpi_point if service else None,
                        'source_extend': crm_line.source_extend_id.name if crm_line.source_extend_id else None,
                        'create_user': sol.create_uid.name,
                        'department_user': sol.create_uid.employee_ids[
                            0].department_id.name if sol.create_uid.employee_ids else None,
                    })
                else:
                    ret_data.append({
                        'date': sale_order.date_order.strftime('%d/%m/%Y'),
                        'code_user_1': code_user_1,
                        'name_user_1': user_name_1,
                        'department_1': department_1,
                        'consulting_role_1': consulting_role_1,
                        'code_user_2': code_user_2,
                        'name_user_2': user_name_2,
                        'department_2': department_2,
                        'consulting_role_2': consulting_role_2,
                        'code_booking': booking.name,
                        'source_booking': booking.source_id.name,
                        'code_customer': sale_order.code_customer,
                        'name_customer': sale_order.partner_id.name,
                        'category_service': service.service_category.name,
                        'service': sol.product_id.name,
                        'service_type': dict_type.get(service.his_service_type),
                        'walkin_num': walkin_num,
                        'uom_price': sol.uom_price,
                        'qty': sol.product_uom_qty,
                        'revenue': sol.price_subtotal,
                        'payment': payment,
                        'kpi_point': service.kpi_point if service else None,
                        'source_extend': crm_line.source_extend_id.name if crm_line.source_extend_id else None,
                        'create_user': sol.create_uid.name,
                        'department_user': sol.create_uid.employee_ids[
                            0].department_id.name if sol.create_uid.employee_ids else None,
                    })
            total_revenue += sol.price_subtotal
        ret_data.append({
            'date': None,
            'code_user_1': None,
            'name_user_1': None,
            'department_1': None,
            'consulting_role_1': None,
            'code_user_2': None,
            'name_user_2': None,
            'department_2': None,
            'consulting_role_2': None,
            'code_booking': None,
            'source_booking': None,
            'code_customer': None,
            'name_customer': None,
            'category_service': None,
            'service': None,
            'service_type': None,
            'walkin_num': None,
            'uom_price': None,
            'qty': None,
            'revenue': total_revenue,
            'payment': None,
            'kpi_point': None,
            'source_extend': None,
            'create_user': None,
            'department_user': None,
        })
        return ret_data

    def create_kpi_service_sale_report(self):
        # get data
        datas = self._get_data_kpi_service_sale_report()

        # in dữ liệu
        report_brand_overview_attachment = self.env['ir.attachment'].browse(
            self.env.ref('report_sale.bao_cao_doanh_thu_ban_dich_vu_attachment').id)
        decode = base64.b64decode(report_brand_overview_attachment.datas)
        wb = load_workbook(BytesIO(decode))
        ws = wb.active
        line_font = Font(name='Times New Roman', size=14)

        code_brand = self.company_id.brand_id.code.lower()
        image_path = get_module_resource('report_sale', 'static/img', 'icon_%s.png' % code_brand)

        img = openpyxl.drawing.image.Image(image_path)
        img.anchor = 'A1'
        ws.add_image(img)
        format_currency = '_(* #,##0_);_(* (#,##0);_(* "-"??_);_(@_)'

        ws['K3'].value = self.start_date.strftime('%d/%m/%Y')
        ws['IN3'].value = self.end_date.strftime('%d/%m/%Y')
        ws['K4'].value = self.company_id.name

        key_col = list(range(1, 27))
        key_list = [
            'date',
            'code_user_1',
            'name_user_1',
            'department_1',
            'consulting_role_1',
            'code_user_2',
            'name_user_2',
            'department_2',
            'consulting_role_2',
            'code_booking',
            'source_booking',
            'code_customer',
            'name_customer',
            'category_service',
            'service',
            'service_type',
            'walkin_num',
            'uom_price',
            'qty',
            'revenue',
            'payment',
            'kpi_point',
            'source_extend',
            'create_user',
            'department_user',
        ]
        row = 7
        if code_brand == 'kn':
            header_fill = ThemeReport.kn_fill
        elif code_brand == 'da':
            header_fill = ThemeReport.da_fill
        elif code_brand == 'pr':
            header_fill = ThemeReport.pr_fill
        elif code_brand == 'hh':
            header_fill = ThemeReport.hh_fill
        else:
            header_fill = ThemeReport.sci_fill
        for data in datas:
            for col, k in zip(key_col, key_list):
                beforeCell = ws.cell(6, col)
                beforeCell.fill = header_fill
                beforeCell.font = Font(name='Times New Roman', size=14, color='FFFFFF')
                cell = ws.cell(row, col)
                cell.value = data[k]
                cell.font = line_font
                cell.border = all_border_thin
                cell.alignment = Alignment(horizontal='left', vertical='center')
                if col == 20:
                    cell.number_format = format_currency
            row += 1
        fp = BytesIO()
        wb.save(fp)
        fp.seek(0)
        report = base64.encodebytes((fp.read()))
        fp.close()
        attachment = self.env['ir.attachment'].sudo().create({
            'name': 'bao_cao_doanh_thu_ban_dich_vu_%s.xlsx' % self.company_id.name,
            'datas': report,
            'res_model': 'temp.creation',
            'public': True,
        })
        url = "/web/content/?model=ir.attachment&id=%s&filename_field=name&field=datas&download=true" \
              % attachment.id
        return {'name': 'BÁO CÁO DOANH SỐ BÁN DỊCH VỤ',
                'type': 'ir.actions.act_url',
                'url': url,
                'target': 'self',
                }
