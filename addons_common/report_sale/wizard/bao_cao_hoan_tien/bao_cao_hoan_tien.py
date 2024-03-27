from odoo import fields, api, models, _
from odoo.exceptions import ValidationError, UserError
from datetime import date, datetime, time
from calendar import monthrange
from openpyxl import load_workbook
from openpyxl.styles import Font, borders, Alignment
import base64
from io import BytesIO
from dateutil.relativedelta import relativedelta
from pytz import timezone, utc
from collections import defaultdict

thin = borders.Side(style='thin')
double = borders.Side(style='double')
all_border_thin = borders.Border(thin, thin, thin, thin)
KEY_LIST = [
    'stt',
    'ngay',
    'ma_phieu_thu',
    'hinh_thuc_thanh_toan',
    'ma_khach_hang',
    'khach_hang',
    'so_tien',
    'noi_dung_giao_dich',
    'nhan_vien_giao_dich',
    'loai_doi_tac',
    'booking',
    'ngay_hen_lich',
    'ngay_den_cua',
    'chi_nhanh'
]

PAYMENT_METHOD = [('tm', 'Tiền mặt'), ('ck', 'Chuyển khoản'), ('nb', 'Thanh toán nội bộ'), ('pos', 'Quẹt thẻ qua POS'),
                  ('vdt', 'Thanh toán qua ví điện tử')]

DICT_PAYMENT_METHOD = dict((key, value) for key, value in PAYMENT_METHOD)

PARTNER_TYPE = [('customer', 'Khách hàng'), ('supplier', 'Nhà cung cấp')]
DICT_PARTNER_TYPE = dict((key, value) for key, value in PARTNER_TYPE)


class ReportLineServiceCancel(models.TransientModel):
    _name = 'report.payment.cancel'
    _description = 'Báo cáo hoàn tiền order by hieupt@scigroup.com.vn'

    start_date = fields.Date('Start date', default=date.today())
    end_date = fields.Date('End date', default=date.today())

    company_ids = fields.Many2many(string='Công ty', comodel_name='res.company', domain=lambda self: [('id', 'in', self.env.user.company_ids.ids)])
    # convert date to datetime for search domain, should be removed if using datetime directly
    start_datetime = fields.Datetime('Start datetime', compute='_compute_datetime')
    end_datetime = fields.Datetime('End datetime', compute='_compute_datetime')

    @api.depends('start_date', 'end_date')
    def _compute_datetime(self):
        self.start_datetime = False
        self.end_datetime = False
        if self.start_date and self.end_date:
            local_tz = timezone(self.env.user.tz or 'Etc/GMT+7')
            start_datetime = datetime(self.start_date.year, self.start_date.month, self.start_date.day, 0, 0, 0)
            end_datetime = datetime(self.end_date.year, self.end_date.month, self.end_date.day, 23, 59, 59)
            start_datetime = local_tz.localize(start_datetime, is_dst=None)
            end_datetime = local_tz.localize(end_datetime, is_dst=None)
            self.start_datetime = start_datetime.astimezone(utc).replace(tzinfo=None)
            self.end_datetime = end_datetime.astimezone(utc).replace(tzinfo=None)

    @api.onchange('start_date')
    def _onchange_start_date(self):
        if self.start_date:
            if self.start_date.month == fields.date.today().month:
                self.end_date = fields.date.today()
            else:
                self.end_date = date(self.start_date.year, self.start_date.month,
                                     monthrange(self.start_date.year, self.start_date.month)[1])

    @api.constrains('start_date', 'end_date')
    def check_dates(self):
        for record in self:
            start_date = fields.Date.from_string(record.start_date)
            end_date = fields.Date.from_string(record.end_date)
            days = (end_date - start_date).days
            if days < 0 or days > 365:
                raise ValidationError(
                    _("Ngày kết thúc không thể ở trước ngày bắt đầu khi xuất báo cáo!!!"))

    def render_form_template(self, pay):
        booking = pay.crm_id

        lang_code = self.env.context.get('lang') or 'en_US'
        date_format = self.env['res.lang']._lang_get(lang_code).date_format
        time_format = self.env['res.lang']._lang_get(lang_code).time_format
        if booking.booking_date:
            booking_date = fields.Datetime.context_timestamp(self.with_context(tz=self.env.user.tz), booking.booking_date).strftime('%d/%m/%Y')
        else:
            booking_date = None

        if booking.arrival_date:
            arrival_date = fields.Datetime.context_timestamp(self.with_context(tz=self.env.user.tz), booking.arrival_date).strftime('%d/%m/%Y')
        else:
            arrival_date = None
        val = {
            'ngay': pay.payment_date.strftime('%d/%m/%Y'),
            'ma_phieu_thu': pay.name,
            'hinh_thuc_thanh_toan': DICT_PAYMENT_METHOD.get(pay.payment_method),
            'ma_khach_hang': pay.partner_id.code_customer,
            'khach_hang': pay.partner_id.name,
            'so_tien': pay.amount_vnd,
            'noi_dung_giao_dich': pay.communication,
            'nhan_vien_giao_dich': pay.user.name,
            'loai_doi_tac': DICT_PARTNER_TYPE.get(pay.partner_type),
            'booking': booking.name,
            'ngay_hen_lich': booking_date,
            'ngay_den_cua': arrival_date,
            'chi_nhanh': pay.journal_id.company_id.name
        }
        return val

    def _get_data_report(self):
        domain = [('payment_type', '=', 'outbound'), ('state', 'not in', ('draft', 'cancelled')), ('company_id', 'in', self.company_ids.ids),
                  ('payment_date', '>=', self.start_date), ('payment_date', '<=', self.end_date)]

        return_val = []
        payments = self.env['account.payment'].sudo().search(domain, order='id')
        count = 0
        for pay in payments:
            count += 1
            temp = self.render_form_template(pay)
            temp['stt'] = count
            return_val.append(temp)
        return return_val

    def create_report_da(self):
        datas = self._get_data_report()
        # in dữ liễu
        report_brand_overview_attachment = self.env['ir.attachment'].browse(
            self.env.ref('report_sale.report_payment_cancel_attachment').id)
        decode = base64.b64decode(report_brand_overview_attachment.datas)
        wb = load_workbook(BytesIO(decode))
        ws = wb.active
        line_font = Font(name='Times New Roman', size=14)

        ws['F4'].value += self.start_date.strftime('%d/%m/%Y')
        ws['G4'].value += self.end_datetime.strftime('%d/%m/%Y')
        ws['F5'].value = ', '.join((element.name for element in self.company_ids))
        key_col = list(range(1, len(KEY_LIST) + 1))
        format_currency = '_(* #,##0_);_(* (#,##0);_(* "-"??_);_(@_)'
        row = 9
        for data in datas:
            for col, k in zip(key_col, KEY_LIST):
                cell = ws.cell(row, col)
                cell.value = data[k]
                cell.font = line_font
                cell.border = all_border_thin
                cell.alignment = Alignment(horizontal='left', vertical='center')
                if col == 7:
                    cell.number_format = format_currency
            row += 1
        fp = BytesIO()
        wb.save(fp)
        fp.seek(0)
        report = base64.encodebytes((fp.read()))
        fp.close()
        attachment = self.env['ir.attachment'].sudo().create({
            'name': 'bao_cao_hoan_tien.xlsx',
            'datas': report,
            'res_model': 'temp.creation',
            'public': True,
        })
        url = "/web/content/?model=ir.attachment&id=%s&filename_field=name&field=datas&download=true" \
              % attachment.id
        return {'name': 'BÁO CÁO HOÀN TIỀN',
                'type': 'ir.actions.act_url',
                'url': url,
                'target': 'self',
                }
