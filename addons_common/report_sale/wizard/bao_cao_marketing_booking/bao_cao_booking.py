from odoo import fields, api, models, _
from odoo.exceptions import ValidationError, UserError
from datetime import date, datetime, time
from calendar import monthrange
from openpyxl import load_workbook
from openpyxl.styles import Font, borders, Alignment
import base64
from io import BytesIO
from dateutil.relativedelta import relativedelta
from pytz import timezone, utc
from collections import defaultdict

thin = borders.Side(style='thin')
double = borders.Side(style='double')
all_border_thin = borders.Border(thin, thin, thin, thin)
KEY_LIST = [
    'thoi_gian_tao',
    'ngay_hen_lich',
    'chi_nhanh',
    'nguoi_tao',
    'phong_ban',
    'ma_booking',
    'nhom_nguon',
    'nguon',
    'chien_dich',
    'trang_thai',
    'khach_den_cua',
    'loai_ban_ghi',
    'ma_dich_vu',
    'ten_dich_vu',
    'nhom_dich_vu',
    'bang_gia',
    'ho_ten_khach_hang',
    'so_dien_thoai',
    'so_dien_thoai_acc',
    'dia_chi',
    'email',
    'da_su_dung',
    'phieu_kham',
    'giai_doan_cua_phieu_kham',
    'trang_thai_line_dich_vu',
    'tong_tien_phai_thu'
]
OPTION = [
    ('01', 'Ngày hẹn lịch'),
    ('02', 'Ngày đến cửa')
]


class SaleMarketingBooking(models.TransientModel):
    _name = 'sale.marketing.booking'
    _description = 'Báo cáo cơ hội'

    start_date = fields.Date('Start date', default=date.today())
    end_date = fields.Date('End date', default=date.today())

    company_ids = fields.Many2many(string='Công ty', comodel_name='res.company', domain=lambda self: [('id', 'in', self.env.user.company_ids.ids)])
    option = fields.Selection(OPTION, string='Tìm kiếm theo', default='01')
    # convert date to datetime for search domain, should be removed if using datetime directly
    start_datetime = fields.Datetime('Start datetime', compute='_compute_datetime')
    end_datetime = fields.Datetime('End datetime', compute='_compute_datetime')

    @api.depends('start_date', 'end_date')
    def _compute_datetime(self):
        self.start_datetime = False
        self.end_datetime = False
        if self.start_date and self.end_date:
            local_tz = timezone(self.env.user.tz or 'Etc/GMT+7')
            start_datetime = datetime(self.start_date.year, self.start_date.month, self.start_date.day, 0, 0, 0)
            end_datetime = datetime(self.end_date.year, self.end_date.month, self.end_date.day, 23, 59, 59)
            start_datetime = local_tz.localize(start_datetime, is_dst=None)
            end_datetime = local_tz.localize(end_datetime, is_dst=None)
            self.start_datetime = start_datetime.astimezone(utc).replace(tzinfo=None)
            self.end_datetime = end_datetime.astimezone(utc).replace(tzinfo=None)

    @api.onchange('start_date')
    def _onchange_start_date(self):
        if self.start_date:
            if self.start_date.month == fields.date.today().month:
                self.end_date = fields.date.today()
            else:
                self.end_date = date(self.start_date.year, self.start_date.month,
                                     monthrange(self.start_date.year, self.start_date.month)[1])

    @api.constrains('start_date', 'end_date')
    def check_dates(self):
        for record in self:
            start_date = fields.Date.from_string(record.start_date)
            end_date = fields.Date.from_string(record.end_date)
            days = (end_date - start_date).days
            if days < 0 or days > 365:
                raise ValidationError(
                    _("Ngày kết thúc không thể ở trước ngày bắt đầu khi xuất báo cáo!!!"))

    def render_form_template(self, crm_line):
        booking = crm_line.crm_id
        walkin_ids = booking.walkin_ids.filtered(lambda element: element.state != 'Cancelled' and crm_line.service_id.id in element.service.ids).sorted(key=lambda element: element.date)

        if len(walkin_ids) >= 1:
            walkin_name_text = walkin_ids[0].name
            walkin_state_text = walkin_ids[0].state
        else:
            walkin_name_text = ''
            walkin_state_text = ''
        str_tuple = (booking.pass_port_address, booking.district_id.name, booking.state_id.name, booking.country_id.name)
        str_convert = (element for element in str_tuple if element != False)
        dia_chi = ', '.join(str_convert)
        product_code = crm_line.service_id.code
        product_name = crm_line.service_id.name
        product_categ = crm_line.service_id.service_category

        lang_code = self.env.context.get('lang') or 'en_US'
        date_format = self.env['res.lang']._lang_get(lang_code).date_format
        time_format = self.env['res.lang']._lang_get(lang_code).time_format

        val = {
            'thoi_gian_tao': '%s' % fields.Datetime.context_timestamp(self.with_context(tz=self.env.user.tz), booking.create_on).strftime(('%s %s') % (date_format, time_format)),
            'ngay_hen_lich': '%s' % fields.Datetime.context_timestamp(self.with_context(tz=self.env.user.tz), booking.booking_date if self.option == '01' else booking.arrival_date).strftime(
                ('%s %s') % (date_format, time_format)),
            'chi_nhanh': booking.company_id.name,
            'nguoi_tao': booking.create_by.name,
            'phong_ban': booking.create_by_department or '',
            'ma_booking': booking.name,
            'nhom_nguon': booking.category_source_id.name or '',
            'nguon': booking.source_id.name,
            'chien_dich': booking.campaign_id.name or '',
            'trang_thai': booking.stage_id.name,
            'khach_den_cua': booking.customer_come,
            'loai_ban_ghi': booking.type_crm_id.name,
            'ma_dich_vu': product_code,
            'ten_dich_vu': product_name,
            'nhom_dich_vu': product_categ.name,
            'bang_gia': booking.price_list_id.name,
            'ho_ten_khach_hang': booking.contact_name,
            'so_dien_thoai': booking.phone,
            'so_dien_thoai_acc': booking.partner_id.phone,
            'dia_chi': dia_chi,
            'email': booking.email_from or '',
            'da_su_dung': crm_line.number_used,
            'phieu_kham': walkin_name_text or '',
            'giai_doan_cua_phieu_kham': walkin_state_text or '',
            'trang_thai_line_dich_vu': crm_line.stage,
            'tong_tien_phai_thu': crm_line.total
        }
        return val

    def _get_data_report(self):
        domain = [('crm_id.type', '=', 'opportunity'), ('crm_id.company_id', 'in', self.company_ids.ids)]

        # Theo ngày hẹn lịch
        if self.option == '01':
            domain.append(('crm_id.booking_date', '>=', self.start_datetime))
            domain.append(('crm_id.booking_date', '<=', self.end_datetime))

        # Theo ngày đến cửa
        elif self.option == '02':
            domain.append(('crm_id.arrival_date', '>=', self.start_datetime))
            domain.append(('crm_id.arrival_date', '<=', self.end_datetime))

        return_val = []
        CrmLine = self.env['crm.line'].sudo().search(domain, order='id')
        for line in CrmLine:
            return_val.append(self.render_form_template(line))
        return return_val

    def create_report(self):
        datas = self._get_data_report()
        # in dữ liễu
        report_brand_overview_attachment = self.env['ir.attachment'].browse(
            self.env.ref('report_sale.bao_cao_booking_attachment').id)
        decode = base64.b64decode(report_brand_overview_attachment.datas)
        wb = load_workbook(BytesIO(decode))
        ws = wb.active
        line_font = Font(name='Times New Roman', size=14)

        ws['J3'].value = self.start_date.strftime('%d/%m/%Y')
        ws['L3'].value = self.end_datetime.strftime('%d/%m/%Y')
        ws['B6'].value = dict(self._fields['option'].selection).get(self.option)
        key_col = list(range(1, len(KEY_LIST) + 1))

        row = 7
        for data in datas:
            for col, k in zip(key_col, KEY_LIST):
                cell = ws.cell(row, col)
                cell.value = data[k]
                cell.font = line_font
                cell.border = all_border_thin
                cell.alignment = Alignment(horizontal='left', vertical='center')
            row += 1
        fp = BytesIO()
        wb.save(fp)
        fp.seek(0)
        report = base64.encodebytes((fp.read()))
        fp.close()
        attachment = self.env['ir.attachment'].sudo().create({
            'name': 'bao_cao_booking.xlsx',
            'datas': report,
            'res_model': 'temp.creation',
            'public': True,
        })
        url = "/web/content/?model=ir.attachment&id=%s&filename_field=name&field=datas&download=true" \
              % attachment.id
        return {'name': 'BÁO CÁO CƠ HỘI',
                'type': 'ir.actions.act_url',
                'url': url,
                'target': 'self',
                }
