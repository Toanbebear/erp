from odoo import fields, api, models, _
from odoo.exceptions import ValidationError, UserError
from datetime import date, datetime, time
from calendar import monthrange
from openpyxl import load_workbook
from openpyxl.styles import Font, borders, Alignment
import base64
from io import BytesIO
from dateutil.relativedelta import relativedelta
from pytz import timezone, utc
from collections import defaultdict

thin = borders.Side(style='thin')
double = borders.Side(style='double')
all_border_thin = borders.Border(thin, thin, thin, thin)
KEY_LIST = [
    'stt',
    'ma_booking',
    'trang_thai_booking',
    'khach_hang',
    'chi_nhanh',
    'nhom_dich_vu_huy',
    'dich_vu_huy',
    'gia_ban_tuong_ung',
    'ly_do_huy',
    'ghi_chu_huy_chi_tiet',
    'ngay_tu_van_dich_vu',
    'nguoi_tao',
    'phong_ban_tu_van',
    'ngay_huy_dich_vu',
    'nguoi_huy',
    'phong_ban_huy_dich_vu'
]
REASON_LINE_CANCEL = [('change_service', 'Đổi sang dịch vụ khác cùng nhóm'), ('consider_more', 'Cân nhắc thêm'),
                      ('due_to_illness', 'Hủy do bệnh lý'), ('create_wrong_service', 'Thao tác tạo sai dịch vụ'),
                      ('not_money', 'Không đủ chi phí'), ('consultant', 'Tham khảo trước'),
                      ('other', 'Lý do khác (Ghi rõ lý do)')]
DICT_REASON_LINE_CANCEL = dict((key, value)for key, value in REASON_LINE_CANCEL)


class ReportLineServiceCancel(models.TransientModel):
    _name = 'report.line.service.cancel'
    _description = 'Bao cao dong dich vu huy order by hieupt@scigroup.com.vn'

    start_date = fields.Date('Start date', default=date.today())
    end_date = fields.Date('End date', default=date.today())

    company_ids = fields.Many2many(string='Công ty', comodel_name='res.company', domain=lambda self: [('id', 'in', self.env.user.company_ids.ids)])
    # convert date to datetime for search domain, should be removed if using datetime directly
    start_datetime = fields.Datetime('Start datetime', compute='_compute_datetime')
    end_datetime = fields.Datetime('End datetime', compute='_compute_datetime')

    @api.depends('start_date', 'end_date')
    def _compute_datetime(self):
        self.start_datetime = False
        self.end_datetime = False
        if self.start_date and self.end_date:
            local_tz = timezone(self.env.user.tz or 'Etc/GMT+7')
            start_datetime = datetime(self.start_date.year, self.start_date.month, self.start_date.day, 0, 0, 0)
            end_datetime = datetime(self.end_date.year, self.end_date.month, self.end_date.day, 23, 59, 59)
            start_datetime = local_tz.localize(start_datetime, is_dst=None)
            end_datetime = local_tz.localize(end_datetime, is_dst=None)
            self.start_datetime = start_datetime.astimezone(utc).replace(tzinfo=None)
            self.end_datetime = end_datetime.astimezone(utc).replace(tzinfo=None)

    @api.onchange('start_date')
    def _onchange_start_date(self):
        if self.start_date:
            if self.start_date.month == fields.date.today().month:
                self.end_date = fields.date.today()
            else:
                self.end_date = date(self.start_date.year, self.start_date.month,
                                     monthrange(self.start_date.year, self.start_date.month)[1])

    @api.constrains('start_date', 'end_date')
    def check_dates(self):
        for record in self:
            start_date = fields.Date.from_string(record.start_date)
            end_date = fields.Date.from_string(record.end_date)
            days = (end_date - start_date).days
            if days < 0 or days > 365:
                raise ValidationError(
                    _("Ngày kết thúc không thể ở trước ngày bắt đầu khi xuất báo cáo!!!"))

    def render_form_template(self, crm_line):
        booking = crm_line.crm_id

        lang_code = self.env.context.get('lang') or 'en_US'
        date_format = self.env['res.lang']._lang_get(lang_code).date_format
        time_format = self.env['res.lang']._lang_get(lang_code).time_format
        if crm_line.create_uid:
            create_user = crm_line.create_uid
            create_employee = self.env['hr.employee'].sudo().search([('work_email', '=', create_user.login)])
        else:
            create_user = False
            create_employee = False

        if crm_line.cancel_user:
            cancel_user = crm_line.cancel_user
            cancel_employee = self.env['hr.employee'].sudo().search([('work_email', '=', create_user.login)])
        else:
            cancel_user = False
            cancel_employee = False

        val = {
            'ma_booking': booking.name,
            'trang_thai_booking': booking.stage_id.name,
            'khach_hang': booking.partner_id.name,
            'chi_nhanh': crm_line.company_id.name,
            'nhom_dich_vu_huy': crm_line.service_id.service_category.name,
            'dich_vu_huy': crm_line.service_id.name,
            'gia_ban_tuong_ung': crm_line.unit_price,
            'ly_do_huy': DICT_REASON_LINE_CANCEL.get(crm_line.reason_line_cancel),
            'ghi_chu_huy_chi_tiet': crm_line.note or None,
            'ngay_tu_van_dich_vu': '%s' % fields.Datetime.context_timestamp(self.with_context(tz=self.env.user.tz), crm_line.create_date).strftime('%d/%m/%Y'),
            'nguoi_tao': create_user.name or None,
            'phong_ban_tu_van': create_employee.department_id.name or None,
            'ngay_huy_dich_vu': '%s' % fields.Datetime.context_timestamp(self.with_context(tz=self.env.user.tz), crm_line.cancel_date).strftime('%d/%m/%Y'),
            'nguoi_huy': cancel_user.name or None,
            'phong_ban_huy_dich_vu': cancel_employee.department_id.name or None
        }
        return val

    def _get_data_report(self):
        domain = [('stage', '=', 'cancel'), ('crm_id.type', '=', 'opportunity'), ('company_id', 'in', self.company_ids.ids),
                  ('cancel_date', '>=', self.start_datetime), ('cancel_date', '<=', self.end_datetime)]

        return_val = []
        CrmLine = self.env['crm.line'].sudo().search(domain, order='id')
        count = 0
        for line in CrmLine:
            count += 1
            temp = self.render_form_template(line)
            temp['stt'] = count
            return_val.append(temp)
        return return_val

    def create_report_da(self):
        datas = self._get_data_report()
        # in dữ liễu
        report_brand_overview_attachment = self.env['ir.attachment'].browse(
            self.env.ref('report_sale.report_line_service_cancel_attachment').id)
        decode = base64.b64decode(report_brand_overview_attachment.datas)
        wb = load_workbook(BytesIO(decode))
        ws = wb.active
        line_font = Font(name='Times New Roman', size=14)

        ws['G4'].value += self.start_date.strftime('%d/%m/%Y')
        ws['H4'].value += self.end_datetime.strftime('%d/%m/%Y')
        ws['G5'].value = ', '.join((element.name for element in self.company_ids))
        key_col = list(range(1, len(KEY_LIST) + 1))
        format_currency = '_(* #,##0_);_(* (#,##0);_(* "-"??_);_(@_)'
        row = 9
        for data in datas:
            for col, k in zip(key_col, KEY_LIST):
                cell = ws.cell(row, col)
                cell.value = data[k]
                cell.font = line_font
                cell.border = all_border_thin
                cell.alignment = Alignment(horizontal='left', vertical='center')
                if col == 8:
                    cell.number_format = format_currency
            row += 1
        fp = BytesIO()
        wb.save(fp)
        fp.seek(0)
        report = base64.encodebytes((fp.read()))
        fp.close()
        attachment = self.env['ir.attachment'].sudo().create({
            'name': 'bao_cao_huy_dong_dich_vu.xlsx',
            'datas': report,
            'res_model': 'temp.creation',
            'public': True,
        })
        url = "/web/content/?model=ir.attachment&id=%s&filename_field=name&field=datas&download=true" \
              % attachment.id
        return {'name': 'BÁO CÁO HỦY DÒNG DỊCH VỤ',
                'type': 'ir.actions.act_url',
                'url': url,
                'target': 'self',
                }
