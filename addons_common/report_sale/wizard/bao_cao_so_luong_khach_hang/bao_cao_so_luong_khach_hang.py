from odoo import fields, api, models, _
from odoo.exceptions import ValidationError, UserError
from datetime import date, datetime, time
from calendar import monthrange
from openpyxl import load_workbook
from openpyxl.styles import Font, borders, Alignment
import base64
from io import BytesIO
from dateutil.relativedelta import relativedelta
from pytz import timezone, utc
from collections import defaultdict
import numpy as np
import pandas as pd
import logging

_logger = logging.getLogger(__name__)
TITLE = {0: 'major', 1: 'minor', 2: 'Laser', 3: 'Spa', 4: 'Odontology', 5: 'Unknown'}
thin = borders.Side(style='thin')
double = borders.Side(style='double')
all_border_thin = borders.Border(thin, thin, thin, thin)


class ReportCountCustomer(models.TransientModel):
    _name = 'report.count.customer'
    _description = 'Báo cáo số lượng khách hàng'

    start_date = fields.Date('Start date', default=date.today())
    end_date = fields.Date('End date', default=date.today())
    company_id = fields.Many2one(string='Công ty', comodel_name='res.company', domain=lambda self: [('id', 'in', self.env.user.company_ids.ids)])
    # convert date to datetime for search domain, should be removed if using datetime directly
    start_datetime = fields.Datetime('Start datetime', compute='_compute_datetime')
    end_datetime = fields.Datetime('End datetime', compute='_compute_datetime')

    @api.depends('start_date', 'end_date')
    def _compute_datetime(self):
        self.start_datetime = False
        self.end_datetime = False
        if self.start_date and self.end_date:
            local_tz = timezone(self.env.user.tz or 'Etc/GMT+7')
            start_datetime = datetime(self.start_date.year, self.start_date.month, self.start_date.day, 0, 0, 0)
            end_datetime = datetime(self.end_date.year, self.end_date.month, self.end_date.day, 23, 59, 59)
            start_datetime = local_tz.localize(start_datetime, is_dst=None)
            end_datetime = local_tz.localize(end_datetime, is_dst=None)
            self.start_datetime = start_datetime.astimezone(utc).replace(tzinfo=None)
            self.end_datetime = end_datetime.astimezone(utc).replace(tzinfo=None)

    @api.onchange('start_date')
    def _onchange_start_date(self):
        if self.start_date:
            if self.start_date.month == fields.date.today().month:
                self.end_date = fields.date.today()
            else:
                self.end_date = date(self.start_date.year, self.start_date.month,
                                     monthrange(self.start_date.year, self.start_date.month)[1])

    @api.constrains('start_date', 'end_date')
    def check_dates(self):
        for record in self:
            start_date = fields.Date.from_string(record.start_date)
            end_date = fields.Date.from_string(record.end_date)
            days = (end_date - start_date).days
            if days < 0 or days > 365:
                raise ValidationError(
                    _("Ngày kết thúc không thể ở trước ngày bắt đầu khi xuất báo cáo!!!"))

    def check_successful_consultation(self, booking_id):

        """
            Hàm check khách hàng tư vấn thành công:
                là khách hàng có ngày đến cửa bằng với ngày làm dịch vụ
        """
        result = False

        # Ngày đến cửa
        booking_date = booking_id.arrival_date.strftime('%d/%m/%Y')
        walkin_ids = booking_id.walkin_ids

        for walkin in walkin_ids:
            walkin_date = walkin.service_date.strftime('%d/%m/%Y')
            if booking_date == walkin_date and walkin.state == 'Completed':
                result = True
                break
        return result

    def check_customer_deposit(self, booking_id):
        """
            Hàm check khách hàng đặt cọc:
                Khách hàng có thanh toán, ngày thanh toán bằng ngày đến cửa, và chưa có phiếu khám hoàn thành.
        """
        result = False
        booking_date = booking_id.arrival_date.strftime('%d/%m/%Y')
        walkin_ids = booking_id.walkin_ids
        payment_ids = booking_id.payment_ids

        if all(not walkin.state == 'Completed' for walkin in walkin_ids) and payment_ids:
            for pay in payment_ids:
                payment_date = pay.payment_date.strftime('%d/%m/%Y')
                if payment_date == booking_date:
                    result = True
                    break
        return result

    def check_number_service(self, booking, service, date):
        specialtys = self.env['sh.medical.specialty'].search([('booking_id', '=', booking.id),
                                                              ('services_date', '<=', datetime(year=date.year,
                                                                                               month=date.month,
                                                                                               day=date.day,
                                                                                               hour=23,
                                                                                               minute=59,
                                                                                               second=59)),
                                                              ('state', '=', 'Done')])
        result = []
        for walkin in specialtys:
            if service in walkin.services:
                result.append(walkin)
        return len(result)

    def walkin_num(self, walkin, service):
        list_done_walkin = walkin.booking_id.walkin_ids.filtered(
            lambda w: service.id in w.service.ids and w.state == 'Completed').sorted('service_date').mapped(
            'service_date')
        index = [i for i, x in enumerate(list_done_walkin) if x == walkin.service_date]

        if len(list_done_walkin) > 0 and len(index) > 0:
            walkin_num = index[0] + 1
            # print("Đã có phiếu khám hoàn thành")
        else:
            # print("Chưa có phiếu khám hoàn thành")
            walkin_num = len(list_done_walkin) + 1
        return walkin_num

    def check_customers_lead(self, booking_id):
        """
            Hàm check khách hàng tiền năng:
                Khách hàng có booking trạng thái out-sold và chưa có payment
        """
        result = False
        payment_ids = booking_id.payment_ids
        if booking_id.stage_id.name == 'Out sold' and len(payment_ids) == 0:
            result = True
        return result

    def check_customers_course(self, booking_id):
        result = False
        if any(line.number_used > 1 and line.stage != 'cancel' for line in booking_id.crm_line_ids):
            result = True
        return result

    def count_service_type_by_walkin(self, walkin_id):
        service_ids = walkin_id.service
        SERVICE_TYPE_FORMAT = {'major': 0, 'minor': 0, 'Laser': 0, 'Spa': 0, 'Odontology': 0, 'Unknown': 0}
        for service in service_ids:
            type_service = service.his_service_type
            surgery_type = service.surgery_type
            if type_service == 'Surgery' and surgery_type == 'major':
                SERVICE_TYPE_FORMAT['major'] += 1
                break
            elif type_service == 'Surgery' and surgery_type == 'minor':
                SERVICE_TYPE_FORMAT['minor'] += 1
                break
            elif type_service == 'Laser':
                SERVICE_TYPE_FORMAT['Laser'] += 1
                break
            elif type_service == 'Spa':
                SERVICE_TYPE_FORMAT['Spa'] += 1
                break
            elif type_service == 'Odontology':
                SERVICE_TYPE_FORMAT['Odontology'] += 1
                break
        return SERVICE_TYPE_FORMAT

    def count_service_type_by_crm_line(self, crm_line_ids):
        SERVICE_TYPE_FORMAT = {'major': 0, 'minor': 0, 'Laser': 0, 'Spa': 0, 'Odontology': 0, 'Unknown': 0}
        for line in crm_line_ids:
            service = line.service_id
            type_service = service.his_service_type
            surgery_type = service.surgery_type

            if type_service == 'Surgery' and surgery_type == 'major':
                SERVICE_TYPE_FORMAT['major'] += 1

                break
            elif type_service == 'Surgery' and surgery_type == 'minor':
                SERVICE_TYPE_FORMAT['minor'] += 1

                break
            elif type_service == 'Laser':
                SERVICE_TYPE_FORMAT['Laser'] += 1
                break
            elif type_service == 'Spa':
                SERVICE_TYPE_FORMAT['Spa'] += 1

                break
            elif type_service == 'Odontology':
                SERVICE_TYPE_FORMAT['Odontology'] += 1

                break
        return SERVICE_TYPE_FORMAT

    def count_type_by_service(self, service):
        SERVICE_TYPE_FORMAT = {'major': 0, 'minor': 0, 'Laser': 0, 'Spa': 0, 'Odontology': 0, 'Unknown': 0}
        type_service = service.his_service_type
        surgery_type = service.surgery_type

        if type_service == 'Surgery' and surgery_type == 'major':
            SERVICE_TYPE_FORMAT['major'] += 1
        elif type_service == 'Surgery' and surgery_type == 'minor':
            SERVICE_TYPE_FORMAT['minor'] += 1
        elif type_service == 'Laser':
            SERVICE_TYPE_FORMAT['Laser'] += 1
        elif type_service == 'Spa':
            SERVICE_TYPE_FORMAT['Spa'] += 1
        elif type_service == 'Odontology':
            SERVICE_TYPE_FORMAT['Odontology'] += 1

        return SERVICE_TYPE_FORMAT

    def count_service_type_by_booking(self, booking_id):
        crm_line = booking_id.crm_line_ids
        SERVICE_TYPE_FORMAT = {'major': 0, 'minor': 0, 'Laser': 0, 'Spa': 0, 'Odontology': 0, 'Unknown': 0}
        for line in crm_line:
            if line.number_used > 1:
                type_service = line.service_id.his_service_type
                surgery_type = line.service_id.surgery_type
                if type_service == 'Surgery' and surgery_type == 'major':
                    SERVICE_TYPE_FORMAT['major'] += 1
                    break
                elif type_service == 'Surgery' and surgery_type == 'minor':
                    SERVICE_TYPE_FORMAT['minor'] += 1
                    break
                elif type_service == 'Laser':
                    SERVICE_TYPE_FORMAT['Laser'] += 1
                    break
                elif type_service == 'Spa':
                    SERVICE_TYPE_FORMAT['Spa'] += 1
                    break
                elif type_service == 'Odontology':
                    SERVICE_TYPE_FORMAT['Odontology'] += 1
                    break
        return SERVICE_TYPE_FORMAT

    def count_serivce_perform(self, services):
        SERVICE_TYPE_FORMAT = {'major': 0, 'minor': 0, 'Laser': 0, 'Spa': 0, 'Odontology': 0, 'Unknown': 0}
        for service in services:
            type_service = service.his_service_type
            surgery_type = service.surgery_type
            if type_service == 'Surgery' and surgery_type == 'major':
                SERVICE_TYPE_FORMAT['major'] += 1

            elif type_service == 'Surgery' and surgery_type == 'minor':
                SERVICE_TYPE_FORMAT['minor'] += 1

            elif type_service == 'Laser':
                SERVICE_TYPE_FORMAT['Laser'] += 1

            elif type_service == 'Spa':
                SERVICE_TYPE_FORMAT['Spa'] += 1

            elif type_service == 'Odontology':
                SERVICE_TYPE_FORMAT['Odontology'] += 1

        return SERVICE_TYPE_FORMAT

    def calculator_serivce_perform(self, services):
        SERVICE_TYPE_FORMAT = {'major': 0, 'minor': 0, 'Laser': 0, 'Spa': 0, 'Odontology': 0, 'Unknown': 0}
        for service in services:
            type_service = service.his_service_type
            surgery_type = service.surgery_type
            kpi = service.kpi_point
            if type_service == 'Surgery' and surgery_type == 'major':
                SERVICE_TYPE_FORMAT['major'] += kpi

            elif type_service == 'Surgery' and surgery_type == 'minor':
                SERVICE_TYPE_FORMAT['minor'] += kpi

            elif type_service == 'Laser':
                SERVICE_TYPE_FORMAT['Laser'] += kpi

            elif type_service == 'Spa':
                SERVICE_TYPE_FORMAT['Spa'] += kpi

            elif type_service == 'Odontology':
                SERVICE_TYPE_FORMAT['Odontology'] += kpi

        return SERVICE_TYPE_FORMAT

    def create_report(self):
        # Config
        datas = np.zeros((6, 6))
        # Phiếu phẫu thuật - thủ thuật
        surgery_ids = self.env['sh.medical.surgery'].search([('surgery_date', '>=', self.start_datetime),
                                                             ('surgery_date', '<=', self.end_datetime),
                                                             ('state', '=', 'Done'), ('institution.his_company', '=', self.company_id.id)], order='id')
        # Phiếu spa - laser - răng hàm mặt
        specialty_ids = self.env['sh.medical.specialty'].search([('services_date', '>=', self.start_datetime),
                                                                 ('services_date', '<=', self.end_datetime),
                                                                 ('state', '=', 'Done'), ('institution.his_company', '=', self.company_id.id)], order='id')
        # Lọc các booking theo ngày đến cửa
        domain = [('customer_come', '=', 'yes'),
                  ('arrival_date', '>=', self.start_datetime),
                  ('arrival_date', '<=', self.end_datetime), ('company_id', '=', self.company_id.id)]

        booking_ids = self.env['crm.lead'].search(domain)
        # Bộ lưu dữ liệu
        specialty_dic = defaultdict(list)
        # ================================================================================================================================
        for booking in booking_ids:
            # Kiểm tra booking là tư vấn thành công
            if self.check_successful_consultation(booking):
                walkin_ids = booking.walkin_ids
                if walkin_ids:
                    for walkin in walkin_ids:
                        count_type = self.count_service_type_by_walkin(walkin)
                        for index in range(len(count_type)):
                            datas[0][index] += count_type[TITLE[index]]
                else:
                    datas[0][5] += 1

            # Kiểm tra booking là khách đặt cọc
            elif self.check_customer_deposit(booking):
                crm_line_ids = booking.crm_line_ids

                if crm_line_ids:
                    count_type = self.count_service_type_by_crm_line(crm_line_ids)
                    for index in range(len(count_type)):
                        datas[1][index] += count_type[TITLE[index]]
                else:
                    datas[1][5] += 1

            # Kiểm tra booking là cơ hội
            elif self.check_customers_lead(booking):
                crm_line_ids = booking.crm_line_ids
                if crm_line_ids:
                    count_type = self.count_service_type_by_crm_line(crm_line_ids)
                    for index in range(len(count_type)):
                        datas[2][index] += count_type[TITLE[index]]
                else:
                    datas[2][5] += 1
        # Lấy ra booking có có nhiều phiếu khám hơn 1
        for specialty in specialty_ids:
            specialty_dic[specialty.walkin].append(specialty.services)

        # Xóa Booking chỉ có một phiếu khám
        # specialty_dic_copy = specialty_dic.copy()
        # for key, value in specialty_dic.items():
        #     if len(value) <= 1:
        #         specialty_dic_copy.pop(key)
        #
        # group_by_booking = defaultdict(dict)
        # for booking, specialty in specialty_dic_copy.items():
        #     # Khách hàng liệu trình
        #     group_by_service = defaultdict(list)
        #     for walkin in specialty:
        #         for serivce in walkin.services:
        #             group_by_service[serivce].append(walkin)
        #         group_by_booking[booking] = group_by_service
        #
        # # Xóa dịch vụ chỉ có một phiếu khám
        # group_by_booking_copy = group_by_booking.copy()
        # for key, value in group_by_booking.items():
        #     temp = value.copy()
        #     for serv, val in value.items():
        #         if len(val) <= 1:
        #             temp.pop(serv)
        #     if len(temp) <= 1:
        #         group_by_booking_copy[key] = temp

        for walkin, services in specialty_dic.items():
            for index in services:
                for service in index:
                    num = self.walkin_num(walkin, service)
                    if num > 1:
                        _logger.info('Liệu trình booking %s, phiếu khám id: %s, dịch vụ: %s', walkin.booking_id.name, walkin.name, service.name)
                        count_type = self.count_type_by_service(service)
                        for index in range(len(count_type)):
                            datas[3][index] += count_type[TITLE[index]]

        # if self.check_customers_course(booking):
        #     crm_line = booking.crm_line_ids.filtered(lambda rec: rec.number_used > 1 and rec.stage != 'cancel')
        #     if crm_line:
        #         count_type = self.count_service_type_by_walkin(booking)
        #         for index in range(len(count_type)):
        #             datas[3][index] += count_type[TITLE[index]]
        #     else:
        #         datas[3][5] += 1
        # ================================================================================================================================
        # Đếm tổng dịch vụ thành công
        for surgery in surgery_ids:
            count_type = self.count_serivce_perform(surgery.services)
            for index in range(len(count_type)):
                datas[4][index] += count_type[TITLE[index]]

            surgery_kpi = self.calculator_serivce_perform(surgery.services)
            for index in range(len(surgery_kpi)):
                datas[5][index] += surgery_kpi[TITLE[index]]

        for specialty in specialty_ids:
            count_type = self.count_serivce_perform(specialty.services)
            for index in range(len(count_type)):
                datas[4][index] += count_type[TITLE[index]]

            specialty_kpi = self.calculator_serivce_perform(specialty.services)
            for index in range(len(specialty_kpi)):
                datas[5][index] += specialty_kpi[TITLE[index]]

        report_brand_overview_attachment = self.env['ir.attachment'].browse(
            self.env.ref('report_sale.bao_cao_so_luong_khach_hang_attachment').id)
        decode = base64.b64decode(report_brand_overview_attachment.datas)
        wb = load_workbook(BytesIO(decode))
        ws = wb.active
        line_font = Font(name='Times New Roman', size=14)
        ws['A2'].value = "TÊN CÔNG TY %s" % self.company_id.name
        ws['C3'].value = ws['C3'].value + ' ' + self.start_date.strftime('%d/%m/%Y')
        ws['F3'].value = ws['F3'].value + ' ' + self.end_date.strftime('%d/%m/%Y')

        row = 7
        for data in datas:
            for index in range(len(data)):
                cell = ws.cell(row, index + 4)
                cell.value = data[index]
                cell.font = line_font
                cell.border = all_border_thin
                cell.alignment = Alignment(horizontal='center', vertical='center')
            row += 1
        fp = BytesIO()
        wb.save(fp)
        fp.seek(0)
        report = base64.encodebytes((fp.read()))
        fp.close()
        attachment = self.env['ir.attachment'].sudo().create({
            'name': 'bao_cao_so_luong_khach_hang_%s.xlsx' % self.company_id.name,
            'datas': report,
            'res_model': 'temp.creation',
            'public': True,
        })
        url = "/web/content/?model=ir.attachment&id=%s&filename_field=name&field=datas&download=true" \
              % attachment.id
        return {'name': 'BÁO CÁO SỐ LƯỢNG KHÁCH HÀNG',
                'type': 'ir.actions.act_url',
                'url': url,
                'target': 'self',
                }
