import base64
import json
import re
from calendar import monthrange
from datetime import date, datetime
from io import BytesIO

import openpyxl
from odoo import fields, api, models, _
from odoo.addons.report_sale.wizard.theme_report import ThemeReport
from odoo.exceptions import ValidationError
from odoo.modules.module import get_module_resource
from openpyxl import load_workbook
from openpyxl.styles import Font, borders, Alignment
from pytz import timezone, utc

thin = borders.Side(style='thin')
double = borders.Side(style='double')
all_border_thin = borders.Border(thin, thin, thin, thin)


class ProductSalesReport(models.TransientModel):
    _name = 'kpi.product.sale.report'
    _description = 'Bao cao doanh so ban san pham'

    start_date = fields.Date('Start date', default=date.today())
    end_date = fields.Date('End date', default=date.today())
    company_id = fields.Many2one('res.company', string='Company', default=lambda self: self.env.company)

    # convert date to datetime for search domain, should be removed if using datetime directly
    start_datetime = fields.Datetime('Start datetime', compute='_compute_datetime')
    end_datetime = fields.Datetime('End datetime', compute='_compute_datetime')

    @api.depends('start_date', 'end_date')
    def _compute_datetime(self):
        self.start_datetime = False
        self.end_datetime = False
        if self.start_date and self.end_date:
            local_tz = timezone(self.env.user.tz or 'Asia/Ho_Chi_Minh')
            start_datetime = datetime(self.start_date.year, self.start_date.month, self.start_date.day, 0, 0, 0)
            end_datetime = datetime(self.end_date.year, self.end_date.month, self.end_date.day, 23, 59, 59)
            start_datetime = local_tz.localize(start_datetime, is_dst=None)
            end_datetime = local_tz.localize(end_datetime, is_dst=None)
            self.start_datetime = start_datetime.astimezone(utc).replace(tzinfo=None)
            self.end_datetime = end_datetime.astimezone(utc).replace(tzinfo=None)

    @api.onchange('start_date')
    def _onchange_start_date(self):
        if self.start_date:
            if self.start_date.month == fields.date.today().month:
                self.end_date = fields.date.today()
            else:
                self.end_date = date(self.start_date.year, self.start_date.month,
                                     monthrange(self.start_date.year, self.start_date.month)[1])

    @api.constrains('start_date', 'end_date')
    def check_dates(self):
        for record in self:
            start_date = fields.Date.from_string(record.start_date)
            end_date = fields.Date.from_string(record.end_date)
            if start_date > end_date:
                raise ValidationError(
                    _("Ngày kết thúc không thể ở trước ngày bắt đầu khi xuất báo cáo!!!"))

    def _get_data_kpi_product_sale_report(self):
        ret_data = []
        total_revenue = 0

        domain = [('state', '=', 'sale'), ('company_id', '=', self.company_id.id),
                  ('date_order', '>=', self.start_datetime),
                  ('date_order', '<=', self.end_datetime), ('pricelist_id.type', '=', 'product')]

        sale_order = self.env['sale.order'].sudo().search(domain, order='date_order')

        for so in sale_order:
            stock_pickings = so.picking_ids.filtered(lambda rec: rec.picking_type_id.code == 'outgoing' and rec.state == 'done').sorted(key=lambda rec: rec.date_done)

            # get invoice
            invoices = so.invoice_ids.filtered(lambda rec: rec.state in ('posted'))
            payment = []
            KEY = 'content'

            for invoice in invoices:
                pay_list = []
                if re.search(KEY, str(invoice.invoice_payments_widget)):
                    list_payments = json.loads(invoice.invoice_payments_widget)['content']
                    list_payment_ids = [e['account_payment_id'] for e in list_payments]
                    pay = self.env['account.payment'].browse(list_payment_ids)
                    pay_list = [e.name for e in pay]
                for pay in pay_list:
                    payment.append(pay)

            # get user
            user = so.user_id
            employee_code = ''
            department = ''
            if user.employee_ids:
                employ = user.employee_ids[0]
                if employ.employee_code:
                    employee_code = employ.employee_code
                else:
                    employee_code = employ.employee_id
                if employ.department_id:
                    department = employ.department_id.name
                for stock_picking in stock_pickings:
                    stock_move_line = stock_picking.move_line_ids_without_package
                    for stock_line in stock_move_line:
                        for sol in so.order_line:
                            if sol.product_id.default_code == stock_line.product_id.default_code:
                                crm_line = sol.crm_line_id
                                consulting_role = None
                                if crm_line.consultants_1.id == user.id:
                                    consulting_role = crm_line.consulting_role_1
                                if crm_line.consultants_2.id == user.id:
                                    consulting_role = crm_line.consulting_role_2
                                val = {
                                    'date': so.date_order.strftime('%d/%m/%Y'),
                                    'code_user': employee_code,
                                    'name_user': user.name,
                                    'department': department,
                                    'code_so': so.name,
                                    'stock_code': stock_picking.name,
                                    'payment_name': ' - '.join(tuple(payment)),
                                    'code_customer': so.code_customer,
                                    'name_customer': so.partner_id.name,
                                    'code_product': stock_line.product_id.default_code,
                                    'product': stock_line.product_id.name,
                                    'unit': stock_line.product_uom_id.name,
                                    'sale_kpi': sol.price_subtotal,
                                    'kpi_point': None,
                                    'consulting_role': consulting_role if consulting_role else None,
                                    'qty': stock_line.qty_done,
                                    'source_id': so.source_id.name or '',
                                }
                                ret_data.append(val)
                                break
        total_revenue = sum([element['sale_kpi'] for element in ret_data])
        ret_data.append({
            'date': None,
            'code_user': None,
            'name_user': None,
            'department': None,
            'code_so': None,
            'stock_code': None,
            'payment_name': None,
            'code_customer': None,
            'name_customer': None,
            'code_product': None,
            'product': None,
            'unit': None,
            'sale_kpi': total_revenue,
            'kpi_point': None,
            'consulting_role': None,
            'qty': None,
            'source_id': None,
        })
        return ret_data

    # Xử lý bằng SQL
    # Cần xử lý lại mã account payment
    def _get_data_kpi_product_sale_report_2(self):
        result = []
        sale_order_line_query = """
                    SELECT 
                        so.name as code_so, 
                        so.date_order,
                        sol.price_subtotal as sale_kpi, 
                        us.name as unit, 
                        rp.code_customer,
                        rp.name as name_customer,
                        us.name as source_id,
                        sol.product_id,
                        sol.id as sol_id,
                        so.id as sale_id,
                        cl.consultants_1,
                        cl.consultants_2,
                        so.user_id
                    FROM sale_order_line sol
                    LEFT JOIN sale_order so on sol.order_id = so.id
                    LEFT JOIN utm_source us on us.id = so.source_id
                    left join product_pricelist ppl on ppl.id = so.pricelist_id
                    left join res_partner rp on rp.id = so.partner_id 
                    left join crm_line cl on cl.id = sol.id
                    where so.state = 'sale' 
                        and so.company_id = %s
                        and so.date_order >= '%s'
                        and so.date_order <= '%s'
                        and ppl.type = 'product'
                    ORDER BY so.date_order
                 """ % (self.company_id.id, self.start_datetime, self.end_datetime)

        self.env.cr.execute(sale_order_line_query)
        sale_orders = self.env.cr.dictfetchall()

        stock_move_line_query = """
                    SELECT 
                        sp.name as stock_code, 
                        so.name as code_so, 
                        pp.id as product_id,
                        pp.default_code as code_product,
                        pt.name as product,
                        uu.name as unit,
                        sml.qty_done as qty,
                        so.user_id,
                        so.id as sale_order_id
                    FROM stock_move_line sml
                    left join stock_picking sp on sp.id = sml.picking_id
                    left join stock_picking_type spt on spt.id = sp.picking_type_id
                    left join product_product pp on pp.id = sml.product_id 
                    left join product_template pt on pt.id = pp.product_tmpl_id  
                    left join uom_uom uu on uu.id = sml.product_uom_id 
                    left join sale_order so on sp.sale_id = so.id
                    where spt.code = 'outgoing' 
                        and sp.state = 'done' 
                        and	so.state = 'sale' 
                        and so.company_id = %s
                        and so.date_order >= '%s'
                        and so.date_order <= '%s'
                        and (sml.package_level_id is null  
                        or spt.show_entire_packs is null)
                 """ % (self.company_id.id, self.start_datetime, self.end_datetime)
        self.env.cr.execute(stock_move_line_query)
        stock_move_lines = self.env.cr.dictfetchall()

        """ Lấy tất cả người dùng liên quan tới phiếu SO để tìm nhân viên liên quan """
        user_ids = set()
        sale_order_ids = set()
        stock_move_lines_dict = {}
        for line in stock_move_lines:
            if line['code_so'] in stock_move_lines_dict:
                stock_move_lines_dict[line['code_so']].append(line)
            else:
                stock_move_lines_dict[line['code_so']] = [line]

            # Thêm id người dùng vào set
            if line['user_id']:
                user_ids.add(line['user_id'])
            sale_order_ids.add(line['sale_order_id'])

        """ Lấy tất cả nhân viên liên quan """
        # fix 4647
        if user_ids:
            query_user = """he.company_id = %s and user_id in %s""" % (self.company_id.id, tuple(user_ids))
            if ",)" in query_user:
                query_user = query_user.replace(',)', ')')
        else:
            query_user = """he.company_id = %s""" % self.company_id.id
        user_employee_query = """
                        SELECT 
                            he.employee_code, 
                            he.employee_id, 
                            he.name as name_user, 
                            hd.name as department,
                            he.user_id 
                        from hr_employee he 
                        left join hr_department hd on hd.id = he.department_id
                        where %s 
                         """ % query_user

        self.env.cr.execute(user_employee_query)
        user_employees = self.env.cr.dictfetchall()

        employees = {}
        for employee in user_employees:
            if employee['user_id'] not in employees:
                employees[employee['user_id']] = employee
            else:
                print('***' * 100)

        order_payments = {}
        total_revenue = 0
        for order_line in sale_orders:
            # get invoice
            payment = []
            if order_line['sale_id'] in order_payments:
                payment = order_payments[order_line['sale_id']]
            else:
                so = self.env['sale.order'].sudo().browse(order_line['sale_id'])
                invoices = so.invoice_ids.filtered(lambda rec: rec.state in ('posted'))

                KEY = 'content'
                for invoice in invoices:
                    pay_list = []
                    if re.search(KEY, str(invoice.invoice_payments_widget)):
                        list_payments = json.loads(invoice.invoice_payments_widget)['content']
                        list_payment_ids = [e['account_payment_id'] for e in list_payments]
                        pay = self.env['account.payment'].browse(list_payment_ids)
                        pay_list = [e.name for e in pay]
                    for pay in pay_list:
                        if pay:
                            payment.append(pay)
                order_payments[order_line['sale_id']] = payment

            if order_line['code_so'] in stock_move_lines_dict:
                stock_line = None
                for move_line in stock_move_lines_dict[order_line['code_so']]:
                    if move_line['product_id'] == order_line['product_id']:
                        stock_line = move_line
                        break
                if stock_line:
                    if stock_line['user_id'] in employees:
                        emp = employees[stock_line['user_id']]
                        employee_code = emp['employee_code'] if emp['employee_code'] else emp['employee_id']
                        name_user = emp['name_user']
                        department = emp['department']

                        consulting_role = None
                        if order_line['consultants_1'] == order_line['user_id']:
                            consulting_role = order_line['consultants_1']
                        if order_line['consultants_2'] == order_line['user_id']:
                            consulting_role = order_line['consultants_2']
                        if payment:
                            payment_name = ' - '.join(tuple(payment))
                        else:
                            payment_name = ''
                        result.append({
                            'date': order_line['date_order'].strftime('%d/%m/%Y'),
                            'code_user': employee_code,
                            'name_user': name_user,
                            'department': department,
                            'code_so': order_line['code_so'],
                            'stock_code': stock_line['stock_code'],
                            'payment_name': payment_name,
                            'code_customer': order_line['code_customer'],
                            'name_customer': order_line['name_customer'],
                            'code_product': stock_line['code_product'],
                            'product': stock_line['product'],
                            'unit': stock_line['unit'],
                            'sale_kpi': order_line['sale_kpi'],
                            'kpi_point': None,
                            'consulting_role': consulting_role,
                            'qty': stock_line['qty'],
                            'source_id': order_line['source_id'] or '',
                        })
                        total_revenue += order_line['sale_kpi']
        result.append({
            'date': None,
            'code_user': None,
            'name_user': None,
            'department': None,
            'code_so': None,
            'stock_code': None,
            'payment_name': None,
            'code_customer': None,
            'name_customer': None,
            'code_product': None,
            'product': None,
            'unit': None,
            'sale_kpi': total_revenue,
            'kpi_point': None,
            'consulting_role': None,
            'qty': None,
            'source_id': None,
        })
        return result

    def create_kpi_product_sale_report(self):
        # get data
        #datas = self._get_data_kpi_product_sale_report()
        datas = self._get_data_kpi_product_sale_report_2()

        # in dữ liệu
        report_brand_overview_attachment = self.env['ir.attachment'].browse(
            self.env.ref('report_sale.bao_cao_doanh_so_ban_san_pham_kpi_attachment').id)
        decode = base64.b64decode(report_brand_overview_attachment.datas)
        wb = load_workbook(BytesIO(decode))
        ws = wb.active
        line_font = Font(name='Times New Roman', size=14)

        ws['H3'].value = self.start_date.strftime('%d/%m/%Y')
        ws['J3'].value = self.end_datetime.strftime('%d/%m/%Y')
        ws['H4'].value = self.company_id.name

        code_brand = self.company_id.brand_id.code.lower()
        image_path = get_module_resource('report_sale', 'static/img', 'icon_%s.png' % code_brand)

        img = openpyxl.drawing.image.Image(image_path)
        img.anchor = 'A1'
        ws.add_image(img)

        key_col_list = list(range(1, 18))
        key_list = [
            'date',
            'code_user',
            'name_user',
            'department',
            'code_so',
            'stock_code',
            'payment_name',
            'code_customer',
            'name_customer',
            'code_product',
            'product',
            'unit',
            'sale_kpi',
            'kpi_point',
            'consulting_role',
            'qty',
            'source_id',
        ]
        row = 7
        if code_brand == 'kn':
            header_fill = ThemeReport.kn_fill
        elif code_brand == 'da':
            header_fill = ThemeReport.da_fill
        elif code_brand == 'pr':
            header_fill = ThemeReport.pr_fill
        elif code_brand == 'hh':
            header_fill = ThemeReport.hh_fill
        else:
            header_fill = ThemeReport.sci_fill
        for data in datas:
            for col, k in zip(key_col_list, key_list):
                beforeCell = ws.cell(6, col)
                beforeCell.fill = header_fill
                beforeCell.font = Font(name='Times New Roman', size=14, color='FFFFFF')
                cell = ws.cell(row, col)
                cell.value = data[k]
                cell.font = line_font
                cell.border = all_border_thin
                cell.alignment = Alignment(horizontal='left', vertical='center')
            row += 1
        fp = BytesIO()
        wb.save(fp)
        fp.seek(0)
        report = base64.encodebytes((fp.read()))
        fp.close()
        attachment = self.env['ir.attachment'].sudo().create({
            'name': 'bao_cao_doanh_so_ban_san_pham %s.xlsx' % self.company_id.name,
            'datas': report,
            'res_model': 'temp.creation',
            'public': True,
        })
        url = "/web/content/?model=ir.attachment&id=%s&filename_field=name&field=datas&download=true" \
              % attachment.id
        return {'name': 'BÁO CÁO DOANH SỐ BÁN SẢN PHẨM',
                'type': 'ir.actions.act_url',
                'url': url,
                'target': 'self',
                }
