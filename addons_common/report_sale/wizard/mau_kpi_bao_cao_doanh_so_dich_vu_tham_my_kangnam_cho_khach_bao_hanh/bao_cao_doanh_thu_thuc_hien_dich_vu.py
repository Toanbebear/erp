from odoo import fields, api, models, _
from odoo.exceptions import ValidationError, UserError
from datetime import date, datetime, time
from calendar import monthrange
from openpyxl import load_workbook
from openpyxl.styles import Font, borders, Alignment
import base64
from io import BytesIO
from dateutil.relativedelta import relativedelta
from pytz import timezone, utc
from odoo.addons.report_sale.wizard.theme_report import ThemeReport
from odoo.modules.module import get_module_resource
import openpyxl

thin = borders.Side(style='thin')
double = borders.Side(style='double')
all_border_thin = borders.Border(thin, thin, thin, thin)

WARD_TYPE = [
    ('All', 'Toàn bộ dịch vụ'),
    ('Surgery', 'Phẫu thuật'),
    ('Spa', 'Spa'),
    ('Laser', 'Laser'),
    ('Odontology', 'Nha khoa')
]

SEARCH_BY = [
    ('01', 'Ngày bắt đầu của phiếu chuyên khoa'),
    ('02', 'Ngày làm dịch vụ của phiếu khám'),
]

SERVICE_HIS_TYPE = [('Spa', 'Spa'), ('Laser', 'Laser'), ('Odontology', 'Nha'), ('Surgery', 'Phẫu thuật'), ('ChiPhi', 'Chi phí khác')]
dict_type = dict((key, value) for key, value in SERVICE_HIS_TYPE)

SURGERY_TYPE = [('minor', 'Tiểu phẫu'), ('major', 'Đại phẫu')]
DICT_SURGERY_TYPE = dict((key, value) for key, value in SURGERY_TYPE)

TEMPLATE = [
    'ma_nv_tu_van',
    'ten_nv_tu_van',
    'bo_phan_nv_tu_van',
    'chuc_danh_nv_tu_van',
    'vai_tro_nv_tu_van',
    'ten_khach_hang',
    'ma_booking',
    'ngay_dat_hang',
    'loai_phau_thuat',
    'loai_dich_vu',
    'nhom_dich_vu',
    'dich_vu',
    'don_vi_xu_ly',
    'buoi_thu',
    'so_luong',
    'nguon_booking',
    'nguon_mo_rong',
    'doanh_so',
    'diem_kpi',
    'ma_khach_hang',
    'thu_tu_tu_van',
    'ma_nv_nguoi_tao',
    'nguoi_tao',
    'phong_ban_nguoi_tao'
]


class InsuranceServicePerformanceReportKangnam(models.TransientModel):
    _name = 'kpi.insurance.service.perform.report.kangnam'
    _description = 'Báo cáo doanh thu thực hiện dịch vụ Kangnam'

    start_date = fields.Date('Start date', default=date.today())
    end_date = fields.Date('End date', default=date.today())
    company_id = fields.Many2one('res.company', string='Chi nhánh', domain="[('brand_id.name', '=', 'Kangnam')]", default=lambda self: self.env.company)
    type = fields.Selection(WARD_TYPE, string='Loại phiếu', default='All')

    # convert date to datetime for search domain, should be removed if using datetime directly
    start_datetime = fields.Datetime('Start datetime', compute='_compute_datetime')
    end_datetime = fields.Datetime('End datetime', compute='_compute_datetime')
    # Kiểu lấy báo cáo theo ngày làm dịch vụ hoặc theo ngày giờ bắt đầu của phiếu chuyên khoa
    search_by = fields.Selection(SEARCH_BY, string='Kiểu tìm kiếm dữ liệu', default='01')

    @api.depends('start_date', 'end_date')
    def _compute_datetime(self):
        self.start_datetime = False
        self.end_datetime = False
        if self.start_date and self.end_date:
            local_tz = timezone(self.env.user.tz or 'Etc/GMT+7')
            start_datetime = datetime(self.start_date.year, self.start_date.month, self.start_date.day, 0, 0, 0)
            end_datetime = datetime(self.end_date.year, self.end_date.month, self.end_date.day, 23, 59, 59)
            start_datetime = local_tz.localize(start_datetime, is_dst=None)
            end_datetime = local_tz.localize(end_datetime, is_dst=None)
            self.start_datetime = start_datetime.astimezone(utc).replace(tzinfo=None)
            self.end_datetime = end_datetime.astimezone(utc).replace(tzinfo=None)

    @api.onchange('start_date')
    def _onchange_start_date(self):
        if self.start_date:
            if self.start_date.month == fields.date.today().month:
                self.end_date = fields.date.today()
            else:
                self.end_date = date(self.start_date.year, self.start_date.month,
                                     monthrange(self.start_date.year, self.start_date.month)[1])

    @api.constrains('start_date', 'end_date')
    def check_dates(self):
        for record in self:
            start_date = fields.Date.from_string(record.start_date)
            end_date = fields.Date.from_string(record.end_date)
            if start_date > end_date:
                raise ValidationError(
                    _("Ngày kết thúc không thể ở trước ngày bắt đầu khi xuất báo cáo!!!"))

    def walkin_num(self, walkin, service):
        list_done_walkin = walkin.booking_id.walkin_ids.filtered(
            lambda w: service.id in w.service.ids and w.state == 'Completed').sorted('service_date').mapped(
            'service_date')
        index = [i for i, x in enumerate(list_done_walkin) if x == walkin.service_date]

        if len(list_done_walkin) > 0 and len(index) > 0:
            walkin_num = index[0] + 1
            # print("Đã có phiếu khám hoàn thành")
        else:
            # print("Chưa có phiếu khám hoàn thành")
            walkin_num = len(list_done_walkin) + 1
        return walkin_num

    def _get_data_report(self):

        if self.search_by == '01':
            domain_surgery = [('surgery_date', '>=', self.start_datetime),
                              ('surgery_date', '<=', self.end_datetime),
                              ('institution.his_company', '=', self.company_id.id),
                              ('walkin.state', '=', 'Completed'),
                              ('walkin.type_crm_id', '=', self.env.ref('crm_base.type_oppor_guarantee').id)]
            domain_specialty = [('services_date', '>=', self.start_datetime),
                                ('services_date', '<=', self.end_datetime),
                                ('institution.his_company', '=', self.company_id.id),
                                ('walkin.state', '=', 'Completed'), ('department.type', '=', 'Odontology'),
                                ('walkin.type_crm_id', '=', self.env.ref('crm_base.type_oppor_guarantee').id)]
        elif self.search_by == '02':
            domain_surgery = [('walkin.service_date', '>=', self.start_datetime),
                              ('walkin.service_date', '<=', self.end_datetime),
                              ('institution.his_company', '=', self.company_id.id),
                              ('walkin.state', '=', 'Completed'),
                              ('walkin.type_crm_id', '=', self.env.ref('crm_base.type_oppor_guarantee').id)]
            domain_specialty = [('walkin.service_date', '>=', self.start_datetime),
                                ('walkin.service_date', '<=', self.end_datetime),
                                ('institution.his_company', '=', self.company_id.id),
                                ('walkin.state', '=', 'Completed'), ('department.type', '=', 'Odontology'),
                                ('walkin.type_crm_id', '=', self.env.ref('crm_base.type_oppor_guarantee').id)]

        if self.type == 'All':
            # Loại phiếu là toàn bộ dịch vụ
            # Là phiếu phẫu thuật
            data_surgery = self._get_data_surgery(domain_surgery)
            # Là phiếu chuyên khoa
            data_specialty = self._get_data_specialty(domain_specialty)
            # Join data
            data = data_surgery + data_specialty
        elif self.type == 'Surgery':
            # Loại phiếu là dịch vụ nhóm phẫu thuật
            data = self._get_data_surgery(domain_surgery)
        else:
            # Loại phiếu là Spa, Lase, Nha khoa
            domain = domain_specialty.append('name.department_type', '=', self.type)
            data = self._get_data_specialty(domain)

        dict_template = dict([(key, None) for key in TEMPLATE])

        dict_template['doanh_so'] = sum([element.get('doanh_so') for element in data]) if data else 0
        return data

    def _get_data_surgery(self, domain):
        data_surgery = self.env['sh.medical.surgery'].sudo().search(domain)
        result = list()
        for surgery in data_surgery:
            Walkin = surgery.walkin
            so = Walkin.sale_order_id
            sol = so.order_line
            for ser_perform in surgery.services:
                sol = sol.filtered(lambda r: r.product_id == ser_perform.product_id and r.price_subtotal > 0)
                if sol:
                    walkin_num = self.walkin_num(Walkin, ser_perform)
                    consultants = []
                    if sol.crm_line_id.consultants_1:
                        consultants.append((sol.crm_line_id.consultants_1.employee_ids[0], sol.crm_line_id.consulting_role_1))
                    if sol.crm_line_id.consultants_2:
                        consultants.append((sol.crm_line_id.consultants_2.employee_ids[0], sol.crm_line_id.consulting_role_2))

                    if len(consultants) == 0:
                        count = 1
                    else:
                        count = len(consultants)

                    for index in range(0, count):
                        result.append({
                            'ma_nv_tu_van': consultants[index][0].employee_code if len(consultants) != 0 else None,
                            'ten_nv_tu_van': consultants[index][0].name if len(consultants) != 0 else None,
                            'bo_phan_nv_tu_van': consultants[index][0].group_job.name if len(consultants) != 0 else None,
                            'chuc_danh_nv_tu_van': consultants[index][0].job_id.name if len(consultants) != 0 else None,
                            'vai_tro_nv_tu_van': consultants[index][1] if len(consultants) != 0 else None,
                            'ten_khach_hang': so.partner_id.name,
                            'ma_booking': so.booking_id.name,
                            'ngay_dat_hang': '%s' % fields.Datetime.context_timestamp(self.with_context(tz=self.env.user.tz), so.date_order).strftime('%d/%m/%Y'),
                            'loai_phau_thuat': DICT_SURGERY_TYPE.get(ser_perform.surgery_type) if ser_perform.surgery_type else None,
                            'loai_dich_vu': dict_type.get(ser_perform.his_service_type),
                            'nhom_dich_vu': ser_perform.service_category.name,
                            'dich_vu': ser_perform.name,
                            'don_vi_xu_ly': sol.uom_price,
                            'buoi_thu': walkin_num,
                            'so_luong': sol.product_uom_qty,
                            'nguon_booking': sol.crm_line_id.crm_id.source_id.name or None,
                            'nguon_mo_rong': sol.crm_line_id.source_extend_id.name or None,
                            'doanh_so': sol.price_subtotal,
                            'diem_kpi': ser_perform.kpi_point,
                            'ma_khach_hang': so.code_customer,
                            'thu_tu_tu_van': 'Tư vấn viên %s' % str(index) if len(consultants) != 0 else None,
                            'ma_nv_nguoi_tao': so.booking_id.create_by.employee_ids[0].employee_code if so.booking_id.create_by.employee_ids else None,
                            'nguoi_tao': so.booking_id.create_by.name,
                            'phong_ban_nguoi_tao': so.booking_id.create_by.employee_ids[0].department_id.name if so.booking_id.create_by.employee_ids else None
                        })
        return result

    def _get_data_specialty(self, domain):
        data_specialty = self.env['sh.medical.specialty'].sudo().search(domain)
        result = list()
        for specialty in data_specialty:
            Walkin = specialty.walkin
            so = specialty.walkin.sale_order_id
            sol = so.order_line
            for spe_perform in specialty.services:
                # get s ale order line for service
                sol = sol.filtered(lambda r: r.product_id == spe_perform.product_id and r.price_subtotal > 0)
                if sol:
                    walkin_num = self.walkin_num(Walkin, spe_perform)
                    consultants = list()
                    if sol.crm_line_id.consultants_1:
                        consultants.append((sol.crm_line_id.consultants_1.employee_ids[0], sol.crm_line_id.consulting_role_1))
                    if sol.crm_line_id.consultants_2:
                        consultants.append((sol.crm_line_id.consultants_2.employee_ids[0], sol.crm_line_id.consulting_role_2))

                    if len(consultants) == 0:
                        count = 1
                    else:
                        count = len(consultants)

                    for index in range(0, count):
                        result.append({
                            'ma_nv_tu_van': consultants[index][0].employee_code if len(consultants) != 0 else None,
                            'ten_nv_tu_van': consultants[index][0].name if len(consultants) != 0 else None,
                            'bo_phan_nv_tu_van': consultants[index][0].group_job.name if len(consultants) != 0 else None,
                            'chuc_danh_nv_tu_van': consultants[index][0].job_id.name if len(consultants) != 0 else None,
                            'vai_tro_nv_tu_van': consultants[index][1] if len(consultants) != 0 else None,
                            'ten_khach_hang': so.partner_id.name,
                            'ma_booking': so.booking_id.name,
                            'ngay_dat_hang': '%s' % fields.Datetime.context_timestamp(self.with_context(tz=self.env.user.tz), so.date_order).strftime('%d/%m/%Y'),
                            'loai_phau_thuat': DICT_SURGERY_TYPE.get(spe_perform.surgery_type) if spe_perform.surgery_type else None,
                            'loai_dich_vu': dict_type.get(spe_perform.his_service_type),
                            'nhom_dich_vu': spe_perform.service_category.name,
                            'dich_vu': spe_perform.name,
                            'don_vi_xu_ly': sol.uom_price,
                            'buoi_thu': walkin_num,
                            'so_luong': sol.product_uom_qty,
                            'nguon_booking': sol.crm_line_id.crm_id.source_id.name or None,
                            'nguon_mo_rong': sol.crm_line_id.source_extend_id.name or None,
                            'doanh_so': sol.price_subtotal,
                            'diem_kpi': spe_perform.kpi_point,
                            'ma_khach_hang': so.code_customer,
                            'thu_tu_tu_van': 'Tư vấn viên %s' % str(index) if len(consultants) != 0 else None,
                            'ma_nv_nguoi_tao': so.booking_id.create_by.employee_ids[0].employee_code if so.booking_id.create_by.employee_ids else None,
                            'nguoi_tao': so.booking_id.create_by.name,
                            'phong_ban_nguoi_tao': so.booking_id.create_by.employee_ids[0].department_id.name if so.booking_id.create_by.employee_ids else None
                        })
        return result

    def create_report_service_perform_report(self):
        # get data
        datas = self._get_data_report()
        # in dữ liễu
        report_brand_overview_attachment = self.env['ir.attachment'].browse(
            self.env.ref('report_sale.bao_cao_doanh_thu_thuc_hien_dich_vu_bao_hanh_kangnam_mau_kpi_attachment').id)
        decode = base64.b64decode(report_brand_overview_attachment.datas)
        wb = load_workbook(BytesIO(decode))
        ws = wb.active
        line_font = Font(name='Times New Roman', size=14)
        if self.type == 'Surgery':
            name = "PHẪU THUẬT"
        elif self.type == 'Spa':
            name = "SPA"
        elif self.type == 'Laser':
            name = "LASER"
        elif self.type == 'Odontology':
            name = "NHA KHOA"
        else:
            name = 'Toàn bộ dịch vụ'.upper()

        code_brand = self.company_id.brand_id.code.lower()
        image_path = get_module_resource('report_sale', 'static/img', 'icon_%s.png' % code_brand)

        img = openpyxl.drawing.image.Image(image_path)
        img.anchor = 'A1'
        ws.add_image(img)

        ws['F2'].value = 'BÁO CÁO DOANH THU THỰC HIỆN DỊCH VỤ %s' % name
        ws['F3'].value += self.start_date.strftime('%d/%m/%Y')
        ws['G3'].value += self.end_datetime.strftime('%d/%m/%Y')
        ws['F4'].value += self.company_id.name

        key_col = list(range(1, len(TEMPLATE) + 1))
        format_currency = '_(* #,##0_);_(* (#,##0);_(* "-"??_);_(@_)'

        if code_brand == 'kn':
            header_fill = ThemeReport.kn_fill
        elif code_brand == 'da':
            header_fill = ThemeReport.da_fill
        elif code_brand == 'pr':
            header_fill = ThemeReport.pr_fill
        elif code_brand == 'hh':
            header_fill = ThemeReport.hh_fill
        else:
            header_fill = ThemeReport.sci_fill
        row = 7
        for data in datas:
            for col, k in zip(key_col, TEMPLATE):
                beforeCell = ws.cell(6, col)
                beforeCell.fill = header_fill
                beforeCell.font = Font(name='Times New Roman', size=14, color='FFFFFF')
                cell = ws.cell(row, col)
                cell.value = data[k]
                cell.font = line_font
                cell.border = all_border_thin
                cell.alignment = Alignment(horizontal='left', vertical='center')
                if col == 18:
                    cell.number_format = format_currency
            row += 1
        fp = BytesIO()
        wb.save(fp)
        fp.seek(0)
        report = base64.encodebytes((fp.read()))
        fp.close()
        attachment = self.env['ir.attachment'].sudo().create({
            'name': 'bao_cao_doanh_thu_ban_dich_vu_bao_hanh_%s.xlsx' % self.company_id.name,
            'datas': report,
            'res_model': 'temp.creation',
            'public': True,
        })
        url = "/web/content/?model=ir.attachment&id=%s&filename_field=name&field=datas&download=true" \
              % attachment.id
        return {'name': 'BÁO CÁO DOANH BÁN DỊCH VỤ CHO KHÁCH BẢO HÀNH',
                'type': 'ir.actions.act_url',
                'url': url,
                'target': 'self',
                }
