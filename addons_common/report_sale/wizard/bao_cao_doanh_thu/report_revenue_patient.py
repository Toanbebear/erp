# -*- coding: utf-8 -*-

from odoo import fields, api, models, _
from odoo.exceptions import ValidationError
from datetime import date, datetime, time, timedelta
from calendar import monthrange
from openpyxl import load_workbook
from openpyxl.styles import Font, borders, Alignment, PatternFill
import base64
from io import BytesIO
from dateutil.relativedelta import relativedelta
from operator import itemgetter
from pytz import timezone

import logging

TEMP_01 = ['stt',
           'ma_benh_nhan',
           'ho_ten_khach_hang',
           'dia_chi',
           'ma_dich_vu',
           'dich_vu',
           'don_gia',
           'tong_so_luong',
           'thanh_tien',
           'mien_giam',
           'tong_tien_da_thu',
           'thoi_gian_thu']
TEMP_02 = [
    'stt',
    'dich_vu',
    'don_gia',
    'tong_so_luong',
    'tong_thanh_tien',
    'mien_giam',
    'tong_da_thu'
]


class SHReportRevenue(models.TransientModel):
    _name = 'sh.report.revenue'
    _description = 'Báo cáo doanh thu'

    institution = fields.Many2one('sh.medical.health.center', string='Chi nhánh', help="Chi nhánh",
                                  required=True, domain=lambda self: [('his_company', 'in', self.env.companies.ids)],
                                  default=lambda self: self.env['sh.medical.health.center'].search(
                                      [('his_company', '=', self.env.companies.ids[0])], limit=1).id)
    start_date = fields.Date('Start date', default=date.today().replace(day=1))
    end_date = fields.Date('End date')

    def _get_report_type(self):
        report_list = [('patient', 'Doanh thu theo bệnh nhân'), ('service', 'Doanh thu theo dịch vụ')]

        return report_list

    report_type = fields.Selection(_get_report_type,
                                   'Report type', default='patient')

    # convert date to datetime for search domain, should be removed if using datetime directly
    start_datetime = fields.Datetime('Start datetime', compute='_compute_datetime')
    end_datetime = fields.Datetime('End datetime', compute='_compute_datetime')

    # convert date to datetime for search domain, should be removed if using datetime directly
    @api.depends('start_date', 'end_date')
    def _compute_datetime(self):
        if self.start_date and self.end_date:
            fmt = "%Y-%m-%d %H:%M:%S"
            start_datetime = datetime(self.start_date.year, self.start_date.month, self.start_date.day, 0, 0, 0)
            end_datetime = datetime(self.end_date.year, self.end_date.month, self.end_date.day, 23, 59, 59)
            now_utc = datetime.now(timezone('UTC'))
            local_tz = self.env.user.tz or 'Etc/GMT-7'
            now_timezone = now_utc.astimezone(timezone(local_tz))
            UTC_OFFSET_TIMEDELTA = datetime.strptime(now_utc.strftime(fmt), fmt) - datetime.strptime(
                now_timezone.strftime(fmt), fmt)
            utc_start_datetime = start_datetime + UTC_OFFSET_TIMEDELTA
            utc_end_datetime = end_datetime + UTC_OFFSET_TIMEDELTA
            self.start_datetime = utc_start_datetime
            self.end_datetime = utc_end_datetime
            # print(self.start_datetime)
            # print(self.end_datetime)

    @api.onchange('start_date')
    def _onchange_start_date(self):
        if self.start_date:
            if self.start_date.month == fields.date.today().month:
                self.end_date = fields.date.today()
            else:
                self.end_date = date(self.start_date.year, self.start_date.month,
                                     monthrange(self.start_date.year, self.start_date.month)[1])

    @api.constrains('start_date', 'end_date')
    def check_dates(self):
        for record in self:
            start_date = fields.Date.from_string(record.start_date)
            end_date = fields.Date.from_string(record.end_date)
            if start_date > end_date:
                raise ValidationError(
                    _("End Date cannot be set before Start Date."))

    # BÁO CÁO DOANH THU THEO DỊCH VỤ
    def revenue_service_report(self):
        template = self.env.ref('shealth_all_in_one.report_revenue_service_template')

        thin = borders.Side(style='thin')
        all_border_thin = borders.Border(left=thin, right=thin, top=thin, bottom=thin)
        datas = self.env['account.move.line'].search(
            [('date', '>=', self.start_date), ('date', '<=', self.end_date), ('parent_state', '=', 'posted'),
             ('move_id.journal_id.type', '=', 'sale'), ('account_id.user_type_id', 'in', [13, 14]), ('product_id.type', 'in', ['service'])])

        return {'name': (_('Báo cáo doanh thu theo dịch vụ')),
                'type': 'ir.actions.act_window',
                'res_model': 'temp.wizard',
                'view_mode': 'form',
                'target': 'inline',
                'view_id': self.env.ref('ms_templates.report_wizard').id,
                'context': {'default_template_id': template.id, 'active_ids': datas.ids,
                            'active_model': 'account.move.line',
                            'external_keys': {'a1': self.institution.name,
                                              'C4': self.start_date.strftime('%d/%m/%Y'),
                                              'E4': self.end_date.strftime('%d/%m/%Y'),
                                              'F10007': self.env.user.name,
                                              'F10003': date.today().strftime('Ngày %d tháng %m năm %Y')}}}

    # BÁO CÁO DOANH THU THEO DỊCH VỤ và CHI TIẾT BỆNH NHÂN
    def report_revenue_patient_excel(self):
        template = self.env.ref('shealth_all_in_one.report_revenue_patient_template')
        domain_payment = [('date', '>=', self.start_date), ('date', '<=', self.end_date), ('parent_state', '=', 'posted'),
                          ('move_id.journal_id.type', '=', 'sale'), ('account_id.user_type_id', 'in', [13, 14]),
                          ('product_id.type', 'in', ['service'])]
        datas = self.env['account.move.line'].search(domain_payment)

        return {'name': (_('Báo cáo doanh thu theo bệnh nhân')),
                'type': 'ir.actions.act_window',
                'res_model': 'temp.wizard',
                'view_mode': 'form',
                'target': 'inline',
                'view_id': self.env.ref('ms_templates.report_wizard').id,
                'context': {'default_template_id': template.id, 'active_ids': datas.ids,
                            'active_model': 'account.move.line',
                            'external_keys': {'a1': self.institution.name,
                                              'e4': self.start_date.strftime('%d/%m/%Y'),
                                              'g4': self.end_date.strftime('%d/%m/%Y'),
                                              'j10007': self.env.user.name, 'j10003': date.today().strftime('Ngày %d tháng %m năm %Y')}}}

    def report_revenue(self):
        domain_payment = [('date', '>=', self.start_date), ('date', '<=', self.end_date), ('parent_state', '=', 'posted'),
                          ('move_id.journal_id.type', '=', 'sale'), ('account_id.user_type_id', 'in', [13, 14]),
                          ('product_id.type', 'in', ['service'])]
        datas = self.env['account.move.line'].search(domain_payment)
        result = list()
        count = 0
        if self.report_type == 'service':
            tmp = dict([(e, None) for e in TEMP_02])
            for data in datas:
                count += 1
                tmp['stt'] = count
                tmp['dich_vu'] = data.product_id.name
                tmp['don_gia'] = data.price_unit
                tmp['tong_so_luong'] = data.quantity
                tmp['tong_thanh_tien'] = data.price_subtotal
                tmp['mien_giam'] = data.discount
                tmp['tong_da_thu'] = data.price_subtotal
                result.append(tmp)
            report_brand_overview_attachment = self.env['ir.attachment'].browse(
                self.env.ref('report_sale.report_revenue_service_template_attachment').id)
            decode = base64.b64decode(report_brand_overview_attachment.datas)
            wb = load_workbook(BytesIO(decode))
            ws = wb.active
            line_font = Font(name='Times New Roman', size=14)
            thin = borders.Side(style='thin')
            all_border_thin = borders.Border(left=thin, right=thin, top=thin, bottom=thin)

            # ws['F2'].value = 'BÁO CÁO DOANH THU THỰC HIỆN DỊCH VỤ %s' % name
            ws['C4'].value = self.start_date.strftime('%d/%m/%Y')
            ws['E4'].value = self.end_datetime.strftime('%d/%m/%Y')
            ws['A1'].value = self.institution.his_company.name.upper()

            key_col = list(range(1, len(TEMP_02) + 1))
            # format_currency = '_(* #,##0_);_(* (#,##0);_(* "-"??_);_(@_)'
            row = 8
            for line in result:
                for col, k in zip(key_col, TEMP_02):
                    cell = ws.cell(row, col)
                    cell.value = line[k]
                    cell.font = line_font
                    cell.border = all_border_thin
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                row += 1
        count = 0
        if self.report_type == 'patient':
            count += 1
            tmp = dict([(e, None) for e in TEMP_01])
            for data in datas:
                count += 1
                tmp['stt'] = count
                tmp['ma_benh_nhan'] = data.partner_id.code_customer
                tmp['ho_ten_khach_hang'] = data.partner_id.name
                tmp['dia_chi'] = data.partner_id.street or None
                tmp['ma_dich_vu'] = data.product_id.default_code
                tmp['dich_vu'] = data.product_id.name
                tmp['don_gia'] = data.price_unit
                tmp['tong_so_luong'] = data.quantity
                tmp['thanh_tien'] = data.price_subtotal
                tmp['mien_giam'] = data.discount
                tmp['tong_tien_da_thu'] = data.price_subtotal
                tmp['thoi_gian_thu'] = data.date.strftime('%d-%m-%Y')
                result.append(tmp)
            result.append(tmp)
            report_brand_overview_attachment = self.env['ir.attachment'].browse(
                self.env.ref('report_sale.report_revenue_patient_template').id)
            decode = base64.b64decode(report_brand_overview_attachment.datas)
            wb = load_workbook(BytesIO(decode))
            ws = wb.active
            line_font = Font(name='Times New Roman', size=14)
            thin = borders.Side(style='thin')
            all_border_thin = borders.Border(left=thin, right=thin, top=thin, bottom=thin)

            # ws['F2'].value = 'BÁO CÁO DOANH THU THỰC HIỆN DỊCH VỤ %s' % name
            ws['E4'].value = self.start_date.strftime('%d/%m/%Y')
            ws['H4'].value = self.end_datetime.strftime('%d/%m/%Y')
            ws['A1'].value = self.institution.his_company.name

            key_col = list(range(1, len(TEMP_01) + 1))
            format_currency = '_(* #,##0_);_(* (#,##0);_(* "-"??_);_(@_)'
            row = 8
            for line in result:
                for col, k in zip(key_col, TEMP_01):
                    cell = ws.cell(row, col)
                    cell.value = line[k]
                    cell.font = line_font
                    cell.border = all_border_thin
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                row += 1

        fp = BytesIO()
        wb.save(fp)
        fp.seek(0)
        report = base64.encodebytes((fp.read()))
        fp.close()
        attachment = self.env['ir.attachment'].sudo().create({
            'name': 'bao_cao_doanh_thu_theo_%s_%s.xlsx' % (self.report_type, self.institution.his_company.name),
            'datas': report,
            'res_model': 'temp.creation',
            'public': True,
        })
        url = "/web/content/?model=ir.attachment&id=%s&filename_field=name&field=datas&download=true" \
              % attachment.id
        return {'name': 'Báo cáo doanh thu theo bệnh nhân',
                'type': 'ir.actions.act_url',
                'url': url,
                'target': 'self',
                }

    # if self.report_type in ('patient'):
    #     return self.report_revenue_patient_excel()
    # else:
    #     return self.revenue_service_report()
