from odoo import fields, api, models, _
from odoo.exceptions import ValidationError, UserError
from datetime import date, datetime, time
from calendar import monthrange
from openpyxl import load_workbook
from openpyxl.styles import Font, borders, Alignment
import base64
from io import BytesIO
from dateutil.relativedelta import relativedelta
from pytz import timezone, utc
from odoo.addons.report_sale.wizard.theme_report import ThemeReport
from odoo.modules.module import get_module_resource
import openpyxl

thin = borders.Side(style='thin')
double = borders.Side(style='double')
all_border_thin = borders.Border(thin, thin, thin, thin)

WARD_TYPE = [
    ('All', 'Toàn bộ dịch vụ'),
    ('Surgery', 'Phẫu thuật'),
    ('Spa', 'Spa'),
    ('Laser', 'Laser'),
    ('Odontology', 'Nha khoa')
]


class ServiceRootBookingReport(models.TransientModel):
    _name = 'kpi.service.root.booking.report'
    _description = 'Bao cao doanh thu tru bao hanh'

    start_date = fields.Date('Start date', default=date.today())
    end_date = fields.Date('End date', default=date.today())
    company_id = fields.Many2one('res.company', string='Company', default=lambda self: self.env.company)
    type = fields.Selection(WARD_TYPE, string='Loại phiếu', default='All')

    # convert date to datetime for search domain, should be removed if using datetime directly
    start_datetime = fields.Datetime('Start datetime', compute='_compute_datetime')
    end_datetime = fields.Datetime('End datetime', compute='_compute_datetime')

    @api.depends('start_date', 'end_date')
    def _compute_datetime(self):
        self.start_datetime = False
        self.end_datetime = False
        if self.start_date and self.end_date:
            local_tz = timezone(self.env.user.tz or 'Etc/GMT+7')
            start_datetime = datetime(self.start_date.year, self.start_date.month, self.start_date.day, 0, 0, 0)
            end_datetime = datetime(self.end_date.year, self.end_date.month, self.end_date.day, 23, 59, 59)
            start_datetime = local_tz.localize(start_datetime, is_dst=None)
            end_datetime = local_tz.localize(end_datetime, is_dst=None)
            self.start_datetime = start_datetime.astimezone(utc).replace(tzinfo=None)
            self.end_datetime = end_datetime.astimezone(utc).replace(tzinfo=None)

    @api.onchange('start_date')
    def _onchange_start_date(self):
        if self.start_date:
            if self.start_date.month == fields.date.today().month:
                self.end_date = fields.date.today()
            else:
                self.end_date = date(self.start_date.year, self.start_date.month,
                                     monthrange(self.start_date.year, self.start_date.month)[1])

    @api.constrains('start_date', 'end_date')
    def check_dates(self):
        for record in self:
            start_date = fields.Date.from_string(record.start_date)
            end_date = fields.Date.from_string(record.end_date)
            if start_date > end_date:
                raise ValidationError(
                    _("Ngày kết thúc không thể ở trước ngày bắt đầu khi xuất báo cáo!!!"))

    def _get_data(self):
        ret_data = []
        total_money = 0
        # search crm_line: kết thúc thuộc booking bảo hành
        crm_line_ids = self.env['crm.line'].search(
            [('stage', '=', 'done'), ('crm_id.type_crm_id', '=', self.env.ref('crm_base.type_oppor_guarantee').id),
             ('crm_id.arrival_date', '>=', self.start_datetime), ('crm_id.arrival_date', '<=', self.end_datetime),
             ('price_list_id.type', '=', 'guarantee'), ('company_id', '=', self.company_id.id)])

        # if self.env.user.has_group('sci_hrms.group_user_doctor'):
        #     domain_surgery.append(('team_member', '=', self.env.user.id))
        #     domain_specialty.append(('team_member', '=', self.env.user.id))
        
        sheet = []

        for crm_line_id in crm_line_ids:
            # get type_guarantee
            if crm_line_id.type_guarantee == '1':
                type_guarantee = 'Một phần trước 01/06/2020'
            elif crm_line_id.type_guarantee == '2':
                type_guarantee = 'Một phần trước 01/10/2020'
            elif crm_line_id.type_guarantee == '3':
                type_guarantee = 'Một phần sau 01/06/2020'
            elif crm_line_id.type_guarantee == '4':
                type_guarantee = 'Một phần sau 01/10/2020'
            elif crm_line_id.type_guarantee == '5':
                type_guarantee = 'Toàn phần trước 01/06/2020'
            elif crm_line_id.type_guarantee == '6':
                type_guarantee = 'Toàn phần trước 01/10/2020'
            elif crm_line_id.type_guarantee == '7':
                type_guarantee = 'Toàn phần sau 01/06/2020'
            else:
                type_guarantee = 'Toàn phần sau 01/10/2020'

            # get service his
            service_his = self.env['sh.medical.health.center.service'].search(
                [('product_id', '=', crm_line_id.product_id.id)])

            # get booking
            booking = crm_line_id.crm_id
            # get phiếu chuyên khoa của booking bảo hành
            walkin_guarantee_ids = booking.walkin_ids
            for walkin_guarantee_id in walkin_guarantee_ids:
                if service_his in walkin_guarantee_id.service and walkin_guarantee_id.id not in sheet and walkin_guarantee_id.state == 'Completed':
                    # get phiếu khám gốc
                    sheet.append(walkin_guarantee_id.id)
                    root_walkin = walkin_guarantee_id.root_walkin

                    # get thành viên
                    # if root_walkin.room_type == 'Surgery':
                    if self.type == 'Surgery':
                        for rec in root_walkin.surgeries_ids:
                            teames = rec.surgery_team
                            # teames = self.env['sh.medical.surgery.team'].search([('name', '=', rec.id)])
                            for team in teames:
                                for ser in team.service_performances:
                                    # get sale order line for service
                                    sol = root_walkin.sale_order_id.order_line.filtered(
                                        lambda r: r.product_id == ser.product_id)
                                    money = sum([rec_sol.price_subtotal for rec_sol in sol])
                                    ret_data.append({
                                        'code_specialty_sheet': team.name.name,
                                        'code_booking': booking.name,
                                        'type_guarantee': type_guarantee,
                                        'customer': booking.partner_id.name,
                                        'service': ser.name,
                                        'kpi_point': ser.kpi_point,
                                        'money': money,
                                        'member': team.team_member.name,
                                        'role': team.role.name,
                                    })
                                    total_money += money
                    elif self.type == 'All':
                        for rec in root_walkin.surgeries_ids:
                            teames = rec.surgery_team
                            for team in teames:
                                for ser in team.service_performances:
                                    # get sale order line for service
                                    sol = root_walkin.sale_order_id.order_line.filtered(
                                        lambda r: r.product_id == ser.product_id)
                                    money = sum([rec_sol.price_subtotal for rec_sol in sol])
                                    ret_data.append({
                                        'code_specialty_sheet': team.name.name,
                                        'code_booking': booking.name,
                                        'type_guarantee': type_guarantee,
                                        'customer': booking.partner_id.name,
                                        'service': ser.name,
                                        'kpi_point': ser.kpi_point,
                                        'money': money,
                                        'member': team.team_member.name,
                                        'role': team.role.name,
                                    })
                                    total_money += money
                        for rec in root_walkin.specialty_ids:
                            teames = rec.specialty_team
                            for team in teames:
                                for ser in team.service_performances:
                                    # get sale order line for service
                                    sol = root_walkin.sale_order_id.order_line.filtered(
                                        lambda r: r.product_id == ser.product_id)
                                    money = sum([rec_sol.price_subtotal for rec_sol in sol])
                                    ret_data.append({
                                        'code_specialty_sheet': team.name.name,
                                        'code_booking': booking.name,
                                        'type_guarantee': type_guarantee,
                                        'customer': booking.partner_id.name,
                                        'service': ser.name,
                                        'kpi_point': ser.kpi_point,
                                        'money': money,
                                        'member': team.team_member.name,
                                        'role': team.role.name,
                                    })
                                    total_money += money
                    else:
                        for rec in root_walkin.specialty_ids:
                            teames = rec.specialty_team
                            # teames = self.env['sh.medical.surgery.team'].search([('name', '=', rec.id)])
                            for team in teames:
                                for ser in team.service_performances:
                                    # get sale order line for service
                                    sol = root_walkin.sale_order_id.order_line.filtered(
                                        lambda r: r.product_id == ser.product_id)
                                    ret_data.append({
                                        'code_specialty_sheet': team.name.name,
                                        'code_booking': booking.name,
                                        'type_guarantee': type_guarantee,
                                        'customer': booking.partner_id.name,
                                        'service': ser.name,
                                        'kpi_point': ser.kpi_point,
                                        'money': sol.price_subtotal,
                                        'member': team.team_member.name,
                                        'role': team.role.name,
                                    })
                                    total_money += sol.price_subtotal
        ret_data.append({
            'code_specialty_sheet': None,
            'code_booking': None,
            'type_guarantee': None,
            'customer': None,
            'service': None,
            'kpi_point': None,
            'money': total_money,
            'member': None,
            'role': None,
        })
        return ret_data

    def create_service_root_booking_report(self):
        # get data
        datas = self._get_data()

        # in dữ liệu
        report_brand_overview_attachment = self.env['ir.attachment'].browse(
            self.env.ref('report_sale.bao_cao_doanh_thu_tru_bao_hanh_attachment').id)
        decode = base64.b64decode(report_brand_overview_attachment.datas)
        wb = load_workbook(BytesIO(decode))
        ws = wb.active
        line_font = Font(name='Times New Roman', size=14)

        code_brand = self.company_id.brand_id.code.lower()
        image_path = get_module_resource('report_sale', 'static/img', 'icon_%s.png' % code_brand)

        img = openpyxl.drawing.image.Image(image_path)
        img.anchor = 'A1'
        ws.add_image(img)

        if self.type == 'Surgery':
            name = "PHẪU THUẬT"
        elif self.type == 'Spa':
            name = "SPA"
        elif self.type == 'Laser':
            name = "LASER"
        elif self.type == 'Odontology':
            name = "NHA KHOA"
        else:
            name = 'Toàn bộ dịch vụ'.upper()

        if code_brand == 'kn':
            header_fill = ThemeReport.kn_fill
        elif code_brand == 'da':
            header_fill = ThemeReport.da_fill
        elif code_brand == 'pr':
            header_fill = ThemeReport.pr_fill
        elif code_brand == 'hh':
            header_fill = ThemeReport.hh_fill
        else:
            header_fill = ThemeReport.sci_fill

        ws['A2'].value = 'BÁO CÁO DOANH THU TRỪ BẢO HÀNH( %s)' % name
        ws['D3'].value = self.start_date.strftime('%d/%m/%Y')
        ws['F3'].value = self.end_datetime.strftime('%d/%m/%Y')
        ws['D4'].value = self.company_id.name

        key_col = list(range(1, 10))
        key_list = [
            'code_specialty_sheet',
            'code_booking',
            'type_guarantee',
            'customer',
            'service',
            'kpi_point',
            'money',
            'member',
            'role',
        ]
        row = 7
        for data in datas:
            for col, k in zip(key_col, key_list):
                beforeCell = ws.cell(6, col)
                beforeCell.fill = header_fill
                beforeCell.font = Font(name='Times New Roman', size=14, color='FFFFFF')
                cell = ws.cell(row, col)
                cell.value = data[k]
                cell.font = line_font
                cell.border = all_border_thin
                cell.alignment = Alignment(horizontal='left', vertical='center')
            row += 1
        fp = BytesIO()
        wb.save(fp)
        fp.seek(0)
        report = base64.encodebytes((fp.read()))
        fp.close()
        attachment = self.env['ir.attachment'].sudo().create({
            'name': 'bao_cao_doanh_thu_tru_bao_hanh_%s.xlsx' % self.company_id.name,
            'datas': report,
            'res_model': 'temp.creation',
            'public': True,
        })
        url = "/web/content/?model=ir.attachment&id=%s&filename_field=name&field=datas&download=true" \
              % attachment.id
        return {'name': 'BÁO CÁO DOANH THU TRỪ BẢO HÀNH',
                'type': 'ir.actions.act_url',
                'url': url,
                'target': 'self',
                }
