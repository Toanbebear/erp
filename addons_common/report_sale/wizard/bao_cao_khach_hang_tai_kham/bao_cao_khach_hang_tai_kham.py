# -*- coding: utf-8 -*-
import openpyxl

from odoo import fields, api, models, _
from odoo.exceptions import ValidationError, UserError
from datetime import date, datetime, time
from calendar import monthrange
from openpyxl import load_workbook
from openpyxl.styles import Font, borders, Alignment, PatternFill
from odoo.modules.module import get_module_resource
import base64
from io import BytesIO
from odoo.addons.report_sale.wizard.theme_report import ThemeReport
from dateutil.relativedelta import relativedelta
from pytz import timezone, utc

import logging

_logger = logging.getLogger(__name__)


class BaoCaoKhachHangTaiKham(models.TransientModel):
    _name = 'report.customer.evaluation'
    _description = 'Bao cao khach hang tai kham'

    start_date = fields.Date('Start date', default=date.today().replace(day=1))
    end_date = fields.Date('End date')
    company_id = fields.Many2one('res.company', string='Company', default=lambda self: self.env.company,
                                 domain=lambda self: [('id', 'in', self.env.companies.ids)])
    # convert date to datetime for search domain, should be removed if using datetime directly
    start_datetime = fields.Datetime('Start datetime', compute='_compute_datetime')
    end_datetime = fields.Datetime('End datetime', compute='_compute_datetime')

    # convert date to datetime for search domain, should be removed if using datetime directly
    @api.depends('start_date', 'end_date')
    def _compute_datetime(self):
        self.start_datetime = False
        self.end_datetime = False
        if self.start_date and self.end_date:
            local_tz = timezone(self.env.user.tz or 'Etc/GMT+7')
            start_datetime = datetime(self.start_date.year, self.start_date.month, self.start_date.day, 0, 0, 0)
            end_datetime = datetime(self.end_date.year, self.end_date.month, self.end_date.day, 23, 59, 59)
            start_datetime = local_tz.localize(start_datetime, is_dst=None)
            end_datetime = local_tz.localize(end_datetime, is_dst=None)
            self.start_datetime = start_datetime.astimezone(utc).replace(tzinfo=None)
            self.end_datetime = end_datetime.astimezone(utc).replace(tzinfo=None)

    @api.onchange('start_date')
    def _onchange_start_date(self):
        if self.start_date:
            if self.start_date.month == fields.date.today().month:
                self.end_date = fields.date.today()
            else:
                self.end_date = date(self.start_date.year, self.start_date.month,
                                     monthrange(self.start_date.year, self.start_date.month)[1])

    @api.constrains('start_date', 'end_date')
    def check_dates(self):
        for record in self:
            start_date = fields.Date.from_string(record.start_date)
            end_date = fields.Date.from_string(record.end_date)
            if start_date > end_date:
                raise ValidationError(
                    _("End Date cannot be set before Start Date."))

    def get_data_report_customer_evaluation(self):
        ret_data = []
        evaluation = self.env['sh.medical.evaluation'].sudo().search(
            [('institution.his_company', '=', self.company_id.id), ('evaluation_start_date', '>=', self.start_date),
             ('evaluation_start_date', '<=', self.end_date)], order='evaluation_start_date asc')
        # phân loại phiếu TK
        dict_patient_level = {
            '5': 'Khách hàng V.I.P',
            '4': 'Đặc biệt',
            '3': 'Quan tâm hơn',
            '2': 'Quan tâm',
            '1': 'Bình thường'
        }

        # đánh giá khách hàng
        customer_reviews = {
            'Normal': 'Bình thường',
            'WellPleased': 'Rất hài lòng',
            'Satisfied': 'Hài lòng',
            'Unsatisfied': 'Không hài lòng',
            'Dissatisfaction': 'Rất không hài lòng'
        }

        # vai trò bác sĩ phẫu thuật
        role_bs_pt = self.env['sh.medical.team.role'].sudo().search([('name', '=', 'Bác sĩ tái khám Phẫu thuật')],
                                                                    limit=1).id
        for rec_evaluation in evaluation:
            # bác sĩ phẫu thuật chính & phụ
            main_doctor = []
            sub_doctor = []
            if rec_evaluation.surgery_history_ids:
                for rec_surgery_history_id_doctor in rec_evaluation.surgery_history_ids.main_doctor:
                    main_doctor.append(rec_surgery_history_id_doctor.name)
                for rec_surgery_history_id_nurse in rec_evaluation.surgery_history_ids.sub_doctor:
                    main_doctor.append(rec_surgery_history_id_nurse.name)

            # dịch vụ
            services = []
            if rec_evaluation.services:
                for rec_service in rec_evaluation.services:
                    services.append(rec_service.name)

            # ngày phẫu thuật
            surgeries_date = []
            if rec_evaluation.walkin.surgeries_ids:
                for rec_surgery_id in rec_evaluation.walkin.surgeries_ids:
                    surgeries_date.append(rec_surgery_id.surgery_date.strftime('%d/%m/%Y'))

            # bác sĩ tái khám
            evaluation_team_doctor = []
            if rec_evaluation.evaluation_team:
                for rec_evaluation_team in rec_evaluation.evaluation_team:
                    if rec_evaluation_team.role.id == role_bs_pt:
                        evaluation_team_doctor.append(rec_evaluation_team.team_member.name)

            # loại tái khám
            evaluation_services = []
            if rec_evaluation.evaluation_services:
                for rec_evaluation_services in rec_evaluation.evaluation_services:
                    evaluation_services.append(rec_evaluation_services.name)

            # điều dưỡng tái khám
            evaluation_team_nurse = []
            if rec_evaluation.evaluation_team:
                for rec_evaluation_team in rec_evaluation.evaluation_team:
                    if 'Điều dưỡng' in rec_evaluation_team.role.name:
                        evaluation_team_nurse.append(rec_evaluation_team.team_member.name)
            value = {
                'evaluation_start_date': rec_evaluation.evaluation_start_date.strftime('%d/%m/%Y'),
                'patient_name': rec_evaluation.patient.name,
                'booking_code': rec_evaluation.walkin.booking_id.name,
                'evaluation_name': rec_evaluation.name,
                'patient_level': dict_patient_level[
                    '%s' % str(rec_evaluation.patient_level)] if rec_evaluation.patient_level else '',
                'main_doctor': ', '.join([a for a in main_doctor if a]),
                'sub_doctor': ', '.join([a for a in sub_doctor if a]),
                'services': ', '.join([a for a in services if a]),
                'surgery_date': ', '.join([a for a in surgeries_date if a]),
                'evaluation_team_doctor': ', '.join([a for a in evaluation_team_doctor if a]),
                'customer_reviews': customer_reviews['%s' % rec_evaluation.customer_reviews],
                'evaluation_services': ', '.join([a for a in evaluation_services if a]),
                'evaluation_team_nurse': ', '.join([a for a in evaluation_team_nurse if a]),
                'notes_complaint': rec_evaluation.notes_complaint if rec_evaluation.notes_complaint else '',
                'chief_complaint': rec_evaluation.chief_complaint if rec_evaluation.chief_complaint else '',
                'next_appointment_date': rec_evaluation.next_appointment_date if rec_evaluation.next_appointment_date else '',
                'doctor_bh': rec_evaluation.doctor_bh.name if rec_evaluation.doctor_bh else '',
                'warranty_appointment_date': rec_evaluation.warranty_appointment_date if rec_evaluation.warranty_appointment_date else '',
                'room': rec_evaluation.room.name if rec_evaluation.room.name else '',
                'institution': rec_evaluation.institution.name,
            }
            ret_data.append(value)
        return ret_data

    def create_report_customer_evaluation(self):
        report_khach_hang_tai_kham_attachment = self.env['ir.attachment'].browse(
            self.env.ref('report_sale.report_khach_hang_tai_kham_attachment').id)
        decode = base64.b64decode(report_khach_hang_tai_kham_attachment.datas)
        wb = load_workbook(BytesIO(decode))
        ws = wb.active

        datas = self.get_data_report_customer_evaluation()

        key_list = [
            'evaluation_start_date',
            'patient_name',
            'booking_code',
            'evaluation_name',
            'patient_level',
            'main_doctor',
            'sub_doctor',
            'services',
            'surgery_date',
            'evaluation_team_doctor',
            'customer_reviews',
            'evaluation_services',
            'evaluation_team_nurse',
            'notes_complaint',
            'chief_complaint',
            'next_appointment_date',
            'doctor_bh',
            'warranty_appointment_date',
            'room',
            'institution',
        ]
        key_list_title = [
            'STT',
            'NGÀY TÁI KHÁM',
            'TÊN KHÁCH HÀNG',
            'MÃ BOOKING',
            'MÃ PHIẾU TÁI KHÁM',
            'PHÂN LOẠI PHIẾU TK',
            'BÁC SĨ PHÃU THUẬT (CHÍNH)',
            'BÁC SĨ PHÃU THUẬT (PHỤ)',
            'DỊCH VỤ',
            'NGÀY PHẪU THUẬT',
            'BÁC SĨ TÁI KHÁM',
            'ĐÁNH GIÁ CỦA KH',
            'LOẠI TÁI KHÁM',
            'ĐIỀU DƯỠNG TÁI KHÁM',
            'TÓM TẮT TÁI KHÁM',
            'TÌNH TRẠNG KHÁCH HÀNG',
            'NGÀY HẸN TK TIẾP',
            'BS CHỈ ĐỊNH BẢO HÀNH',
            'NGÀY HẸN BẢO HÀNH',
            'PHÒNG',
            'CỚ SỞ Y TẾ'
        ]

        key_col_list = list(range(2, len(key_list) + 2))
        key_col_list_title = list(range(1, len(key_list) + 2))

        code_brand = self.company_id.brand_id.code.lower()
        if code_brand == 'kn':
            header_fill = ThemeReport.kn_fill
        elif code_brand == 'da':
            header_fill = ThemeReport.da_fill
        elif code_brand == 'pr':
            header_fill = ThemeReport.pr_fill
        elif code_brand == 'hh':
            header_fill = ThemeReport.hh_fill
        else:
            header_fill = ThemeReport.sci_fill

        # in tiêu đề
        row = 5
        index_row = 0
        for col, k in zip(key_col_list_title, key_list_title):
            cell = ws.cell(row, col)
            cell.fill = header_fill
            cell.value = k
            cell.font = Font(name='Times New Roman', size=12, color='FFFFFF')
            cell.border = ThemeReport.all_border_thin

        # in data
        row = 6
        for line_data in datas:
            ws.cell(row, 1).border = ThemeReport.all_border_thin
            ws.cell(row, 1).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(row, 1).value = index_row + 1
            for col, k in zip(key_col_list, key_list):
                cell = ws.cell(row, col)
                cell.value = line_data[k]
                cell.font = ThemeReport.line_font
                cell.border = ThemeReport.all_border_thin
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            row += 1
            index_row += 1

        ws['I2'].value = 'Đơn vị: %s' % self.company_id.name
        ws['I2'].font = Font(name='Times New Roman', size=13)
        ws['I3'].value = 'Từ ngày: %s đến ngày: %s' % (
            self.start_date.strftime('%d/%m/%Y'), self.end_date.strftime('%d/%m/%Y'))
        ws['I3'].font = Font(name='Times New Roman', size=13)

        code_brand = self.company_id.brand_id.code.lower()
        image_path = get_module_resource('report_sale', 'static/img', 'icon_%s.png' % code_brand)
        img = openpyxl.drawing.image.Image(image_path)
        img.anchor = 'A1'
        ws.add_image(img)

        fp = BytesIO()
        wb.save(fp)
        fp.seek(0)
        report = base64.encodebytes((fp.read()))
        fp.close()
        attachment = self.env['ir.attachment'].with_user(1).create({'name': 'bao_cao_khach_hang_tai_kham.xlsx',
                                                                    'datas': report,
                                                                    'res_model': 'temp.creation',
                                                                    'public': True})
        url = "/web/content/?model=ir.attachment&id=%s&filename_field=name&field=datas&download=true" \
              % attachment.id
        return {'name': 'Báo cáo khách hàng tái khám',
                'type': 'ir.actions.act_url',
                'url': url,
                'target': 'self',
                }
