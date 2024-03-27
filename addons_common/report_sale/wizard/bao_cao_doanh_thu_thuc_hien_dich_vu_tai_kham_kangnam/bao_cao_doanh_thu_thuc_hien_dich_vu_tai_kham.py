from odoo import fields, api, models, _
from odoo.exceptions import ValidationError, UserError
from datetime import date, datetime, time
from calendar import monthrange
from openpyxl import load_workbook
from openpyxl.styles import Font, borders, Alignment
import base64
from io import BytesIO
from dateutil.relativedelta import relativedelta
from pytz import timezone, utc
from odoo.addons.report_sale.wizard.theme_report import ThemeReport
from odoo.modules.module import get_module_resource
import openpyxl

thin = borders.Side(style='thin')
double = borders.Side(style='double')
all_border_thin = borders.Border(thin, thin, thin, thin)
SERVICE_HIS_TYPE = [('Spa', 'Spa'), ('Laser', 'Laser'), ('Odontology', 'Nha'), ('Surgery', 'Phẫu thuật'), ('ChiPhi', 'Chi phí khác')]
dict_type = dict((key, value) for key, value in SERVICE_HIS_TYPE)
SURGERY_TYPE = [('minor', 'Tiểu phẫu'), ('major', 'Đại phẫu')]
DICT_SURGERY_TYPE = dict((key, value) for key, value in SURGERY_TYPE)
TEMP = [
    'employee_code',
    'member',
    'group_job',
    'job_id',
    'role',
    'name_customer',
    'code_booking',
    'code_evaluation',
    'start_date',
    'surgery_type',
    'his_service_type',
    'service_category',
    'service',
    'state',
    'gender',
    'birth_date',
    'street',
    'patient_level',
    'notes_complaint',
    'doctor_bh',
    'warranty_appointment_date',
    # 'type_service',
    'end_date',
    'revenue',
    'code_walkin',
    'type_booking',
    'institution'
]
LEVEL = [
    ('4', 'Đặc biệt'),
    ('3', 'Quan tâm hơn'),
    ('2', 'Quan tâm'),
    ('1', 'Bình thường')
]
DICT_LEVEL = dict((key, value) for key, value in LEVEL)


class ServiceEvaluationPerformanceReportKangnam(models.TransientModel):
    _name = 'kpi.service.evaluation.perform.report.kangnam'
    _description = 'Bao cao doanh thu thuc hien dich vu tái khám'

    start_date = fields.Date('Start date', default=date.today())
    end_date = fields.Date('End date', default=date.today())
    company_id = fields.Many2one('res.company', domain="[('brand_id.name', '=', 'Kangnam')]", string='Company', default=lambda self: self.env.company)

    # convert date to datetime for search domain, should be removed if using datetime directly
    start_datetime = fields.Datetime('Start datetime', compute='_compute_datetime')
    end_datetime = fields.Datetime('End datetime', compute='_compute_datetime')

    @api.depends('start_date', 'end_date')
    def _compute_datetime(self):
        self.start_datetime = False
        self.end_datetime = False
        if self.start_date and self.end_date:
            local_tz = timezone(self.env.user.tz or 'Etc/GMT+7')
            start_datetime = datetime(self.start_date.year, self.start_date.month, self.start_date.day, 0, 0, 0)
            end_datetime = datetime(self.end_date.year, self.end_date.month, self.end_date.day, 23, 59, 59)
            start_datetime = local_tz.localize(start_datetime, is_dst=None)
            end_datetime = local_tz.localize(end_datetime, is_dst=None)
            self.start_datetime = start_datetime.astimezone(utc).replace(tzinfo=None)
            self.end_datetime = end_datetime.astimezone(utc).replace(tzinfo=None)

    @api.onchange('start_date')
    def _onchange_start_date(self):
        if self.start_date:
            if self.start_date.month == fields.date.today().month:
                self.end_date = fields.date.today()
            else:
                self.end_date = date(self.start_date.year, self.start_date.month,
                                     monthrange(self.start_date.year, self.start_date.month)[1])

    @api.constrains('start_date', 'end_date')
    def check_dates(self):
        for record in self:
            start_date = fields.Date.from_string(record.start_date)
            end_date = fields.Date.from_string(record.end_date)
            if start_date > end_date:
                raise ValidationError(
                    _("Ngày kết thúc không thể ở trước ngày bắt đầu khi xuất báo cáo!!!"))

    def _get_data(self):
        ret_data = []
        datas = self.env['sh.medical.evaluation.team'].search(
            [('name.evaluation_start_date', '>=', self.start_datetime),
             ('name.evaluation_start_date', '<=', self.end_datetime),
             ('name.institution.his_company', '=', self.company_id.id), ('name.state', '=', 'Completed')])

        total_revenue = 0
        for data in datas:
            evaluation = data.name
            walkin = evaluation.walkin
            partner = evaluation.patient.partner_id
            booking = walkin.booking_id
            list_so = booking.order_ids.filtered(lambda r: r.state == 'sale')

            for ser_evaluation in data.service_performances:
                # get revenue of service from all so
                revenue_ser = 0
                for so in list_so:
                    sol = so.order_line.filtered(lambda r: r.product_id == ser_evaluation.product_id)
                    if sol:
                        for rec_sol in sol:
                            revenue_ser += rec_sol.price_subtotal

                ret_data.append({
                    'employee_code': data.team_member.employee_code,
                    'member': data.team_member.name,
                    'group_job': data.team_member.group_job.name or None,
                    'job_id': data.team_member.job_id.name or None,
                    'role': data.role.name,
                    'name_customer': booking.partner_id.name,
                    'code_booking': booking.name,
                    'code_evaluation': data.name.name,
                    'start_date': data.name.evaluation_start_date.strftime('%d/%m/%Y'),
                    'surgery_type': DICT_SURGERY_TYPE.get(ser_evaluation.surgery_type) if ser_evaluation.surgery_type else None,
                    'his_service_type': dict_type.get(ser_evaluation.his_service_type),
                    'service_category': ser_evaluation.service_category.name,
                    'service': ser_evaluation.name,
                    'state': data.name.state,
                    'gender': evaluation.patient.gender,
                    'birth_date': partner.birth_date,
                    'street': partner.street,
                    'patient_level': DICT_LEVEL.get(evaluation.patient_level, None),
                    'notes_complaint': evaluation.notes_complaint or None,
                    'doctor_bh': evaluation.doctor_bh.name or None,
                    'warranty_appointment_date': evaluation.warranty_appointment_date.strftime('%d/%m/%Y') if evaluation.warranty_appointment_date else None,
                    # 'type_service': dict_type.get(ser_evaluation.his_service_type, None),
                    'end_date': data.name.evaluation_end_date.strftime('%d/%m/%Y') if data.name.evaluation_end_date else None,
                    'revenue': revenue_ser,
                    'code_walkin': walkin.name,
                    'type_booking': booking.type_crm_id.name,
                    'institution': data.name.institution.name,
                })
                # add total revenue
                total_revenue += revenue_ser
        val = dict((e, None)for e in TEMP)
        val['revenue'] = total_revenue
        ret_data.append(val)
        return ret_data

    def create_report_service_evaluation_perform_report(self):
        # get data
        datas = self._get_data()

        # in dữ liệu
        report_brand_overview_attachment = self.env['ir.attachment'].browse(
            self.env.ref('report_sale.bao_cao_doanh_thu_thuc_hien_dich_vu_tai_kham_kangnam_attachment').id)
        decode = base64.b64decode(report_brand_overview_attachment.datas)
        wb = load_workbook(BytesIO(decode))
        ws = wb.active
        line_font = Font(name='Times New Roman', size=14)
        ws['F4'].value += self.start_date.strftime('%d/%m/%Y')
        ws['G4'].value += self.end_datetime.strftime('%d/%m/%Y')
        ws['G5'].value = self.company_id.name
        format_currency = '_(* #,##0_);_(* (#,##0);_(* "-"??_);_(@_)'

        code_brand = self.company_id.brand_id.code.lower()
        image_path = get_module_resource('report_sale', 'static/img', 'icon_%s.png' % code_brand)

        img = openpyxl.drawing.image.Image(image_path)
        img.anchor = 'A1'
        ws.add_image(img)

        key_col = list(range(1, len(TEMP) + 1))

        row = 8
        if code_brand == 'kn':
            header_fill = ThemeReport.kn_fill
        elif code_brand == 'da':
            header_fill = ThemeReport.da_fill
        elif code_brand == 'pr':
            header_fill = ThemeReport.pr_fill
        elif code_brand == 'hh':
            header_fill = ThemeReport.hh_fill
        else:
            header_fill = ThemeReport.sci_fill
        for data in datas:
            for col, k in zip(key_col, TEMP):
                beforeCell = ws.cell(7, col)
                beforeCell.fill = header_fill
                beforeCell.font = Font(name='Times New Roman', size=14, color='FFFFFF')
                cell = ws.cell(row, col)
                cell.value = data[k]
                cell.font = line_font
                cell.border = all_border_thin
                cell.alignment = Alignment(horizontal='left', vertical='center')
                if col == 24:
                    cell.number_format = format_currency
            row += 1
        # cell = ws.cell(row, 1)
        # cell.value = 'TỔNG TIỀN'
        # cell.font = line_font
        # cell.border = all_border_thin
        # cell.alignment = Alignment(horizontal='center', vertical='center')
        # cell = ws.cell(row, 7)
        # cell.value = int(datas['total_revenue'])
        # cell.font = line_font
        # cell.border = all_border_thin
        # cell.alignment = Alignment(horizontal='left', vertical='center')
        # ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)

        fp = BytesIO()
        wb.save(fp)
        fp.seek(0)
        report = base64.encodebytes((fp.read()))
        fp.close()
        attachment = self.env['ir.attachment'].sudo().create({
            'name': 'bao_cao_doanh_thu_thuc_hien_dich_vu_tai_kham_%s.xlsx' % self.company_id.name,
            'datas': report,
            'res_model': 'temp.creation',
            'public': True,
        })
        url = "/web/content/?model=ir.attachment&id=%s&filename_field=name&field=datas&download=true" \
              % attachment.id
        return {'name': 'BÁO CÁO DOANH THU THỰC HIỆN DỊCH VỤ TÁI KHÁM',
                'type': 'ir.actions.act_url',
                'url': url,
                'target': 'self',
                }
