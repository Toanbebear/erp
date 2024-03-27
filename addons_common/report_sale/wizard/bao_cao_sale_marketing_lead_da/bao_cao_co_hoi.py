from odoo import fields, api, models, _
from odoo.exceptions import ValidationError, UserError
from datetime import date, datetime, time
from calendar import monthrange
from openpyxl import load_workbook
from openpyxl.styles import Font, borders, Alignment
import base64
from io import BytesIO
from dateutil.relativedelta import relativedelta
from pytz import timezone, utc

thin = borders.Side(style='thin')
double = borders.Side(style='double')
all_border_thin = borders.Border(thin, thin, thin, thin)


class SaleMarketingLeadDA(models.TransientModel):
    _name = 'sale.marketing.lead.da'
    _description = 'Báo cáo cơ hội'

    start_date = fields.Date('Start date', default=date.today())
    end_date = fields.Date('End date', default=date.today())
    # brand_id = fields.Many2one(string='Thương hiệu', comodel_name='res.brand', domain=lambda self: [('id', 'in', self.env.user.company_ids.mapped('brand_id').ids)])
    company_ids = fields.Many2many(string='Chi nhánh', comodel_name='res.company', domain=lambda self: [('id', 'in', self.env.user.company_ids.ids)])

    # convert date to datetime for search domain, should be removed if using datetime directly
    start_datetime = fields.Datetime('Start datetime', compute='_compute_datetime')
    end_datetime = fields.Datetime('End datetime', compute='_compute_datetime')

    @api.depends('start_date', 'end_date')
    def _compute_datetime(self):
        self.start_datetime = False
        self.end_datetime = False
        if self.start_date and self.end_date:
            local_tz = timezone(self.env.user.tz or 'Etc/GMT+7')
            start_datetime = datetime(self.start_date.year, self.start_date.month, self.start_date.day, 0, 0, 0)
            end_datetime = datetime(self.end_date.year, self.end_date.month, self.end_date.day, 23, 59, 59)
            start_datetime = local_tz.localize(start_datetime, is_dst=None)
            end_datetime = local_tz.localize(end_datetime, is_dst=None)
            self.start_datetime = start_datetime.astimezone(utc).replace(tzinfo=None)
            self.end_datetime = end_datetime.astimezone(utc).replace(tzinfo=None)

    @api.onchange('start_date')
    def _onchange_start_date(self):
        if self.start_date:
            if self.start_date.month == fields.date.today().month:
                self.end_date = fields.date.today()
            else:
                self.end_date = date(self.start_date.year, self.start_date.month,
                                     monthrange(self.start_date.year, self.start_date.month)[1])

    @api.constrains('start_date', 'end_date')
    def check_dates(self):
        for record in self:
            start_date = fields.Date.from_string(record.start_date)
            end_date = fields.Date.from_string(record.end_date)
            days = (end_date - start_date).days
            if days < 0 or days > 365:
                raise ValidationError(
                    _("Ngày kết thúc không thể ở trước ngày bắt đầu khi xuất báo cáo!!!"))

    def render_form_template(self, crm_line):
        Crm = crm_line.crm_id
        str_tuple = (Crm.pass_port_address, Crm.district_id.name, Crm.state_id.name, Crm.country_id.name)
        str_convert = (element for element in str_tuple if element != False)
        dia_chi = ', '.join(str_convert)
        product_code = crm_line.service_id.code
        product_name = crm_line.service_id.name
        product_categ = crm_line.service_id.service_category

        lang_code = self.env.context.get('lang') or 'en_US'
        date_format = self.env['res.lang']._lang_get(lang_code).date_format
        time_format = self.env['res.lang']._lang_get(lang_code).time_format

        val = {
            'thoi_gian_tao': '%s' % fields.Datetime.context_timestamp(self.with_context(tz=self.env.user.tz), Crm.create_on).strftime(('%s %s') % (date_format, time_format)),
            'chi_nhanh': Crm.company_id.name,
            'nguoi_tao': Crm.create_by.name,
            'phong_ban': self.env['hr.employee'].sudo().search([('user_id', '=', Crm.create_by.id)], limit=1).department_id.name or None,
            'kieu_du_lieu': Crm.type_data,
            'nhom_nguon': Crm.category_source_id.name,
            'nguon': Crm.source_id.name,
            'chien_dich': Crm.campaign_id.name,
            'trang_thai': Crm.stage_id.name,
            'ma_dich_vu': product_code,
            'ten_dich_vu': product_name,
            'nhom_dich_vu': product_categ.name,
            'ho_ten_khach_hang': Crm.contact_name,
            # 'so_dien_thoai': Crm.phone,
            'so_dien_thoai': None,
            'dia_chi': dia_chi,
            'tinh_tp': Crm.state_id.name or None,
        }
        return val

    def _get_data_report(self):
        domain = [('stage', '!=', 'cancel'), ('crm_id.type', '=', 'lead'),
                  ('crm_id.company_id', 'in', self.company_ids.ids),
                  ('crm_id.create_on', '>=', self.start_datetime), ('crm_id.create_on', '<=', self.end_datetime)]
        CrmLine = self.env['crm.line'].sudo().search(domain, order='id')
        return_val = []
        for line in CrmLine:
            return_val.append(self.render_form_template(line))
        return return_val

    def create_report_da(self):
        datas = self._get_data_report()
        # in dữ liễu
        report_brand_overview_attachment = self.env['ir.attachment'].browse(
            self.env.ref('report_sale.bao_cao_co_hoi_attachment').id)
        decode = base64.b64decode(report_brand_overview_attachment.datas)
        wb = load_workbook(BytesIO(decode))
        ws = wb.active
        line_font = Font(name='Times New Roman', size=14)

        ws['H3'].value = self.start_date.strftime('%d/%m/%Y')
        ws['J3'].value = self.end_datetime.strftime('%d/%m/%Y')
        key_col = list(range(1, 17))

        key_list = [
            'thoi_gian_tao',
            'chi_nhanh',
            'nguoi_tao',
            'phong_ban',
            'kieu_du_lieu',
            'nhom_nguon',
            'nguon',
            'chien_dich',
            'trang_thai',
            'ma_dich_vu',
            'ten_dich_vu',
            'nhom_dich_vu',
            'ho_ten_khach_hang',
            'so_dien_thoai',
            'dia_chi',
            'tinh_tp',
        ]
        row = 7
        for data in datas:
            for col, k in zip(key_col, key_list):
                cell = ws.cell(row, col)
                cell.value = data[k]
                cell.font = line_font
                cell.border = all_border_thin
                cell.alignment = Alignment(horizontal='left', vertical='center')
            row += 1
        fp = BytesIO()
        wb.save(fp)
        fp.seek(0)
        report = base64.encodebytes((fp.read()))
        fp.close()
        attachment = self.env['ir.attachment'].sudo().create({
            'name': 'bao_cao_co_hoi.xlsx',
            'datas': report,
            'res_model': 'temp.creation',
            'public': True,
        })
        url = "/web/content/?model=ir.attachment&id=%s&filename_field=name&field=datas&download=true" \
              % attachment.id
        return {'name': 'BÁO CÁO CƠ HỘI',
                'type': 'ir.actions.act_url',
                'url': url,
                'target': 'self',
                }
