from odoo import fields, api, models, _
from odoo.exceptions import ValidationError, UserError
from datetime import date, datetime, time
from calendar import monthrange
from openpyxl import load_workbook
from openpyxl.styles import Font, borders, Alignment
import base64
from io import BytesIO
from dateutil.relativedelta import relativedelta
from pytz import timezone, utc
from odoo.addons.report_sale.wizard.theme_report import ThemeReport
from odoo.modules.module import get_module_resource
import openpyxl


thin = borders.Side(style='thin')
double = borders.Side(style='double')
all_border_thin = borders.Border(thin, thin, thin, thin)

WARD_TYPE = [
    ('All', 'Toàn bộ dịch vụ'),
    ('Surgery', 'Phẫu thuật'),
    ('Spa', 'Spa'),
    ('Laser', 'Laser'),
    ('Odontology', 'Nha khoa')
]
TYPE_GUARANTEE = [('1', 'Một phần trước 01/06/2020'), ('2', 'Một phần trước 01/10/2020'),
                  ('3', 'Một phần sau 01/06/2020'), ('4', 'Một phần sau 01/10/2020'),
                  ('5', 'Toàn phần trước 01/06/2020'), ('6', 'Toàn phần trước 01/10/2020'),
                  ('7', 'Toàn phần sau 01/06/2020'), ('8', 'Toàn phần sau 01/10/2020'), ('9', 'Bảo hành không do lỗi chuyên môn'), ('10', 'Bảo hành chung (TH Paris)')]
DICT_TYPE_GUARANTEE = dict((key, value) for key, value in TYPE_GUARANTEE)

SERVICE_HIS_TYPE = [('Spa', 'Spa'), ('Laser', 'Laser'), ('Odontology', 'Nha'), ('Surgery', 'Phẫu thuật'), ('ChiPhi', 'Chi phí khác')]
dict_type = dict((key, value) for key, value in SERVICE_HIS_TYPE)


class ServiceInsurancePerformanceReport(models.TransientModel):
    _name = 'kpi.service.insurance.perform.report'
    _description = 'Bao cao doanh thu thuc hien dich vu bao hanh'

    start_date = fields.Date('Start date', default=date.today())
    end_date = fields.Date('End date', default=date.today())
    company_id = fields.Many2one('res.company', string='Company', default=lambda self: self.env.company)
    type = fields.Selection(WARD_TYPE, string='Loại phiếu', default='All')

    # convert date to datetime for search domain, should be removed if using datetime directly
    start_datetime = fields.Datetime('Start datetime', compute='_compute_datetime')
    end_datetime = fields.Datetime('End datetime', compute='_compute_datetime')

    @api.depends('start_date', 'end_date')
    def _compute_datetime(self):
        self.start_datetime = False
        self.end_datetime = False
        if self.start_date and self.end_date:
            local_tz = timezone(self.env.user.tz or 'Etc/GMT+7')
            start_datetime = datetime(self.start_date.year, self.start_date.month, self.start_date.day, 0, 0, 0)
            end_datetime = datetime(self.end_date.year, self.end_date.month, self.end_date.day, 23, 59, 59)
            start_datetime = local_tz.localize(start_datetime, is_dst=None)
            end_datetime = local_tz.localize(end_datetime, is_dst=None)
            self.start_datetime = start_datetime.astimezone(utc).replace(tzinfo=None)
            self.end_datetime = end_datetime.astimezone(utc).replace(tzinfo=None)

    @api.onchange('start_date')
    def _onchange_start_date(self):
        if self.start_date:
            if self.start_date.month == fields.date.today().month:
                self.end_date = fields.date.today()
            else:
                self.end_date = date(self.start_date.year, self.start_date.month,
                                     monthrange(self.start_date.year, self.start_date.month)[1])

    @api.constrains('start_date', 'end_date')
    def check_dates(self):
        for record in self:
            start_date = fields.Date.from_string(record.start_date)
            end_date = fields.Date.from_string(record.end_date)
            if start_date > end_date:
                raise ValidationError(
                    _("Ngày kết thúc không thể ở trước ngày bắt đầu khi xuất báo cáo!!!"))

    def _get_data_report(self):
        ret_data = []
        total_revenue = 0

        domain_surgery = [('name.surgery_date', '>=', self.start_datetime),
                          ('name.surgery_date', '<=', self.end_datetime),
                          ('name.institution.his_company', '=', self.company_id.id),
                          ('name.walkin.state', '=', 'Completed'),
                          ('name.walkin.type_crm_id', '=', self.env.ref('crm_base.type_oppor_guarantee').id)]

        domain_specialty = [('name.services_date', '>=', self.start_datetime),
                            ('name.services_date', '<=', self.end_datetime),
                            ('name.institution.his_company', '=', self.company_id.id),
                            ('name.walkin.type_crm_id', '=', self.env.ref('crm_base.type_oppor_guarantee').id)]
        
        # if self.env.user.has_group('sci_hrms.group_user_doctor'):
        #     domain_surgery.append(('team_member', '=', self.env.user.id))
        #     domain_specialty.append(('team_member', '=', self.env.user.id))

        Surgery = self.env['sh.medical.surgery.team'].sudo().with_context(company_ids=[company.id for company in self.env['res.company']])
        Specialty = self.env['sh.medical.specialty.team'].sudo().with_context(company_ids=[company.id for company in self.env['res.company']])

        if self.type == 'All':
            data_surgery = Surgery.search(domain_surgery)
            for data in data_surgery:
                # get walkin
                walkin = data.name.walkin.root_walkin

                # get so
                so = walkin.sale_order_id

                for ser_perform in data.service_performances:
                    sol = so.order_line.filtered(lambda r: r.product_id == ser_perform.product_id)
                    money = sum([rec_sol.price_subtotal for rec_sol in sol])
                    ret_data.append({
                        'code_booking': data.name.booking_id.name,
                        'name_customer': data.name.booking_id.partner_id.name,
                        'code_walkin': walkin.name,
                        'code_specialty_sheet': data.name.name,
                        'his_service_type': dict_type.get(ser_perform.his_service_type),
                        'service_category': ser_perform.service_category.name,
                        'service': ser_perform.name,
                        'qty': sol.crm_line_id.quantity if len(sol) == 1 else None,
                        # 'uom_price': sol.crm_line_id.uom_price if len(sol) == 1 else None,
                        'uom_price': sum([rec.crm_line_id.uom_price for rec in sol]),
                        'number_used': sol.crm_line_id.number_used if len(sol) == 1 else None,
                        'initial_product_id': sol.crm_line_id.initial_product_id.name or None,
                        'type_guarantee': DICT_TYPE_GUARANTEE.get(sol.crm_line_id.type_guarantee) if sol.crm_line_id.type_guarantee else None,
                        'start_datetime': data.name.surgery_date.strftime('%d/%m/%Y'),
                        'end_datetime': data.name.surgery_end_date.strftime('%d/%m/%Y'),
                        'money': money if sol else None,
                        'point_service': ser_perform.kpi_point,
                        'type_booking': walkin.booking_id.type_crm_id.name,
                        'employee_code': data.team_member.employee_code,
                        'member': data.team_member.name,
                        'group_job': data.team_member.group_job.name or None,
                        'role': data.role.name,
                        'institution': data.name.institution.name,
                    })
                    total_revenue += money

            domain_specialty.append(('name.department_type', 'in', ('Spa', 'Laser', 'Odontology')))
            data_surgery = Specialty.search(domain_specialty)
            for data in data_surgery:
                # get walkin
                walkin = data.name.walkin.root_walkin

                # get so
                so = walkin.sale_order_id

                for ser_perform in data.service_performances:
                    # get s ale order line for service
                    sol = so.order_line.filtered(lambda r: r.product_id == ser_perform.product_id)
                    money = sum([rec_sol.price_subtotal for rec_sol in sol])
                    ret_data.append({
                        'code_booking': data.name.booking_id.name,
                        'name_customer': data.name.booking_id.partner_id.name,
                        'code_walkin': walkin.name,
                        'code_specialty_sheet': data.name.name,
                        'his_service_type': dict_type.get(ser_perform.his_service_type),
                        'service_category': ser_perform.service_category.name,
                        'service': ser_perform.name,
                        'qty': sol.crm_line_id.quantity if len(sol) == 1 else None,
                        # 'uom_price': sol.crm_line_id.uom_price if len(sol) == 1 else None,
                        'uom_price': sum([rec.crm_line_id.uom_price for rec in sol]),
                        'number_used': sol.crm_line_id.number_used if len(sol) == 1 else None,
                        'initial_product_id': sol.crm_line_id.initial_product_id.name or None,
                        'type_guarantee': DICT_TYPE_GUARANTEE.get(sol.crm_line_id.type_guarantee) if sol.crm_line_id.type_guarantee else None,
                        'start_datetime': data.name.services_date.strftime('%d/%m/%Y'),
                        'end_datetime': data.name.services_end_date.strftime('%d/%m/%Y'),
                        'money': money if sol else None,
                        'point_service': ser_perform.kpi_point,
                        'type_booking': walkin.booking_id.type_crm_id.name,
                        'employee_code': data.team_member.employee_code,
                        'member': data.team_member.name,
                        'group_job': data.team_member.group_job.name or None,
                        'role': data.role.name,
                        'institution': data.name.institution.name,
                    })
                    total_revenue += money

            ret_data.append({
                'code_booking': None,
                'name_customer': None,
                'code_walkin': None,
                'code_specialty_sheet': None,
                'his_service_type': None,
                'service_category': None,
                'service': None,
                'qty': None,
                'uom_price': None,
                'number_used': None,
                'initial_product_id': None,
                'type_guarantee': None,
                'start_datetime': None,
                'end_datetime': None,
                'money': total_revenue,
                'point_service': None,
                'type_booking': None,
                'employee_code': None,
                'member': None,
                'group_job': None,
                'role': None,
                'institution': None,
            })
        elif self.type == 'Surgery':
            data_surgery = Surgery.search(domain_surgery)
            for data in data_surgery:
                # get walkin
                walkin = data.name.walkin.root_walkin

                # get so
                so = walkin.sale_order_id

                for ser_perform in data.service_performances:
                    sol = so.order_line.filtered(lambda r: r.product_id == ser_perform.product_id)
                    money = sum([rec_sol.price_subtotal for rec_sol in sol])
                    ret_data.append({
                        'code_booking': data.name.booking_id.name,
                        'name_customer': data.name.booking_id.partner_id.name,
                        'code_walkin': walkin.name,
                        'code_specialty_sheet': data.name.name,
                        'his_service_type': dict_type.get(ser_perform.his_service_type),
                        'service_category': ser_perform.service_category.name,
                        'service': ser_perform.name,
                        'qty': sol.crm_line_id.quantity if len(sol) == 1 else None,
                        # 'uom_price': sol.crm_line_id.uom_price if len(sol) == 1 else None,
                        'uom_price': sum([rec.crm_line_id.uom_price for rec in sol]),
                        'number_used': sol.crm_line_id.number_used if len(sol) == 1 else None,
                        'initial_product_id': sol.crm_line_id.initial_product_id.name or None,
                        'type_guarantee': DICT_TYPE_GUARANTEE.get(sol.crm_line_id.type_guarantee) if sol.crm_line_id.type_guarantee else None,
                        'start_datetime': data.name.surgery_date.strftime('%d/%m/%Y'),
                        'end_datetime': data.name.surgery_end_date.strftime('%d/%m/%Y'),
                        'money': money if sol else None,
                        'point_service': ser_perform.kpi_point,
                        'type_booking': walkin.booking_id.type_crm_id.name,
                        'employee_code': data.team_member.employee_code,
                        'member': data.team_member.name,
                        'group_job': data.team_member.group_job.name or None,
                        'role': data.role.name,
                        'institution': data.name.institution.name,
                    })
                    total_revenue += money
            ret_data.append({
                'code_booking': None,
                'name_customer': None,
                'code_walkin': None,
                'code_specialty_sheet': None,
                'his_service_type': None,
                'service_category': None,
                'service': None,
                'qty': None,
                'uom_price': None,
                'number_used': None,
                'initial_product_id': None,
                'type_guarantee': None,
                'start_datetime': None,
                'end_datetime': None,
                'money': total_revenue,
                'point_service': None,
                'type_booking': None,
                'employee_code': None,
                'member': None,
                'group_job': None,
                'role': None,
                'institution': None,
            })
        else:
            domain_specialty.append(('name.department_type', '=', self.type))
            data_surgery = Specialty.search(domain_specialty)
            for data in data_surgery:
                # get walkin
                walkin = data.name.walkin.root_walkin

                # get so
                so = walkin.sale_order_id

                for ser_perform in data.service_performances:
                    # get s ale order line for service
                    sol = so.order_line.filtered(lambda r: r.product_id == ser_perform.product_id)
                    money = sum([rec_sol.price_subtotal for rec_sol in sol])
                    ret_data.append({
                        'code_booking': data.name.booking_id.name,
                        'name_customer': data.name.booking_id.partner_id.name,
                        'code_walkin': walkin.name,
                        'code_specialty_sheet': data.name.name,
                        'his_service_type': dict_type.get(ser_perform.his_service_type),
                        'service_category': ser_perform.service_category.name,
                        'service': ser_perform.name,
                        'qty': sol.crm_line_id.quantity if len(sol) == 1 else None,
                        # 'uom_price': sol.crm_line_id.uom_price if len(sol) == 1 else None,
                        'uom_price': sum([rec.crm_line_id.uom_price for rec in sol]),
                        'number_used': sol.crm_line_id.number_used if len(sol) == 1 else None,
                        'initial_product_id': sol.crm_line_id.initial_product_id.name or None,
                        'type_guarantee': DICT_TYPE_GUARANTEE.get(sol.crm_line_id.type_guarantee) if sol.crm_line_id.type_guarantee else None,
                        'start_datetime': data.name.services_date.strftime('%d/%m/%Y'),
                        'end_datetime': data.name.services_end_date.strftime('%d/%m/%Y'),
                        'money': money if sol else None,
                        'point_service': ser_perform.kpi_point,
                        'type_booking': walkin.booking_id.type_crm_id.name,
                        'employee_code': data.team_member.employee_code,
                        'member': data.team_member.name,
                        'group_job': data.team_member.group_job.name or None,
                        'role': data.role.name,
                        'institution': data.name.institution.name,
                    })
                    total_revenue += money

            ret_data.append({
                'code_booking': None,
                'name_customer': None,
                'code_walkin': None,
                'code_specialty_sheet': None,
                'his_service_type': None,
                'service_category': None,
                'service': None,
                'qty': None,
                'uom_price': None,
                'number_used': None,
                'initial_product_id': None,
                'type_guarantee': None,
                'start_datetime': None,
                'end_datetime': None,
                'money': total_revenue,
                'point_service': None,
                'type_booking': None,
                'employee_code': None,
                'member': None,
                'group_job': None,
                'role': None,
                'institution': None,
            })
        return ret_data

    def create_report_service_insurance_perform_report(self):
        # get data
        datas = self._get_data_report()

        # in dữ liễu
        report_brand_overview_attachment = self.env['ir.attachment'].browse(
            self.env.ref('report_sale.bao_cao_doanh_thu_thuc_hien_dich_vu_bao_hanh_attachment').id)
        decode = base64.b64decode(report_brand_overview_attachment.datas)
        wb = load_workbook(BytesIO(decode))
        ws = wb.active
        line_font = Font(name='Times New Roman', size=14)
        if self.type == 'Surgery':
            name = "PHẪU THUẬT"
        elif self.type == 'Spa':
            name = "SPA"
        elif self.type == 'Laser':
            name = "LASER"
        elif self.type == 'Odontology':
            name = "NHA KHOA"
        else:
            name = 'Toàn bộ dịch vụ'.upper()

        code_brand = self.company_id.brand_id.code.lower()
        image_path = get_module_resource('report_sale', 'static/img', 'icon_%s.png' % code_brand)

        img = openpyxl.drawing.image.Image(image_path)
        img.anchor = 'A1'
        ws.add_image(img)

        ws['A2'].value = 'BÁO CÁO DOANH THU THỰC HIỆN DỊCH VỤ BẢO HÀNH %s' % name
        ws['M4'].value = self.start_date.strftime('%d/%m/%Y')
        ws['O4'].value = self.end_datetime.strftime('%d/%m/%Y')
        ws['M5'].value = self.company_id.name
        format_currency = '_(* #,##0_);_(* (#,##0);_(* "-"??_);_(@_)'

        key_col = list(range(1, 25))
        format_currency = '_(* #,##0_);_(* (#,##0);_(* "-"??_);_(@_)'

        key_list = [
            'code_booking',
            'name_customer',
            'code_walkin',
            'code_specialty_sheet',
            'his_service_type',
            'service_category',
            'service',
            'qty',
            'uom_price',
            'number_used',
            'initial_product_id',
            'type_guarantee',
            'start_datetime',
            'end_datetime',
            'money',
            'point_service',
            'type_booking',
            'employee_code',
            'member',
            'group_job',
            'role',
            'institution',
        ]
        row = 8
        if code_brand == 'kn':
            header_fill = ThemeReport.kn_fill
        elif code_brand == 'da':
            header_fill = ThemeReport.da_fill
        elif code_brand == 'pr':
            header_fill = ThemeReport.pr_fill
        elif code_brand == 'hh':
            header_fill = ThemeReport.hh_fill
        else:
            header_fill = ThemeReport.sci_fill
        for data in datas:
            for col, k in zip(key_col, key_list):
                beforeCell = ws.cell(7, col)
                beforeCell.fill = header_fill
                beforeCell.font = Font(name='Times New Roman', size=14, color='FFFFFF')
                cell = ws.cell(row, col)
                cell.value = data[k]
                cell.font = line_font
                cell.border = all_border_thin
                cell.alignment = Alignment(horizontal='left', vertical='center')
                if col == 15:
                    cell.number_format = format_currency
            row += 1
        fp = BytesIO()
        wb.save(fp)
        fp.seek(0)
        report = base64.encodebytes((fp.read()))
        fp.close()
        attachment = self.env['ir.attachment'].sudo().create({
            'name': 'bao_cao_doanh_thu_thuc_hien_dich_vu_bao_hanh_%s.xlsx' % self.company_id.name,
            'datas': report,
            'res_model': 'temp.creation',
            'public': True,
        })
        url = "/web/content/?model=ir.attachment&id=%s&filename_field=name&field=datas&download=true" \
              % attachment.id
        return {'name': 'BÁO CÁO DOANH THU THỰC HIỆN DỊCH VỤ BẢO HÀNH',
                'type': 'ir.actions.act_url',
                'url': url,
                'target': 'self',
                }
