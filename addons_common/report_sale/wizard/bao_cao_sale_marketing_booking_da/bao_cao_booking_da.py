from odoo import fields, api, models, _
from odoo.exceptions import ValidationError, UserError
from datetime import date, datetime, time
from calendar import monthrange
from openpyxl import load_workbook
from openpyxl.styles import Font, borders, Alignment
import base64
from io import BytesIO
from dateutil.relativedelta import relativedelta
from pytz import timezone, utc
from collections import defaultdict

thin = borders.Side(style='thin')
double = borders.Side(style='double')
all_border_thin = borders.Border(thin, thin, thin, thin)
KEY_LIST = [
    'thoi_gian_tao',
    'ngay_hen_lich',
    'chi_nhanh',
    'nguoi_tao',
    'phong_ban',
    'ma_booking',
    'nhom_nguon',
    'nguon',
    'ghi_chu',
    'trang_thai',
    'khach_den_cua',
    'loai_ban_ghi',
    'ma_dich_vu',
    'ten_dich_vu',
    'nhom_dich_vu',
    'kieu_du_lieu',
    'ho_ten_khach_hang',
    'so_dien_thoai',
    'so_dien_thoai_acc',
    'dia_chi',
    'tinh_tp',
    'da_su_dung',
    'phieu_kham',
    'giai_doan_cua_phieu_kham',
    'trang_thai_line_dich_vu',
    'tong_tien_phai_thu',
    'tong_tien_khach_tra',
]
OPTION = [
    ('01', 'Ngày hẹn lịch'),
    ('02', 'Ngày đến cửa'),
    ('03', 'Ngày tạo')
]


class SaleMarketingBookingDA(models.TransientModel):
    _name = 'sale.marketing.booking.da'
    _description = 'Báo cáo cơ hội'

    start_date = fields.Date('Start date', default=date.today())
    end_date = fields.Date('End date', default=date.today())

    company_ids = fields.Many2many(string='Công ty', comodel_name='res.company', domain=lambda self: [('id', 'in', self.env.user.company_ids.ids)])
    option = fields.Selection(OPTION, string='Tìm kiếm theo', default='01')
    # convert date to datetime for search domain, should be removed if using datetime directly
    start_datetime = fields.Datetime('Start datetime', compute='_compute_datetime')
    end_datetime = fields.Datetime('End datetime', compute='_compute_datetime')

    @api.depends('start_date', 'end_date')
    def _compute_datetime(self):
        self.start_datetime = False
        self.end_datetime = False
        if self.start_date and self.end_date:
            local_tz = timezone(self.env.user.tz or 'Etc/GMT+7')
            start_datetime = datetime(self.start_date.year, self.start_date.month, self.start_date.day, 0, 0, 0)
            end_datetime = datetime(self.end_date.year, self.end_date.month, self.end_date.day, 23, 59, 59)
            start_datetime = local_tz.localize(start_datetime, is_dst=None)
            end_datetime = local_tz.localize(end_datetime, is_dst=None)
            self.start_datetime = start_datetime.astimezone(utc).replace(tzinfo=None)
            self.end_datetime = end_datetime.astimezone(utc).replace(tzinfo=None)

    @api.onchange('start_date')
    def _onchange_start_date(self):
        if self.start_date:
            if self.start_date.month == fields.date.today().month:
                self.end_date = fields.date.today()
            else:
                self.end_date = date(self.start_date.year, self.start_date.month,
                                     monthrange(self.start_date.year, self.start_date.month)[1])

    @api.constrains('start_date', 'end_date')
    def check_dates(self):
        for record in self:
            start_date = fields.Date.from_string(record.start_date)
            end_date = fields.Date.from_string(record.end_date)
            days = (end_date - start_date).days
            if days < 0 or days > 365:
                raise ValidationError(
                    _("Ngày kết thúc không thể ở trước ngày bắt đầu khi xuất báo cáo!!!"))

    # def render_form_template(self, crm_line_id):
    #
    #     crm_line = self.sudo().env['crm.line'].with_context(
    #         company_ids=[c.id for c in self.env['res.company'].search([])]).browse(crm_line_id)
    #     booking = crm_line.crm_id
    #     # type_crm_id = booking.type_crm_id.id
    #     # crm_line_stage = crm_line.stage
    #     # if crm_line_stage == 'cancel':
    #     #     return {}
    #     # else:
    #     if self.env.ref('crm_base.type_oppor_new').id:
    #         walkin_ids = booking.walkin_ids.filtered(
    #             lambda element: element.state != 'Cancelled' and crm_line.service_id.id in element.service.ids).sorted(
    #             key=lambda element: element.date)
    #
    #         if self.option == '01':
    #             date_filter = booking.booking_date
    #         elif self.option == '02':
    #             date_filter = booking.arrival_date
    #         elif self.option == '03':
    #             date_filter = booking.create_on
    #
    #         if len(walkin_ids) >= 1:
    #             walkin_name_text = walkin_ids[0].name
    #             walkin_state_text = walkin_ids[0].state
    #         else:
    #             walkin_name_text = None
    #             walkin_state_text = None
    #
    #         str_tuple = (
    #         booking.pass_port_address, booking.district_id.name, booking.state_id.name, booking.country_id.name)
    #         str_convert = (element for element in str_tuple if element != False)
    #         dia_chi = ', '.join(str_convert)
    #         product_code = crm_line.service_id.code
    #         product_name = crm_line.service_id.name
    #         product_categ = crm_line.service_id.service_category
    #
    #         lang_code = self.env.context.get('lang') or 'en_US'
    #         date_format = self.env['res.lang']._lang_get(lang_code).date_format
    #         time_format = self.env['res.lang']._lang_get(lang_code).time_format
    #
    #         val = {
    #             'thoi_gian_tao': '%s' % fields.Datetime.context_timestamp(self.with_context(tz=self.env.user.tz),
    #                                                                       booking.create_on).strftime(
    #                 ('%s %s') % (date_format, time_format)),
    #             'ngay_hen_lich': '%s' % fields.Datetime.context_timestamp(self.with_context(tz=self.env.user.tz),
    #                                                                       date_filter).strftime(
    #                 ('%s %s') % (date_format, time_format)),
    #             'chi_nhanh': booking.company_id.name,
    #             'nguoi_tao': booking.create_by.name,
    #             'phong_ban': self.env['hr.employee'].sudo().search([('user_id', '=', booking.create_by.id)],
    #                                                                limit=1).department_id.name or None,
    #             'ma_booking': booking.name,
    #             'nhom_nguon': booking.category_source_id.name or None,
    #             'nguon': booking.source_id.name,
    #             'ghi_chu': booking.note or None,
    #             'trang_thai': booking.stage_id.name,
    #             'khach_den_cua': booking.customer_come,
    #             'loai_ban_ghi': booking.type_crm_id.name,
    #             'ma_dich_vu': product_code,
    #             'ten_dich_vu': product_name,
    #             'nhom_dich_vu': product_categ.name,
    #             'kieu_du_lieu': booking.type_data,
    #             'ho_ten_khach_hang': booking.contact_name,
    #             'so_dien_thoai': booking.phone,
    #             'so_dien_thoai_acc': booking.partner_id.phone,
    #             'dia_chi': dia_chi,
    #             'tinh_tp': booking.state_id.name or None,
    #             'da_su_dung': crm_line.number_used,
    #             'phieu_kham': walkin_name_text or None,
    #             'giai_doan_cua_phieu_kham': walkin_state_text or None,
    #             'trang_thai_line_dich_vu': crm_line.stage,
    #             'tong_tien_phai_thu': crm_line.total,
    #             'tong_tien_khach_tra': booking.amount_paid
    #         }
    #         return val
    #     else:
    #         return {}

    def _get_data_report(self):
        # Theo ngày hẹn lịch
        if self.option == '01':
            query = """
                SELECT cl.id, cld.create_on, cld.booking_date, rc.name, rp.name, hd.name, cld.name, ccs.name, us.name, cld.note, 
                cs.name, cld.customer_come, ct.name, pt.default_code, pt.name, pc.name, cld.type_data, cld.contact_name,
                cld.phone, rp.phone, cld.pass_port_address, rcd.name, rcs.name, rc1.name, cl.number_used, sma.name, sma.state,
                cl.stage, cl.total, cld.amount_paid
                FROM crm_line cl
                FULL JOIN crm_lead cld on cl.crm_id = cld.id
                FULL JOIN crm_stage cs on cld.stage_id = cs.id
                FULL join res_company rc on rc.id = cld.company_id
                FULL join res_users ru on ru.id = cld.create_by
                FULL join res_partner rp on rp.id = ru.partner_id
                FULL join hr_employee he on he.user_id = ru.id
                FULL join hr_department hd on hd.id = he.department_id
                FULL join crm_category_source ccs on ccs.id = cld.category_source_id
                FULL join utm_source us on us.id = cld.source_id
                FULL join crm_type ct on ct.id = cld.type_crm_id
                FULL join sh_medical_health_center_service smh on smh.id = cl.service_id
                FULL join product_product pp on pp.id = smh.product_id
                FULL join product_template pt on pt.id = pp.product_tmpl_id
                FULL join product_category pc on pc.id = pt.categ_id
                FULL join res_country_district rcd on rcd.id = cld.district_id
                FULL join res_country_state rcs on rcs.id = cld.state_id
                FULL join res_country rc1 on rc1.id = cld.country_id
                FULL join sh_medical_appointment_register_walkin sma on sma.booking_id = cld.id
                WHERE cld.company_id in %s AND 
                    cs.name != 'Cancel' AND 
                    cld.type = 'opportunity' AND
                    cld.booking_date BETWEEN %s AND %s
                ORDER BY cld.booking_date, sma.id
            """
        # Theo ngày đến cửa
        elif self.option == '02':
            query = """
                SELECT cl.id, cld.create_on, cld.arrival_date, rc.name, rp.name, hd.name, cld.name, ccs.name, us.name, cld.note, 
                cs.name, cld.customer_come, ct.name, pt.default_code, pt.name, pc.name, cld.type_data, cld.contact_name,
                cld.phone, rp.phone, cld.pass_port_address, rcd.name, rcs.name, rc1.name, cl.number_used, sma.name, sma.state,
                cl.stage, cl.total, cld.amount_paid
                FROM crm_line cl
                FULL JOIN crm_lead cld on cl.crm_id = cld.id
                FULL JOIN crm_stage cs on cld.stage_id = cs.id
                FULL join res_company rc on rc.id = cld.company_id
                FULL join res_users ru on ru.id = cld.create_by
                FULL join res_partner rp on rp.id = ru.partner_id
                FULL join hr_employee he on he.user_id = ru.id
                FULL join hr_department hd on hd.id = he.department_id
                FULL join crm_category_source ccs on ccs.id = cld.category_source_id
                FULL join utm_source us on us.id = cld.source_id
                FULL join crm_type ct on ct.id = cld.type_crm_id
                FULL join sh_medical_health_center_service smh on smh.id = cl.service_id
                FULL join product_product pp on pp.id = smh.product_id
                FULL join product_template pt on pt.id = pp.product_tmpl_id
                FULL join product_category pc on pc.id = pt.categ_id
                FULL join res_country_district rcd on rcd.id = cld.district_id
                FULL join res_country_state rcs on rcs.id = cld.state_id
                FULL join res_country rc1 on rc1.id = cld.country_id
                FULL join sh_medical_appointment_register_walkin sma on sma.booking_id = cld.id
                WHERE cld.company_id in %s AND 
                    cs.name != 'Cancel' AND 
                    cld.type = 'opportunity' AND
                    cld.arrival_date BETWEEN %s AND %s
                ORDER BY cld.arrival_date, sma.id
            """
        # Theo ngày tạo booking
        elif self.option == '03':
            query = """
                SELECT cl.id, cld.create_on, cld.create_on, rc.name, rp.name, hd.name, cld.name, ccs.name, us.name, cld.note, 
                cs.name, cld.customer_come, ct.name, pt.default_code, pt.name, pc.name, cld.type_data, cld.contact_name,
                cld.phone, rp.phone, cld.pass_port_address, rcd.name, rcs.name, rc1.name, cl.number_used, sma.name, sma.state,
                cl.stage, cl.total, cld.amount_paid
                FROM crm_line cl
                FULL JOIN crm_lead cld on cl.crm_id = cld.id
                FULL JOIN crm_stage cs on cld.stage_id = cs.id
                FULL join res_company rc on rc.id = cld.company_id
                FULL join res_users ru on ru.id = cld.create_by
                FULL join res_partner rp on rp.id = ru.partner_id
                FULL join hr_employee he on he.user_id = ru.id
                FULL join hr_department hd on hd.id = he.department_id
                FULL join crm_category_source ccs on ccs.id = cld.category_source_id
                FULL join utm_source us on us.id = cld.source_id
                FULL join crm_type ct on ct.id = cld.type_crm_id
                FULL join sh_medical_health_center_service smh on smh.id = cl.service_id
                FULL join product_product pp on pp.id = smh.product_id
                FULL join product_template pt on pt.id = pp.product_tmpl_id
                FULL join product_category pc on pc.id = pt.categ_id
                FULL join res_country_district rcd on rcd.id = cld.district_id
                FULL join res_country_state rcs on rcs.id = cld.state_id
                FULL join res_country rc1 on rc1.id = cld.country_id
                FULL join sh_medical_appointment_register_walkin sma on sma.booking_id = cld.id
                WHERE cld.company_id in %s AND 
                cs.name != 'Cancel' AND 
                cld.type = 'opportunity' AND
                cld.create_on BETWEEN %s AND %s
                ORDER BY cld.create_on, sma.id
            """
        return_val = []
        self.flush()
        self.clear_caches()
        self._cr.execute(query, (tuple(self.company_ids.ids), self.start_datetime, self.end_datetime))
        result = self.env.cr.fetchall()
        list_line = []
        lang_code = self.env.context.get('lang') or 'en_US'
        date_format = self.env['res.lang']._lang_get(lang_code).date_format
        time_format = self.env['res.lang']._lang_get(lang_code).time_format
        for value in result:
            if value[0] not in list_line:
                dia_chi = value[20] + '; 'if value[20] else ''
                dia_chi += value[21] + '; 'if value[21] else ''
                dia_chi += value[22] + '; 'if value[22] else ''
                dia_chi += value[23] if value[23] else ''
                val = {
                    'thoi_gian_tao': '%s' % fields.Datetime.context_timestamp(self.with_context(tz=self.env.user.tz),
                                                                              value[1]).strftime(
                        ('%s %s') % (date_format, time_format)),
                    'ngay_hen_lich': '%s' % fields.Datetime.context_timestamp(self.with_context(tz=self.env.user.tz),
                                                                              value[2]).strftime(
                        ('%s %s') % (date_format, time_format)),
                    'chi_nhanh': value[3],
                    'nguoi_tao': value[4],
                    'phong_ban': value[5],
                    'ma_booking': value[6],
                    'nhom_nguon': value[7],
                    'nguon': value[8],
                    'ghi_chu': value[9],
                    'trang_thai': value[10],
                    'khach_den_cua': value[11],
                    'loai_ban_ghi': value[12],
                    'ma_dich_vu': value[13],
                    'ten_dich_vu': value[14],
                    'nhom_dich_vu': value[15],
                    'kieu_du_lieu': value[16],
                    'ho_ten_khach_hang': value[17],
                    # 'so_dien_thoai': value[18],
                    # 'so_dien_thoai_acc': value[19],
                    'so_dien_thoai': None,
                    'so_dien_thoai_acc': None,
                    'dia_chi': dia_chi,
                    'tinh_tp': value[22],
                    'da_su_dung': value[24],
                    'phieu_kham': value[25],
                    'giai_doan_cua_phieu_kham': value[26],
                    'trang_thai_line_dich_vu': value[27],
                    'tong_tien_phai_thu': value[28],
                    'tong_tien_khach_tra': value[29]
                }
                list_line.append(value[0])
            if val:
                return_val.append(val)
        return return_val

    def create_report_da(self):
        datas = self._get_data_report()
        # in dữ liễu
        report_brand_overview_attachment = self.env['ir.attachment'].browse(
            self.env.ref('report_sale.bao_cao_booking_attachment').id)
        decode = base64.b64decode(report_brand_overview_attachment.datas)
        wb = load_workbook(BytesIO(decode))
        ws = wb.active
        line_font = Font(name='Times New Roman', size=14)
        format_currency = '_(* #,##0_);_(* (#,##0);_(* "-"??_);_(@_)'
        ws['M3'].value = ws['M3'].value + self.start_date.strftime('%d/%m/%Y')
        ws['O3'].value = ws['O3'].value + self.end_datetime.strftime('%d/%m/%Y')
        ws['B6'].value = dict(self._fields['option'].selection).get(self.option)
        key_col = list(range(1, len(KEY_LIST) + 1))

        row = 7
        for data in datas:
            for col, k in zip(key_col, KEY_LIST):
                cell = ws.cell(row, col)
                cell.value = data[k]
                cell.font = line_font
                cell.border = all_border_thin
                cell.alignment = Alignment(horizontal='left', vertical='center')
                if col in (26, 27):
                    cell.number_format = format_currency
            row += 1
        fp = BytesIO()
        wb.save(fp)
        fp.seek(0)
        report = base64.encodebytes((fp.read()))
        fp.close()
        attachment = self.env['ir.attachment'].sudo().create({
            'name': 'bao_cao_booking_%s.xlsx'% datetime.now().strftime('%d_%m_%Y'),
            'datas': report,
            'res_model': 'temp.creation',
            'public': True,
        })
        url = "/web/content/?model=ir.attachment&id=%s&filename_field=name&field=datas&download=true" \
              % attachment.id
        return {'name': 'BÁO CÁO CƠ HỘI',
                'type': 'ir.actions.act_url',
                'url': url,
                'target': 'self',
                }