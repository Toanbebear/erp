import base64
from calendar import monthrange
from datetime import date, datetime
from io import BytesIO

from openpyxl import load_workbook
from openpyxl.styles import Font, borders, Alignment
from pytz import timezone, utc

from odoo import fields, api, models, _
from odoo.exceptions import ValidationError

thin = borders.Side(style='thin')
double = borders.Side(style='double')
all_border_thin = borders.Border(thin, thin, thin, thin)


class CustomerDepositReport(models.TransientModel):
    _name = 'kpi.customer.deposit.report'
    _description = 'Doanh thu Khách hàng Đặt cọc'

    start_date = fields.Date('Start date', default=date.today())
    end_date = fields.Date('End date', default=date.today())
    company_id = fields.Many2one('res.company', string='Company', default=lambda self: self.env.company)

    # convert date to datetime for search domain, should be removed if using datetime directly
    start_datetime = fields.Datetime('Start datetime', compute='_compute_datetime')
    end_datetime = fields.Datetime('End datetime', compute='_compute_datetime')

    @api.depends('start_date', 'end_date')
    def _compute_datetime(self):
        self.start_datetime = False
        self.end_datetime = False
        if self.start_date and self.end_date:
            local_tz = timezone(self.env.user.tz or 'Etc/GMT+7')
            start_datetime = datetime(self.start_date.year, self.start_date.month, self.start_date.day, 0, 0, 0)
            end_datetime = datetime(self.end_date.year, self.end_date.month, self.end_date.day, 23, 59, 59)
            start_datetime = local_tz.localize(start_datetime, is_dst=None)
            end_datetime = local_tz.localize(end_datetime, is_dst=None)
            self.start_datetime = start_datetime.astimezone(utc).replace(tzinfo=None)
            self.end_datetime = end_datetime.astimezone(utc).replace(tzinfo=None)

    @api.onchange('start_date')
    def _onchange_start_date(self):
        if self.start_date:
            if self.start_date.month == fields.date.today().month:
                self.end_date = fields.date.today()
            else:
                self.end_date = date(self.start_date.year, self.start_date.month,
                                     monthrange(self.start_date.year, self.start_date.month)[1])

    @api.constrains('start_date', 'end_date')
    def check_dates(self):
        for record in self:
            start_date = fields.Date.from_string(record.start_date)
            end_date = fields.Date.from_string(record.end_date)
            if start_date > end_date:
                raise ValidationError(
                    _("Ngày kết thúc không thể ở trước ngày bắt đầu khi xuất báo cáo!!!"))

    def _get_data_report(self):
        data = []

        payments = self.env['account.payment'].sudo().search(
            [('state', '=', 'posted'),
             ('company_id', '=', self.company_id.id),
             ('payment_date', '>=', self.start_datetime),
             ('payment_date', '<=', self.end_datetime)])

        for payment in payments:
            data.append({
                'payment_date': payment.payment_date.strftime('%d/%m/%Y'),
                'name': payment.name,
                'booking': payment.crm_id.name,
                'customer': payment.partner_id.name,
                'amount': payment.amount,
                'company': payment.company_id.name,
            })
        return data

    def create_report(self):

        datas = self._get_data_report()

        # in dữ liệu
        attachment = self.env['ir.attachment'].browse(self.env.ref('report_sale.report_customer_deposit_attachment').id)
        decode = base64.b64decode(attachment.datas)

        wb = load_workbook(BytesIO(decode))
        ws = wb.active
        line_font = Font(name='Times New Roman', size=14)

        keys = [
            'payment_date',
            'name',
            'booking',
            'customer',
            'amount',
            'company',
        ]

        row = 9
        for data in datas:
            col = 1
            for key in keys:
                cell = ws.cell(row, col)
                cell.value = data[key]
                cell.font = line_font
                cell.border = all_border_thin
                cell.alignment = Alignment(horizontal='left', vertical='center')
                col += 1
            row += 1

        ws['F%s' % (row + 2)].value = '(Báo cáo xuất ngày: %s)' % date.today().strftime('%d/%m/%Y')

        fp = BytesIO()
        wb.save(fp)
        fp.seek(0)
        report = base64.encodebytes((fp.read()))
        fp.close()

        attachment = self.env['ir.attachment'].sudo().create({
            'name': 'Báo cáo Doanh thu Khách hàng Đặt cọc %s.xlsx' % datetime.now(),
            'datas': report,
            'res_model': 'temp.creation',
            'public': True,
        })
        url = "/web/content/?model=ir.attachment&id=%s&filename_field=name&field=datas&download=true" \
              % attachment.id
        return {'name': 'Doanh thu Khách hàng Đặt cọc',
                'type': 'ir.actions.act_url',
                'url': url,
                'target': 'self',
                }
