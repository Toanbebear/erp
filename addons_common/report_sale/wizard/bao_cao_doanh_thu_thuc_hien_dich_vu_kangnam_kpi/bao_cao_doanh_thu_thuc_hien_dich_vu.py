from odoo import fields, api, models, _
from odoo.exceptions import ValidationError, UserError
from datetime import date, datetime, time
from calendar import monthrange
from openpyxl import load_workbook
from openpyxl.styles import Font, borders, Alignment
import base64
from io import BytesIO
from dateutil.relativedelta import relativedelta
from pytz import timezone, utc
from odoo.addons.report_sale.wizard.theme_report import ThemeReport
from odoo.modules.module import get_module_resource
import openpyxl

thin = borders.Side(style='thin')
double = borders.Side(style='double')
all_border_thin = borders.Border(thin, thin, thin, thin)

WARD_TYPE = [
    ('All', 'Toàn bộ dịch vụ'),
    ('Surgery', 'Phẫu thuật'),
    ('Spa', 'Spa'),
    ('Laser', 'Laser'),
    ('Odontology', 'Nha khoa')
]

SEARCH_BY = [
    ('01', 'Ngày bắt đầu của phiếu chuyên khoa'),
    ('02', 'Ngày làm dịch vụ của phiếu khám'),
]

SERVICE_HIS_TYPE = [('Spa', 'Spa'), ('Laser', 'Laser'), ('Odontology', 'Nha'), ('Surgery', 'Phẫu thuật'), ('ChiPhi', 'Chi phí khác')]
dict_type = dict((key, value) for key, value in SERVICE_HIS_TYPE)

SURGERY_TYPE = [('minor', 'Tiểu phẫu'), ('major', 'Đại phẫu')]
DICT_SURGERY_TYPE = dict((key, value) for key, value in SURGERY_TYPE)
TEMP = [
    'stt',
    'employee_code',  # Mã nhân viên
    'member',  # Thành viên
    'group_job',  # Bộ phận
    'job_id',  # Chức danh
    'role',  # Vai trò
    'name_customer',  # Tên khách hàng
    'code_booking',  # Mã booking
    'so_code',  # Số sale order
    'start_datetime',  # Ngày giờ bắt đầu
    'surgery_type',  # Loại phẫu thuật
    'his_service_type',  # Nhóm dịch vụ cấp 1
    'service_category',  # Nhóm dịch vụ
    'service',  # Dịch vụ
    'uom_price',  # Đơn vị xử lý
    'qty',  # Số lượng
    'number_used',  # Đã sử dụng
    'money',  # Doanh thu thực
    'code_walkin',  # Mã phiếu khám
    'service_date',  # Ngày phiếu khám
    'code_specialty_sheet',  # Mã phiếu chuyên khoa
    'department',  # Khoa phòng
    'end_datetime',  # Ngày giờ kết thúc
    'point_service',  # Điểm KPI
    'type_booking',  # Phân loại booking
    'institution',  # Đơn vị
]


class ServicePerformanceReportKangnam(models.TransientModel):
    _name = 'service.perform.report.kangnam.temp.kpi'
    _description = 'Báo cáo doanh thu thực hiện dịch vụ Kangnam mau KPI task #810'

    start_date = fields.Date('Start date', default=date.today())
    end_date = fields.Date('End date', default=date.today())
    company_id = fields.Many2one('res.company', string='Chi nhánh', domain="[('brand_id.name', '=', 'Kangnam')]", default=lambda self: self.env.company)
    type = fields.Selection(WARD_TYPE, string='Loại phiếu', default='All')

    # convert date to datetime for search domain, should be removed if using datetime directly
    start_datetime = fields.Datetime('Start datetime', compute='_compute_datetime')
    end_datetime = fields.Datetime('End datetime', compute='_compute_datetime')
    # Kiểu lấy báo cáo theo ngày làm dịch vụ hoặc theo ngày giờ bắt đầu của phiếu chuyên khoa
    # search_by = fields.Selection(SEARCH_BY, string='Kiểu tìm kiếm dữ liệu', default='01')
    expected_date = fields.Selection([('conf_walkin', 'Hoàn thành phiếu khám'), ('conf_surgery', 'Hoàn thành phiếu chuyên khoa/phẫu thuật')],
                                     string='Điều kiện xuất', default='conf_walkin')
    @api.depends('start_date', 'end_date')
    def _compute_datetime(self):
        self.start_datetime = False
        self.end_datetime = False
        if self.start_date and self.end_date:
            local_tz = timezone(self.env.user.tz or 'Etc/GMT+7')
            start_datetime = datetime(self.start_date.year, self.start_date.month, self.start_date.day, 0, 0, 0)
            end_datetime = datetime(self.end_date.year, self.end_date.month, self.end_date.day, 23, 59, 59)
            start_datetime = local_tz.localize(start_datetime, is_dst=None)
            end_datetime = local_tz.localize(end_datetime, is_dst=None)
            self.start_datetime = start_datetime.astimezone(utc).replace(tzinfo=None)
            self.end_datetime = end_datetime.astimezone(utc).replace(tzinfo=None)

    @api.onchange('start_date')
    def _onchange_start_date(self):
        if self.start_date:
            if self.start_date.month == fields.date.today().month:
                self.end_date = fields.date.today()
            else:
                self.end_date = date(self.start_date.year, self.start_date.month,
                                     monthrange(self.start_date.year, self.start_date.month)[1])

    @api.constrains('start_date', 'end_date')
    def check_dates(self):
        for record in self:
            start_date = fields.Date.from_string(record.start_date)
            end_date = fields.Date.from_string(record.end_date)
            if start_date > end_date:
                raise ValidationError(
                    _("Ngày kết thúc không thể ở trước ngày bắt đầu khi xuất báo cáo!!!"))

    def _get_data_report(self):
        ret_data = []
        total_revenue = 0
        if self.expected_date == 'conf_walkin':
            # Ngày hoàn thành của phiếu khám.
            # Cớ sở y tế là công ty chọn.
            # Trạng thái là hoàn thành.
            # Loại booking khác bảo hành
            domain_surgery = [('name.walkin.service_date', '>=', self.start_datetime),
                              ('name.walkin.service_date', '<=', self.end_datetime),
                              ('name.institution.his_company', '=', self.company_id.id),
                              ('name.walkin.state', '=', 'Completed'),
                              ('role', '!=', 'Bác sĩ kiểm soát chuyên môn'),
                              ('name.walkin.type_crm_id', '!=', self.env.ref('crm_base.type_oppor_guarantee').id)]

            domain_specialty = [('name.walkin.service_date', '>=', self.start_datetime),
                                ('name.walkin.service_date', '<=', self.end_datetime),
                                ('name.institution.his_company', '=', self.company_id.id),
                                ('name.walkin.state', '=', 'Completed'),
                                ('role', '!=', 'Bác sĩ kiểm soát chuyên môn'),
                                ('name.walkin.type_crm_id', '!=', self.env.ref('crm_base.type_oppor_guarantee').id)]
        elif self.expected_date == 'conf_surgery':
            domain_surgery = [('name.surgery_date', '>=', self.start_datetime),
                              ('name.surgery_date', '<=', self.end_datetime),
                              ('name.institution.his_company', '=', self.company_id.id),
                              ('role', '!=', 'Bác sĩ kiểm soát chuyên môn'),
                              ('name.state', '=', 'Done'),
                              ('name.walkin.type_crm_id', '!=', self.env.ref('crm_base.type_oppor_guarantee').id)]

            domain_specialty = [('name.services_date', '>=', self.start_datetime),
                                ('name.services_date', '<=', self.end_datetime),
                                ('name.institution.his_company', '=', self.company_id.id),
                                ('role', '!=', 'Bác sĩ kiểm soát chuyên môn'),
                                ('name.state', '=', 'Done'),
                                ('name.walkin.type_crm_id', '!=', self.env.ref('crm_base.type_oppor_guarantee').id)]

        Surgery = self.env['sh.medical.surgery.team'].sudo().with_context(company_ids=[company.id for company in self.env['res.company']])
        Specialty = self.env['sh.medical.specialty.team'].sudo().with_context(company_ids=[company.id for company in self.env['res.company']])

        lang_code = self.env.context.get('lang') or 'en_US'
        date_format = self.env['res.lang']._lang_get(lang_code).date_format
        time_format = self.env['res.lang']._lang_get(lang_code).time_format

        if self.type == 'All':
            data_surgery = Surgery.search(domain_surgery)
            for data in data_surgery:
                # get walkin
                walkin = data.name.walkin
                # get khoa phòng
                # department = data.name.department.name
                # get so
                so = walkin.sale_order_id
                # ngày phiếu khám
                # service_date = walkin.service_date

                for ser_perform in data.service_performances:
                    sol = so.order_line.filtered(lambda r: r.product_id == ser_perform.product_id)
                    # money = sum([rec_sol.price_subtotal for rec_sol in sol])
                    ret_data.append({
                        'employee_code': data.team_member.employee_code,
                        'member': data.team_member.name,
                        'group_job': data.team_member.group_job.name or None,
                        'job_id': data.team_member.job_id.name or None,
                        'role': data.role.name,
                        'name_customer': data.name.booking_id.partner_id.name,
                        'code_booking': data.name.booking_id.name,
                        'so_code': so.name,
                        'start_datetime': '%s' % fields.Datetime.context_timestamp(self.with_context(tz=self.env.user.tz), data.name.surgery_date).strftime('%d/%m/%Y'),
                        'surgery_type': DICT_SURGERY_TYPE.get(ser_perform.surgery_type) if ser_perform.surgery_type else None,
                        'his_service_type': dict_type.get(ser_perform.his_service_type),
                        'service_category': ser_perform.service_category.name,
                        'service': ser_perform.name,
                        'uom_price': sum([rec.crm_line_id.uom_price for rec in sol]),
                        'qty': sol.crm_line_id.quantity if len(sol) == 1 else None,
                        'number_used': None, #sol.crm_line_id.number_used if len(sol) == 1 else None,
                        'money': None, #money if sol else None,
                        'code_walkin': None, #walkin.name,
                        'service_date': None, #'%s' % fields.Datetime.context_timestamp(self.with_context(tz=self.env.user.tz), service_date).strftime('%d/%m/%Y'),
                        'code_specialty_sheet': None, #data.name.name,
                        'department': None, #department,
                        # 'uom_price': sol.crm_line_id.uom_price if len(sol) == 1 else None,
                        'end_datetime': None, #'%s' % fields.Datetime.context_timestamp(self.with_context(tz=self.env.user.tz), data.name.surgery_end_date).strftime('%d/%m/%Y'),
                        'point_service': None, #ser_perform.kpi_point,
                        'type_booking': None, #walkin.booking_id.type_crm_id.name,
                        'institution': None, #data.name.institution.name,
                    })
                    # total_revenue += money

            domain_specialty.append(('name.department_type', 'in', ('Spa', 'Laser', 'Odontology')))
            data_surgery = Specialty.search(domain_specialty)
            for data in data_surgery:
                # get walkin
                walkin = data.name.walkin
                # get khoa phòng
                # department = data.name.department.name
                # get so
                so = walkin.sale_order_id
                # ngày phiếu khám
                # service_date = walkin.service_date

                for ser_perform in data.service_performances:
                    # get s ale order line for service
                    sol = so.order_line.filtered(lambda r: r.product_id == ser_perform.product_id)
                    # money = sum([rec_sol.price_subtotal for rec_sol in sol])
                    ret_data.append({
                        'employee_code': data.team_member.employee_code,
                        'member': data.team_member.name,
                        'group_job': data.team_member.group_job.name or None,
                        'job_id': data.team_member.job_id.name or None,
                        'role': data.role.name,
                        'name_customer': data.name.booking_id.partner_id.name,
                        'code_booking': data.name.booking_id.name,
                        'so_code': so.name,
                        'start_datetime': '%s' % fields.Datetime.context_timestamp(self.with_context(tz=self.env.user.tz), data.name.services_date).strftime('%d/%m/%Y'),
                        'surgery_type': ser_perform.surgery_type or None,
                        'his_service_type': dict_type.get(ser_perform.his_service_type),
                        'service_category': ser_perform.service_category.name,
                        'service': ser_perform.name,
                        'uom_price': sum([rec.crm_line_id.uom_price for rec in sol]),
                        'qty': sol.crm_line_id.quantity if len(sol) == 1 else None,
                        'number_used': None, #sol.crm_line_id.number_used if len(sol) == 1 else None,
                        'money': None, #money if sol else None,
                        'code_walkin': None, #walkin.name,
                        'service_date': None, #'%s' % fields.Datetime.context_timestamp(self.with_context(tz=self.env.user.tz), service_date).strftime('%d/%m/%Y'),
                        'code_specialty_sheet': None, #data.name.name,
                        'department': None, #department,
                        # 'uom_price': sol.crm_line_id.uom_price if len(sol) == 1 else None,
                        'end_datetime': None, #'%s' % fields.Datetime.context_timestamp(self.with_context(tz=self.env.user.tz), data.name.services_end_date).strftime('%d/%m/%Y'),
                        'point_service': None, #ser_perform.kpi_point,
                        'type_booking': None, #walkin.booking_id.type_crm_id.name,
                        'institution': None, #data.name.institution.name,
                    })
            #         total_revenue += money
            #
            # val = dict([(e, None) for e in TEMP])
            # val['money'] = total_revenue
            # ret_data.append(val)
        elif self.type == 'Surgery':
            data_surgery = Surgery.search(domain_surgery)
            for data in data_surgery:
                # get walkin
                walkin = data.name.walkin
                # get khoa phòng
                # department = data.name.department.name
                # get so
                so = walkin.sale_order_id
                # ngày phiếu khám
                # service_date = walkin.service_date

                for ser_perform in data.service_performances:
                    sol = so.order_line.filtered(lambda r: r.product_id == ser_perform.product_id)
                    # money = sum([rec_sol.price_subtotal for rec_sol in sol])
                    ret_data.append({
                        'employee_code': data.team_member.employee_code,
                        'member': data.team_member.name,
                        'group_job': data.team_member.group_job.name or None,
                        'job_id': data.team_member.job_id.name or None,
                        'role': data.role.name,
                        'name_customer': data.name.booking_id.partner_id.name,
                        'code_booking': data.name.booking_id.name,
                        'so_code': so.name,
                        'start_datetime': '%s' % fields.Datetime.context_timestamp(self.with_context(tz=self.env.user.tz), data.name.surgery_date).strftime('%d/%m/%Y'),
                        'surgery_type': DICT_SURGERY_TYPE.get(ser_perform.surgery_type) if ser_perform.surgery_type else None,
                        'his_service_type': dict_type.get(ser_perform.his_service_type),
                        'service_category': ser_perform.service_category.name,
                        'service': ser_perform.name,
                        'uom_price': sum([rec.crm_line_id.uom_price for rec in sol]),
                        'qty': sol.crm_line_id.quantity if len(sol) == 1 else None,
                        'number_used': None, #sol.crm_line_id.number_used if len(sol) == 1 else None,
                        'money': None, #money if sol else None,
                        'code_walkin': None, #walkin.name,
                        'service_date': None, #'%s' % fields.Datetime.context_timestamp(self.with_context(tz=self.env.user.tz), service_date).strftime('%d/%m/%Y'),
                        'code_specialty_sheet': None, #data.name.name,
                        'department': None, #department,
                        # 'uom_price': sol.crm_line_id.uom_price if len(sol) == 1 else None,
                        'end_datetime': None, #'%s' % fields.Datetime.context_timestamp(self.with_context(tz=self.env.user.tz), data.name.surgery_end_date).strftime('%d/%m/%Y'),
                        'point_service': None, #ser_perform.kpi_point,
                        'type_booking': None, #walkin.booking_id.type_crm_id.name,
                        'institution': None, #data.name.institution.name,
                    })
                    # total_revenue += money

            # val = dict([(e, None) for e in TEMP])
            # val['money'] = total_revenue
            # ret_data.append(val)
        else:
            domain_specialty.append(('name.department_type', '=', self.type))
            data_surgery = Specialty.search(domain_specialty)
            for data in data_surgery:
                # get walkin
                walkin = data.name.walkin
                # get khoa phòng
                # department = data.name.department.name
                # get so
                so = walkin.sale_order_id
                # ngày phiếu khám
                # service_date = walkin.service_date

                for ser_perform in data.service_performances:
                    # get s ale order line for service
                    sol = so.order_line.filtered(lambda r: r.product_id == ser_perform.product_id)
                    # money = sum([rec_sol.price_subtotal for rec_sol in sol])
                    ret_data.append({
                        'employee_code': data.team_member.employee_code,
                        'member': data.team_member.name,
                        'group_job': data.team_member.group_job.name or None,
                        'job_id': data.team_member.job_id.name or None,
                        'role': data.role.name,
                        'name_customer': data.name.booking_id.partner_id.name,
                        'code_booking': data.name.booking_id.name,
                        'so_code': so.name,
                        'start_datetime': '%s' % fields.Datetime.context_timestamp(self.with_context(tz=self.env.user.tz), data.name.services_date).strftime('%d/%m/%Y'),
                        'surgery_type': DICT_SURGERY_TYPE.get(ser_perform.surgery_type) if ser_perform.surgery_type else None,
                        'his_service_type': dict_type.get(ser_perform.his_service_type),
                        'service_category': ser_perform.service_category.name,
                        'service': ser_perform.name,
                        'uom_price': sum([rec.crm_line_id.uom_price for rec in sol]),
                        'qty': sol.crm_line_id.quantity if len(sol) == 1 else None,
                        'number_used': None, #sol.crm_line_id.number_used if len(sol) == 1 else None,
                        'money': None, #money if sol else None,
                        'code_walkin': None, #walkin.name,
                        'service_date': None, #'%s' % fields.Datetime.context_timestamp(self.with_context(tz=self.env.user.tz), service_date).strftime('%d/%m/%Y'),
                        'code_specialty_sheet': None, #data.name.name,
                        'department': None, #department,
                        # 'uom_price': sol.crm_line_id.uom_price if len(sol) == 1 else None,
                        'end_datetime': None, #'%s' % fields.Datetime.context_timestamp(self.with_context(tz=self.env.user.tz), data.name.services_end_date).strftime('%d/%m/%Y'),
                        'point_service': None, #ser_perform.kpi_point,
                        'type_booking': None, #walkin.booking_id.type_crm_id.name,
                        'institution': None, #data.name.institution.name,
                    })
                    # total_revenue += money

            # val = dict([(e, None) for e in TEMP])
            # val['money'] = total_revenue
            # ret_data.append(val)
        return ret_data

    def create_report_service_perform_report(self):
        # get data
        datas = self._get_data_report()

        # in dữ liễu
        report_brand_overview_attachment = self.env['ir.attachment'].browse(
            self.env.ref('report_sale.bao_cao_doanh_thu_thuc_hien_dich_vu_kangnam_mau_kpi_attachment').id)
        decode = base64.b64decode(report_brand_overview_attachment.datas)
        wb = load_workbook(BytesIO(decode))
        ws = wb.get_sheet_by_name('Lam_DichVu')
        # line_font = Font(name='Times New Roman', size=14)
        if self.type == 'Surgery':
            name = "PHẪU THUẬT"
        elif self.type == 'Spa':
            name = "SPA"
        elif self.type == 'Laser':
            name = "LASER"
        elif self.type == 'Odontology':
            name = "NHA KHOA"
        else:
            name = 'Toàn bộ dịch vụ'.upper()

        code_brand = self.company_id.brand_id.code.lower()
        image_path = get_module_resource('report_sale', 'static/img', 'icon_%s.png' % code_brand)

        img = openpyxl.drawing.image.Image(image_path)
        img.anchor = 'A1'
        ws.add_image(img)

        ws['A2'].value = 'BÁO CÁO DOANH THU THỰC HIỆN DỊCH VỤ %s' % name
        ws['N4'].value = self.start_date.strftime('%d/%m/%Y')
        ws['P4'].value = self.end_datetime.strftime('%d/%m/%Y')
        ws['N5'].value = self.company_id.name

        key_col = list(range(1, len(TEMP) + 1))
        # format_currency = '_(* #,##0_);_(* (#,##0);_(* "-"??_);_(@_)'

        # if code_brand == 'kn':
        #     header_fill = ThemeReport.kn_fill
        # elif code_brand == 'da':
        #     header_fill = ThemeReport.da_fill
        # elif code_brand == 'pr':
        #     header_fill = ThemeReport.pr_fill
        # elif code_brand == 'hh':
        #     header_fill = ThemeReport.hh_fill
        # else:
        #     header_fill = ThemeReport.sci_fill

        row = 10
        for data in datas:
            for col, k in zip(key_col, TEMP):
                # beforeCell = ws.cell(9, col)
                # beforeCell.fill = header_fill
                # beforeCell.font = Font(name='Times New Roman', size=14, color='FFFFFF')
                if col not in range(17, 25):
                    if col != 1:
                        cell = ws.cell(row, col)
                        cell.value = data[k]
                        ws.cell(row, 16).value = 'Khách hàng mới'
                    else:
                        ws.cell(row, 1).value = row - 9
                    # cell.font = line_font
                    # cell.border = all_border_thin
                    # cell.alignment = Alignment(horizontal='left', vertical='center')
                # if col == 17:
                #     cell.number_format = format_currency
            row += 1
        fp = BytesIO()
        wb.save(fp)
        fp.seek(0)
        report = base64.encodebytes((fp.read()))
        fp.close()
        attachment = self.env['ir.attachment'].sudo().create({
            'name': 'bao_cao_doanh_thu_thuc_hien_dich_vu_%s.xlsx' % self.company_id.name,
            'datas': report,
            'res_model': 'temp.creation',
            'public': True,
        })
        url = "/web/content/?model=ir.attachment&id=%s&filename_field=name&field=datas&download=true" \
              % attachment.id
        return {'name': 'BÁO CÁO DOANH THU THỰC HIỆN DỊCH VỤ',
                'type': 'ir.actions.act_url',
                'url': url,
                'target': 'self',
                }
