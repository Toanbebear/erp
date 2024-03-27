from odoo import fields, api, models, _
from odoo.exceptions import ValidationError, UserError
from datetime import date, datetime, time
from calendar import monthrange
import openpyxl
from openpyxl.drawing.image import Image
from openpyxl import load_workbook
from openpyxl.styles import Font, borders, Alignment, PatternFill
from odoo.addons.report_sale.wizard.theme_report import ThemeReport
import base64
from io import BytesIO
from dateutil.relativedelta import relativedelta
from pytz import timezone, utc
from odoo.modules.module import get_module_resource


class SpecialTreatmentReport(models.TransientModel):
    _name = 'report.paris.phone.call.cham.soc.khach.hang'
    _description = 'Báo cáo chăm sóc Phone call'

    start_date = fields.Date('Start date', default=date.today())
    end_date = fields.Date('End date', default=date.today())
    care_type = fields.Selection(
        [('All', 'Tất cả đơn vị'), ('Spa', 'Spa'), ('Laser', 'Laser'), ('Odontology', 'Nha'), ('Surgery', 'Phẫu thuật'),
         ('DVKH', 'Dịch vụ khách hàng'), ('LT', 'Lễ tân')], 'Bộ phận chăm sóc', default='All')
    company_id = fields.Many2one('res.company', string='Chi nhánh',
                                 domain=lambda self: [('id', 'in', self.env.user.company_ids.ids)])

    # convert date to datetime for search domain, should be removed if using datetime directly
    start_datetime = fields.Datetime('Start datetime', compute='_compute_datetime')
    end_datetime = fields.Datetime('End datetime', compute='_compute_datetime')

    @api.depends('start_date', 'end_date')
    def _compute_datetime(self):
        self.start_datetime = False
        self.end_datetime = False
        if self.start_date and self.end_date:
            local_tz = timezone(self.env.user.tz or 'Etc/GMT+7')
            start_datetime = datetime(self.start_date.year, self.start_date.month, self.start_date.day, 0, 0, 0)
            end_datetime = datetime(self.end_date.year, self.end_date.month, self.end_date.day, 23, 59, 59)
            start_datetime = local_tz.localize(start_datetime, is_dst=None)
            end_datetime = local_tz.localize(end_datetime, is_dst=None)
            self.start_datetime = start_datetime.astimezone(utc).replace(tzinfo=None)
            self.end_datetime = end_datetime.astimezone(utc).replace(tzinfo=None)

    @api.onchange('start_date')
    def _onchange_start_date(self):
        if self.start_date:
            if self.start_date.month == fields.date.today().month:
                self.end_date = fields.date.today()
            else:
                self.end_date = date(self.start_date.year, self.start_date.month,
                                     monthrange(self.start_date.year, self.start_date.month)[1])

    @api.constrains('start_date', 'end_date')
    def check_dates(self):
        for record in self:
            start_date = fields.Date.from_string(record.start_date)
            end_date = fields.Date.from_string(record.end_date)
            if start_date > end_date:
                raise ValidationError(
                    _("Ngày kết thúc không thể ở trước ngày bắt đầu khi xuất báo cáo!!!"))

    def create_report_paris_bao_cao_phone_cal_cham_soc_khach_hang(self):
        # get data
        datas = self._get_data_report_paris_bao_cao_phone_cal_cham_soc_khach_hang()

        # get temp excel
        report_brand_overview_attachment = self.env['ir.attachment'].browse(
            self.env.ref('report_sale.report_paris_phone_call_cham_soc_khach_hang_attachment').id)
        decode = base64.b64decode(report_brand_overview_attachment.datas)
        wb = load_workbook(BytesIO(decode))
        ws = wb.active

        # in thông tin chung
        code_brand = self.company_id.brand_id.code.lower()
        if code_brand == 'kn':
            header_fill = ThemeReport.kn_fill
        elif code_brand == 'da':
            header_fill = ThemeReport.da_fill
        elif code_brand == 'pr':
            header_fill = ThemeReport.pr_fill
        elif code_brand == 'hh':
            header_fill = ThemeReport.hh_fill
        else:
            header_fill = ThemeReport.sci_fill
        image_path = get_module_resource('report_sale', 'static/img', 'icon_%s.png' % code_brand)
        img = openpyxl.drawing.image.Image(image_path)
        img.anchor = 'A1'
        ws.add_image(img)
        ws['I2'].value = self.company_id.name
        ws['I3'].value = self.start_date.strftime('%d/%m/%Y')
        ws['K3'].value = self.end_datetime.strftime('%d/%m/%Y')

        # in tiêu đề và giá trị
        key_list = [
            'type_phone_call',
            'name_phone_call',
            'partner',
            'code_booking',
            'booking_services',
            'walkin_services',
            'service_date',
            'surgery_team',
            'call_date',
            'desc',
            'note',
            'booking_date',
            'user_updated',
            'care_type',
            'type_pc',
            'support_rating',
            'direction',
            'state',
            'company_id',
        ]
        key_list_title = [
            'STT',
            'LOẠI PHONE CALL',
            'TÊN PHONE CALL',
            'TÊN KHÁCH HÀNG',
            'MÃ BOOKING',
            'DỊCH VỤ BOOKING',
            'DỊCH VỤ THỰC HIỆN (PHIẾU KHÁM)',
            'NGÀY LÀM DV',
            'BÁC SĨ PHẪU THUẬT',
            'NGÀY GỌI',
            'MÔ TẢ',
            'NỘI DUNG GHI CHÚ',
            'NGÀY HẸN LỊCH',
            'CẬP NHẬT LẦN CUỐI BỞI',
            'ĐƠN VỊ CHĂM SÓC',
            'LOẠI',
            'ĐÁNH GIÁ HÀI LÒNG',
            'HƯỚNG GỌI',
            'TRẠNG THÁI',
            'CÔNG TY',
        ]
        key_col_list = list(range(2, len(key_list) + 2))
        key_col_list_title = list(range(1, len(key_list) + 2))
        line_font = Font(name='Times New Roman', size=12)
        row = 6
        index_row = 0

        for col, k in zip(key_col_list_title, key_list_title):
            cell = ws.cell(5, col)
            cell.fill = header_fill
            cell.value = k
            cell.font = Font(name='Times New Roman', size=12, color='FFFFFF')
            cell.border = ThemeReport.all_border_thin
            cell.alignment = Alignment(horizontal='left', vertical='center')

        for data in datas:
            ws.cell(row, 1).border = ThemeReport.all_border_thin
            ws.cell(row, 1).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(row, 1).value = index_row + 1
            for col, k in zip(key_col_list, key_list):
                cell = ws.cell(row, col)
                cell.value = data[k]
                cell.font = line_font
                cell.border = ThemeReport.all_border_thin
                if col in [6, 7, 9, 11, 12]:
                    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                else:
                    cell.alignment = Alignment(horizontal='left', vertical='center')
            row += 1
            index_row += 1
        fp = BytesIO()
        wb.save(fp)
        fp.seek(0)
        report = base64.encodebytes((fp.read()))
        fp.close()
        attachment = self.env['ir.attachment'].sudo().create({
            'name': 'bao_cao_phone_call_cham_soc_khach_hang_%s.xlsx' % self.company_id.name,
            'datas': report,
            'res_model': 'temp.creation',
            'public': True,
        })
        url = "/web/content/?model=ir.attachment&id=%s&filename_field=name&field=datas&download=true" \
              % attachment.id
        return {'name': 'BÁO CÁO PHONE CALL CHĂM SÓC KHÁCH HÀNG',
                'type': 'ir.actions.act_url',
                'url': url,
                'target': 'self',
                }

    def _get_data_report_paris_bao_cao_phone_cal_cham_soc_khach_hang(self):
        ret_data = []
        # search data
        domain = [('call_date', '>=', self.start_datetime), ('call_date', '<=', self.end_datetime),
                  ('company_id', '=', self.company_id.id)]
        if self.care_type != 'All':
            domain += [('care_type', '=', self.care_type)]
        phones_call = self.env['crm.phone.call'].sudo().search(domain, order='call_date asc')

        # lấy thông tin care type
        care_type = {
            "Spa": 'Spa', "Laser": 'Laser', "Odontology": 'Nha',
            "Surgery": 'Phẫu thuật',
            "DVKH": 'Dịch vụ khách hàng', "LT": 'Lễ tân'
        }

        # lấy thông tin type_pc
        type_pc = {
            'Check': 'Chăm sóc lần 1',
            'Check1': 'Chăm sóc lần 2',
            'Check2': 'Chăm sóc lần 3',
            'Check3': 'Chăm sóc lần 4',
            'Check8': 'Chăm sóc lần 5',
            'Check4': 'Chăm sóc kết thúc liệu trình 1',
            'Check5': 'Chăm sóc kết thúc liệu trình 2',
            'Check6': 'Chăm sóc kết thúc liệu trình 3',
            'Check7': 'Chăm sóc kết thúc liệu trình 4',
            'Change': 'Thay băng lần 1',
            'Change1': 'Thay băng lần 2',
            'Change2': 'Thay băng lần 3',
            'Change3': 'Thay băng lần 4',
            'Change4': 'Thay băng lần 5',
            'Change5': 'Thay băng lần 6',
            'ReCheck': 'Cắt chỉ',
            'ReCheck1': 'Hút dịch',
            'ReCheck2': 'Rút ống mũi',
            'ReCheck3': 'Thay nẹp mũi',
            'ReCheck4': 'Tái khám lần 1',
            'ReCheck5': 'Tái khám lần 2',
            'ReCheck6': 'Tái khám lần 3',
            'ReCheck7': 'Tái khám lần 4',
            'ReCheck8': 'Tái khám lần 5',
            'ReCheck11': 'Tái khám lần 6',
            'ReCheck9': 'Tái khám định kì',
            'ReCheck10': 'Nhắc liệu trình',
            'Potential': 'Khai thác dịch vụ tiềm năng'
        }

        # lấy thông tin rating
        rating = {
            '1': 'Rất tệ',
            '2': 'Không hài lòng',
            '3': 'Bình thường',
            '4': 'Hài lòng',
            '5': 'Rất hài lòng',
        }

        # lấy thông tin direction
        direction = {
            'out': 'Gọi ra',
            'in': 'Gọi vào'
        }

        # lấy thông tin state
        state = {
            'draft': 'Chưa xử lý',
            'not_connect': 'Chưa kết nối',
            'connected': 'Đã xử lý',
            'later': 'Hẹn gọi lại sau',
            'duplicate': 'Trùng KH',
            'connected_1': 'Đã hăm sóc',
            'not_connect_1': 'Chuyển lịch',
            'error_phone': 'Sai Số',
            'connected_2': 'Xác nhận lịch',
            'later_1': 'Hẹn lịch',
            'cancelled': 'Hủy lịch'
        }

        for phone_call in phones_call:
            # lấy thông tin dịch vụ booking
            booking_services = []
            if phone_call.crm_line_id:
                for rec_crm_line_id in phone_call.crm_line_id:
                    booking_services.append(rec_crm_line_id.product_id.name)

            # lấy thông tin dịch vụ phiếu khám
            walkin_services = []
            if phone_call.service_id:
                for rec_service_id in phone_call.service_id:
                    walkin_services.append(rec_service_id.name)

            # lấy thông tin dịch vụ phiếu khám
            surgery_team = []
            if phone_call.medical_id:
                if phone_call.medical_id.surgeries_ids:
                    for rec_surgery_id in phone_call.medical_id.surgeries_ids:
                        if rec_surgery_id.surgery_team:
                            for rec_surgery_team in rec_surgery_id.surgery_team:
                                if rec_surgery_team.role.name == 'BS mổ chính':
                                    surgery_team.append(rec_surgery_team.team_member.name)
            ret_data.append({
                'type_phone_call': phone_call.type_crm_id.name,
                'name_phone_call': phone_call.name,
                'partner': phone_call.partner_id.name,
                'code_booking': phone_call.crm_id.name,
                'booking_services': ','.join(booking_services) if booking_services else '',
                'walkin_services': ','.join(walkin_services) if walkin_services else '',
                'service_date': phone_call.date.strftime('%d/%m/%Y') if phone_call.date else '',
                'surgery_team': ','.join(surgery_team) if surgery_team else '',
                'call_date': phone_call.call_date.strftime('%d/%m/%Y') if phone_call.call_date else '',
                'desc': phone_call.desc if phone_call.desc else '',
                'note': phone_call.note if phone_call.note else '',
                'booking_date': phone_call.crm_id.booking_date.strftime(
                    '%d/%m/%Y') if phone_call.crm_id.booking_date else '',
                'user_updated': phone_call.write_uid.name if phone_call.write_uid else '',
                'care_type': care_type['%s' % phone_call.care_type] if phone_call.care_type else '',
                'type_pc': type_pc['%s' % phone_call.type_pc] if phone_call.type_pc else '',
                'support_rating': rating['%s' % phone_call.support_rating] if phone_call.support_rating else '',
                'direction': direction['%s' % phone_call.direction] if phone_call.direction else '',
                'state': state['%s' % phone_call.state] if phone_call.state else '',
                'company_id': phone_call.company_id.name,
            })
        return ret_data
