from odoo import fields, api, models, _
from odoo.exceptions import ValidationError, UserError
from datetime import date, datetime, time
from calendar import monthrange
import openpyxl
from openpyxl.drawing.image import Image
from openpyxl import load_workbook
from openpyxl.styles import Font, borders, Alignment, PatternFill
import base64
from io import BytesIO
from dateutil.relativedelta import relativedelta
from pytz import timezone, utc
from odoo.modules.module import get_module_resource

thin = borders.Side(style='thin')
double = borders.Side(style='double')
all_border_thin = borders.Border(thin, thin, thin, thin)
da_fill = PatternFill(start_color='0e7661', end_color='0e7661', fill_type='solid')
kn_fill = PatternFill(start_color='003471', end_color='003471', fill_type='solid')
pr_fill = PatternFill(start_color='012C5F', end_color='012C5F', fill_type='solid')
hh_fill = PatternFill(start_color='053D7C', end_color='053D7C', fill_type='solid')
sci_fill = PatternFill(start_color='003471', end_color='003471', fill_type='solid')


class SpecialTreatmentReport(models.TransientModel):
    _name = 'special.treatment.report'
    _description = 'Báo cáo khách hàng chăm sóc đặc biệt'

    start_date = fields.Date('Start date', default=date.today())
    end_date = fields.Date('End date', default=date.today())
    company_id = fields.Many2one('res.company', string='Chi nhánh', domain=lambda self: [('id', 'in', self.env.user.company_ids.ids)])

    # convert date to datetime for search domain, should be removed if using datetime directly
    start_datetime = fields.Datetime('Start datetime', compute='_compute_datetime')
    end_datetime = fields.Datetime('End datetime', compute='_compute_datetime')

    @api.depends('start_date', 'end_date')
    def _compute_datetime(self):
        self.start_datetime = False
        self.end_datetime = False
        if self.start_date and self.end_date:
            local_tz = timezone(self.env.user.tz or 'Etc/GMT+7')
            start_datetime = datetime(self.start_date.year, self.start_date.month, self.start_date.day, 0, 0, 0)
            end_datetime = datetime(self.end_date.year, self.end_date.month, self.end_date.day, 23, 59, 59)
            start_datetime = local_tz.localize(start_datetime, is_dst=None)
            end_datetime = local_tz.localize(end_datetime, is_dst=None)
            self.start_datetime = start_datetime.astimezone(utc).replace(tzinfo=None)
            self.end_datetime = end_datetime.astimezone(utc).replace(tzinfo=None)

    @api.onchange('start_date')
    def _onchange_start_date(self):
        if self.start_date:
            if self.start_date.month == fields.date.today().month:
                self.end_date = fields.date.today()
            else:
                self.end_date = date(self.start_date.year, self.start_date.month,
                                     monthrange(self.start_date.year, self.start_date.month)[1])

    @api.constrains('start_date', 'end_date')
    def check_dates(self):
        for record in self:
            start_date = fields.Date.from_string(record.start_date)
            end_date = fields.Date.from_string(record.end_date)
            days = (end_date - start_date).days
            if days < 0 or days > 365:
                raise ValidationError(
                    _("Ngày kết thúc không thể ở trước ngày bắt đầu khi xuất báo cáo!!!"))

    def render_form_template(self, evaluation):
        surgery_date = ''
        main_doctor = ''

        id_history_surgery_date= []
        for s_date in evaluation.surgery_history_ids.mapped('surgery_date'):
            if s_date.strftime('%d/%m/%Y') not in id_history_surgery_date:
                id_history_surgery_date.append(s_date.strftime('%d/%m/%Y'))
                surgery_date += s_date.strftime('%d/%m/%Y') + "\n"

        for m_doctor in evaluation.surgery_history_ids.mapped('main_doctor'):
            main_doctor += m_doctor.name + "\n"
        val = {
            'evaluation_start_date': evaluation.evaluation_start_date.strftime('%d/%m/%Y'),
            'patient': evaluation.patient.name,
            'code_evaluation': evaluation.name,
            'code_booking': evaluation.walkin.booking_id.name,
            'services': '\n'.join(evaluation.services.mapped('name')),
            'surgery_date': surgery_date,
            'main_doctor': main_doctor,
            'evaluation_doctor': evaluation.doctor.name or '',
            'patient_level': str(dict(evaluation._fields['patient_level']._description_selection(self.env)).get(evaluation.patient_level)),
            'reason_bh': evaluation.walkin.booking_id.note if evaluation.walkin.booking_id.type_crm_id.id == self.env.ref('crm_base.type_oppor_guarantee').id else '',
            'notes_complaint': evaluation.notes_complaint or '',
            'notes': evaluation.notes or '',
            'chief_complaint': evaluation.chief_complaint or ''
        }
        return val

    def _get_data_report(self):
        domain = [('state', '=', 'Completed'),('patient_level', '=', '4'), ('institution.his_company', '=', self.company_id.id),
                  ('evaluation_start_date', '>=', self.start_datetime), ('evaluation_start_date', '<=', self.end_datetime)]
        evaluations = self.env['sh.medical.evaluation'].sudo().search(domain, order='id')
        return_val = []
        for evaluation in evaluations:
            return_val.append(self.render_form_template(evaluation))
        return return_val

    def create_report(self):
        datas = self._get_data_report()
        # in dữ liễu
        report_khach_hang_dac_biet_attachment = self.env['ir.attachment'].browse(
            self.env.ref('report_sale.bao_cao_khach_hang_dac_biet_attachment').id)
        decode = base64.b64decode(report_khach_hang_dac_biet_attachment.datas)
        wb = load_workbook(BytesIO(decode))
        ws = wb.active
        line_font = Font(name='Times New Roman', size=14)

        code_brand = self.company_id.brand_id.code.lower()
        image_path = get_module_resource('report_sale', 'static/img', 'icon_%s.png' % code_brand)

        img = openpyxl.drawing.image.Image(image_path)
        img.anchor = 'A1'
        ws.add_image(img)

        ws['E4'].value = 'Từ: %s' % (self.start_date.strftime('%d/%m/%Y'))
        ws['F4'].value = 'Đến %s ' % (self.end_date.strftime('%d/%m/%Y'))
        ws['E5'].value = 'Đơn vị: %s ' % (self.company_id.name)
        key_col = list(range(2, 14))
        key_list = [
            'evaluation_start_date',
            'patient',
            'code_evaluation',
            'code_booking',
            'services',
            'surgery_date',
            'main_doctor',
            'evaluation_doctor',
            'patient_level',
            'reason_bh',
            'notes_complaint',
            'chief_complaint'
        ]
        row = 8
        index_row = 0
        if code_brand == 'kn':
            header_fill = kn_fill
        elif code_brand == 'da':
            header_fill = da_fill
        elif code_brand == 'pr':
            header_fill = pr_fill
        elif code_brand == 'hh':
            header_fill = hh_fill
        else:
            header_fill = sci_fill

        for data in datas:
            cell_index = ws.cell(row, 1)
            cell_index.border, cell_index.alignment, cell_index.value = all_border_thin, Alignment(horizontal='center', vertical='center', wrap_text=True), index_row + 1

            ws.cell(7, 1).fill, ws.cell(7, 1).font = header_fill, Font(name='Times New Roman', size=14, color='FFFFFF')
            for col, k in zip(key_col, key_list):
                beforeCell = ws.cell(7, col)
                beforeCell.fill = header_fill
                beforeCell.font = Font(name='Times New Roman', size=14, color='FFFFFF')

                cell = ws.cell(row, col)
                cell.value = data[k]
                cell.font = line_font
                cell.border = all_border_thin
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            row += 1
            index_row += 1
        fp = BytesIO()
        wb.save(fp)
        fp.seek(0)
        report = base64.encodebytes((fp.read()))
        fp.close()
        attachment = self.env['ir.attachment'].sudo().create({
            'name': 'bao_cao_khach_hang_dac_biet.xlsx',
            'datas': report,
            'res_model': 'temp.creation',
            'public': True,
        })
        url = "/web/content/?model=ir.attachment&id=%s&filename_field=name&field=datas&download=true" \
              % attachment.id
        return {'name': 'BÁO CÁO DOANH THU BỆNH NHÂN LƯU',
                'type': 'ir.actions.act_url',
                'url': url,
                'target': 'self',
                }
