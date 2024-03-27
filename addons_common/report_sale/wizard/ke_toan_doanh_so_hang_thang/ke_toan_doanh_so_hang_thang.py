from odoo import fields, api, models, _
from odoo.exceptions import ValidationError, UserError
from datetime import date, datetime, time
from calendar import monthrange
from openpyxl import load_workbook
from openpyxl.styles import Font, borders, Alignment
import base64
from io import BytesIO
from dateutil.relativedelta import relativedelta
from pytz import timezone, utc
from collections import defaultdict
import numpy as np
import pandas as pd
import logging

_logger = logging.getLogger(__name__)
TITLE = {0: 'major', 1: 'minor', 2: 'Laser', 3: 'Spa', 4: 'Odontology', 5: 'Unknown'}
thin = borders.Side(style='thin')
double = borders.Side(style='double')
all_border_thin = borders.Border(thin, thin, thin, thin)
"""
    Chị ngocbb@scigroup.com.vn order theo task#1005
    Email gửi ngày: 18/04/2022
"""
TEMP = [
    'ten_phieu',
    'ngay_phieu_thu',
    'ma_booking',
    'bang_gia',
    'cong_ty',
    'noi_dung',
    'loai_thanh_toan',
    'so_nhat_ky',
    'ma_khach_hang',
    'ten_khach_hang',
    'so_tien_bang_so',
    'trang_thai',
    'ma_san_pham',
    'ten_san_pham',
    'don_vi_tinh',
    'so_luong',
    'thanh_tien'
]


class AccountReportEveryMonth(models.TransientModel):
    _name = 'account.payment.report'
    _description = 'Báo cáo tổng hợp Doanh số tháng'

    start_date = fields.Date('Start date', default=date.today())
    end_date = fields.Date('End date', default=date.today())
    company_id = fields.Many2many(string='Công ty', comodel_name='res.company', domain=lambda self: [('id', 'in', self.env.user.company_ids.ids)])
    # convert date to datetime for search domain, should be removed if using datetime directly
    start_datetime = fields.Datetime('Start datetime', compute='_compute_datetime')
    end_datetime = fields.Datetime('End datetime', compute='_compute_datetime')

    @api.depends('start_date', 'end_date')
    def _compute_datetime(self):
        self.start_datetime = False
        self.end_datetime = False
        if self.start_date and self.end_date:
            local_tz = timezone(self.env.user.tz or 'Etc/GMT+7')
            start_datetime = datetime(self.start_date.year, self.start_date.month, self.start_date.day, 0, 0, 0)
            end_datetime = datetime(self.end_date.year, self.end_date.month, self.end_date.day, 23, 59, 59)
            start_datetime = local_tz.localize(start_datetime, is_dst=None)
            end_datetime = local_tz.localize(end_datetime, is_dst=None)
            self.start_datetime = start_datetime.astimezone(utc).replace(tzinfo=None)
            self.end_datetime = end_datetime.astimezone(utc).replace(tzinfo=None)

    @api.onchange('start_date')
    def _onchange_start_date(self):
        if self.start_date:
            if self.start_date.month == fields.date.today().month:
                self.end_date = fields.date.today()
            else:
                self.end_date = date(self.start_date.year, self.start_date.month,
                                     monthrange(self.start_date.year, self.start_date.month)[1])

    @api.constrains('start_date', 'end_date')
    def check_dates(self):
        for record in self:
            start_date = fields.Date.from_string(record.start_date)
            end_date = fields.Date.from_string(record.end_date)
            days = (end_date - start_date).days
            if days < 0 or days > 365:
                raise ValidationError(
                    _("Ngày kết thúc không thể ở trước ngày bắt đầu khi xuất báo cáo!!!"))

    def create_report(self):
        domain = [('state', 'in', ('posted', 'reconciled')), ('crm_id', '!=', None),
                  ('payment_date', '>=', self.start_date), ('payment_date', '<=', self.end_date),
                  ('journal_id.company_id', 'in', self.company_id.ids)]
        payments = self.sudo().env['account.payment'].with_context(company_ids=self.env['res.company'].search([('active', '=', True)]).ids).search(domain, order='payment_date')
        datas = list()

        grouped_payment = defaultdict(list)
        grouped_sale_order = defaultdict(list)

        for pay in payments:
            grouped_payment[pay.crm_id].append(pay)
        for key in grouped_payment.keys():
            sale_order = self.sudo().env['sale.order'].with_context(company_ids=self.env['res.company'].search([]).ids).search(
                [('booking_id', '=', key.id), ('state', 'not in', ('cancel',))])
            if sale_order:
                for so in sale_order:
                    grouped_sale_order[key].append(so)

        result = list()
        for key in grouped_payment.keys():
            if key not in grouped_sale_order.keys():
                for pay in grouped_payment.get(key):
                    val = {
                        'payment': pay,
                        'sale_order': None,
                    }
                    result.append(val)
            else:
                list_payment = list()
                for pay in grouped_payment.get(key):
                    list_payment.append((pay, pay.amount_vnd))
                list_sale_order = list()
                for so in grouped_sale_order.get(key):
                    if so.pricelist_id.type == 'product':
                        list_sale_order.append((so, so.amount_total))
                # sort list_sale_order
                list_payment.sort(key=lambda x: x[1], reverse=False)
                list_sale_order.sort(key=lambda x: x[1], reverse=False)

                temp_list_sale_order = list_sale_order.copy()

                for element in list_payment:
                    amount_sale_order = element[1]
                    count = 0
                    if temp_list_sale_order:
                        for ele in temp_list_sale_order:
                            amount_payment = ele[1]
                            if amount_sale_order == amount_payment:
                                val = {
                                    'payment': element[0],
                                    'sale_order': ele[0]
                                }
                                result.append(val)
                                temp_list_sale_order.pop(temp_list_sale_order.index(ele))
                                count += 1
                                break
                        if count == 0:
                            val = {
                                'payment': element[0],
                                'sale_order': None,
                            }
                            result.append(val)
                    else:
                        val = {
                            'payment': element[0],
                            'sale_order': None,
                        }
                        result.append(val)

        def sort_by_element(element):
            return element.get('payment').id

        result.sort(key=sort_by_element, reverse=False)
        for element in result:
            payment = element.get('payment')
            sale_order = element.get('sale_order')

            val = {
                'ten_phieu': payment.name,
                'ngay_phieu_thu': payment.payment_date.strftime('%d/%m/%Y'),
                'ma_booking': payment.crm_id.name,
                'bang_gia': sale_order.pricelist_id.type if sale_order else None,
                'cong_ty': payment.journal_id.company_id.name,
                'noi_dung': payment.communication,
                'loai_thanh_toan': payment.payment_type,
                'so_nhat_ky': payment.journal_id.name,
                'ma_khach_hang': payment.partner_id.code_customer,
                'ten_khach_hang': payment.partner_id.name,
                'so_tien_bang_so': payment.amount_vnd,
                'trang_thai': payment.state,
                'ma_san_pham': None,
                'ten_san_pham': None,
                'don_vi_tinh': None,
                'so_luong': None,
                'thanh_tien': None
            }
            if sale_order:
                sol = sale_order.order_line
                count = 0
                for line in sol:
                    count += 1
                    if count == 1:
                        val['ma_san_pham'] = line.product_id.default_code
                        val['ten_san_pham'] = line.product_id.name
                        val['don_vi_tinh'] = line.product_uom.name
                        val['so_luong'] = line.product_uom_qty
                        val['thanh_tien'] = line.price_subtotal
                        datas.append(val)
                    else:
                        val = {
                            'ten_phieu': None,
                            'ngay_phieu_thu': None,
                            'ma_booking': None,
                            'bang_gia': None,
                            'cong_ty': None,
                            'noi_dung': None,
                            'loai_thanh_toan': None,
                            'so_nhat_ky': None,
                            'ma_khach_hang': None,
                            'ten_khach_hang': None,
                            'so_tien_bang_so': None,
                            'trang_thai': None,
                            'ma_san_pham': line.product_id.default_code,
                            'ten_san_pham': line.product_id.name,
                            'don_vi_tinh': line.product_uom.name,
                            'so_luong': line.product_uom_qty,
                            'thanh_tien': line.price_subtotal
                        }
                        datas.append(val)
            else:
                datas.append(val)

        report_brand_overview_attachment = self.env['ir.attachment'].browse(
            self.env.ref('report_sale.ke_toan_doanh_so_hang_thang_attachment').id)
        decode = base64.b64decode(report_brand_overview_attachment.datas)
        wb = load_workbook(BytesIO(decode))
        ws = wb.active
        line_font = Font(name='Times New Roman', size=14)
        ws['A2'].value = "TÊN CÔNG TY %s" % '; '.join(c.name for c in self.company_id)
        ws['F3'].value += self.start_date.strftime('%d/%m/%Y')
        ws['I3'].value += self.end_date.strftime('%d/%m/%Y')
        key_col = list(range(1, len(TEMP) + 1))
        format_currency = '_(* #,##0_);_(* (#,##0);_(* "-"??_);_(@_)'
        row = 7
        for data in datas:
            for col, k in zip(key_col, TEMP):
                cell = ws.cell(row, col)
                cell.value = data[k]
                cell.font = line_font
                cell.border = all_border_thin
                cell.alignment = Alignment(horizontal='left', vertical='center')
                if col in (11, 16, 17):
                    cell.number_format = format_currency
            row += 1
        fp = BytesIO()
        wb.save(fp)
        fp.seek(0)
        report = base64.encodebytes((fp.read()))
        fp.close()
        attachment = self.env['ir.attachment'].sudo().create({
            'name': 'ke_toan_doanh_so_hang_thang_%s.xlsx' % self.end_date.month,
            'datas': report,
            'res_model': 'temp.creation',
            'public': True,
        })
        url = "/web/content/?model=ir.attachment&id=%s&filename_field=name&field=datas&download=true" \
              % attachment.id
        return {'name': 'BÁO CÁO KẾ TOÁN DOANH SỐ HÀNG THÁNG',
                'type': 'ir.actions.act_url',
                'url': url,
                'target': 'self',
                }
