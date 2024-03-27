from odoo import fields, api, models, _
from odoo.exceptions import ValidationError, UserError
from datetime import date, datetime, time
from calendar import monthrange
from openpyxl import load_workbook
from openpyxl.styles import Font, borders, Alignment
import base64
from io import BytesIO
from dateutil.relativedelta import relativedelta
from pytz import timezone, utc

thin = borders.Side(style='thin')
double = borders.Side(style='double')
all_border_thin = borders.Border(thin, thin, thin, thin)

KEY_LIST = [
    'ngay_dat_hang',
    'chi_nhanh',
    'ma_booking',
    'SO',
    'nguoi_tao',
    'phong_ban_nguoi_tao',
    'ma_dich_vu',
    'ten_dich_vu',
    'nhom_dich_vu',
    'bang_gia',
    'nhom_nguon',
    'nguon',
    'ma_khach_hang',
    'ten_khach_hang',
    'so_dien_thoai',
    'dia_chi',
    'email',
    'tong_tien',
    'trang_thai'
]


class SaleMarketingSaleOrderDA(models.TransientModel):
    _name = 'sale.marketing.sale.order.da'
    _description = 'Báo cáo bán hàng'

    start_date = fields.Date('Start date', default=date.today())
    end_date = fields.Date('End date', default=date.today())
    # brand_id = fields.Many2one(string='Thương hiệu', comodel_name='res.brand', domain=lambda self: [('id', 'in', self.env.user.company_ids.mapped('brand_id').ids)])
    company_ids = fields.Many2many(string='Chi nhánh', comodel_name='res.company', domain=lambda self: [('id', 'in', self.env.user.company_ids.ids)])

    # convert date to datetime for search domain, should be removed if using datetime directly
    start_datetime = fields.Datetime('Start datetime', compute='_compute_datetime')
    end_datetime = fields.Datetime('End datetime', compute='_compute_datetime')

    @api.depends('start_date', 'end_date')
    def _compute_datetime(self):
        self.start_datetime = False
        self.end_datetime = False
        if self.start_date and self.end_date:
            local_tz = timezone(self.env.user.tz or 'Etc/GMT+7')
            start_datetime = datetime(self.start_date.year, self.start_date.month, self.start_date.day, 0, 0, 0)
            end_datetime = datetime(self.end_date.year, self.end_date.month, self.end_date.day, 23, 59, 59)
            start_datetime = local_tz.localize(start_datetime, is_dst=None)
            end_datetime = local_tz.localize(end_datetime, is_dst=None)
            self.start_datetime = start_datetime.astimezone(utc).replace(tzinfo=None)
            self.end_datetime = end_datetime.astimezone(utc).replace(tzinfo=None)

    @api.onchange('start_date')
    def _onchange_start_date(self):
        if self.start_date:
            if self.start_date.month == fields.date.today().month:
                self.end_date = fields.date.today()
            else:
                self.end_date = date(self.start_date.year, self.start_date.month,
                                     monthrange(self.start_date.year, self.start_date.month)[1])

    @api.constrains('start_date', 'end_date')
    def check_dates(self):
        for record in self:
            start_date = fields.Date.from_string(record.start_date)
            end_date = fields.Date.from_string(record.end_date)
            days = (end_date - start_date).days
            if days < 0 or days > 365:
                raise ValidationError(
                    _("Ngày kết thúc không thể ở trước ngày bắt đầu khi xuất báo cáo!!!"))

    def render_form_template(self, sol):
        so = sol.order_id
        crm_line = sol.crm_line_id
        crm = crm_line.crm_id

        code_customer = so.code_customer
        partner = self.env['res.partner'].search([('code_customer', '=', code_customer)], limit=1)

        str_tuple = (partner.pass_port_address, partner.district_id.name, partner.state_id.name, partner.country_id.name)
        str_convert = (element for element in str_tuple if element != False)
        dia_chi = ', '.join(str_convert)

        product_code = crm_line.service_id.code or sol.product_id.code
        product_name = crm_line.service_id.name or sol.product_id.name
        product_categ = crm_line.service_id.service_category or sol.product_id.categ_id

        lang_code = self.env.context.get('lang') or 'en_US'
        date_format = self.env['res.lang']._lang_get(lang_code).date_format
        time_format = self.env['res.lang']._lang_get(lang_code).time_format

        val = {
            'ngay_dat_hang': '%s' % fields.Datetime.context_timestamp(self.with_context(tz=self.env.user.tz), so.date_order).strftime(('%s %s') % (date_format, time_format)),
            'chi_nhanh': so.company_id.name,
            'ma_booking': so.booking_id.name or '',
            'SO': so.name,
            'nguoi_tao': so.user_id.name or '',
            'phong_ban_nguoi_tao': self.env['hr.employee'].sudo().search([('user_id', '=', so.user_id.id)], limit=1).department_id.name,
            'ma_dich_vu': product_code,
            'ten_dich_vu': product_name,
            'nhom_dich_vu': product_categ.name,
            'bang_gia': so.pricelist_id.name,
            'nhom_nguon': crm.category_source_id.name or '',
            'nguon': crm.source_id.name or '',
            'ma_khach_hang': code_customer,
            'ten_khach_hang': partner.display_name,
            'so_dien_thoai': so.phone_customer,
            'dia_chi': dia_chi,
            'email': partner.email or '',
            'tong_tien': sol.price_subtotal,
            'trang_thai': sol.state
        }
        return val

    def _get_data_report(self):
        domain = [('stage_sol', '!=', 'cancel'), ('order_id.state', '=', 'sale'), ('order_id.company_id', 'in', self.company_ids.ids),
                  ('order_id.date_order', '>=', self.start_datetime), ('order_id.date_order', '<=', self.end_datetime)]
        sale_order_lines = self.env['sale.order.line'].sudo().search(domain, order='id')
        return_val = []
        for line in sale_order_lines:
            return_val.append(self.render_form_template(line))
        return return_val

    def create_report_da(self):
        datas = self._get_data_report()
        # in dữ liễu
        report_brand_overview_attachment = self.env['ir.attachment'].browse(
            self.env.ref('report_sale.bao_cao_sale_order_attachment').id)
        decode = base64.b64decode(report_brand_overview_attachment.datas)
        wb = load_workbook(BytesIO(decode))
        ws = wb.active
        line_font = Font(name='Times New Roman', size=14)

        ws['J3'].value = self.start_date.strftime('%d/%m/%Y')
        ws['L3'].value = self.end_datetime.strftime('%d/%m/%Y')
        key_col = list(range(1, len(KEY_LIST) + 1))

        row = 7
        for data in datas:
            for col, k in zip(key_col, KEY_LIST):
                cell = ws.cell(row, col)
                cell.value = data[k]
                cell.font = line_font
                cell.border = all_border_thin
                cell.alignment = Alignment(horizontal='left', vertical='center')
            row += 1
        fp = BytesIO()
        wb.save(fp)
        fp.seek(0)
        report = base64.encodebytes((fp.read()))
        fp.close()
        attachment = self.env['ir.attachment'].sudo().create({
            'name': 'bao_cao_sale_order.xlsx',
            'datas': report,
            'res_model': 'temp.creation',
            'public': True,
        })
        url = "/web/content/?model=ir.attachment&id=%s&filename_field=name&field=datas&download=true" \
              % attachment.id
        return {'name': 'BÁO CÁO SALE ORDER',
                'type': 'ir.actions.act_url',
                'url': url,
                'target': 'self',
                }
