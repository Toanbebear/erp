import base64
from calendar import monthrange
from datetime import date
from io import BytesIO

from openpyxl import load_workbook
from openpyxl.styles import Font, borders

from odoo import fields, api, models, _
from odoo.exceptions import ValidationError

thin = borders.Side(style='thin')
double = borders.Side(style='double')
all_border_thin = borders.Border(thin, thin, thin, thin)

GROUP_REPORT_CHECK = ['Surgery', 'Laser', 'Spa', 'Odontology', 'Other']

chua_xu_ly = ('draft', 'not_connect', 'later', 'not_connect_1')
da_xu_ly = ('connected', 'duplicate', 'zalo', 'sms', 'connected_2', 'later_1', 'cancelled', 'before', 'connected_1', 'not_connect_1')
sai_so = ('error_phone',)

type_pc = {
	'Check': 'Chăm sóc lần 1',
	'Check1': 'Chăm sóc lần 2',
	'Check2': 'Chăm sóc lần 3',
	'Check3': 'Chăm sóc lần 4',
	'Check8': 'Chăm sóc lần 5',
	'Check4': 'Chăm sóc kết thúc liệu trình 1',
	'Check5': 'Chăm sóc kết thúc liệu trình 2',
	'Check6': 'Chăm sóc kết thúc liệu trình 3',
	# 'Check7': 'Chăm sóc kết thúc liệu trình 4',
	'Change': 'Thay băng lần 1',
	'Change1': 'Thay băng lần 2',
	'Change2': 'Thay băng lần 3',
	'Change3': 'Thay băng lần 4',
	'Change4': 'Thay băng lần 5',
	# 'Change5': 'Thay băng lần 6',
	'ReCheck4': 'Tái khám lần 1',
	'ReCheck5': 'Tái khám lần 2',
	'ReCheck6': 'Tái khám lần 3',
	'ReCheck7': 'Tái khám lần 4',
	'ReCheck8': 'Tái khám lần 5',
	# 'ReCheck11': 'Tái khám lần 6',
	'ReCheck9': 'Tái khám định kì',
	'ReCheck': 'Cắt chỉ',
	'ReCheck1': 'Hút dịch',
	'ReCheck2': 'Rút ống mũi',
	'ReCheck3': 'Thay nẹp mũi',
	'ReCheck10': 'Nhắc liệu trình',
	'Potential': 'Khai thác dịch vụ tiềm năng'
}

query_company = ''' select count(*) from crm_phone_call cpc where cpc.group_report = %s and cpc.state in %s and cpc.company_id = %s and cpc.type_pc = %s and cpc.call_date >= %s and cpc.call_date <= %s '''
query_brand = ''' select count(*) from crm_phone_call cpc where cpc.group_report = %s and cpc.state in %s and cpc.company_id in %s and cpc.type_pc = %s and cpc.call_date >= %s and cpc.call_date <= %s '''


class ReportPhoneCall(models.TransientModel):
	_name = 'report.phone.call'
	_description = 'Báo cáo phone call'

	start_date = fields.Date('Start date', default=date.today())
	end_date = fields.Date('End date', default=date.today())
	company_id = fields.Many2one(string='Công ty', comodel_name='res.company')
	brand_id = fields.Many2one(string='Thương hiệu', comodel_name='res.brand')
	report_type = fields.Selection([('company', 'Chi nhánh'), ('brand', 'Thương hiệu')], string='Lựa chọn tải')

	@api.onchange('start_date')
	def _onchange_start_date(self):
		if self.start_date:
			if self.start_date.month == fields.date.today().month:
				self.end_date = fields.date.today()
			else:
				self.end_date = date(self.start_date.year, self.start_date.month,
									 monthrange(self.start_date.year, self.start_date.month)[1])

	@api.constrains('start_date', 'end_date')
	def check_dates(self):
		for record in self:
			start_date = fields.Date.from_string(record.start_date)
			end_date = fields.Date.from_string(record.end_date)
			days = (end_date - start_date).days
			if days < 0 or days > 365:
				raise ValidationError(
					_("Ngày kết thúc không thể ở trước ngày bắt đầu khi xuất báo cáo!!!"))

	def data_type_brand(self):
		data = []
		company_ids = self.env['res.company'].search([('brand_id', '=', self.brand_id.id)])
		for rec_type_pc in type_pc.keys():
			line = []
			for rec_group_pc in GROUP_REPORT_CHECK:
				# nhóm chưa xử lý
				self.env.cr.execute(query_brand, (
				rec_group_pc, chua_xu_ly, tuple(company_ids.ids), rec_type_pc, self.start_date, self.end_date))
				result_cxl = self.env.cr.fetchall()
				line.append(result_cxl[0][0])
				# nhom đã xử lý
				self.env.cr.execute(query_brand, (
					rec_group_pc, da_xu_ly, tuple(company_ids.ids), rec_type_pc, self.start_date, self.end_date))
				result_dxl = self.env.cr.fetchall()
				line.append(result_dxl[0][0])
				# nhom sai số
				self.env.cr.execute(query_brand, (
					rec_group_pc, sai_so, tuple(company_ids.ids), rec_type_pc, self.start_date, self.end_date))
				result_ss = self.env.cr.fetchall()
				line.append(result_ss[0][0])
			data.append(line)
		return data

	def data_type_company(self):
		data = []
		for rec_type_pc in type_pc.keys():
			line = []
			for rec_group_pc in GROUP_REPORT_CHECK:
				# nhóm chưa xử lý
				self.env.cr.execute(query_company, (rec_group_pc, chua_xu_ly, self.company_id.id, rec_type_pc, self.start_date, self.end_date))
				result_cxl = self.env.cr.fetchall()
				line.append(result_cxl[0][0])
				# nhom đã xử lý
				self.env.cr.execute(query_company, (
				rec_group_pc, da_xu_ly, self.company_id.id, rec_type_pc, self.start_date, self.end_date))
				result_dxl = self.env.cr.fetchall()
				line.append(result_dxl[0][0])
				# nhom sai số
				self.env.cr.execute(query_company, (
					rec_group_pc, sai_so, self.company_id.id, rec_type_pc, self.start_date, self.end_date))
				result_ss = self.env.cr.fetchall()
				line.append(result_ss[0][0])
			data.append(line)
		return data

	def report(self):
		inventory_attachment = self.env['ir.attachment'].sudo().browse(
			self.env.ref('report_sale.bao_cao_phone_call_attachment').id)
		decode = base64.b64decode(inventory_attachment.datas)
		wb = load_workbook(BytesIO(decode))
		ws = wb.active

		if self.report_type == 'brand':
			datas = self.data_type_brand()
			ws['f3'].value = 'Thương hiệu:'
			ws['g3'].value = self.brand_id.name
		else:
			datas = self.data_type_company()
			ws['f3'].value = 'Chi nhánh'
			ws['g3'].value = self.company_id.name
		ws['g4'].value = self.start_date
		ws['i4'].value = self.end_date

		ws['f3'].font = Font(name='Times New Roman', size=13)
		ws['g3'].font = Font(name='Times New Roman', size=13)
		ws['g4'].font = Font(name='Times New Roman', size=13)
		ws['i4'].font = Font(name='Times New Roman', size=13)
		row = 9
		col_run = 2
		for group in datas:
			for i in range(0, len(group)):
				ws.cell(row, col_run).value = group[i]
				col_run += 1
			row += 1
			col_run = 2
		fp = BytesIO()
		wb.save(fp)
		fp.seek(0)
		report = base64.encodebytes((fp.read()))
		fp.close()
		attachment = self.env['ir.attachment'].with_user(1).create({'name': 'phone_call_report.xlsx',
																	'datas': report,
																	'res_model': 'temp.creation',
																	'public': True})

		url = "/web/content/?model=ir.attachment&id=%s&filename_field=name&field=datas&download=true" \
			  % attachment.id

		return {
			'name': 'Báo cáo Phone Call',
			'type': 'ir.actions.act_url',
			'url': url,
			'target': 'self',
		}
