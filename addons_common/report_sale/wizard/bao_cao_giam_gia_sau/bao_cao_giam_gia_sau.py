from odoo import fields, api, models, _
from odoo.exceptions import ValidationError, UserError
from datetime import date, datetime, time
from calendar import monthrange
from openpyxl import load_workbook
from openpyxl.styles import Font, borders, Alignment
import base64
from io import BytesIO
from dateutil.relativedelta import relativedelta
from pytz import timezone, utc
from odoo.addons.report_sale.wizard.theme_report import ThemeReport
from odoo.modules.module import get_module_resource
import openpyxl

thin = borders.Side(style='thin')
double = borders.Side(style='double')
all_border_thin = borders.Border(thin, thin, thin, thin)


class DiscountReviewReport(models.TransientModel):
    _name = 'discount.review.report'
    _description = 'Báo cáo duyệt giảm giá sâu'

    start_date = fields.Date('Start date', default=date.today())
    end_date = fields.Date('End date', default=date.today())
    # convert date to datetime for search domain, should be removed if using datetime directly
    start_datetime = fields.Datetime('Start datetime', compute='_compute_datetime')
    end_datetime = fields.Datetime('End datetime', compute='_compute_datetime')

    company_id = fields.Many2one('res.company', string='Company', default=lambda self: self.env.company, domain=lambda self: [('id', 'in', self.env.companies.ids)])

    @api.depends('start_date', 'end_date')
    def _compute_datetime(self):
        self.start_datetime = False
        self.end_datetime = False
        if self.start_date and self.end_date:
            local_tz = timezone(self.env.user.tz or 'Etc/GMT+7')
            start_datetime = datetime(self.start_date.year, self.start_date.month, self.start_date.day, 0, 0, 0)
            end_datetime = datetime(self.end_date.year, self.end_date.month, self.end_date.day, 23, 59, 59)
            start_datetime = local_tz.localize(start_datetime, is_dst=None)
            end_datetime = local_tz.localize(end_datetime, is_dst=None)
            self.start_datetime = start_datetime.astimezone(utc).replace(tzinfo=None)
            self.end_datetime = end_datetime.astimezone(utc).replace(tzinfo=None)

    @api.onchange('start_date')
    def _onchange_start_date(self):
        if self.start_date:
            if self.start_date.month == fields.date.today().month:
                self.end_date = fields.date.today()
            else:
                self.end_date = date(self.start_date.year, self.start_date.month,
                                     monthrange(self.start_date.year, self.start_date.month)[1])

    @api.constrains('start_date', 'end_date')
    def check_dates(self):
        for record in self:
            start_date = fields.Date.from_string(record.start_date)
            end_date = fields.Date.from_string(record.end_date)
            if start_date > end_date:
                raise ValidationError(
                    _("Ngày kết thúc không thể ở trước ngày bắt đầu khi xuất báo cáo!!!"))

    def _get_data(self):
        ret_data = []
        # nếu có filter công ty
        if self.company_id:
            domain = [('write_date', '>=', self.start_datetime), ('write_date', '<=', self.end_datetime),
                      ('stage_id', '=', 'approve'), ('company_id', 'in', self.company_id.ids)]
        else:
            domain = [('write_date', '>=', self.start_datetime), ('write_date', '<=', self.end_datetime),
                      ('stage_id', '=', 'approve'), ('company_id', 'in', self.env.companies.ids)]

        discount_review_list = self.env['crm.discount.review'].search(domain)
        for line in discount_review_list:
            line = line.sudo()
            if line.crm_line_id:
                product = line.crm_line_id.service_id.name
                total_before_discount = line.crm_line_id.total_before_discount
            else:
                product = line.order_line_id.product_id.name
                total_before_discount = line.order_line_id.price_unit

            ret_data.append({
                'date': line.write_date.strftime('%d/%m/%Y'),
                'partner_id': line.partner_id.name,
                'crm_line_id': product,
                'create_uid': line.create_uid.name,
                'write_uid': line.write_uid.name,
                'name': line.name,
                'total_before_discount': total_before_discount,
                'total_amount_before_deep_discount': line.total_amount_before_deep_discount,
                'total_discount_cash': line.total_discount_cash,
                'total_amount_after_discount': line.total_amount_after_discount,
                'company_id': line.company_id.name,
            })

        return ret_data

    def create_report_discount_review(self):
        # get data
        datas = self._get_data()

        # in báo cáo
        report_discount_review_attachment = self.env['ir.attachment'].browse(
            self.env.ref('report_sale.report_discount_review_attachment').id)
        decode = base64.b64decode(report_discount_review_attachment.datas)
        wb = load_workbook(BytesIO(decode))
        ws = wb.active

        code_brand = self.company_id.brand_id.code.lower()
        image_path = get_module_resource('report_sale', 'static/img', 'icon_%s.png' % code_brand)

        img = openpyxl.drawing.image.Image(image_path)
        img.anchor = 'A1'
        ws.add_image(img)

        line_font = Font(name='Times New Roman', size=14)
        ws['A2'].value = 'BÁO CÁO DUYỆT GIẢM GIÁ SÂU'
        ws['A2'].font = Font(name='Times New Roman', size=20)

        ws['E3'].value = 'Ngày duyệt từ: '
        ws['E3'].font = line_font
        ws['F3'].value = self.start_date.strftime('%d/%m/%Y')
        ws['F3'].font = line_font
        ws['G3'].value = 'đến: '
        ws['G3'].font = line_font
        ws['H3'].value = self.end_date.strftime('%d/%m/%Y')
        ws['H3'].font = line_font

        # ws['E4'].value = 'Chi nhánh: '
        # ws['E4'].font = line_font
        # ws['F4'].value = self.company_id.name
        # ws['F4'].font = line_font
        key_col_list = list(range(2, 13))
        key_list = [
            'date',
            'partner_id',
            'crm_line_id',
            'create_uid',
            'write_uid',
            'name',
            'total_before_discount',
            'total_amount_before_deep_discount',
            'total_discount_cash',
            'total_amount_after_discount',
            'company_id',
        ]
        row = 7
        if code_brand == 'kn':
            header_fill = ThemeReport.kn_fill
        elif code_brand == 'da':
            header_fill = ThemeReport.da_fill
        elif code_brand == 'pr':
            header_fill = ThemeReport.pr_fill
        elif code_brand == 'hh':
            header_fill = ThemeReport.hh_fill
        else:
            header_fill = ThemeReport.sci_fill
        index_row = 1
        for line_data in datas:
            cell1 = ws.cell(row, 1)
            cell1.value,cell1.font,cell1.border,cell1.alignment = index_row,line_font,all_border_thin,Alignment(horizontal='center', vertical='center')
            index_row += 1

            for col, k in zip(key_col_list, key_list):
                beforeCell = ws.cell(6, col)
                beforeCell.fill = header_fill
                beforeCell.font = Font(name='Times New Roman', size=14, color='FFFFFF')
                cell = ws.cell(row, col)
                cell = ws.cell(row, col)
                cell.value = line_data[k]
                cell.font = line_font
                cell.border = all_border_thin
                cell.alignment = Alignment(horizontal='left', vertical='center')

                if 7 < col < 12:
                    cell.number_format = '_(* #,##0_);_(* (#,##0);_(* "-"??_);_(@_)'
            row += 1

        fp = BytesIO()
        wb.save(fp)
        fp.seek(0)
        report = base64.encodebytes((fp.read()))
        fp.close()
        attachment = self.env['ir.attachment'].sudo().create({
            'name': 'bao_cao_duyet_giam_gia_sau.xlsx',
            'datas': report,
            'res_model': 'temp.creation',
            'public': True,
        })
        url = "/web/content/?model=ir.attachment&id=%s&filename_field=name&field=datas&download=true" \
              % attachment.id
        return {'name': 'BÁO CÁO DUYỆT GIẢM GIÁ SÂU',
                'type': 'ir.actions.act_url',
                'url': url,
                'target': 'self',
                }
        # return {
        #     'name': 'Báo cáo duyệt giảm giá sâu',
        #     'type': 'ir.actions.act_window',
        #     'res_model': 'temp.wizard',
        #     'view_mode': 'form',
        #     'view_type': 'form',
        #     'target': 'inline',
        #     'view_id': self.env.ref('ms_templates.report_wizard').id,
        #     'context': {'attachment_id': attachment.id}
        # }
