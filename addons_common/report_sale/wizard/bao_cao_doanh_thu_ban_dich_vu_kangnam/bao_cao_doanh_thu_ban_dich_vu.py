import openpyxl

from odoo import fields, api, models, _
from odoo.exceptions import ValidationError, UserError
from datetime import date, datetime, time
from calendar import monthrange
from openpyxl import load_workbook
from openpyxl.styles import Font, borders, Alignment
import base64
from io import BytesIO
from dateutil.relativedelta import relativedelta
from pytz import timezone, utc

from odoo.modules.module import get_module_resource
from odoo.addons.report_sale.wizard.theme_report import ThemeReport
from datetime import datetime

thin = borders.Side(style='thin')
double = borders.Side(style='double')
all_border_thin = borders.Border(thin, thin, thin, thin)
SERVICE_HIS_TYPE = [('Spa', 'Spa'), ('Laser', 'Laser'), ('Odontology', 'Nha'), ('Surgery', 'Phẫu thuật'),
                    ('ChiPhi', 'Chi phí khác')]
dict_type = dict((key, value) for key, value in SERVICE_HIS_TYPE)
TEMPLATE = [
    'ma_nv_tu_van',
    'ten_nv_tu_van',
    'bo_phan_nv_tu_van',
    'chuc_danh_nv_tu_van',
    'vai_tro_nv_tu_van',
    'ten_khach_hang',
    'ma_booking',
    'ngay_dat_hang',
    'loai_phau_thuat',
    'loai_dich_vu',
    'nhom_dich_vu',
    'dich_vu',
    'don_vi_xu_ly',
    'buoi_thu',
    'so_luong',
    'nguon_booking',
    'nguon_mo_rong',
    'doanh_so',
    'diem_kpi',
    'ma_khach_hang',
    'gioi_tinh',
    'ngay_sinh',
    'dia_chi',
    'ma_dich_vu',
    'thu_tu_tu_van',
    'ma_nv_nguoi_tao',
    'nguoi_tao',
    'phong_ban_nguoi_tao'
]
SERVICE_HIS_TYPE = [('Spa', 'Spa'), ('Laser', 'Laser'), ('Odontology', 'Nha'), ('Surgery', 'Phẫu thuật'),
                    ('ChiPhi', 'Chi phí khác')]
dict_type = dict((key, value) for key, value in SERVICE_HIS_TYPE)

SURGERY_TYPE = [('minor', 'Tiểu phẫu'), ('major', 'Đại phẫu')]
DICT_SURGERY_TYPE = dict((key, value) for key, value in SURGERY_TYPE)

CONSULTING_ROLE = [('1', 'Tư vấn độc lập'), ('2', 'Tư vấn đồng thời'), ('3', 'Lễ tân - CVTV cùng tư vấn'),
                   ('4', 'BS da liễu - KTV cùng tư vấn'), ('5', 'Tư vấn chính'), ('6', 'Tư vấn phụ')]
DICT_CONSULTING_ROLE = dict((key, value) for key, value in CONSULTING_ROLE)
GENDER = [('male', 'Nam'), ('female', 'Nữ'), ('transguy', 'Người chuyển giới nam'), ('transgirl', 'người chuyển giới nữ'), ('other', 'Khác')]
DICT_GENDER = dict((key, value) for key, value in GENDER)


class KangnamServiceSaleReport(models.TransientModel):
    _name = 'kangnam.kpi.service.sale.report'
    _description = 'Bao cao doanh thu ban dich vu kangnam'

    start_date = fields.Date('Start date', default=date.today())
    end_date = fields.Date('End date', default=date.today())
    company_id = fields.Many2one('res.company', domain="[('brand_id.name', '=', 'Kangnam')]", string='Company',
                                 default=lambda self: self.env.company)
    # domain="[('brand_id.code', '=', 'KN')]",
    # convert date to datetime for search domain, should be removed if using datetime directly
    start_datetime = fields.Datetime('Start datetime', compute='_compute_datetime')
    end_datetime = fields.Datetime('End datetime', compute='_compute_datetime')
    expected_date = fields.Selection([('conf_walkin', 'Hoàn thành phiếu khám'), ('conf_surgery', 'Hoàn thành phiếu chuyên khoa/phẫu thuật')],
                                     string='Điều kiện xuất', default='conf_walkin')

    @api.depends('start_date', 'end_date')
    def _compute_datetime(self):
        self.start_datetime = False
        self.end_datetime = False
        if self.start_date and self.end_date:
            local_tz = timezone(self.env.user.tz or 'Etc/GMT+7')
            start_datetime = datetime(self.start_date.year, self.start_date.month, self.start_date.day, 0, 0, 0)
            end_datetime = datetime(self.end_date.year, self.end_date.month, self.end_date.day, 23, 59, 59)
            start_datetime = local_tz.localize(start_datetime, is_dst=None)
            end_datetime = local_tz.localize(end_datetime, is_dst=None)
            self.start_datetime = start_datetime.astimezone(utc).replace(tzinfo=None)
            self.end_datetime = end_datetime.astimezone(utc).replace(tzinfo=None)

    @api.onchange('start_date')
    def _onchange_start_date(self):
        if self.start_date:
            if self.start_date.month == fields.date.today().month:
                self.end_date = fields.date.today()
            else:
                self.end_date = date(self.start_date.year, self.start_date.month,
                                     monthrange(self.start_date.year, self.start_date.month)[1])

    def walkin_num(self, walkin, service):
        list_done_walkin = walkin.booking_id.walkin_ids.filtered(
            lambda w: service.id in w.service.ids and w.state == 'Completed').sorted('service_date').mapped(
            'service_date')
        index = [i for i, x in enumerate(list_done_walkin) if x == walkin.service_date]

        if len(list_done_walkin) > 0 and len(index) > 0:
            walkin_num = index[0] + 1
            # print("Đã có phiếu khám hoàn thành")
        else:
            # print("Chưa có phiếu khám hoàn thành")
            walkin_num = len(list_done_walkin) + 1
        return walkin_num

    @api.constrains('start_date', 'end_date')
    def check_dates(self):
        for record in self:
            start_date = fields.Date.from_string(record.start_date)
            end_date = fields.Date.from_string(record.end_date)
            if start_date > end_date:
                raise ValidationError(
                    _("Ngày kết thúc không thể ở trước ngày bắt đầu khi xuất báo cáo!!!"))

    def _get_line_data_form_walkin(self, medical):
        walkin = medical.walkin
        so = walkin.sale_order_id
        booking = so.booking_id
        health_service = self.env['sh.medical.health.center.service']

        if medical._name == 'sh.medical.surgery':
            date_order = fields.Datetime.context_timestamp(self.with_context(tz=self.env.user.tz), medical.surgery_end_date).strftime('%d/%m/%Y')
        elif medical._name == 'sh.medical.specialty':
            date_order = fields.Datetime.context_timestamp(self.with_context(tz=self.env.user.tz), medical.services_end_date).strftime('%d/%m/%Y')
        else:
            date_order = None
        result = list()
        for sol in so.order_line:
            crm_line = sol.crm_line_id
            service = health_service.search([('product_id', '=', sol.product_id.id)])
            walkin_num = self.walkin_num(walkin, service)
            consultants = list()
            if crm_line.consultants_1:
                consultants.append((crm_line.consultants_1, crm_line.consulting_role_1))
            if crm_line.consultants_2:
                consultants.append((crm_line.consultants_2, crm_line.consulting_role_2))
            for index in range(1, len(consultants) + 1):
                work_email = consultants[index - 1][0].login
                # user = consultants[index - 1][0]
                employee = self.env['hr.employee'].sudo().search([('work_email', '=', str(work_email))])

                if service.his_service_type in ['Spa', 'Laser', 'Odontology', 'Surgery']:
                    result.append({
                        'ma_nv_tu_van': employee.employee_code if employee else None,
                        'ten_nv_tu_van': employee.name if employee else None,
                        'bo_phan_nv_tu_van': employee.group_job.name if employee and employee.group_job else None,
                        'chuc_danh_nv_tu_van': employee.job_id.name if employee and employee.job_id else None,
                        'vai_tro_nv_tu_van': DICT_CONSULTING_ROLE.get(consultants[index - 1][1]) if len(consultants) != 0 and consultants[index - 1][1] != 'FALSE' else None,
                        'ten_khach_hang': so.partner_id.name,
                        'ma_booking': booking.name,
                        'ngay_dat_hang': date_order,
                        'loai_phau_thuat': DICT_SURGERY_TYPE.get(service.surgery_type) if service.surgery_type else None,
                        'loai_dich_vu': dict_type.get(service.his_service_type),
                        'nhom_dich_vu': service.service_category.name,
                        'dich_vu': service.name,
                        'don_vi_xu_ly': sol.uom_price,
                        'buoi_thu': walkin_num,
                        'so_luong': sol.product_uom_qty,
                        'nguon_booking': booking.source_id.name or None,
                        'nguon_mo_rong': crm_line.source_extend_id.name or None,
                        'doanh_so': sol.price_subtotal,
                        'diem_kpi': service.kpi_point,
                        'gioi_tinh': DICT_GENDER.get(so.partner_id.gender),
                        'ngay_sinh': so.partner_id.birth_date or None,
                        'dia_chi': so.partner_id.street or None,
                        'ma_dich_vu': service.default_code,
                        'ma_khach_hang': so.code_customer,
                        'thu_tu_tu_van': 'Tư vấn viên %s' % str(index) if len(consultants) != 0 else None,
                        'ma_nv_nguoi_tao': booking.create_by.employee_ids[0].employee_code if booking.create_by.employee_ids else None,
                        'nguoi_tao': booking.create_by.name,
                        'phong_ban_nguoi_tao': booking.create_by.employee_ids[0].department_id.name if booking.create_by.employee_ids else None
                    })
        return result

    def _get_data_kpi_service_sale_report(self):
        ret_data = []
        domain_surgery = [('state', '=', 'Done'), ('institution.his_company', '=', self.company_id.id),
                          ('surgery_end_date', '>=', self.start_datetime), ('surgery_end_date', '<=', self.end_datetime)]

        domain_specialty = [('state', '=', 'Done'), ('institution.his_company', '=', self.company_id.id),
                            ('services_end_date', '>=', self.start_datetime), ('services_end_date', '<=', self.end_datetime)]

        list_medical = list()
        if self.expected_date == 'conf_walkin':
            medical_surgery = self.sudo().env['sh.medical.surgery'].with_context(company_ids=[c.id for c in self.env['res.company']]).search(domain_surgery).filtered(lambda x: x.walkin.state == 'Completed')
            medical_specialty = self.sudo().env['sh.medical.specialty'].with_context(company_ids=[c.id for c in self.env['res.company']]).search(domain_specialty).filtered(lambda x: x.walkin.state == 'Completed')
            for surgery in medical_surgery:
                list_medical.append(surgery)
            for specialty in medical_specialty:
                list_medical.append(specialty)
        elif self.expected_date == 'conf_surgery':
            medical_surgery = self.sudo().env['sh.medical.surgery'].with_context(company_ids=[c.id for c in self.env['res.company']]).search(domain_surgery)
            medical_specialty = self.sudo().env['sh.medical.specialty'].with_context(company_ids=[c.id for c in self.env['res.company']]).search(domain_specialty)
            for surgery in medical_surgery:
                list_medical.append(surgery)
            for specialty in medical_specialty:
                list_medical.append(specialty)

        def sorted_model(element):
            return element.walkin.sale_order_id.date_order

        list_medical = sorted(list_medical, key=lambda element: sorted_model(element))
        for medical in list_medical:
            lines = self._get_line_data_form_walkin(medical)
            for line in lines:
                ret_data.append(line)

        total_revenue = sum([e.get('doanh_so') for e in ret_data])
        to_dict = dict((key, None) for key in TEMPLATE)
        to_dict['doanh_so'] = total_revenue
        ret_data.append(to_dict)
        return ret_data

    def create_kpi_service_sale_report(self):
        # get data
        datas = self._get_data_kpi_service_sale_report()
        # in dữ liệu
        report_brand_overview_attachment = self.env['ir.attachment'].browse(
            self.env.ref('report_sale.bao_cao_doanh_thu_ban_dich_vu_kangnam_attachment').id)
        decode = base64.b64decode(report_brand_overview_attachment.datas)
        wb = load_workbook(BytesIO(decode))
        ws = wb.active
        line_font = Font(name='Times New Roman', size=14)

        code_brand = self.company_id.brand_id.code.lower()
        image_path = get_module_resource('report_sale', 'static/img', 'icon_%s.png' % code_brand)

        img = openpyxl.drawing.image.Image(image_path)
        img.anchor = 'A1'
        ws.add_image(img)
        format_currency = '_(* #,##0_);_(* (#,##0);_(* "-"??_);_(@_)'

        ws['F3'].value += self.start_date.strftime('%d/%m/%Y')
        ws['G3'].value += self.end_date.strftime('%d/%m/%Y')
        ws['F4'].value += self.company_id.name

        key_col = list(range(1, len(TEMPLATE) + 1))

        row = 7
        if code_brand == 'kn':
            header_fill = ThemeReport.kn_fill
        elif code_brand == 'da':
            header_fill = ThemeReport.da_fill
        elif code_brand == 'pr':
            header_fill = ThemeReport.pr_fill
        elif code_brand == 'hh':
            header_fill = ThemeReport.hh_fill
        else:
            header_fill = ThemeReport.sci_fill
        for data in datas:
            for col, k in zip(key_col, TEMPLATE):
                beforeCell = ws.cell(6, col)
                beforeCell.fill = header_fill
                beforeCell.font = Font(name='Times New Roman', size=14, color='FFFFFF')
                cell = ws.cell(row, col)
                cell.value = data[k]
                cell.font = line_font
                cell.border = all_border_thin
                cell.alignment = Alignment(horizontal='left', vertical='center')
                if col == 18:
                    cell.number_format = format_currency
            row += 1
        fp = BytesIO()
        wb.save(fp)
        fp.seek(0)
        report = base64.encodebytes((fp.read()))
        fp.close()
        attachment = self.env['ir.attachment'].sudo().create({
            'name': 'bao_cao_doanh_thu_ban_dich_vu_kangnam_%s.xlsx' % self.company_id.name,
            'datas': report,
            'res_model': 'temp.creation',
            'public': True,
        })
        url = "/web/content/?model=ir.attachment&id=%s&filename_field=name&field=datas&download=true" \
              % attachment.id
        return {'name': 'BÁO CÁO DOANH SỐ BÁN DỊCH VỤ KANGNAM',
                'type': 'ir.actions.act_url',
                'url': url,
                'target': 'self',
                }
