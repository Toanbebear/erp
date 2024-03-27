from odoo import fields, api, models, _
from odoo.exceptions import ValidationError, UserError
from datetime import date, datetime, time
from calendar import monthrange
import openpyxl
from openpyxl.drawing.image import Image
from openpyxl import load_workbook
from openpyxl.styles import Font, borders, Alignment, PatternFill
import base64
from io import BytesIO
from dateutil.relativedelta import relativedelta
from pytz import timezone, utc
from odoo.modules.module import get_module_resource

thin = borders.Side(style='thin')
double = borders.Side(style='double')
all_border_thin = borders.Border(thin, thin, thin, thin)
da_fill = PatternFill(start_color='0e7661', end_color='0e7661', fill_type='solid')
kn_fill = PatternFill(start_color='003471', end_color='003471', fill_type='solid')
pr_fill = PatternFill(start_color='012C5F', end_color='012C5F', fill_type='solid')
hh_fill = PatternFill(start_color='053D7C', end_color='053D7C', fill_type='solid')
sci_fill = PatternFill(start_color='003471', end_color='003471', fill_type='solid')

class SaleMarketingWalkin(models.TransientModel):
    _name = 'sale.marketing.walkin'
    _description = 'Báo cáo dịch vụ thành công'

    start_date = fields.Date('Start date', default=date.today())
    end_date = fields.Date('End date', default=date.today())
    # brand_id = fields.Many2one(string='Thương hiệu', comodel_name='res.brand', domain=lambda self: [('id', 'in', self.env.user.company_ids.mapped('brand_id').ids)])
    company_ids = fields.Many2many(string='Chi nhánh', comodel_name='res.company', domain=lambda self: [('id', 'in', self.env.user.company_ids.ids)])

    # convert date to datetime for search domain, should be removed if using datetime directly
    start_datetime = fields.Datetime('Start datetime', compute='_compute_datetime')
    end_datetime = fields.Datetime('End datetime', compute='_compute_datetime')

    @api.depends('start_date', 'end_date')
    def _compute_datetime(self):
        self.start_datetime = False
        self.end_datetime = False
        if self.start_date and self.end_date:
            local_tz = timezone(self.env.user.tz or 'Etc/GMT+7')
            start_datetime = datetime(self.start_date.year, self.start_date.month, self.start_date.day, 0, 0, 0)
            end_datetime = datetime(self.end_date.year, self.end_date.month, self.end_date.day, 23, 59, 59)
            start_datetime = local_tz.localize(start_datetime, is_dst=None)
            end_datetime = local_tz.localize(end_datetime, is_dst=None)
            self.start_datetime = start_datetime.astimezone(utc).replace(tzinfo=None)
            self.end_datetime = end_datetime.astimezone(utc).replace(tzinfo=None)

    @api.onchange('start_date')
    def _onchange_start_date(self):
        if self.start_date:
            if self.start_date.month == fields.date.today().month:
                self.end_date = fields.date.today()
            else:
                self.end_date = date(self.start_date.year, self.start_date.month,
                                     monthrange(self.start_date.year, self.start_date.month)[1])

    @api.constrains('start_date', 'end_date')
    def check_dates(self):
        for record in self:
            start_date = fields.Date.from_string(record.start_date)
            end_date = fields.Date.from_string(record.end_date)
            days = (end_date - start_date).days
            if days < 0 or days > 365:
                raise ValidationError(
                    _("Ngày kết thúc không thể ở trước ngày bắt đầu khi xuất báo cáo!!!"))

    def _get_data_report(self):
        domain = [('state', '=', ('Completed')), ('institution.his_company', 'in', self.company_ids.ids),
                  ('service_date', '>=', self.start_datetime), ('service_date', '<=', self.end_datetime)]
        walkins = self.env['sh.medical.appointment.register.walkin'].sudo().search(domain, order='id')
        return_val = []
        for walkin in walkins:
            walkin = walkin.sudo()
            for service in walkin.service:
                amount_line = 0
                # TÍNH TOÁN BUỔI LÀM DỊCH VỤ LÀ BUỔI THỨ MẤY
                list_done_walkin = walkin.booking_id.walkin_ids.filtered(
                    lambda w: service.id in w.service.ids and w.state == 'Completed').sorted('service_date').mapped('service_date')
                index = [i for i, x in enumerate(list_done_walkin) if x == walkin.service_date]
                if len(list_done_walkin) > 0 and len(index) > 0:
                    walkin_num = index[0] + 1
                    # print("Đã có phiếu khám hoàn thành")
                else:
                    # print("Chưa có phiếu khám hoàn thành")
                    walkin_num = len(list_done_walkin) + 1

                amount_line = sum(walkin.sale_order_id.order_line.filtered(
                    lambda sl: sl.product_id == service.product_id).mapped('price_subtotal'))
                val = {
                    'arrival_date': walkin.booking_id.arrival_date.strftime('%d/%m/%Y') if walkin.booking_id.arrival_date else '-',
                    'service_date': walkin.service_date.strftime('%d/%m/%Y'),
                    'name': walkin.name,
                    'SO_name': walkin.sale_order_id.name,
                    'walkin_type': walkin.type_crm_id.name,
                    'services': service.display_name,
                    'walkin_num': walkin_num,
                    'institution': walkin.institution.name,
                    'code_booking': walkin.booking_id.name,
                    'amount_total': walkin.booking_id.amount_total,
                    'source_booking': walkin.booking_id.source_id.name,
                    'patient': walkin.patient.display_name,
                    # 'phone': walkin.patient.phone,
                    'phone': None,
                    'price_list': walkin.booking_id.price_list_id.name,
                    'amount_line': amount_line,
                    'country': walkin.booking_id.country_id.name or '',
                    'state': walkin.booking_id.state_id.name or '',
                    'district': walkin.booking_id.district_id.name or '',
                    'email': walkin.booking_id.email_from or '',
                    'street': walkin.booking_id.street or ''
                }
                return_val.append(val)
        return return_val

    def create_report(self):
        datas = self._get_data_report()
        # in dữ liễu
        report_brand_overview_attachment = self.env['ir.attachment'].browse(
            self.env.ref('report_sale.bao_cao_sale_walkin_attachment').id)
        decode = base64.b64decode(report_brand_overview_attachment.datas)
        wb = load_workbook(BytesIO(decode))
        ws = wb.active
        line_font = Font(name='Times New Roman', size=14)

        code_brand = self.company_ids[0].brand_id.code.lower()
        if len(self.company_ids) > 1:
            image_path = get_module_resource('report_sale', 'static/img', 'icon_sci.png')
        else:
            image_path = get_module_resource('report_sale', 'static/img', 'icon_%s.png' % code_brand)

        img = openpyxl.drawing.image.Image(image_path)
        img.anchor = 'A1'
        ws.add_image(img)

        ws['B2'].value = 'Từ ngày: %s đến ngày %s ' % (self.start_date.strftime('%d/%m/%Y'),self.end_date.strftime('%d/%m/%Y'))
        key_col = list(range(1, 21))

        format_currency = '_(* #,##0_);_(* (#,##0);_(* "-"??_);_(@_)'
        key_list = [
            'arrival_date',
            'service_date',
            'name',
            'SO_name',
            'walkin_type',
            'services',
            'walkin_num',
            'institution',
            'code_booking',
            'amount_total',
            'source_booking',
            'patient',
            'phone',
            'price_list',
            'amount_line',
            'country',
            'state',
            'district',
            'email',
            'street'
        ]
        row = 5
        if code_brand == 'kn':
            header_fill = kn_fill
        elif code_brand == 'da':
            header_fill = da_fill
        elif code_brand == 'pr':
            header_fill = pr_fill
        elif code_brand == 'hh':
            header_fill = hh_fill
        else:
            header_fill = sci_fill
        for data in datas:
            for col, k in zip(key_col, key_list):
                beforeCell = ws.cell(4, col)
                beforeCell.fill = header_fill
                beforeCell.font = Font(name='Times New Roman', size=14, color='FFFFFF')
                cell = ws.cell(row, col)
                cell.value = data[k]
                cell.font = line_font
                cell.border = all_border_thin
                cell.alignment = Alignment(horizontal='left', vertical='center')

                if col == 10 or col == 15:
                    cell.number_format = format_currency

            row += 1
        fp = BytesIO()
        wb.save(fp)
        fp.seek(0)
        report = base64.encodebytes((fp.read()))
        fp.close()
        attachment = self.env['ir.attachment'].sudo().create({
            'name': 'bao_cao_sale_dich_vu_thanh_cong.xlsx',
            'datas': report,
            'res_model': 'temp.creation',
            'public': True,
        })
        url = "/web/content/?model=ir.attachment&id=%s&filename_field=name&field=datas&download=true" \
              % attachment.id
        return {'name': 'BÁO CÁO SALE DỊCH VỤ THÀNH CÔNG',
                'type': 'ir.actions.act_url',
                'url': url,
                'target': 'self',
                }
