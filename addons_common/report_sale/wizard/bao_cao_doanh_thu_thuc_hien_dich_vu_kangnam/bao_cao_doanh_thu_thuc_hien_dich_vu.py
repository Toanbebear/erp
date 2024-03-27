from odoo import fields, api, models, _
from odoo.exceptions import ValidationError, UserError
from datetime import date, datetime, time
from calendar import monthrange
from openpyxl import load_workbook
from openpyxl.styles import Font, borders, Alignment
import base64
from io import BytesIO
from dateutil.relativedelta import relativedelta
from pytz import timezone, utc
from odoo.addons.report_sale.wizard.theme_report import ThemeReport
from odoo.modules.module import get_module_resource
import openpyxl

thin = borders.Side(style='thin')
double = borders.Side(style='double')
all_border_thin = borders.Border(thin, thin, thin, thin)

WARD_TYPE = [
    ('All', 'Toàn bộ dịch vụ'),
    ('Surgery', 'Phẫu thuật'),
    ('Spa', 'Spa'),
    ('Laser', 'Laser'),
    ('Odontology', 'Nha khoa')
]

SEARCH_BY = [
    ('01', 'Ngày bắt đầu của phiếu chuyên khoa'),
    ('02', 'Ngày làm dịch vụ của phiếu khám'),
]

SERVICE_HIS_TYPE = [('Spa', 'Spa'), ('Laser', 'Laser'), ('Odontology', 'Nha'), ('Surgery', 'Phẫu thuật'), ('ChiPhi', 'Chi phí khác')]
dict_type = dict((key, value) for key, value in SERVICE_HIS_TYPE)

SURGERY_TYPE = [('minor', 'Tiểu phẫu'), ('major', 'Đại phẫu')]
DICT_SURGERY_TYPE = dict((key, value) for key, value in SURGERY_TYPE)
TEMP = [
    'employee_code',  # Mã nhân viên
    'member',  # Thành viên
    'group_job',  # Bộ phận
    'job_id',  # Chức danh
    'role',  # Vai trò
    'name_customer',  # Tên khách hàng
    'code_booking',  # Mã booking
    'so_code',  # Số sale order
    'start_datetime',  # Ngày giờ bắt đầu
    'surgery_type',  # Loại phẫu thuật
    'his_service_type',  # Nhóm dịch vụ cấp 1
    'service_category',  # Nhóm dịch vụ
    'service',  # Dịch vụ
    'uom_price',  # Đơn vị xử lý
    'qty',  # Số lượng
    'number_used',  # Đã sử dụng
    'money',  # Doanh thu thực
    'code_walkin',  # Mã phiếu khám
    'service_date',  # Ngày phiếu khám
    'code_specialty_sheet',  # Mã phiếu chuyên khoa
    'department',  # Khoa phòng
    'end_datetime',  # Ngày giờ kết thúc
    'point_service',  # Điểm KPI
    'type_booking',  # Phân loại booking
    'institution'  # Đơn vị
]


class ServicePerformanceReportKangnam(models.TransientModel):
    _name = 'kpi.service.perform.report.kangnam'
    _description = 'Báo cáo doanh thu thực hiện dịch vụ Kangnam'

    start_date = fields.Date('Start date', default=date.today())
    end_date = fields.Date('End date', default=date.today())
    # company_id = fields.Many2one('res.company', string='Chi nhánh', domain="[('brand_id.name', '=', 'Kangnam')]", default=lambda self: self.env.company)
    company_id = fields.Many2one('res.company', string='Chi nhánh', default=lambda self: self.env.company)
    type = fields.Selection(WARD_TYPE, string='Loại phiếu', default='All')

    # convert date to datetime for search domain, should be removed if using datetime directly
    start_datetime = fields.Datetime('Start datetime', compute='_compute_datetime')
    end_datetime = fields.Datetime('End datetime', compute='_compute_datetime')
    # Kiểu lấy báo cáo theo ngày làm dịch vụ hoặc theo ngày giờ bắt đầu của phiếu chuyên khoa
    # search_by = fields.Selection(SEARCH_BY, string='Kiểu tìm kiếm dữ liệu', default='01')
    expected_date = fields.Selection([('conf_walkin', 'Hoàn thành phiếu khám'), ('conf_surgery', 'Hoàn thành phiếu chuyên khoa/phẫu thuật')],
                                     string='Điều kiện xuất', default='conf_walkin')
    is_query = fields.Boolean('Sử dụng truy vấn', default=True, help='Sử dụng truy vấn tăng tốc độ tạo báo cáo')

    @api.depends('start_date', 'end_date')
    def _compute_datetime(self):
        self.start_datetime = False
        self.end_datetime = False
        if self.start_date and self.end_date:
            local_tz = timezone(self.env.user.tz or 'Etc/GMT+7')
            start_datetime = datetime(self.start_date.year, self.start_date.month, self.start_date.day, 0, 0, 0)
            end_datetime = datetime(self.end_date.year, self.end_date.month, self.end_date.day, 23, 59, 59)
            start_datetime = local_tz.localize(start_datetime, is_dst=None)
            end_datetime = local_tz.localize(end_datetime, is_dst=None)
            self.start_datetime = start_datetime.astimezone(utc).replace(tzinfo=None)
            self.end_datetime = end_datetime.astimezone(utc).replace(tzinfo=None)

    @api.onchange('start_date')
    def _onchange_start_date(self):
        if self.start_date:
            if self.start_date.month == fields.date.today().month:
                self.end_date = fields.date.today()
            else:
                self.end_date = date(self.start_date.year, self.start_date.month,
                                     monthrange(self.start_date.year, self.start_date.month)[1])

    @api.constrains('start_date', 'end_date')
    def check_dates(self):
        for record in self:
            start_date = fields.Date.from_string(record.start_date)
            end_date = fields.Date.from_string(record.end_date)
            if start_date > end_date:
                raise ValidationError(
                    _("Ngày kết thúc không thể ở trước ngày bắt đầu khi xuất báo cáo!!!"))

    def _get_data_report_v2(self):
        query_state = ""
        if self.expected_date == 'conf_walkin':
            query_state = " and pk.state = 'Completed' "

        if self.type == 'All':
            query_type = " and pb.type in ('Spa', 'Laser', 'Odontology')"
        else:
            query_type = " and pb.type = '%s' " % self.type
        # Odontology = dentistry  Nha khoa

        ret_data = []
        sale_order_ids = []
        total_revenue = 0

        if self.type in ['All', 'Surgery']:
            # Lấy phiếu phẫu thuật
            query = """
                    select so.id,
                        e.employee_code as employee_code,
                        e.name as member, 
                        gj.name as group_job, 
                        j.name as job_id,
                        r.name as role,
                        kh.name as name_customer,
                        bk.name as code_booking,
                        so.name as so_code,
                        pt.surgery_date as start_datetime,
                        dv.surgery_type as surgery_type,
                        dv.his_service_type as his_service_type,
                        pc.name as service_category,
                        sp.name as service,
                        pk.name as code_walkin,
                        pk.service_date as service_date,
                        pt.name as code_specialty_sheet,
                        pb.name as department,
                        pt.surgery_end_date as end_datetime,
                        dv.kpi_point as point_service,
                        ct.name as type_booking,
                        rc.name as institution,
                        pp.id
                    from sh_surgery_team_services_rel ds
                    left join sh_medical_surgery_team dpt on dpt.id = ds.surgery_team_id 
                    left join sh_medical_surgery pt on pt.id = dpt.name
                    left join sh_medical_appointment_register_walkin pk on pk.id = pt.walkin
                    left join sh_medical_health_center cs on cs.id = pt.institution 
                    left join res_company rc on rc.id = cs.his_company  
                    left join sh_medical_health_center_service dv on dv.id = ds.service_id  
                    left join product_product pp on pp.id = dv.product_id  
                    left join product_template sp on sp.id = pp.product_tmpl_id  
                    left join sh_medical_physician p on p.id = dpt.team_member  
                    left join hr_employee e on e.id = p.employee_id  
                    left join hr_job j on j.id = e.job_id  
                    left join hr_group_job gj on gj.id = j.group_job  
                    left join hr_job_position hjp on hjp.id = j.position_id  
                    left join sh_medical_team_role r on r.id = dpt.role 
                    left join crm_lead bk on bk.id = pk.booking_id  
                    left join crm_type ct on ct.id = bk.type_crm_id  
                    left join res_partner kh on kh.id = bk.partner_id
                    left join sale_order so on so.id = pk.sale_order_id 
                    left join sh_medical_health_center_service_category sc on sc.id = dv.service_category 
                    left join product_category pc on pc.id = sc.product_cat_id  
                    left join sh_medical_health_center_ward pb on pb.id = pt.department 
                    where pt.state = 'Done' 
                        and pt.surgery_end_date >= %s 
                        and pt.surgery_end_date <= %s """ + query_state + """
                        and pk.type_crm_id <> 3
                        and cs.his_company = %s"""
            self.env.cr.execute(query, (self.start_datetime,
                                        self.end_datetime,
                                        self.company_id.id))
            datas = self.env.cr.fetchall()

            # Phẫu thuật: Sale order
            if len(datas) > 0:
                for data in datas:
                    sale_order_ids.append(data[0])

        if self.type or self.type == ['All']:
            # Lấy dữ liệu phiếu chuyên khoa
            specialty_query = """
                        select so.id,
                            e.employee_code as employee_code,
                            e.name as member, 
                            gj.name as group_job, 
                            j.name as job_id,
                            r.name as role,
                            kh.name as name_customer,
                            bk.name as code_booking,
                            so.name as so_code,
                            pt.services_date as start_datetime,
                            dv.surgery_type as surgery_type,
                            dv.his_service_type as his_service_type,
                            pc.name as service_category,
                            sp.name as service,
                            pk.name as code_walkin,
                            pk.service_date as service_date,
                            pt.name as code_specialty_sheet,
                            pb.name as department,
                            pt.services_end_date as end_datetime,
                            dv.kpi_point as point_service,
                            ct.name as type_booking,
                            rc.name as institution,
                            pp.id
                        from sh_specialty_team_services_rel ds
                        left join sh_medical_specialty_team dpt on dpt.id = ds.specialty_team_id
                        left join sh_medical_specialty pt on pt.id = dpt.name
                        left join sh_medical_appointment_register_walkin pk on pk.id = pt.walkin
                        left join sh_medical_health_center cs on cs.id = pt.institution 
                        left join res_company rc on rc.id = cs.his_company  
                        left join sh_medical_health_center_service dv on dv.id = ds.service_id  
                        left join product_product pp on pp.id = dv.product_id  
                        left join product_template sp on sp.id = pp.product_tmpl_id  
                        left join sh_medical_physician p on p.id = dpt.team_member  
                        left join hr_employee e on e.id = p.employee_id  
                        left join hr_job j on j.id = e.job_id  
                        left join hr_group_job gj on gj.id = j.group_job  
                        left join hr_job_position hjp on hjp.id = j.position_id  
                        left join sh_medical_team_role r on r.id = dpt.role 
                        left join crm_lead bk on bk.id = pk.booking_id  
                        left join crm_type ct on ct.id = bk.type_crm_id  
                        left join res_partner kh on kh.id = bk.partner_id
                        left join sale_order so on so.id = pk.sale_order_id 
                        left join sh_medical_health_center_service_category sc on sc.id = dv.service_category 
                        left join product_category pc on pc.id = sc.product_cat_id  
                        left join sh_medical_health_center_ward pb on pb.id = pt.department
                        where pt.state = 'Done' 
                            and pt.services_end_date >= %s 
                            and pt.services_end_date <= %s """ + query_state + """
                            and pk.type_crm_id <> 3
                            and cs.his_company = %s
                            """ + query_type
            self.env.cr.execute(specialty_query, (self.start_datetime,
                                                  self.end_datetime,
                                                  self.company_id.id))

            specialty_datas = self.env.cr.fetchall()
            # Chuyên khoa: Sale order
            if len(specialty_datas) > 0:
                for data in specialty_datas:
                    sale_order_ids.append(data[0])

        if sale_order_ids:
            # Lấy dữ liệu Sale order
            query_sol = """
                    select so.id, 
                    sol.product_id, 
                    sol.price_subtotal, 
                    cl.uom_price, 
                    cl.quantity as qty,
                    cl.number_used 
                    from sale_order_line sol
                    left join sale_order so on so.id = sol.order_id 
                    left join crm_line cl on cl.id = sol.crm_line_id 
                    where so.id in %s"""
            self.env.cr.execute(query_sol, [tuple(sale_order_ids)])
            sales = self.env.cr.fetchall()
            sale_order_lines = {}
            for line in sales:
                if line[0] in sale_order_lines:
                    sale_order_lines[line[0]].append((line[1],
                        line[2],
                        line[3],
                        line[4],
                        line[5]))
                else:
                    sale_order_lines[line[0]] = [
                        (line[1],
                        line[2],
                        line[3],
                        line[4],
                        line[5])
                    ]

        if self.type in ['All', 'Surgery']:
            # Xử lý phiếu phẫu thuật
            for row in datas:
                uom_price = 0
                qty = 0
                number_used = 0
                money = 0
                index = 0
                if row[0] in sale_order_lines:
                    lines = sale_order_lines[row[0]]
                    for line in lines:
                        if row[22] == line[0]:
                            money += line[1]
                            price = line[2] if line[2] else 0
                            uom_price += price
                            q = line[3] if line[3] else 0
                            qty += q
                            used = line[4] if line[4] else 0
                            number_used += used
                ret_data.append({
                    'employee_code': row[1],
                    'member': row[2],
                    'group_job': row[3] or None,
                    'job_id': row[4] or None,
                    'role': row[5],
                    'name_customer': row[6],
                    'code_booking': row[7],
                    'so_code': row[8],
                    'start_datetime': '%s' % fields.Datetime.context_timestamp(
                                            self.with_context(tz=self.env.user.tz),
                                            row[9]).strftime('%d/%m/%Y'),
                    'surgery_type': DICT_SURGERY_TYPE.get(row[10]) if row[10] else None,
                    'his_service_type': dict_type.get(row[11]),
                    'service_category': row[12],
                    'service': row[13],
                    'uom_price': uom_price if uom_price else None,
                    # 'qty': sol.crm_line_id.quantity if len(sol) == 1 else None,
                    'qty': qty if qty else None,
                    # 'number_used': sol.crm_line_id.number_used if len(sol) == 1 else None,
                    'number_used': number_used if number_used else None,
                    'money': money if money else None,
                    'code_walkin': row[14],
                    'service_date': '%s' % fields.Datetime.context_timestamp(
                                            self.with_context(tz=self.env.user.tz),
                                            row[15]).strftime('%d/%m/%Y'),
                    'code_specialty_sheet': row[16],
                    'department': row[17],
                    'end_datetime': '%s' % fields.Datetime.context_timestamp(
                                            self.with_context(tz=self.env.user.tz),
                                            row[18]).strftime('%d/%m/%Y'),
                    'point_service': row[19],
                    'type_booking': row[20],
                    'institution': row[21],
                })
                total_revenue += money

        if self.type or self.type == ['All']:
            # Xử lý phiếu chuyên khoa
            for row in specialty_datas:
                uom_price = 0
                qty = 0
                number_used = 0
                money = 0
                if row[0] in sale_order_lines:
                    lines = sale_order_lines[row[0]]
                    for line in lines:
                        if row[22] == line[0]:
                            money += line[1]
                            price = line[2] if line[2] else 0
                            uom_price += price
                            q = line[3] if line[3] else 0
                            qty += q
                            used = line[4] if line[4] else 0
                            number_used += used
                ret_data.append({
                    'employee_code': row[1],
                    'member': row[2],
                    'group_job': row[3] or None,
                    'job_id': row[4] or None,
                    'role': row[5],
                    'name_customer': row[6],
                    'code_booking': row[7],
                    'so_code': row[8],
                    'start_datetime': '%s' % fields.Datetime.context_timestamp(
                        self.with_context(tz=self.env.user.tz),
                        row[9]).strftime('%d/%m/%Y'),
                    'surgery_type': DICT_SURGERY_TYPE.get(row[10]) if row[10] else None,
                    'his_service_type': dict_type.get(row[11]),
                    'service_category': row[12],
                    'service': row[13],
                    'uom_price': uom_price if uom_price else None,
                    # 'qty': sol.crm_line_id.quantity if len(sol) == 1 else None,
                    'qty': qty if qty else None,
                    # 'number_used': sol.crm_line_id.number_used if len(sol) == 1 else None,
                    'number_used': number_used if number_used else None,
                    'money': money if money else None,
                    'code_walkin': row[14],
                    'service_date': '%s' % fields.Datetime.context_timestamp(
                        self.with_context(tz=self.env.user.tz),
                        row[15]).strftime('%d/%m/%Y'),
                    'code_specialty_sheet': row[16],
                    'department': row[17],
                    'end_datetime': '%s' % fields.Datetime.context_timestamp(
                        self.with_context(tz=self.env.user.tz),
                        row[18]).strftime('%d/%m/%Y'),
                    'point_service': row[19],
                    'type_booking': row[20],
                    'institution': row[21],
                })
                total_revenue += money

        # Tạo row hiện tổng cho dict
        val = dict((e, None) for e in TEMP)
        val['money'] = total_revenue
        ret_data.append(val)

        return ret_data
    def _get_data_report(self):
        ret_data = []
        total_revenue = 0

        Surgery = self.env['sh.medical.surgery.team'].sudo().with_context(company_ids=[c.id for c in self.env['res.company'].search([])])
        Specialty = self.env['sh.medical.specialty.team'].sudo().with_context(company_ids=[c.id for c in self.env['res.company'].search([])])
        if self.expected_date == 'conf_walkin':
            domain_surgery = [('name.state', '=', 'Done'), ('name.institution.his_company', '=', self.company_id.id),
                              ('name.surgery_end_date', '>=', self.start_datetime), ('name.surgery_end_date', '<=', self.end_datetime),
                              ('name.walkin.state', '=', 'Completed'),
                              ('name.walkin.type_crm_id', '!=', self.env.ref('crm_base.type_oppor_guarantee').id)]

            domain_specialty = [('name.state', '=', 'Done'), ('name.institution.his_company', '=', self.company_id.id),
                                ('name.services_end_date', '>=', self.start_datetime), ('name.services_end_date', '<=', self.end_datetime),
                                ('name.walkin.state', '=', 'Completed'),
                                ('name.walkin.type_crm_id', '!=', self.env.ref('crm_base.type_oppor_guarantee').id)]

        elif self.expected_date == 'conf_surgery':
            domain_surgery = [('name.state', '=', 'Done'), ('name.institution.his_company', '=', self.company_id.id),
                              ('name.surgery_end_date', '>=', self.start_datetime), ('name.surgery_end_date', '<=', self.end_datetime),
                              ('name.walkin.type_crm_id', '!=', self.env.ref('crm_base.type_oppor_guarantee').id)]

            domain_specialty = [('name.state', '=', 'Done'), ('name.institution.his_company', '=', self.company_id.id),
                                ('name.services_end_date', '>=', self.start_datetime), ('name.services_end_date', '<=', self.end_datetime),
                                ('name.walkin.type_crm_id', '!=', self.env.ref('crm_base.type_oppor_guarantee').id)]

        lang_code = self.env.context.get('lang') or 'en_US'
        date_format = self.env['res.lang']._lang_get(lang_code).date_format
        time_format = self.env['res.lang']._lang_get(lang_code).time_format

        if self.type == 'All':
            data_surgery = Surgery.search(domain_surgery)
            for data in data_surgery:
                # get walkin
                walkin = data.name.walkin
                # get khoa phòng
                department = data.name.department.name
                # get so
                so = walkin.sale_order_id
                # ngày phiếu khám
                service_date = walkin.service_date

                for ser_perform in data.service_performances:
                    sol = so.order_line.filtered(lambda r: r.product_id == ser_perform.product_id)
                    money = sum([rec_sol.price_subtotal for rec_sol in sol])
                    ret_data.append({
                        'employee_code': data.team_member.employee_code,
                        'member': data.team_member.name,
                        'group_job': data.team_member.group_job.name or None,
                        'job_id': data.team_member.job_id.name or None,
                        'role': data.role.name,
                        'name_customer': data.name.booking_id.partner_id.name,
                        'code_booking': data.name.booking_id.name,
                        'so_code': so.name,
                        'start_datetime': '%s' % fields.Datetime.context_timestamp(self.with_context(tz=self.env.user.tz), data.name.surgery_date).strftime('%d/%m/%Y'),
                        'surgery_type': DICT_SURGERY_TYPE.get(ser_perform.surgery_type) if ser_perform.surgery_type else None,
                        'his_service_type': dict_type.get(ser_perform.his_service_type),
                        'service_category': ser_perform.service_category.name,
                        'service': ser_perform.name,
                        'uom_price': sum([rec.crm_line_id.uom_price for rec in sol]),
                        'qty': sol.crm_line_id.quantity if len(sol) == 1 else None,
                        'number_used': sol.crm_line_id.number_used if len(sol) == 1 else None,
                        'money': money if sol else None,
                        'code_walkin': walkin.name,
                        'service_date': '%s' % fields.Datetime.context_timestamp(self.with_context(tz=self.env.user.tz), service_date).strftime('%d/%m/%Y'),
                        'code_specialty_sheet': data.name.name,
                        'department': department,
                        # 'uom_price': sol.crm_line_id.uom_price if len(sol) == 1 else None,
                        'end_datetime': '%s' % fields.Datetime.context_timestamp(self.with_context(tz=self.env.user.tz), data.name.surgery_end_date).strftime('%d/%m/%Y'),
                        'point_service': ser_perform.kpi_point,
                        'type_booking': walkin.booking_id.type_crm_id.name,
                        'institution': data.name.institution.name,
                    })
                    total_revenue += money

            domain_specialty.append(('name.department_type', 'in', ('Spa', 'Laser', 'Odontology')))
            data_surgery = Specialty.search(domain_specialty)
            for data in data_surgery:
                # get walkin
                walkin = data.name.walkin
                # get khoa phòng
                department = data.name.department.name
                # get so
                so = walkin.sale_order_id
                # ngày phiếu khám
                service_date = walkin.service_date

                for ser_perform in data.service_performances:
                    # get s ale order line for service
                    sol = so.order_line.filtered(lambda r: r.product_id == ser_perform.product_id)
                    money = sum([rec_sol.price_subtotal for rec_sol in sol])
                    ret_data.append({
                        'employee_code': data.team_member.employee_code,
                        'member': data.team_member.name,
                        'group_job': data.team_member.group_job.name or None,
                        'job_id': data.team_member.job_id.name or None,
                        'role': data.role.name,
                        'name_customer': data.name.booking_id.partner_id.name,
                        'code_booking': data.name.booking_id.name,
                        'so_code': so.name,
                        'start_datetime': '%s' % fields.Datetime.context_timestamp(self.with_context(tz=self.env.user.tz), data.name.services_date).strftime('%d/%m/%Y'),
                        'surgery_type': ser_perform.surgery_type or None,
                        'his_service_type': dict_type.get(ser_perform.his_service_type),
                        'service_category': ser_perform.service_category.name,
                        'service': ser_perform.name,
                        'uom_price': sum([rec.crm_line_id.uom_price for rec in sol]),
                        'qty': sol.crm_line_id.quantity if len(sol) == 1 else None,
                        'number_used': sol.crm_line_id.number_used if len(sol) == 1 else None,
                        'money': money if sol else None,
                        'code_walkin': walkin.name,
                        'service_date': '%s' % fields.Datetime.context_timestamp(self.with_context(tz=self.env.user.tz), service_date).strftime('%d/%m/%Y'),
                        'code_specialty_sheet': data.name.name,
                        'department': department,
                        # 'uom_price': sol.crm_line_id.uom_price if len(sol) == 1 else None,
                        'end_datetime': '%s' % fields.Datetime.context_timestamp(self.with_context(tz=self.env.user.tz), data.name.services_end_date).strftime('%d/%m/%Y'),
                        'point_service': ser_perform.kpi_point,
                        'type_booking': walkin.booking_id.type_crm_id.name,
                        'institution': data.name.institution.name,
                    })
                    total_revenue += money
            val = dict((e, None) for e in TEMP)
            val['money'] = total_revenue
            ret_data.append(val)
        elif self.type == 'Surgery':
            data_surgery = Surgery.search(domain_surgery)
            for data in data_surgery:
                # get walkin
                walkin = data.name.walkin
                # get khoa phòng
                department = data.name.department.name
                # get so
                so = walkin.sale_order_id
                # ngày phiếu khám
                service_date = walkin.service_date

                for ser_perform in data.service_performances:
                    sol = so.order_line.filtered(lambda r: r.product_id == ser_perform.product_id)
                    money = sum([rec_sol.price_subtotal for rec_sol in sol])
                    ret_data.append({
                        'employee_code': data.team_member.employee_code,
                        'member': data.team_member.name,
                        'group_job': data.team_member.group_job.name or None,
                        'job_id': data.team_member.job_id.name or None,
                        'role': data.role.name,
                        'name_customer': data.name.booking_id.partner_id.name,
                        'code_booking': data.name.booking_id.name,
                        'so_code': so.name,
                        'start_datetime': '%s' % fields.Datetime.context_timestamp(self.with_context(tz=self.env.user.tz), data.name.surgery_date).strftime('%d/%m/%Y'),
                        'surgery_type': DICT_SURGERY_TYPE.get(ser_perform.surgery_type) if ser_perform.surgery_type else None,
                        'his_service_type': dict_type.get(ser_perform.his_service_type),
                        'service_category': ser_perform.service_category.name,
                        'service': ser_perform.name,
                        'uom_price': sum([rec.crm_line_id.uom_price for rec in sol]),
                        'qty': sol.crm_line_id.quantity if len(sol) == 1 else None,
                        'number_used': sol.crm_line_id.number_used if len(sol) == 1 else None,
                        'money': money if sol else None,
                        'code_walkin': walkin.name,
                        'service_date': '%s' % fields.Datetime.context_timestamp(self.with_context(tz=self.env.user.tz), service_date).strftime('%d/%m/%Y'),
                        'code_specialty_sheet': data.name.name,
                        'department': department,
                        # 'uom_price': sol.crm_line_id.uom_price if len(sol) == 1 else None,
                        'end_datetime': '%s' % fields.Datetime.context_timestamp(self.with_context(tz=self.env.user.tz), data.name.surgery_end_date).strftime('%d/%m/%Y'),
                        'point_service': ser_perform.kpi_point,
                        'type_booking': walkin.booking_id.type_crm_id.name,
                        'institution': data.name.institution.name,
                    })
                    total_revenue += money
            val = dict((e, None) for e in TEMP)
            val['money'] = total_revenue
            ret_data.append(val)
        else:
            domain_specialty.append(('name.department_type', '=', self.type))
            data_surgery = Specialty.search(domain_specialty)
            for data in data_surgery:
                # get walkin
                walkin = data.name.walkin
                # get khoa phòng
                department = data.name.department.name
                # get so
                so = walkin.sale_order_id
                # ngày phiếu khám
                service_date = walkin.service_date

                for ser_perform in data.service_performances:
                    # get s ale order line for service
                    sol = so.order_line.filtered(lambda r: r.product_id == ser_perform.product_id)
                    money = sum([rec_sol.price_subtotal for rec_sol in sol])
                    ret_data.append({
                        'employee_code': data.team_member.employee_code,
                        'member': data.team_member.name,
                        'group_job': data.team_member.group_job.name or None,
                        'job_id': data.team_member.job_id.name or None,
                        'role': data.role.name,
                        'name_customer': data.name.booking_id.partner_id.name,
                        'code_booking': data.name.booking_id.name,
                        'so_code': so.name,
                        'start_datetime': '%s' % fields.Datetime.context_timestamp(self.with_context(tz=self.env.user.tz), data.name.services_date).strftime('%d/%m/%Y'),
                        'surgery_type': DICT_SURGERY_TYPE.get(ser_perform.surgery_type) if ser_perform.surgery_type else None,
                        'his_service_type': dict_type.get(ser_perform.his_service_type),
                        'service_category': ser_perform.service_category.name,
                        'service': ser_perform.name,
                        'uom_price': sum([rec.crm_line_id.uom_price for rec in sol]),
                        'qty': sol.crm_line_id.quantity if len(sol) == 1 else None,
                        'number_used': sol.crm_line_id.number_used if len(sol) == 1 else None,
                        'money': money if sol else None,
                        'code_walkin': walkin.name,
                        'service_date': '%s' % fields.Datetime.context_timestamp(self.with_context(tz=self.env.user.tz), service_date).strftime('%d/%m/%Y'),
                        'code_specialty_sheet': data.name.name,
                        'department': department,
                        # 'uom_price': sol.crm_line_id.uom_price if len(sol) == 1 else None,
                        'end_datetime': '%s' % fields.Datetime.context_timestamp(self.with_context(tz=self.env.user.tz), data.name.services_end_date).strftime('%d/%m/%Y'),
                        'point_service': ser_perform.kpi_point,
                        'type_booking': walkin.booking_id.type_crm_id.name,
                        'institution': data.name.institution.name,
                    })
                    total_revenue += money
            val = dict((e, None) for e in TEMP)
            val['money'] = total_revenue
            ret_data.append(val)
        return ret_data

    def create_report_service_perform_report(self):
        # get data
        if self.is_query:
            datas = self._get_data_report_v2()
        else:
            datas = self._get_data_report()

        # in dữ liễu
        report_brand_overview_attachment = self.env['ir.attachment'].browse(
            self.env.ref('report_sale.bao_cao_doanh_thu_thuc_hien_dich_vu_kangnam_attachment').id)
        decode = base64.b64decode(report_brand_overview_attachment.datas)
        wb = load_workbook(BytesIO(decode))
        ws = wb.active
        line_font = Font(name='Times New Roman', size=14)
        if self.type == 'Surgery':
            name = "PHẪU THUẬT"
        elif self.type == 'Spa':
            name = "SPA"
        elif self.type == 'Laser':
            name = "LASER"
        elif self.type == 'Odontology':
            name = "NHA KHOA"
        else:
            name = 'Toàn bộ dịch vụ'.upper()

        code_brand = self.company_id.brand_id.code.lower()
        image_path = get_module_resource('report_sale', 'static/img', 'icon_%s.png' % code_brand)

        img = openpyxl.drawing.image.Image(image_path)
        img.anchor = 'A1'
        ws.add_image(img)

        ws['A2'].value = 'BÁO CÁO DOANH THU THỰC HIỆN DỊCH VỤ %s' % name
        ws['N4'].value += self.start_date.strftime('%d/%m/%Y')
        ws['P4'].value += self.end_datetime.strftime('%d/%m/%Y')
        ws['N5'].value += self.company_id.name

        key_col = list(range(1, len(TEMP) + 1))
        format_currency = '_(* #,##0_);_(* (#,##0);_(* "-"??_);_(@_)'


        if code_brand == 'kn':
            header_fill = ThemeReport.kn_fill
        elif code_brand == 'da':
            header_fill = ThemeReport.da_fill
        elif code_brand == 'pr':
            header_fill = ThemeReport.pr_fill
        elif code_brand == 'hh':
            header_fill = ThemeReport.hh_fill
        else:
            header_fill = ThemeReport.sci_fill

        row = 8
        for data in datas:
            for col, k in zip(key_col, TEMP):
                beforeCell = ws.cell(7, col)
                beforeCell.fill = header_fill
                beforeCell.font = Font(name='Times New Roman', size=14, color='FFFFFF')
                cell = ws.cell(row, col)
                cell.value = data[k]
                cell.font = line_font
                cell.border = all_border_thin
                cell.alignment = Alignment(horizontal='left', vertical='center')
                if col == 17:
                    cell.number_format = format_currency
            row += 1
        fp = BytesIO()
        wb.save(fp)
        fp.seek(0)
        report = base64.encodebytes((fp.read()))
        fp.close()
        attachment = self.env['ir.attachment'].sudo().create({
            'name': 'bao_cao_doanh_thu_thuc_hien_dich_vu_%s.xlsx' % self.company_id.name,
            'datas': report,
            'res_model': 'temp.creation',
            'public': True,
        })
        url = "/web/content/?model=ir.attachment&id=%s&filename_field=name&field=datas&download=true" \
              % attachment.id
        return {'name': 'BÁO CÁO DOANH THU THỰC HIỆN DỊCH VỤ',
                'type': 'ir.actions.act_url',
                'url': url,
                'target': 'self',
                }
