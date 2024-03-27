from odoo import fields, api, models, _
from odoo.exceptions import ValidationError, UserError
from datetime import date, datetime, time
from calendar import monthrange
import openpyxl
from openpyxl.drawing.image import Image
from openpyxl import load_workbook
from openpyxl.styles import Font, borders, Alignment,PatternFill
import base64
from io import BytesIO
from dateutil.relativedelta import relativedelta
from pytz import timezone, utc
from odoo.modules.module import get_module_resource

thin = borders.Side(style='thin')
double = borders.Side(style='double')
all_border_thin = borders.Border(thin, thin, thin, thin)
da_fill = PatternFill(start_color='0e7661', end_color='0e7661', fill_type='solid')
kn_fill = PatternFill(start_color='003471', end_color='003471', fill_type='solid')
pr_fill = PatternFill(start_color='012C5F', end_color='012C5F', fill_type='solid')
hh_fill = PatternFill(start_color='053D7C', end_color='053D7C', fill_type='solid')
sci_fill = PatternFill(start_color='003471', end_color='003471', fill_type='solid')
SERVICE_HIS_TYPE = [('Spa', 'Spa'), ('Laser', 'Laser'), ('Odontology', 'Nha'), ('Surgery', 'Phẫu thuật'), ('ChiPhi', 'Chi phí khác')]
dict_type = dict((key, value) for key, value in SERVICE_HIS_TYPE)

class RoundingsReport(models.TransientModel):
    _name = 'roundings.report'
    _description = 'Báo cáo bệnh nhân lưu'

    start_date = fields.Date('Start date', default=date.today())
    end_date = fields.Date('End date', default=date.today())
    company_ids = fields.Many2many(string='Chi nhánh', comodel_name='res.company', domain=lambda self: [('id', 'in', self.env.user.company_ids.ids)])

    # convert date to datetime for search domain, should be removed if using datetime directly
    start_datetime = fields.Datetime('Start datetime', compute='_compute_datetime')
    end_datetime = fields.Datetime('End datetime', compute='_compute_datetime')

    @api.depends('start_date', 'end_date')
    def _compute_datetime(self):
        self.start_datetime = False
        self.end_datetime = False
        if self.start_date and self.end_date:
            local_tz = timezone(self.env.user.tz or 'Etc/GMT+7')
            start_datetime = datetime(self.start_date.year, self.start_date.month, self.start_date.day, 0, 0, 0)
            end_datetime = datetime(self.end_date.year, self.end_date.month, self.end_date.day, 23, 59, 59)
            start_datetime = local_tz.localize(start_datetime, is_dst=None)
            end_datetime = local_tz.localize(end_datetime, is_dst=None)
            self.start_datetime = start_datetime.astimezone(utc).replace(tzinfo=None)
            self.end_datetime = end_datetime.astimezone(utc).replace(tzinfo=None)

    @api.onchange('start_date')
    def _onchange_start_date(self):
        if self.start_date:
            if self.start_date.month == fields.date.today().month:
                self.end_date = fields.date.today()
            else:
                self.end_date = date(self.start_date.year, self.start_date.month,
                                     monthrange(self.start_date.year, self.start_date.month)[1])

    @api.constrains('start_date', 'end_date')
    def check_dates(self):
        for record in self:
            start_date = fields.Date.from_string(record.start_date)
            end_date = fields.Date.from_string(record.end_date)
            days = (end_date - start_date).days
            if days < 0 or days > 365:
                raise ValidationError(
                    _("Ngày kết thúc không thể ở trước ngày bắt đầu khi xuất báo cáo!!!"))

    def _get_data_report(self):
        domain = [('name.state', '=', 'Completed'), ('name.inpatient_id.institution.his_company', 'in', self.company_ids.ids),
                  ('name.evaluation_start_date', '>=', self.start_datetime), ('name.evaluation_start_date', '<=', self.end_datetime)]
        roundingteams = self.env['sh.medical.rounding.team'].sudo().search(domain, order='id')
        return_val = []
        for roundingteam in roundingteams:
            roundingteam = roundingteam.sudo()
            for service in roundingteam.service_performances:
                so_line = self.env['sale.order.line'].sudo().search([('order_id', '=', roundingteam.name.inpatient_id.walkin.sale_order_id.id),('product_id', '=', service.product_id.id)], order='id')

                val = {
                    'code_rounding': roundingteam.name.name,
                    'member': roundingteam.team_member.name,
                    'role': roundingteam.role.name,
                    'his_service_type': dict_type.get(service.his_service_type),
                    'service_category': service.service_category.name,
                    'services': service.name,
                    'code_booking': roundingteam.name.inpatient_id.walkin.booking_id.name,
                    'patient': roundingteam.name.patient.name,
                    'evaluation_start_date': roundingteam.name.evaluation_start_date.strftime('%d/%m/%Y'),
                    'code_walkin': roundingteam.name.inpatient_id.walkin.name,
                    'walkin_type': roundingteam.name.inpatient_id.walkin.type_crm_id.name,
                    'institution': roundingteam.name.inpatient_id.institution.name,
                    'so_amount': sum(so_line.mapped('price_subtotal')) or 0
                }
                return_val.append(val)
        return return_val

    def create_report(self):
        datas = self._get_data_report()
        # in dữ liễu
        report_inpatient_attachment = self.env['ir.attachment'].browse(
            self.env.ref('report_sale.bao_cao_doanh_thu_cshp_attachment').id)
        decode = base64.b64decode(report_inpatient_attachment.datas)
        wb = load_workbook(BytesIO(decode))
        ws = wb.active
        line_font = Font(name='Times New Roman', size=14)

        code_brand = self.company_ids[0].brand_id.code.lower()
        if len(self.company_ids) > 1:
            image_path = get_module_resource('report_sale', 'static/img', 'icon_sci.png')
        else:
            image_path = get_module_resource('report_sale', 'static/img', 'icon_%s.png' % code_brand)

        img = openpyxl.drawing.image.Image(image_path)
        img.anchor = 'A1'
        ws.add_image(img)

        ws['C2'].value = 'Từ ngày: %s đến ngày %s ' % (self.start_date.strftime('%d/%m/%Y'),self.end_date.strftime('%d/%m/%Y'))
        key_col = list(range(1, 14))
        key_list = [
            'code_rounding',
            'member',
            'role',
            'his_service_type',
            'service_category',
            'services',
            'code_booking',
            'patient',
            'evaluation_start_date',
            'code_walkin',
            'walkin_type',
            'institution',
            'so_amount'
        ]
        row = 5
        if code_brand == 'kn':
            header_fill = kn_fill
        elif code_brand == 'da':
            header_fill = da_fill
        elif code_brand == 'pr':
            header_fill = pr_fill
        elif code_brand == 'hh':
            header_fill = hh_fill
        else:
            header_fill = sci_fill
        format_currency = '_(* #,##0_);_(* (#,##0);_(* "-"??_);_(@_)'
        for data in datas:
            for col, k in zip(key_col, key_list):
                beforeCell = ws.cell(4, col)
                beforeCell.fill = header_fill
                beforeCell.font = Font(name='Times New Roman', size=14, color='FFFFFF')
                cell = ws.cell(row, col)
                cell.value = data[k]
                cell.font = line_font
                cell.border = all_border_thin
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                if col == 13:
                    cell.number_format = format_currency
            row += 1
        fp = BytesIO()
        wb.save(fp)
        fp.seek(0)
        report = base64.encodebytes((fp.read()))
        fp.close()
        attachment = self.env['ir.attachment'].sudo().create({
            'name': 'bao_cao_doanh_thu_cshp.xlsx',
            'datas': report,
            'res_model': 'temp.creation',
            'public': True,
        })
        url = "/web/content/?model=ir.attachment&id=%s&filename_field=name&field=datas&download=true" \
              % attachment.id
        return {'name': 'BÁO CÁO DOANH THU CSHP',
                'type': 'ir.actions.act_url',
                'url': url,
                'target': 'self',
                }
