from odoo import fields, api, models, _
from odoo.exceptions import ValidationError, UserError
from datetime import date, datetime, time
from calendar import monthrange
from openpyxl import load_workbook
from openpyxl.styles import Font, borders, Alignment
import base64
from io import BytesIO
from pytz import timezone, utc
from collections import defaultdict
from operator import itemgetter
from odoo.tools.profiler import profile

CONSULTING_ROLE = [('1', 'Tư vấn độc lập'), ('2', 'Tư vấn đồng thời'), ('3', 'Lễ tân - CVTV cùng tư vấn'),
                   ('4', 'BS da liễu - KTV cùng tư vấn'), ('5', 'Tư vấn chính'), ('6', 'Tư vấn phụ')]
DICT_CONSULTING_ROLE = dict((key, value) for key, value in CONSULTING_ROLE)

thin = borders.Side(style='thin')
double = borders.Side(style='double')
all_border_thin = borders.Border(thin, thin, thin, thin)
'''
    #1370
    [Báo cáo ERP - HOT] Báo cáo Doanh số Bán Dịch vụ, Sản phẩm - Đông Á, Paris
'''

TEMP = [
    'stt',
    'ma_nv_tu_van',
    'tu_van_vien',
    'bo_phan_tu_van',
    'chuc_danh_tu_van',
    'vai_tro_tu_van',
    'ten_khach_hang',
    'ma_booking',
    'ngay_thanh_toan',
    'loai_phau_thuat',
    'loai_dich_vu',
    'nhom_dich_vu_cap_2',
    'ten_dich_vu',
    'don_vi_xu_ly',
    'buoi_thu',
    'so_luong',
    'tong_tien_ban_dau',
    'tong_tien_da_tra',
    'tong_tien_phai_thu_con_lai',
    'tien_thu',
    'phan_bo_doanh_so',
    'loai_thanh_toan',
    'phan_loai_khach_hang',
    'nguon_booking',
    'nguon_mo_rong',
    'phan_loai_mo_rong',
    'thu_tu_tu_van',
    'ma_dich_vu',
    'ma_khach_hang',
    'phong_ban_tao_booking',
    'emp_unique',
    'order_emp'
]

SERVICE_HIS_TYPE = [('Spa', 'Spa'), ('Laser', 'Laser'), ('Odontology', 'Nha'), ('Surgery', 'Phẫu thuật'), ('ChiPhi', 'Chi phí khác')]
dict_type = dict((key, value) for key, value in SERVICE_HIS_TYPE)

SURGERY_TYPE = [('minor', 'Tiểu phẫu'), ('major', 'Đại phẫu')]
DICT_SURGERY_TYPE = dict((key, value) for key, value in SURGERY_TYPE)


class ProductServiceSalesReport(models.TransientModel):
    _name = 'sale.product.report.payment'
    _description = '#1370 [Báo cáo ERP - HOT] Báo cáo Doanh số Bán Dịch vụ, Sản phẩm - Đông Á, Paris'

    start_date = fields.Date('Start date', default=date.today())
    end_date = fields.Date('End date', default=date.today())
    company_ids = fields.Many2many('res.company', string='Company', domain=lambda self: [('id', 'in', self.env.user.company_ids.ids)])

    # convert date to datetime for search domain, should be removed if using datetime directly
    start_datetime = fields.Datetime('Start datetime', compute='_compute_datetime')
    end_datetime = fields.Datetime('End datetime', compute='_compute_datetime')

    @api.depends('start_date', 'end_date')
    def _compute_datetime(self):
        self.start_datetime = False
        self.end_datetime = False
        if self.start_date and self.end_date:
            local_tz = timezone(self.env.user.tz or 'Etc/GMT+7')
            start_datetime = datetime(self.start_date.year, self.start_date.month, self.start_date.day, 0, 0, 0)
            end_datetime = datetime(self.end_date.year, self.end_date.month, self.end_date.day, 23, 59, 59)
            start_datetime = local_tz.localize(start_datetime, is_dst=None)
            end_datetime = local_tz.localize(end_datetime, is_dst=None)
            self.start_datetime = start_datetime.astimezone(utc).replace(tzinfo=None)
            self.end_datetime = end_datetime.astimezone(utc).replace(tzinfo=None)

    @api.onchange('start_date')
    def _onchange_start_date(self):
        if self.start_date:
            if self.start_date.month == fields.date.today().month:
                self.end_date = fields.date.today()
            else:
                self.end_date = date(self.start_date.year, self.start_date.month,
                                     monthrange(self.start_date.year, self.start_date.month)[1])

    @api.constrains('start_date', 'end_date')
    def check_dates(self):
        for record in self:
            start_date = fields.Date.from_string(record.start_date)
            end_date = fields.Date.from_string(record.end_date)
            if start_date > end_date:
                raise ValidationError(
                    _("Ngày kết thúc không thể ở trước ngày bắt đầu khi xuất báo cáo!!!"))

    # def get_consultants_in_line(self, crm_line):
    #     consultants = []
    #     if crm_line.consultants_1:
    #         consultants.append((crm_line.consultants_1, crm_line.consulting_role_1))
    #     elif crm_line.consultants_2:
    #         consultants.append((crm_line.consultants_2, crm_line.consulting_role_2))
    #     return consultants
    def get_info_from_line_service(self, crm_line_service, crm_lead, payment):
        query = """
                   SELECT
                        CASE WHEN consultants_2 IS NOT NULL THEN consultants_2 ELSE consultants_1 END,
                        CASE WHEN consultants_2 IS NOT NULL THEN consulting_role_2 ELSE consulting_role_1 END            
                    FROM crm_line_product
                    where id = %s
               """
        self.env.cr.execute(query, (tuple(crm_line_service.ids)))
        query_result = self.env.cr.fetchall()
        # list_consultants = self.get_consultants_in_line(crm_line_service)
        result = []
        count = 0
        list_done_walkin = crm_lead.walkin_ids.filtered(lambda w: crm_line_service.service_id.id in w.service.ids and w.state == 'Completed').sorted('service_date').mapped('service_date')
        index = [i for i, x in enumerate(list_done_walkin) if datetime.date(x) <= payment[0]]

        if len(list_done_walkin) > 0 and len(index) > 0:
            walkin_num = index[0] + 1
        else:
            walkin_num = len(list_done_walkin) + 1
        if len(query_result) > 0:
            for element in query_result:
                count += 1
                user = self.env['res.users'].sudo().browse(int(element[0]))
                work_email = user.login
                employee = self.env['hr.employee'].sudo().search([('work_email', '=', str(work_email))])
                result.append({
                    'id': crm_line_service.id,
                    'ma_nv_tu_van': employee.employee_code if employee else None,
                    'tu_van_vien': employee.name if employee else None,
                    'bo_phan_tu_van': employee.group_job.name if employee and employee.group_job else None,
                    'chuc_danh_tu_van': employee.job_id.name if employee and employee.job_id else None,
                    'vai_tro_tu_van': DICT_CONSULTING_ROLE.get(element[1], None),
                    'ten_khach_hang': crm_lead.partner_id.name,
                    'ma_booking': crm_lead.name,
                    'ngay_thanh_toan': None,
                    'loai_phau_thuat': DICT_SURGERY_TYPE.get(crm_line_service.service_id.surgery_type, 'Sản phẩm'),
                    'loai_dich_vu': dict_type.get(crm_line_service.service_id.his_service_type, 'Sản phẩm'),
                    'nhom_dich_vu_cap_2': crm_line_service.service_id.service_category.name,
                    'ten_dich_vu': crm_line_service.service_id.name,
                    'don_vi_xu_ly': crm_line_service.uom_price,
                    'buoi_thu': walkin_num,
                    'so_luong': crm_line_service.quantity,
                    'tong_tien_ban_dau': None,
                    'tong_tien_da_tra': None,
                    'tong_tien_phai_thu_con_lai': None,
                    'tien_thu': None,
                    'phan_bo_doanh_so': None,
                    'loai_thanh_toan': None,
                    'phan_loai_khach_hang': None,
                    'nguon_booking': crm_lead.source_id.name,
                    'nguon_mo_rong': crm_line_service.source_extend_id.name,
                    'phan_loai_mo_rong': None,
                    'thu_tu_tu_van': 'Tư vấn viên %s' % count,
                    'ma_dich_vu': crm_line_service.service_id.default_code,
                    'ma_khach_hang': crm_lead.partner_id.code_customer,
                    'phong_ban_tao_booking': crm_lead.create_by_department or None,
                    'emp_unique': None,
                    'order_emp': None
                })
        else:
            result.append({
                'id': crm_line_service.id,
                'ma_nv_tu_van': None,
                'tu_van_vien': None,
                'bo_phan_tu_van': None,
                'chuc_danh_tu_van': None,
                'vai_tro_tu_van': None,
                'ten_khach_hang': crm_lead.partner_id.name,
                'ma_booking': crm_lead.name,
                'ngay_thanh_toan': None,
                'loai_phau_thuat': DICT_SURGERY_TYPE.get(crm_line_service.service_id.surgery_type, 'Sản phẩm'),
                'loai_dich_vu': dict_type.get(crm_line_service.service_id.his_service_type, 'Sản phẩm'),
                'nhom_dich_vu_cap_2': crm_line_service.service_id.service_category.name,
                'ten_dich_vu': crm_line_service.service_id.name,
                'don_vi_xu_ly': crm_line_service.uom_price,
                'buoi_thu': walkin_num,
                'so_luong': crm_line_service.quantity,
                'tong_tien_ban_dau': None,
                'tong_tien_da_tra': None,
                'tong_tien_phai_thu_con_lai': None,
                'tien_thu': None,
                'phan_bo_doanh_so': None,
                'loai_thanh_toan': None,
                'phan_loai_khach_hang': None,
                'nguon_booking': crm_lead.source_id.name,
                'nguon_mo_rong': crm_line_service.source_extend_id.name,
                'phan_loai_mo_rong': None,
                'thu_tu_tu_van': 'Tư vấn viên %s' % count,
                'ma_dich_vu': crm_line_service.service_id.default_code,
                'ma_khach_hang': crm_lead.partner_id.code_customer,
                'phong_ban_tao_booking': crm_lead.create_by_department or None,
                'emp_unique': None,
                'order_emp': None
            })
        return result

    def get_info_from_line_product(self, crm_line_product, crm_lead):
        query = """
           SELECT
                CASE WHEN consultants_2 IS NOT NULL THEN consultants_2 ELSE consultants_1 END,
                CASE WHEN consultants_2 IS NOT NULL THEN consulting_role_2 ELSE consulting_role_1 END            
            FROM crm_line_product
            where id = %s
       """
        self.env.cr.execute(query, (tuple(crm_line_product.ids)))
        query_result = self.env.cr.fetchall()
        # list_consultants = self.get_consultants_in_line(crm_line_product)
        result = []
        count = 0
        if len(query_result) > 0:
            for element in query_result:
                count += 1
                user = self.env['res.users'].sudo().browse(int(element[0]))
                work_email = user.login
                employee = self.env['hr.employee'].sudo().search([('work_email', '=', str(work_email))])
                result.append({
                    'ma_nv_tu_van': employee.employee_code if employee else None,
                    'tu_van_vien': employee.name if employee else None,
                    'bo_phan_tu_van': employee.group_job.name if employee and employee.group_job else None,
                    'chuc_danh_tu_van': employee.job_id.name if employee and employee.job_id else None,
                    'vai_tro_tu_van': DICT_CONSULTING_ROLE.get(element[1], None),
                    'ten_khach_hang': crm_lead.partner_id.name,
                    'ma_booking': crm_lead.name,
                    'ngay_thanh_toan': None,
                    'loai_phau_thuat': 'Sản phẩm',
                    'loai_dich_vu': 'Sản phẩm',
                    'nhom_dich_vu_cap_2': 'Sản phẩm',
                    'ten_dich_vu': crm_line_product.product_id.name,
                    'don_vi_xu_ly': crm_line_product.product_uom_qty,
                    'buoi_thu': None,
                    'so_luong': crm_line_product.product_uom_qty,
                    'tong_tien_ban_dau': None,
                    'tong_tien_da_tra': None,
                    'tong_tien_phai_thu_con_lai': None,
                    'tien_thu': None,
                    'phan_bo_doanh_so': None,
                    'loai_thanh_toan': None,
                    'phan_loai_khach_hang': None,
                    'nguon_booking': crm_lead.source_id.name,
                    'nguon_mo_rong': crm_line_product.source_extend_id.name,
                    'phan_loai_mo_rong': None,
                    'thu_tu_tu_van': 'Tư vấn viên %s' % count,
                    'ma_dich_vu': crm_line_product.product_id.default_code,
                    'ma_khach_hang': crm_lead.partner_id.code_customer,
                    'phong_ban_tao_booking': crm_lead.create_by_department or None,
                    'emp_unique': None,
                    'order_emp': None
                })
        else:
            result.append({
                'ma_nv_tu_van': None,
                'tu_van_vien': None,
                'bo_phan_tu_van': None,
                'chuc_danh_tu_van': None,
                'vai_tro_tu_van': None,
                'ten_khach_hang': crm_lead.partner_id.name,
                'ma_booking': crm_lead.name,
                'ngay_thanh_toan': None,
                'loai_phau_thuat': 'Sản phẩm',
                'loai_dich_vu': 'Sản phẩm',
                'nhom_dich_vu_cap_2': 'Sản phẩm',
                'ten_dich_vu': crm_line_product.product_id.name,
                'don_vi_xu_ly': crm_line_product.product_uom_qty,
                'buoi_thu': None,
                'so_luong': crm_line_product.product_uom_qty,
                'tong_tien_ban_dau': None,
                'tong_tien_da_tra': None,
                'tong_tien_phai_thu_con_lai': None,
                'tien_thu': None,
                'phan_bo_doanh_so': None,
                'loai_thanh_toan': None,
                'phan_loai_khach_hang': None,
                'nguon_booking': crm_lead.source_id.name,
                'nguon_mo_rong': crm_line_product.source_extend_id.name,
                'phan_loai_mo_rong': None,
                'thu_tu_tu_van': 'Tư vấn viên %s' % count,
                'ma_dich_vu': crm_line_product.product_id.default_code,
                'ma_khach_hang': crm_lead.partner_id.code_customer,
                'phong_ban_tao_booking': crm_lead.create_by_department or None,
                'emp_unique': None,
                'order_emp': None
            })
        return result

    def get_data_report(self):
        ret_data = []
        query = """
            SELECT pay.id, cld.id
            FROM account_payment as pay
            LEFT JOIN account_journal as jour ON pay.journal_id = jour.id
            LEFT JOIN res_company as cmp ON jour.company_id = cmp.id
            left join crm_lead cld on cld.id = pay.crm_id
            WHERE 
                cmp.id in %s AND state = 'posted' AND 
                payment_date BETWEEN %s AND %s and pay.crm_id is not null
        """
        self.env.cr.execute(query, (tuple(self.company_ids.ids), self.start_date, self.end_date))
        query_result = self.env.cr.fetchall()

        dict_by_booking = defaultdict(list)
        payment = self.env['account.payment'].sudo()
        for element in query_result:
            pay = payment.browse(element[0])
            dict_by_booking[element[1]].append(pay)
        data_line_crm = []
        dict_cl = {}
        dict_clp = {}
        query = """
                    SELECT crm_id, cl.id
                    FROM crm_line as cl
                    WHERE crm_id in %s and stage != %s
                    """
        self.env.cr.execute(query, (tuple(dict_by_booking), 'cancel'))
        query_result = self.env.cr.fetchall()
        for c in query_result:
            if c[0] not in dict_cl:
                dict_cl[c[0]] = []
            # else:
            dict_cl[c[0]].append(c[1])
        query = """
                    SELECT booking_id, cl.id
                    FROM crm_line_product as cl
                    WHERE booking_id in %s and stage_line_product != %s
                    """
        self.env.cr.execute(query, (tuple(dict_by_booking), 'cancel'))
        query_result = self.env.cr.fetchall()
        for c in query_result:
            if c[0] not in dict_clp:
                dict_clp[c[0]] = []
            # else:
            dict_clp[c[0]].append(c[1])
        for crm_lead, pay_list in dict_by_booking.items():
            '''
            ==========================================
                Group các phiếu thu
                1. Lấy thông tin các phiếu payment thành list các thông tin cần thiết [(date, type, amount)]
                2. Group các phiếu thu theo tiêu thức ngày tháng
            ==========================================
            '''
            payment_list = []
            crm_id = self.env['crm.lead'].sudo().browse(crm_lead)
            for pay in pay_list:
                if pay.payment_type == 'inbound':
                    payment_list.append((pay.payment_date, pay.amount_vnd))
                if pay.payment_type == 'outbound':
                    payment_list.append((pay.payment_date, -1 * pay.amount_vnd))

            set_by_payment_date = set(map(lambda x: x[0], payment_list))

            new_payment_list = [(day, sum([element[1] for element in payment_list if element[0] == day])) for day in set_by_payment_date]

            result_payment = []
            for element in new_payment_list:
                if element[1] < 0:
                    result_payment.append((element[0], 'Hoàn tiền', element[1] * -1))
                if element[1] > 0:
                    result_payment.append((element[0], 'Nhận tiền', element[1]))
            for payment in result_payment:
                count = 0
                crm_lines_service = []
                if crm_lead in dict_cl:
                    crm_lines_service = self.env['crm.line'].sudo().search([('id','in',dict_cl[crm_lead])])
                crm_lines_product = []
                if crm_lead in dict_clp:
                    crm_lines_product = self.env['crm.line.product'].sudo().search([('id','in',dict_clp[crm_lead])])
                if crm_lines_service:
                    for line in crm_lines_service:
                        count += 1
                        val = self.get_info_from_line_service(line, crm_id, payment)
                        for e in val:
                            e['ngay_thanh_toan'] = payment[0].strftime('%d/%m/%Y')
                        if count == 1:
                            val[0]['tien_thu'] = payment[2]
                            val[0]['loai_thanh_toan'] = payment[1]
                            val[0]['tong_tien_ban_dau'] = crm_id.amount_total
                            val[0]['tong_tien_da_tra'] = crm_id.amount_paid
                            val[0]['tong_tien_phai_thu_con_lai'] = crm_id.amount_total - crm_id.amount_paid
                        data_line_crm.append(val)
                if crm_lines_product:
                    for line in crm_lines_product:
                        count += 1
                        val = self.get_info_from_line_product(line, crm_id)
                        for e in val:
                            e['ngay_thanh_toan'] = payment[0].strftime('%d/%m/%Y')
                        if count == 1:
                            val[0]['tien_thu'] = payment[2]
                            val[0]['loai_thanh_toan'] = payment[1]
                            val[0]['tong_tien_ban_dau'] = crm_id.amount_total
                            val[0]['tong_tien_da_tra'] = crm_id.amount_paid
                            val[0]['tong_tien_phai_thu_con_lai'] = crm_id.amount_total - crm_id.amount_paid
                        data_line_crm.append(val)

        count = 0
        for element in data_line_crm:
            for line in element:
                count += 1
                line['stt'] = count
                ret_data.append(line)
        # Sort theo ngày thanh toán, và tên khách hàng
        ret_data.sort(key=itemgetter('ngay_thanh_toan', 'ten_khach_hang'), reverse=False)
        return ret_data

    def create_kpi_product_sale_report(self):
        # get data
        datas = self.get_data_report()

        # in dữ liệu
        report_brand_overview_attachment = self.env['ir.attachment'].browse(
            self.env.ref('report_sale.bao_cao_doanh_so_ban_san_pham_attachment_1370').id)
        decode = base64.b64decode(report_brand_overview_attachment.datas)
        wb = load_workbook(BytesIO(decode))
        ws = wb.active
        line_font = Font(name='Times New Roman', size=12)

        ws['G3'].value += self.start_date.strftime('%d/%m/%Y')
        ws['H3'].value += self.end_datetime.strftime('%d/%m/%Y')
        ws['G2'].value = "; ".join(c.name.upper() for c in self.company_ids)

        # code_brand = self.company_id.brand_id.code.lower()
        # image_path = get_module_resource('report_sale', 'static/img', 'icon_%s.png' % code_brand)
        #
        # img = openpyxl.drawing.image.Image(image_path)
        # img.anchor = 'A1'
        # ws.add_image(img)

        key_col_list = list(range(1, len(TEMP) + 1))

        row = 8
        # if code_brand == 'kn':
        #     header_fill = ThemeReport.kn_fill
        # elif code_brand == 'da':
        #     header_fill = ThemeReport.da_fill
        # elif code_brand == 'pr':
        #     header_fill = ThemeReport.pr_fill
        # elif code_brand == 'hh':
        #     header_fill = ThemeReport.hh_fill
        # else:
        #     header_fill = ThemeReport.sci_fill
        format_currency = '_(* #,##0_);_(* (#,##0);_(* "-"??_);_(@_)'
        alignment = Alignment(horizontal='left', vertical='center')
        for data in datas:
            for col, k in zip(key_col_list, TEMP):
                # beforeCell = ws.cell(8, col)
                # beforeCell.fill = header_fill
                # beforeCell.font = Font(name='Times New Roman', size=14, color='FFFFFF')
                cell = ws.cell(row, col)
                cell.value = data[k]
                cell.font = line_font
                cell.border = all_border_thin
                if col in range(14, 21):
                    cell.number_format = format_currency
                if col == 31:
                    cell.value = '=IF(COUNTIF(OFFSET($B${0}, 0, 0, ROW()-6, 1), INDIRECT(ADDRESS(ROW(), 2)))=1, COUNTIF(OFFSET($B${0}, 0, 0, ROW()-6, 1), INDIRECT(ADDRESS(ROW(), 2))), "")'.format(row)
                if col == 32:
                    cell.value = '=SUM(OFFSET(cell_left, 7 - ROW(), 0, ROW()-6, 1))'
            row += 1
        fp = BytesIO()
        wb.save(fp)
        fp.seek(0)
        report = base64.encodebytes((fp.read()))
        fp.close()
        attachment = self.env['ir.attachment'].sudo().create({
            'name': 'bao_cao_doanh_so_ban_san_pham.xlsx',
            'datas': report,
            'res_model': 'temp.creation',
            'public': True,
        })
        url = "/web/content/?model=ir.attachment&id=%s&filename_field=name&field=datas&download=true" \
              % attachment.id
        return {'name': 'BÁO CÁO DOANH SỐ BÁN SẢN PHẨM',
                'type': 'ir.actions.act_url',
                'url': url,
                'target': 'self',
                }
