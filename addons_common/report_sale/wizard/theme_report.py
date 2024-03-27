from openpyxl.styles import Font, borders, Alignment, PatternFill


class ThemeReport:
    thin = borders.Side(style='thin')
    double = borders.Side(style='double')
    all_border_thin = borders.Border(thin, thin, thin, thin)
    line_font = Font(name='Times New Roman', size=12)
    da_fill = PatternFill(start_color='0e7661', end_color='0e7661', fill_type='solid')
    kn_fill = PatternFill(start_color='003471', end_color='003471', fill_type='solid')
    pr_fill = PatternFill(start_color='012C5F', end_color='012C5F', fill_type='solid')
    hh_fill = PatternFill(start_color='053D7C', end_color='053D7C', fill_type='solid')
    sci_fill = PatternFill(start_color='003471', end_color='003471', fill_type='solid')
