from odoo import fields, api, models, _
from odoo.exceptions import ValidationError, UserError
from datetime import date, datetime, time
from calendar import monthrange
from openpyxl import load_workbook
from openpyxl.styles import Font, borders, Alignment
import base64
from io import BytesIO
from dateutil.relativedelta import relativedelta
from pytz import timezone, utc
from collections import defaultdict
from operator import itemgetter

thin = borders.Side(style='thin')
double = borders.Side(style='double')
all_border_thin = borders.Border(thin, thin, thin, thin)
KEY_LIST = [
    'ngay_tao',
    'ngay_thu_tien',
    'so_chung_tu',
    'trang_thai_payment',
    'loai_thanh_toan',
    'ma_khach_hang',
    'ten_khach_hang',
    'noi_dung',
    'so_tien_thanh_toan',
    'ngay_tao_booking',
    'ngay_hen_lich',
    'ma_booking',
    'nguoi_tao_booking',
    'phong_ban_nguoi_tao',
    'bang_gia',
    'trang_thai_booking',
    'nguon_booking',
    'chi_nhanh',
]
PAYMENT_TYPE = [('outbound', 'Hoàn tiền'), ('inbound', 'Nhận tiền'), ('transfer', 'Giao dịch nội bộ')]
DIC_PAYMENT_TYPE = dict((key, value) for key, value in PAYMENT_TYPE)
PAYMENT_STATE = [('draft', 'Nháp'), ('posted', 'Đã xác nhận'), ('sent', 'Đã gửi'), ('reconciled', 'Đã được đối soát'), ('cancelled', 'Đã hủy')]
DIC_PAYMENT_STATE = dict((key, value) for key, value in PAYMENT_STATE)

class SaleMarketingPaymentDA(models.TransientModel):
    _name = 'sale.marketing.payment.da'
    _description = 'Báo cáo viện phí'

    start_date = fields.Date('Start date', default=date.today())
    end_date = fields.Date('End date', default=date.today())

    company_ids = fields.Many2many(string='Công ty', comodel_name='res.company', domain=lambda self: [('id', 'in', self.env.user.company_ids.ids)])
    # convert date to datetime for search domain, should be removed if using datetime directly
    start_datetime = fields.Datetime('Start datetime', compute='_compute_datetime')
    end_datetime = fields.Datetime('End datetime', compute='_compute_datetime')
    report_type = fields.Selection([('01', 'Theo ngày tạo phiếu thu tiền'), ('02', 'Theo ngày thu tiền')], string='Lựa chọn tải')

    @api.depends('start_date', 'end_date')
    def _compute_datetime(self):
        self.start_datetime = False
        self.end_datetime = False
        if self.start_date and self.end_date:
            local_tz = timezone(self.env.user.tz or 'Etc/GMT+7')
            start_datetime = datetime(self.start_date.year, self.start_date.month, self.start_date.day, 0, 0, 0)
            end_datetime = datetime(self.end_date.year, self.end_date.month, self.end_date.day, 23, 59, 59)
            start_datetime = local_tz.localize(start_datetime, is_dst=None)
            end_datetime = local_tz.localize(end_datetime, is_dst=None)
            self.start_datetime = start_datetime.astimezone(utc).replace(tzinfo=None)
            self.end_datetime = end_datetime.astimezone(utc).replace(tzinfo=None)

    @api.onchange('start_date')
    def _onchange_start_date(self):
        if self.start_date:
            if self.start_date.month == fields.date.today().month:
                self.end_date = fields.date.today()
            else:
                self.end_date = date(self.start_date.year, self.start_date.month,
                                     monthrange(self.start_date.year, self.start_date.month)[1])

    @api.constrains('start_date', 'end_date')
    def check_dates(self):
        for record in self:
            start_date = fields.Date.from_string(record.start_date)
            end_date = fields.Date.from_string(record.end_date)
            days = (end_date - start_date).days
            if days < 0 or days > 365:
                raise ValidationError(
                    _("Ngày kết thúc không thể ở trước ngày bắt đầu khi xuất báo cáo!!!"))

    def render_form_template(self, pay):
        booking = pay.crm_id
        partner = pay.partner_id
        user = pay.user
        employee = booking.create_by.employee_ids


        # str_tuple = (booking.pass_port_address, booking.district_id.name, booking.state_id.name, booking.country_id.name)
        # str_convert = (element for element in str_tuple if element != False)
        # dia_chi = ', '.join(str_convert)

        lang_code = self.env.context.get('lang') or 'en_US'
        date_format = self.env['res.lang']._lang_get(lang_code).date_format
        time_format = self.env['res.lang']._lang_get(lang_code).time_format

        if booking:
            booking_create = fields.Datetime.context_timestamp(self.with_context(tz=self.env.user.tz), booking.create_on).strftime(('%s %s') % (date_format, time_format))
            booking_date = fields.Datetime.context_timestamp(self.with_context(tz=self.env.user.tz), booking.booking_date).strftime(('%s %s') % (date_format, time_format))
        else:
            booking_create = None
            booking_date = None
        val = {
            'ngay_tao': '%s' % fields.Datetime.context_timestamp(self.with_context(tz=self.env.user.tz), pay.create_date).strftime(('%s %s') % (date_format, time_format)) or None,
            'ngay_thu_tien': pay.payment_date or '',
            'so_chung_tu': pay.name or '',
            'trang_thai_payment': DIC_PAYMENT_STATE.get(pay.state, None),
            'loai_thanh_toan': DIC_PAYMENT_TYPE.get(pay.payment_type, None),
            'ma_khach_hang': partner.code_customer or None,
            'ten_khach_hang': partner.name or None,
            'noi_dung': pay.communication or None,
            'so_tien_thanh_toan': pay.amount or None,
            'ngay_tao_booking': booking_create or None,
            'ngay_hen_lich': booking_date or None,
            'ma_booking': booking.name or None,
            'nguoi_tao_booking': booking.create_by.name or None,
            'phong_ban_nguoi_tao': employee.department_id.name or None,
            'bang_gia': booking.price_list_id.name or None,
            'trang_thai_booking': booking.state_id.name or None,
            'nguon_booking': booking.source_id.name or None,
            'chi_nhanh': pay.company_id.name or None,
        }
        return val

    def _get_data_report(self):
        if self.report_type == '01':
            domain = [('company_id', 'in', self.company_ids.ids), ('create_date', '>=', self.start_datetime), ('create_date', '<=', self.end_datetime)]
        elif self.report_type == '02':
            domain = [('company_id', 'in', self.company_ids.ids), ('payment_date', '>=', self.start_date), ('payment_date', '<=', self.end_date)]

        return_val = []
        payment_ids = self.env['account.payment'].sudo().search(domain, order='id')
        for pay in payment_ids:
            return_val.append(self.render_form_template(pay))

        if self.report_type == '01':
            return_val.sort(key=itemgetter('ngay_tao', 'so_chung_tu'), reverse=False)
        elif self.report_type == '02':
            return_val.sort(key=itemgetter('ngay_thu_tien', 'so_chung_tu'), reverse=False)
        return return_val

    def create_report_da(self):
        datas = self._get_data_report()
        # in dữ liễu
        report_brand_overview_attachment = self.env['ir.attachment'].browse(
            self.env.ref('report_sale.bao_cao_payment_attachment').id)
        decode = base64.b64decode(report_brand_overview_attachment.datas)
        wb = load_workbook(BytesIO(decode))
        ws = wb.active
        line_font = Font(name='Times New Roman', size=14)

        ws['H3'].value = self.start_date.strftime('%d/%m/%Y')
        ws['J3'].value = self.end_datetime.strftime('%d/%m/%Y')
        # ws['B6'].value = dict(self._fields['option'].selection).get(self.option)
        key_col = list(range(1, len(KEY_LIST) + 1))
        format_currency = '_(* #,##0_);_(* (#,##0);_(* "-"??_);_(@_)'
        row = 7
        for data in datas:
            for col, k in zip(key_col, KEY_LIST):
                cell = ws.cell(row, col)
                cell.value = data[k]
                cell.font = line_font
                cell.border = all_border_thin
                cell.alignment = Alignment(horizontal='left', vertical='center')
                if col == 9:
                    cell.number_format = format_currency
            row += 1
        fp = BytesIO()
        wb.save(fp)
        fp.seek(0)
        report = base64.encodebytes((fp.read()))
        fp.close()
        attachment = self.env['ir.attachment'].sudo().create({
            'name': 'bao_cao_payment.xlsx',
            'datas': report,
            'res_model': 'temp.creation',
            'public': True,
        })
        url = "/web/content/?model=ir.attachment&id=%s&filename_field=name&field=datas&download=true" \
              % attachment.id
        return {'name': 'BÁO CÁO PAYMENT',
                'type': 'ir.actions.act_url',
                'url': url,
                'target': 'self',
                }
