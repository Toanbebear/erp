from odoo import fields, api, models, _
from odoo.exceptions import ValidationError, UserError
from datetime import date, datetime, time, timedelta
from calendar import monthrange
from openpyxl import load_workbook
from openpyxl.styles import Font, borders, Alignment
import base64
from io import BytesIO
from dateutil.relativedelta import relativedelta
from pytz import timezone, utc
from collections import defaultdict

thin = borders.Side(style='thin')
double = borders.Side(style='double')
all_border_thin = borders.Border(thin, thin, thin, thin)

TEMPLATE_BK = [
    'Ngày',
    'Mã Booking',
    'Mã khách hàng',
    'Tên khách hàng',
    'Mã phiếu khám',
    'Mã phiếu chuyên khoa',
    'Dịch vụ',
    'Đơn hàng',
    'Số lượng',
    'Đơn vị xử lý',
    'Ngày giờ bắt đầu',
    'Ngày giờ kết thúc',
    'Doanh thu thực hiện dịch vụ đã làm',
    'Phòng thực hiện',
    'Đơn vị',
]
TEMPLATE = [
    'ngay',
    'ma_booking',
    'ma_khach',
    'ten_khach',
    'ma_phieu_kham',
    'ma_phieu_chuyen_khoa',
    'dich_vu',
    'don_hang',
    'so_luong_xu_ly',
    'don_vi_xu_ly',
    'ngay_bat_dau',
    'ngay_ket_thuc',
    'doanh_thu_thuc_hien',
    'khoa_phong',
    'don_vi_thuc_hien',

]


class ServicePerformanceAccountReport(models.TransientModel):
    _name = 'service.performance.report'
    _description = 'Báo cáo doanh thu thực hiện toàn bộ dịch vụ'

    start_date = fields.Date('Start date', default=date.today())
    end_date = fields.Date('End date', default=date.today())
    company_ids = fields.Many2many(string='Công ty', comodel_name='res.company',
                                   domain=lambda self: [('id', 'in', self.env.user.company_ids.ids)])

    # convert date to datetime for search domain, should be removed if using datetime directly
    start_datetime = fields.Datetime('Start datetime', compute='_compute_datetime')
    end_datetime = fields.Datetime('End datetime', compute='_compute_datetime')

    @api.depends('start_date', 'end_date')
    def _compute_datetime(self):
        self.start_datetime = False
        self.end_datetime = False
        if self.start_date and self.end_date:
            local_tz = timezone(self.env.user.tz or 'Etc/GMT+7')
            start_datetime = datetime(self.start_date.year, self.start_date.month, self.start_date.day, 0, 0, 0)
            end_datetime = datetime(self.end_date.year, self.end_date.month, self.end_date.day, 23, 59, 59)
            start_datetime = local_tz.localize(start_datetime, is_dst=None)
            end_datetime = local_tz.localize(end_datetime, is_dst=None)
            self.start_datetime = start_datetime.astimezone(utc).replace(tzinfo=None)
            self.end_datetime = end_datetime.astimezone(utc).replace(tzinfo=None)

    @api.onchange('start_date')
    def _onchange_start_date(self):
        if self.start_date:
            if self.start_date.month == fields.date.today().month:
                self.end_date = fields.date.today()
            else:
                self.end_date = date(self.start_date.year, self.start_date.month,
                                     monthrange(self.start_date.year, self.start_date.month)[1])

    @api.constrains('start_date', 'end_date')
    def check_dates(self):
        for record in self:
            start_date = fields.Date.from_string(record.start_date)
            end_date = fields.Date.from_string(record.end_date)
            if start_date > end_date:
                raise ValidationError(
                    _("Ngày kết thúc không thể ở trước ngày bắt đầu khi xuất báo cáo!!!"))

    def create_report_service_perform_report(self):
        result_data = self._get_data()
        report_brand_overview_attachment = self.env['ir.attachment'].browse(
            self.env.ref('report_sale.bao_cao_ke_toan_doanh_thu_thuc_hien_dich_vu_attachment').id)
        decode = base64.b64decode(report_brand_overview_attachment.datas)
        wb = load_workbook(BytesIO(decode))
        ws = wb.active
        line_font = Font(name='Times New Roman', size=14)

        ws['G4'].value = ws['G4'].value + self.start_date.strftime('%d/%m/%Y')
        ws['I4'].value = ws['I4'].value + self.end_datetime.strftime('%d/%m/%Y')
        ws['H5'].value = '; '.join((company.name for company in self.company_ids))
        format_currency = '_(* #,##0_);_(* (#,##0);_(* "-"??_);_(@_)'
        key_col = list(range(1, len(TEMPLATE) + 1))
        row = 8
        for data in result_data:
            for col, k in zip(key_col, TEMPLATE):
                cell = ws.cell(row, col)
                cell.value = data[k]
                cell.font = line_font
                cell.border = all_border_thin
                cell.alignment = Alignment(horizontal='left', vertical='center')
                if col == 13:
                    cell.number_format = format_currency
            row += 1
        fp = BytesIO()
        wb.save(fp)
        fp.seek(0)
        report = base64.encodebytes((fp.read()))
        fp.close()
        attachment = self.env['ir.attachment'].sudo().create({
            'name': 'bao_cao_ke_toan_doanh_thu_thuc_hien_dich_vu_%s.xlsx' % datetime.now().strftime('%d_%m_%Y'),
            'datas': report,
            'res_model': 'temp.creation',
            'public': True,
        })
        url = "/web/content/?model=ir.attachment&id=%s&filename_field=name&field=datas&download=true" \
              % attachment.id
        return {'name': 'BÁO CÁO DOANH THU THỰC HIỆN DỊCH VỤ TOÀN BỘ DỊCH VỤ',
                'type': 'ir.actions.act_url',
                'url': url,
                'target': 'self',
                }

    def create_report_service_perform_report_bk(self):
        # get data
        result_data = []
        # Từ SO theo trạng thái "Đơn bán hàng" và "Ngày đặt hàng" nằm trong khoảng thời gian xuất báo cáo.
        domain = [('date_order', '>=', self.start_datetime),
                  ('date_order', '<=', self.end_datetime),
                  ('booking_id', '!=', None),
                  ('company_id', 'in', self.company_ids.ids),
                  ('pricelist_id.type', '=', 'service'),
                  ('state', '=', 'sale')]
        sale_order = self.env['sale.order'].search(domain, order='date_order')
        list_booking = [so.booking_id.id for so in sale_order]
        # Phiếu chuyên khoa
        specialty = self.env['sh.medical.specialty'].search_read([('booking_id', 'in', list_booking)], ['walkin'])
        surgery = self.env['sh.medical.surgery'].search_read([('booking_id', 'in', list_booking)], ['walkin'])
        result_dic = defaultdict(list)
        Walkin = self.env['sh.medical.appointment.register.walkin']
        for element in specialty:
            result_dic[Walkin.browse(element.get('walkin')[0]).sale_order_id].append(
                self.env['sh.medical.specialty'].browse(element.get('id'))
            )

        for element in surgery:
            result_dic[Walkin.browse(element.get('walkin')[0]).sale_order_id].append(
                self.env['sh.medical.surgery'].browse(element.get('id'))
            )

        for so, info in result_dic.items():
            booking = so.booking_id
            medical = info[0]
            ngay = so.date_order + timedelta(hours=7)
            ngay_gio_bat_dau = medical.surgery_date if 'PTTT' in medical.name else medical.services_date
            ngay_gio_ket_thuc = medical.surgery_end_date if 'PTTT' in medical.name else medical.services_end_date
            phong_thuc_hien = medical.operating_room.name if 'PTTT' in medical.name else medical.perform_room.name
            order_lines = so.order_line
            for line in order_lines:
                vals = {
                    'Ngày': '%s' % fields.Datetime.context_timestamp(self.with_context(tz=self.env.user.tz),
                                                                     ngay).strftime('%d/%m/%Y'),
                    'Mã Booking': booking.name,
                    'Mã khách hàng': so.code_customer,
                    'Tên khách hàng': so.partner_id.name,
                    'Mã phiếu khám': medical.walkin.name,
                    'Mã phiếu chuyên khoa': medical.name,
                    'Dịch vụ': line.product_id.name,
                    'Đơn hàng': so.name,
                    'Số lượng': line.product_uom_qty,
                    'Đơn vị xử lý': line.uom_price,
                    'Ngày giờ bắt đầu': '%s' % fields.Datetime.context_timestamp(self.with_context(tz=self.env.user.tz),
                                                                                 ngay_gio_bat_dau).strftime('%d/%m/%Y'),
                    'Ngày giờ kết thúc': '%s' % fields.Datetime.context_timestamp(
                        self.with_context(tz=self.env.user.tz), ngay_gio_ket_thuc).strftime('%d/%m/%Y'),
                    'Doanh thu thực hiện dịch vụ đã làm': line.price_subtotal,
                    'Phòng thực hiện': phong_thuc_hien,
                    'Đơn vị': medical.institution.name,
                }
                result_data.append(vals)
        # in dữ liễu
        raise ValidationError('dừng')
        report_brand_overview_attachment = self.env['ir.attachment'].browse(
            self.env.ref('report_sale.bao_cao_ke_toan_doanh_thu_thuc_hien_dich_vu_attachment').id)
        decode = base64.b64decode(report_brand_overview_attachment.datas)
        wb = load_workbook(BytesIO(decode))
        ws = wb.active
        line_font = Font(name='Times New Roman', size=14)

        ws['G4'].value = ws['G4'].value + self.start_date.strftime('%d/%m/%Y')
        ws['I4'].value = ws['I4'].value + self.end_datetime.strftime('%d/%m/%Y')
        ws['H5'].value = '; '.join((company.name for company in self.company_ids))
        format_currency = '_(* #,##0_);_(* (#,##0);_(* "-"??_);_(@_)'
        key_col = list(range(1, len(TEMPLATE) + 1))
        row = 8
        for data in result_data:
            for col, k in zip(key_col, TEMPLATE):
                cell = ws.cell(row, col)
                cell.value = data[k]
                cell.font = line_font
                cell.border = all_border_thin
                cell.alignment = Alignment(horizontal='left', vertical='center')
                if col == 13:
                    cell.number_format = format_currency
            row += 1
        fp = BytesIO()
        wb.save(fp)
        fp.seek(0)
        report = base64.encodebytes((fp.read()))
        fp.close()
        attachment = self.env['ir.attachment'].sudo().create({
            'name': 'bao_cao_ke_toan_doanh_thu_thuc_hien_dich_vu_%s.xlsx' % datetime.now().strftime('%d_%m_%Y'),
            'datas': report,
            'res_model': 'temp.creation',
            'public': True,
        })
        url = "/web/content/?model=ir.attachment&id=%s&filename_field=name&field=datas&download=true" \
              % attachment.id
        return {'name': 'BÁO CÁO DOANH THU THỰC HIỆN DỊCH VỤ TOÀN BỘ DỊCH VỤ',
                'type': 'ir.actions.act_url',
                'url': url,
                'target': 'self',
                }

    def _get_data(self):
        query = """
            with surgery_specialty as (
            select surg_serv.service_id,surgery.walkin,surgery.name form,surgery.surgery_date service_start_date,surgery.surgery_end_date service_end_date, smhco."name" room
            from sh_surgery_service_rel surg_serv
            left join sh_medical_surgery surgery on surg_serv.surgery_id = surgery.id 
            left join sh_medical_health_center_ot smhco on smhco.id = surgery.operating_room 
            union all
            select spec_serv.service_id,specialty.walkin, specialty.name form,specialty.services_date service_start_date,specialty.services_end_date service_end_date, smhco.name room
            from sh_specialty_service_rel spec_serv
            left join sh_medical_specialty specialty on spec_serv.specialty_id = specialty.id
            left join sh_medical_health_center_ot smhco on smhco.id = specialty.perform_room)
            select TO_CHAR(so.date_order + INTERVAL '7 hours', 'DD/MM/YYYY HH24:MI:SS') as ngay,
            cl."name" as ma_booking,
            rp.code_customer as ma_khach,
            rp.name as ten_khach,
            smarw.name as ma_phieu_kham,
            ss.form as ma_phieu_chuyen_khoa,
            pt."name" as dich_vu,
            so.name as don_hang,
            sol.product_uom_qty as so_luong_xu_ly,
            sol.uom_price as don_vi_xu_ly,
            TO_CHAR(ss.service_start_date + INTERVAL '7 hours', 'DD/MM/YYYY HH24:MI:SS') as ngay_bat_dau,
            TO_CHAR(ss.service_end_date + INTERVAL '7 hours', 'DD/MM/YYYY HH24:MI:SS') as ngay_ket_thuc,
            sol.price_subtotal as doanh_thu_thuc_hien,
            ss.room as khoa_phong,
            rc.name as don_vi_thuc_hien
            from sale_order_line sol 
            left join sale_order so on so.id = sol.order_id 
            left join res_partner rp on rp.id = so.partner_id 
            left join crm_lead cl on cl.id = so.booking_id  
            left join product_product pp on pp.id = sol.product_id 
            left join product_template pt on pt.id = pp.product_tmpl_id 
            left join res_company rc on rc.id = so.company_id 
            left join product_pricelist as price on so.pricelist_id = price.id
            left join sh_medical_appointment_register_walkin smarw on smarw.sale_order_id = so.id 
            left join ( select surgery_specialty.*,
            smhcs.product_id 
            from surgery_specialty
            left join sh_medical_health_center_service smhcs on surgery_specialty.service_id = smhcs.id
            ) ss on ss.walkin = smarw.id and ss.product_id = pp.id
            where price.type != 'product' 
                and so.state in ('sale','done') 
                and so.company_id in %s 
                and so.date_order >= %s
                and so.date_order <= %s
            order by so.date_order asc;
        """
        start_datetime = datetime(year=(self.start_date - timedelta(days=1)).year, month=(self.start_date - timedelta(days=1)).month, day=(self.start_date - timedelta(days=1)).day, hour=17, minute=0, second=0)
        end_datetime = datetime(year=self.end_date.year, month=self.end_date.month, day=self.end_date.day, hour=17, minute=0, second=0)
        self.env.cr.execute(query, (tuple(self.company_ids.ids), start_datetime, end_datetime))
        query_result = self.env.cr.dictfetchall()
        return query_result
