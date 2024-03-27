from odoo import fields, api, models, _
from odoo.exceptions import ValidationError, UserError
from datetime import date, datetime, time
from calendar import monthrange
import openpyxl
from openpyxl.drawing.image import Image
from openpyxl import load_workbook
from openpyxl.styles import Font, borders, Alignment, PatternFill
from odoo.addons.report_sale.wizard.theme_report import ThemeReport
import base64
from io import BytesIO
from dateutil.relativedelta import relativedelta
from pytz import timezone, utc
from odoo.modules.module import get_module_resource


class SpecialTreatmentReport(models.TransientModel):
    _name = 'crm.hh.ehc.report.deposit'
    _description = 'Báo cáo doanh số khách hàng đặt cọc'

    start_date = fields.Date('Start date', default=date.today())
    end_date = fields.Date('End date', default=date.today())

    # convert date to datetime for search domain, should be removed if using datetime directly
    start_datetime = fields.Datetime('Start datetime', compute='_compute_datetime')
    end_datetime = fields.Datetime('End datetime', compute='_compute_datetime')

    @api.depends('start_date', 'end_date')
    def _compute_datetime(self):
        self.start_datetime = False
        self.end_datetime = False
        if self.start_date and self.end_date:
            local_tz = timezone(self.env.user.tz or 'Etc/GMT+7')
            start_datetime = datetime(self.start_date.year, self.start_date.month, self.start_date.day, 0, 0, 0)
            end_datetime = datetime(self.end_date.year, self.end_date.month, self.end_date.day, 23, 59, 59)
            start_datetime = local_tz.localize(start_datetime, is_dst=None)
            end_datetime = local_tz.localize(end_datetime, is_dst=None)
            self.start_datetime = start_datetime.astimezone(utc).replace(tzinfo=None)
            self.end_datetime = end_datetime.astimezone(utc).replace(tzinfo=None)

    @api.onchange('start_date')
    def _onchange_start_date(self):
        if self.start_date:
            if self.start_date.month == fields.date.today().month:
                self.end_date = fields.date.today()
            else:
                self.end_date = date(self.start_date.year, self.start_date.month,
                                     monthrange(self.start_date.year, self.start_date.month)[1])

    @api.constrains('start_date', 'end_date')
    def check_dates(self):
        for record in self:
            start_date = fields.Date.from_string(record.start_date)
            end_date = fields.Date.from_string(record.end_date)
            if start_date > end_date:
                raise ValidationError(
                    _("Ngày kết thúc không thể ở trước ngày bắt đầu khi xuất báo cáo!!!"))

    def create_report_bao_cao_doanh_so_khach_hang_dat_coc(self):
        # get data

        template = self.env['ir.attachment'].browse(self.env.ref('report_sale.report_ehc_doanh_so_kh_dat_coc_attachment').id)
        decode = base64.b64decode(template.datas)
        wb = load_workbook(BytesIO(decode))
        ws = wb.active
        thin = borders.Side(style='thin')
        all_border_thin = borders.Border(left=thin, right=thin, top=thin, bottom=thin)
        line_font = Font(name='Times New Roman', size=13)

        # in tiêu đề và giá trị
        # key_list = [
        #     'code_booking',
        #     'name_customer',
        #     'category_source',
        #     'communication',
        #     'payment_money',
        #     'type_payment',
        #     'state_payment',
        # ]
        key_list_title = [
            'STT',
            'Mã Booking',
            'Họ tên khách hàng',
            'Nhóm nguồn',
            'Nội dung',
            'Số tiền',
            'Phương thức thanh toán',
            'Trạng thái Payment',
            'Ngày thanh toán'
        ]
        select = """ 
                            SELECT 
                                cl.name,
                                cl.name,
                                cl.contact_name,
                                ccs.name,
                                ap.communication,
                                ap.amount,
                                ap.payment_method,
                                ap.state,
                                ap.payment_date
                            FROM 
                                account_payment ap
                            LEFT JOIN crm_lead cl ON cl.id = ap.crm_id  
                            LEFT JOIN res_company rc ON rc.id = ap.company_id
                            LEFT JOIN crm_category_source ccs ON ccs.id = cl.category_source_id
                            WHERE 
                                (ap.payment_date between %s and %s) and
                                rc.id = 24 and ap.is_deposit = 'true'
                        """
        self.env.cr.execute(select,
                            [self.start_datetime.strftime('%Y-%m-%d %H:%M:%S'),
                             self.end_datetime.strftime('%Y-%m-%d %H:%M:%S')
                             ])
        datas = self.env.cr.fetchall()
        if not datas:
            context = dict(self._context or {})
            context['message'] = 'Không tồn tại dữ liệu từ ngày: %s tới ngày: %s' % (
                self.start_date.strftime('%d/%m/%Y'), self.end_date.strftime('%d/%m/%Y'))
            return {
                'name': _('Thông báo'),  # label
                'type': 'ir.actions.act_window',
                'view_mode': 'form',
                'view_id': self.env.ref('sh_message.sh_message_wizard').id,
                'res_model': 'sh.message.wizard',  # model want to display
                'target': 'new',  # if you want popup
                'context': context,
            }
        row = 3
        stt = 1
        format_currency = '_(* #,##0_);_(* (#,##0);_(* "-"??_);_(@_)'

        for data in datas:
            col = 1
            for index in range(0, len(key_list_title)):
                cell = ws.cell(row, col)
                if index == 0:
                    cell.value = stt
                elif index == 8:
                    cell.value = data[index].strftime('%d-%m-%Y')
                elif index == 5:
                    cell.value = data[index]
                    cell.number_format = format_currency
                elif index == 6:
                    if data[index] == 'tm':
                        cell.value = 'Tiền mặt'
                    elif data[index] == 'ck':
                        cell.value = 'Chuyển khoản'
                    elif data[index] == 'nb':
                        cell.value = 'Thanh tóán nội bộ'
                    elif data[index] == 'ck':
                        cell.value = 'Quẹt thẻ qua POS'
                    elif data[index] == 'ck':
                        cell.value = 'Thanh toán qua ví điện tử'
                elif index == 7:
                    if data[index] == 'draft':
                        cell.value = 'Nháp'
                    elif data[index] == 'posted':
                        cell.value = 'Đã xác nhận'
                    elif data[index] == 'sent':
                        cell.value = 'Đã gửi'
                    elif data[index] == 'reconciled':
                        cell.value = 'Đã được đối soát'
                    elif data[index] == 'cancelled':
                        cell.value = 'Đã hủy'
                else:
                    if not data[index]:
                        cell.value = '-'
                    else:
                        cell.value = data[index]
                cell.font = Font(name='Times New Roman', size=13)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = all_border_thin
                col += 1
            stt += 1
            row += 1
        fp = BytesIO()
        wb.save(fp)
        fp.seek(0)
        report = base64.encodebytes((fp.read()))
        fp.close()

        attachment = self.env['ir.attachment'].sudo().create({
            'name': 'bao_cao_doanh_so_khach_hang_dat_coc.xlsx',
            'datas': report,
            'res_model': 'crm.hh.ehc.report.deposit',
            'public': True,
        })

        url = "/web/content/?model=ir.attachment&id=%s&filename_field=name&field=datas&download=true" \
              % attachment.id

        return {'name': 'Báo cáo doanh số khách hàng đặt cọc',
                'type': 'ir.actions.act_url',
                'url': url,
                'target': 'self',
                }