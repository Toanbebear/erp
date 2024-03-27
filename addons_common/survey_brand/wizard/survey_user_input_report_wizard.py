# -*- coding: utf-8 -*-
import base64
from calendar import monthrange
from datetime import date, datetime, timedelta
from io import BytesIO

from pytz import utc, timezone

from odoo import fields, models, api, _
from odoo.exceptions import ValidationError
from openpyxl import load_workbook


class SurveyUserInputWizard(models.TransientModel):
    _name = 'survey.user.input.report.wizard'
    _description = 'Báo cáo khảo sát'

    start_date = fields.Date('Start date', default=date.today().replace(day=1), required=True)
    end_date = fields.Date('End date', default=date.today(), required=True)
    brand_id = fields.Many2one('res.brand')
    start_datetime = fields.Datetime('Start datetime', compute='_compute_datetime')
    end_datetime = fields.Datetime('End datetime', compute='_compute_datetime')
    company_ids = fields.Many2many('res.company', string='Chi nhánh',
                                   domain="[('brand_id', '=', brand_id)]", required=True)
    stage = fields.Selection([('new', 'Chưa khởi động'), ('skip', 'Hoàn thiện một phần'), ('done', 'Đã hoàn thành'), ('all', 'Tất cả'), ('all_2', 'Tất cả (Hoàn thành + Hoàn thiện 1 phần)')], string = 'Trạng thái khảo sát', required=True, default='all')
    type = fields.Selection([('auto', 'Khiếu nại'), ('manually', 'Khách hàng trà lời'), ('all', 'Tất cả')], string='Phân loại khảo sát', required=True, default='all')
    survey_id = fields.Many2one('survey.survey', required=True)
    fetch_all = fields.Boolean(string='Tất cả công ty', default=True)

    @api.onchange('fetch_all', 'survey_id')
    def _onchange_companies(self):
        if self.fetch_all:
            if self.brand_id:
                companies = self.env['res.company'].search([('brand_id', '=', self.brand_id.id), ('active', '=', True)])
                self.company_ids = companies.ids
        else:
            self.company_ids = None

    @api.onchange('start_date')
    def _onchange_start_date(self):
        if self.start_date:
            if self.start_date.month == fields.date.today().month:
                self.end_date = fields.date.today()
            else:
                self.end_date = date(self.start_date.year, self.start_date.month,
                                     monthrange(self.start_date.year, self.start_date.month)[1])

    @api.constrains('start_date', 'end_date')
    def check_dates(self):
        for record in self:
            start_date = fields.Date.from_string(record.start_date)
            end_date = fields.Date.from_string(record.end_date)

            if start_date > end_date:
                raise ValidationError(
                    _("Ngày kết thúc phải sau ngày bắt đầu."))

    @api.depends('start_date', 'end_date')
    def _compute_datetime(self):
        self.start_datetime = False
        self.end_datetime = False
        if self.start_date and self.end_date:
            local_tz = timezone(self.env.user.tz or 'Etc/GMT+7')
            start_datetime = datetime(self.start_date.year, self.start_date.month, self.start_date.day, 0, 0, 0)
            end_datetime = datetime(self.end_date.year, self.end_date.month, self.end_date.day, 23, 59, 59)
            start_datetime = local_tz.localize(start_datetime, is_dst=None)
            end_datetime = local_tz.localize(end_datetime, is_dst=None)
            self.start_datetime = start_datetime.astimezone(utc).replace(tzinfo=None)
            self.end_datetime = end_datetime.astimezone(utc).replace(tzinfo=None)

    def _get_columns(self):
        # Lấy ra tất cả câu trả lời của survey_id
        cols = []
        index = 1
        question_answers_dict = {}
        answers_question_dict = {}
        question_text_dict = {}
        for question in self.survey_id.question_ids:
            if question.question_type in ['simple_choice', 'multiple_choice', 'textbox', 'free_text', 'numerical_box', 'date', 'datetime']:
                col = "%s.%s" % (index, question.title)
                cols.append(col)
                question_answers_dict[col] = "%s" % question.id
                if question.comments_allowed:
                    col_text = "%s.%s" % (index, 'Khác')
                    cols.append(col_text)
                    question_answers_dict[col_text] = "%s.text" % question.id
                    question_text_dict[question.id] = question.id
                if question.has_note:
                    col_note = "%s.%s" % (index, 'Lí do khách hàng lựa chọn')
                    cols.append(col_note)
                    question_answers_dict[col_note] = "%s.note" % question.id
                    question_text_dict[question.id] = question.id
                answers_question_dict["%s" % question.id] = col
            elif question.question_type == 'matrix':
                # cols.append("%s.%s" % (index, question.title))
                # print('Xử lý kiểu gì đây')
                columns = question.labels_ids
                rows = question.labels_ids_2
                for row in rows:
                    col = "%s.%s_%s" % (index,question.title, row.value)
                    cols.append(col)
                    question_answers_dict[col] = "%s_%s" % (question.id, row.id)
                        # answers_question_dict["%s_%s_%s" % (question.id, column.id, row.id)] = col
            # elif question.question_type in ['simple_choice', 'multiple_choice']:
            #     answers = question.labels_ids
            #     for answer in answers:
            #         col = "%s.%s_%s" % (index, question.title, answer.value)
            #         cols.append(col)
            #         question_answers_dict[col] = "%s_%s" % (question.id, answer.id)
            #         answers_question_dict["%s_%s" % (question.id, answer.id)] = col
            else:
                print('Khỏi xử lý')
            index += 1
        return cols, question_answers_dict, answers_question_dict, question_text_dict

    def _get_data(self, fix_cols, answer_columns, question_answers_dict, question_text_dict):

        list_type = ('auto', 'manually')
        if self.type != 'all':
            list_type = (self.type,)
        list_stage = ('new', 'skip', 'done')
        if self.stage not in ['all', 'all_2']:
            list_stage = (self.stage,)
        elif self.stage == 'all_2':
            list_stage = ('skip', 'done')
        select = """ 
                    SELECT 
                        sui.create_date timestamp1,
                        rb.name as brand,
                        rc.name as branch,
                        cl.name as booking,
                        rp.name as name,
                        rp.phone as phone,
                        sui.create_date timestamp1,
                        pc.name as group_service,
                        sui.service_type,
                        sst.name,
                        sui.state,
                        sui.score_user_input,
                        sui.id
                    FROM 
                        survey_user_input sui
                    LEFT JOIN res_partner rp ON rp.id = sui.partner_id  
                    LEFT JOIN res_brand rb ON rb.id = sui.brand_id  
                    LEFT JOIN res_company rc ON rc.id = sui.company_id  
                    LEFT JOIN crm_lead cl ON cl.id = sui.crm_id  
                    LEFT JOIN survey_survey_type sst ON sst.id = sui.survey_time_id  
                    LEFT JOIN sh_medical_health_center_service_category csc ON csc.id = sui.group_service_id  
                    LEFT JOIN product_category pc ON pc.id = csc.product_cat_id
                    WHERE 
                        (sui.create_date between %s and %s) and
                        sui.survey_id = %s and sui.company_id in %s and sui.brand_id = %s and
                        sui.state in %s and sui.input_type in %s
                """
        self.env.cr.execute(select,
                            [self.start_datetime.strftime('%Y-%m-%d %H:%M:%S'),
                             self.end_datetime.strftime('%Y-%m-%d %H:%M:%S'),
                             self.survey_id.id,
                             tuple(self.company_ids.mapped('id')),
                             self.brand_id.id,
                             list_stage,
                             list_type])
        datas = self.env.cr.fetchall()

        # Lấy tất cả id của user_input cho vào 1 mảng để query tất cả
        user_input_ids = []
        for data in datas:
            # Id của user_input
            user_input_ids.append(data[12])

        user_input_lines = []
        if user_input_ids:
            query_input_line = """ 
                                SELECT 
                                    sul.question_id,
                                    sul.value_suggested,
                                    sul.value_suggested_row,
                                    sul.value_free_text,
                                    sul.answer_type,
                                    sul.answer_value,
                                    sul.value_text,
                                    sul.user_input_id,
                                    sul.value_datetime,
                                    sul.value_date,
                                    sul.value_comment,
                                    sl.value
                                FROM 
                                    survey_user_input_line sul
                                Left join survey_label sl on sl.id = sul.value_suggested
                                WHERE 
                                    sul.user_input_id in %s
                            """
            params = [tuple(user_input_ids)]
            self.env.cr.execute(query_input_line, params)
            user_input_lines = self.env.cr.fetchall()

        user_input_ids_dict = {}
        for line in user_input_lines:
            # Kiểm tra kết quả khảo sát có trong dict chưa? user_input_id
            if line[7] in user_input_ids_dict:
                # answer_type
                if line[4] == 'suggestion':
                    if line[1]:
                        if line[2]:
                            key2 = "%s_%s" % (line[0], line[2])
                            user_input_ids_dict[line[7]][key2] = line[11]
                        else:
                            key1 = "%s" % (line[0])
                            if key1 in user_input_ids_dict[line[7]]:
                                val = user_input_ids_dict[line[7]][key1]
                                if line[5]:
                                    user_input_ids_dict[line[7]][key1] = "%s,%s" % (val, line[5])
                            else:
                                user_input_ids_dict[line[7]][key1] = line[5]
                        if line[10]:
                            key_note = "%s.note" % (line[0])
                            if key_note in user_input_ids_dict[line[7]]:
                                val_note = user_input_ids_dict[line[7]][key_note]
                                user_input_ids_dict[line[7]][key_note] = "%s,%s" % (val_note, line[10])
                            else:
                                user_input_ids_dict[line[7]][key_note] = line[10]
                elif line[4] == 'text':
                    # Tìm điều kiện để tạo 1 key mới
                    if line[0] in question_text_dict:
                        key_text = "%s.text" % (line[0])
                        user_input_ids_dict[line[7]][key_text] = line[6]
                    else:
                        key0 = "%s" % (line[0])
                        user_input_ids_dict[line[7]][key0] = line[6]
                elif line[4] == 'free_text':
                    key_f0 = "%s" % (line[0])
                    user_input_ids_dict[line[7]][key_f0] = line[3]
                elif line[4] == 'datetime':
                    key_datetime = "%s" % (line[0])
                    user_input_ids_dict[line[7]][key_datetime] = line[8]
                elif line[4] == 'date':
                    key_date = "%s" % (line[0])
                    user_input_ids_dict[line[7]][key_date] = line[9]
            # Chưa có câu hỏi trong dict thì thêm mới
            else:
                user_input_line_dict = {}
                if line[4] == 'suggestion':
                    if line[1]:
                        if line[2]:
                            user_input_line_dict["%s_%s" % (line[0], line[2])] = line[11]
                        else:
                            user_input_line_dict["%s" % (line[0])] = line[5]
                        if line[10]:
                            key_note = "%s.note" % (line[0])
                            user_input_line_dict[key_note] = line[10]
                elif line[4] == 'text':
                    # Tìm điều kiện để tạo 1 key mới
                    if line[0] in question_text_dict:
                        key_text = "%s.text" % (line[0])
                        user_input_line_dict[key_text] = line[6]
                    else:
                        user_input_line_dict["%s" % (line[0])] = line[6]
                elif line[4] == 'free_text':
                    user_input_line_dict["%s" % (line[0])] = line[3]
                elif line[4] == 'datetime':
                    user_input_line_dict["%s" % (line[0])] = line[8]
                elif line[4] == 'date':
                    user_input_line_dict["%s" % (line[0])] = line[9]

                # if line[0]:
                #     user_input_lines_dict[line[0]] = user_input_line_dict
                user_input_ids_dict[line[7]] = user_input_line_dict

        # Xử lý kết quả
        results = []
        for data in datas:
            result = []
            # Những cột fix cứng
            length_fix_cols = len(fix_cols)
            for i in range(length_fix_cols):
                result.append(data[i])

            user_input_lines = {}

            if data[12] in user_input_ids_dict:
                user_input_lines = user_input_ids_dict[data[12]]

            # Thêm giá trị cột tùy biến
            for answer_column in answer_columns:
                if answer_column in question_answers_dict:
                    k = question_answers_dict[answer_column]
                    if k in user_input_lines:
                        result.append(user_input_lines[k])
                    else:
                        result.append('')
            results.append(result)
        return results

    def report_survey(self):
        answer_columns, question_answers_dict, answers_question_dict, question_text_dict = self._get_columns()
        # Tính toán số lượng cột
        fix_cols = [
            'STT',
            'Thương hiệu',
            'Chi nhánh (Tên cơ sở)',
            'Booking',
            'Tên khách hàng',
            'Số điện thoại',
            'Ngày khảo sát',
            'Nhóm dịch vụ',
            'Loại dịch vụ',
            'Phân loại khảo sát (Tiêu chí)',
            'Tiến độ',
            'Điểm quy đổi'
        ]
        columns = fix_cols + answer_columns
        datas = self._get_data(fix_cols, answer_columns, question_answers_dict, question_text_dict)
        if not datas:
            context = dict(self._context or {})
            context['message'] = 'Không tồn tại dữ liệu từ ngày: %s tới ngày: %s' % (
                self.start_date.strftime('%d/%m/%Y'), self.end_date.strftime('%d/%m/%Y'))
            return {
                'name': _('Thông báo'),  # label
                'type': 'ir.actions.act_window',
                'view_mode': 'form',
                'view_id': self.env.ref('sh_message.sh_message_wizard').id,
                'res_model': 'sh.message.wizard',  # model want to display
                'target': 'new',  # if you want popup
                'context': context,
            }

        # Khởi tạo đối tượng làm việc worksheet
        template = self.env['ir.attachment'].browse(
            self.env.ref('survey_brand.survey_user_input_report_wizard_attachment').id)
        decode = base64.b64decode(template.datas)
        wb = load_workbook(BytesIO(decode))
        ws = wb.active

        # line_font = Font(name='Times New Roman', size=14)

        # ws['H3'].value = self.start_date.strftime('%d/%m/%Y')
        # ws['J3'].value = self.end_datetime.strftime('%d/%m/%Y')
        # ws['H4'].value = self.company_id.name
        col_index = 1
        row = 1
        for c in columns:
            cell = ws.cell(row, col_index)
            cell.value = c
            col_index += 1
        row = 2

        for data in datas:
            # print("datadatadatadata")
            # for col, k in zip(key_col_list, key_list):
            col = 1
            length = len(columns)
            for index in range(0, length):
                # print("col")
                # print(column_name)
                # beforeCell = ws.cell(6, col)
                # beforeCell.fill = header_fill
                # beforeCell.font = Font(name='Times New Roman', size=14, color='FFFFFF')
                cell = ws.cell(row, col)
                if index == 6:
                    cell.value = fields.Datetime.context_timestamp(self.with_context(tz=self.env.user.tz),data[index]).strftime('%d/%m/%Y %H:%M:%S')
                elif index == 5:
                    if data[index] and len(data[index]) > 7:
                        cell.value = data[index][0:3] + 'xxxx' + data[index][7:]
                    else:
                        cell.value = ''
                elif index == 10:
                    if data[index] == 'skip':
                        cell.value = 'Hoàn thiện một phần'
                    elif data[index] == 'new':
                        cell.value = 'Chưa khởi động'
                    elif data[index] == 'done':
                        cell.value = 'Đã hoàn thành'
                else:
                    cell.value = data[index]
                # cell.font = line_font
                # cell.border = all_border_thin
                # cell.alignment = Alignment(horizontal='left', vertical='center')
                col += 1
            row += 1

        stt = 0
        for i in range(2, len(datas) + 2):
            stt += 1
            cell = ws.cell(i, 1)
            cell.value = str(stt)

        fp = BytesIO()
        wb.save(fp)
        fp.seek(0)
        report = base64.encodebytes((fp.read()))
        fp.close()

        attachment = self.env['ir.attachment'].sudo().create({
            'name': 'bao_cao_ket_qua_khao_sat_%s.xlsx' % self.brand_id.name,
            'datas': report,
            'res_model': 'temp.creation',
            'public': True,
        })

        url = "/web/content/?model=ir.attachment&id=%s&filename_field=name&field=datas&download=true" \
              % attachment.id

        return {'name': 'Báo cáo kết quả khảo sát',
                'type': 'ir.actions.act_url',
                'url': url,
                'target': 'self',
                }
