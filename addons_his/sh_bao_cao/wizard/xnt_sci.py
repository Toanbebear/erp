import base64
from datetime import date, datetime
from io import BytesIO

from openpyxl import load_workbook
from openpyxl.styles import Font, borders, Alignment
from pytz import timezone

from odoo import fields, api, models
from odoo.exceptions import ValidationError

thin = borders.Side(style='thin')
thick = borders.Side(style='medium')
dotted = borders.Side(style='hair')
gray_thin = borders.Side(style='thin', color='808080')
all_border_thin = borders.Border(left=thin, right=thin, top=thin, bottom=thin)
all_border_gray = borders.Border(left=gray_thin, right=gray_thin, top=gray_thin, bottom=gray_thin)
dotted_top_bot = borders.Border(left=thin, right=thin, top=dotted, bottom=dotted)
thick_dotted_top_bot = borders.Border(left=thick, right=thick, top=dotted, bottom=dotted)
all_border_thick = borders.Border(left=thick, right=thick, top=thick, bottom=thick)
center_alm = Alignment(horizontal='center', vertical='center')


class XNTReportSci(models.TransientModel):
    _name = "bc.xnt.sci"
    _description = "Báo cáo xuất/nhập tồn SCI"

    company = fields.Many2one('res.company', default=1)
    start_date = fields.Date('Ngày bắt đầu', default=date.today().replace(day=1))
    end_date = fields.Date('Ngày kết thúc', default=date.today())
    start_datetime = fields.Datetime('Start datetime', compute='_compute_datetime')
    end_datetime = fields.Datetime('End datetime', compute='_compute_datetime')
    preview = fields.Boolean('Bạn có muốn xem trước ?')
    location = fields.Many2one('stock.location', string='Kho')

    @api.constrains('end_date')
    def _constraint_end_datetime(self):
        now = date.today()
        if self.end_date > now:
            raise ValidationError('Ngày kết thúc không được quá ngày hiện tại!')

    @api.depends('start_date', 'end_date')
    def _compute_datetime(self):
        self.start_datetime = False
        self.end_datetime = False
        if self.start_date and self.end_date:
            local_tz = timezone(self.env.user.tz or 'Etc/GMT+7')
            start_datetime = datetime(self.start_date.year, self.start_date.month, self.start_date.day, 0, 0, 0)
            end_datetime = datetime(self.end_date.year, self.end_date.month, self.end_date.day, 23, 59, 59)
            # start_datetime = local_tz.localize(start_datetime, is_dst=None)
            # end_datetime = local_tz.localize(end_datetime, is_dst=None)
            # self.start_datetime = start_datetime.astimezone(utc).replace(tzinfo=None)
            # self.end_datetime = end_datetime.astimezone(utc).replace(tzinfo=None)
            self.start_datetime = start_datetime
            self.end_datetime = end_datetime

    # def create_report(self):
    #     if self.env.company.id != 1:
    #         raise ValidationError('Bạn cần đứng tại chi nhánh SCI để xuất báo cáo !')
    #     template = self.env['ir.attachment'].browse(self.env.ref('sh_bao_cao.bao_cao_xnt_sci_attachment').id)
    #     decode = base64.b64decode(template.datas)
    #     wb = load_workbook(BytesIO(decode))
    #     ws = wb.active
    #     thin = borders.Side(style='thin')
    #     all_border_thin = borders.Border(left=thin, right=thin, top=thin, bottom=thin)
    #     line_font = Font(name='Times New Roman', size=13)
    #     # ws['a1'].value = 'Công ty: SCI GROUP'
    #     # ws['a1'].font = Font(name='Times New Roman', size=15, bold=True)
    #     # ws['a1'].alignment = Alignment(horizontal='left', vertical='center')
    #     ws['a2'].value = 'Kho: %s' % self.location.name
    #     ws['a2'].font = Font(name='Times New Roman', size=15, bold=True)
    #     ws['a2'].alignment = Alignment(horizontal='left', vertical='center')
    #     date_range = 'Từ ngày: %s đến ngày: %s' % (
    #         self.start_datetime.strftime('%d/%m/%Y %H:%M:%S'), self.end_datetime.strftime('%d/%m/%Y %H:%M:%S'))
    #
    #     ws['a5'].value = date_range
    #     ws['a5'].font = Font(name='Times New Roman', size=15, bold=True)
    #     ws['a5'].alignment = Alignment(horizontal='left', vertical='center')
    #     datas = self.get_data()
    #     row = 10
    #     for line in datas:
    #         for col, val in enumerate(line, 1):
    #             cell = ws.cell(row, col)
    #             cell.value = val
    #             cell.border = all_border_thin
    #             if isinstance(line[0], str):
    #                 cell.font = Font(name='Times New Roman', size=10, bold=True)
    #             else:
    #                 cell.font = Font(name='Times New Roman', size=10)
    #             if col > 4:
    #                 cell.number_format = '_(* #,##0_);_(* (#,##0);_(* "-"??_);_(@_)'
    #         row += 1
    #
    #     # sign
    #     row += 1
    #     ws.cell(row, 9).value = date.today().strftime('Ngày %d tháng %m năm %Y')
    #     ws.cell(row, 9).font = Font(name='Times New Roman', size=10, italic=True)
    #     ws.cell(row, 9).alignment = center_alm
    #
    #     row += 2
    #     ws.cell(row, 3).value = 'Kế toán'
    #     ws.cell(row, 9).value = 'Người lập'
    #     ws.cell(row + 1, 3).value = ws.cell(row + 1, 9).value = '(Ký, họ tên)'
    #     ws.cell(row, 3).font = ws.cell(row, 9).font = Font(name='Times New Roman', size=10, bold=True)
    #     ws.cell(row + 1, 3).font = ws.cell(row + 1, 9).font = Font(name='Times New Roman', size=10, italic=True)
    #     ws.cell(row, 3).alignment = ws.cell(row, 9).alignment = ws.cell(row + 1, 3).alignment = ws.cell(row + 1,
    #                                                                                                     9).alignment = center_alm
    #
    #     fp = BytesIO()
    #     wb.save(fp)
    #     fp.seek(0)
    #     report = base64.encodebytes((fp.read()))
    #     fp.close()
    #     attachment = self.env['ir.attachment'].sudo().create({
    #         'name': 'Báo cáo XNT kho SCI.xlsx',
    #         'datas': report,
    #         'res_model': 'temp.creation',
    #         'public': True,
    #     })
    #     if self.preview:
    #         return {'name': 'Báo cáo XNT SCI.xlsx',
    #                 'type': 'ir.actions.act_window',
    #                 'res_model': 'temp.wizard',
    #                 'view_mode': 'form',
    #                 'view_type': 'form',
    #                 'target': 'inline',
    #                 'view_id': self.env.ref('ms_templates.report_wizard').id,
    #                 'context': {'attachment_id': attachment.id}}
    #     else:
    #         url = "/web/content/?model=ir.attachment&id=%s&filename_field=name&field=datas&download=true" \
    #               % attachment.id
    #         return {'name': 'Báo cáo xuất/nhập tồn SCI.xlsx',
    #                 'type': 'ir.actions.act_url',
    #                 'url': url,
    #                 'target': 'self',
    #                 }

    # def get_data(self):
    #     datas = []
    #     data_sum = ['Tổng', '', '', '', 0, 0, 0, 0, 0, 0, 0, 0]
    #     Move = self.env['stock.move']
    #     Quant = self.env['stock.quant']
    #     Product = self.env['product.product']
    #     main_location = self.location
    #     supply_location = self.env.ref('stock.stock_location_suppliers')
    #     customer_location = self.env.ref('stock.stock_location_customers')
    #     false_and_return_picks = self.env['stock.picking'].search(
    #         ['&', '&', ('date_done', '>=', self.start_datetime), ('date_done', '<=', self.end_datetime),
    #          '|', ('name', 'ilike', '-FP'), ('origin', 'ilike', '-FP')])
    #     moves_in_domain = [('date', '>=', self.start_datetime), ('state', '=', 'done'),
    #                        ('picking_id', 'not in', false_and_return_picks.ids),
    #                        ('company_id', '=', self.company.id),
    #                        ('location_dest_id', 'child_of', main_location.id), '!',
    #                        ('location_id', 'child_of', main_location.id)]
    #     moves_out_domain = [('date', '>=', self.start_datetime), ('state', '=', 'done'),
    #                         ('picking_id', 'not in', false_and_return_picks.ids),
    #                         ('company_id', '=', self.company.id),
    #                         ('location_id', 'child_of', main_location.id), '!',
    #                         ('location_dest_id', 'child_of', main_location.id)]
    #     categs = self.env['sh.medical.medicines.category'].search([], order='type')
    #     index = 1
    #     i = 1
    #     for categ in categs:
    #         categ_datas = [[''] * 12, ['Nhóm hàng: %s' % categ.name, '', '', '', 0, 0, 0, 0, 0, 0, 0, 0]]
    #         meds = self.env['sh.medical.medicines'].with_context(active_test=False).search(
    #             [('medicine_category_id', '=', categ.id)])
    #         product_ids = meds.mapped('product_id.id')
    #         c_moves_in_domain = moves_in_domain + [('date', '<=', self.end_datetime), ('product_id', 'in', product_ids)]
    #         c_moves_out_domain = moves_out_domain + [('date', '<=', self.end_datetime),
    #                                                  ('product_id', 'in', product_ids)]
    #         move_in_res_new = []
    #         move_in = self.env['stock.move'].search(c_moves_in_domain, order="id asc")
    #         move_out_res_new = []
    #         move_out = self.env['stock.move'].search(c_moves_out_domain, order="id asc")
    #         quants_res_new = []
    #         quants = Quant.search([('product_id', 'in', product_ids), ('location_id', 'child_of', main_location.id)],
    #                               order="id asc")
    #         for product in product_ids:
    #             move_in_ids = move_in.filtered(lambda sm: sm.product_id.id == product)
    #             if move_in_ids:
    #                 move_in_res_new.append({
    #                     'product': product,
    #                     # 'quantity': sum(move_in_ids.mapped('stock_valuation_layer_ids').mapped('quantity')) if move_in_ids.mapped('stock_valuation_layer_ids') else sum(move_in_ids.mapped('quantity_done')),
    #                     'quantity': sum(move_in_ids.mapped('quantity_done')),
    #                     'value': sum(move_in_ids.mapped('stock_valuation_layer_ids').mapped('value'))
    #                 })
    #             move_out_ids = move_out.filtered(lambda sm: sm.product_id.id == product)
    #             if move_out_ids:
    #                 move_out_res_new.append({
    #                     'product': product,
    #                     # 'quantity': sum(move_out_ids.mapped('stock_valuation_layer_ids').mapped('quantity')) if move_out_ids.mapped('stock_valuation_layer_ids') else sum(move_out_ids.mapped('quantity_done')),
    #                     'quantity': sum(move_out_ids.mapped('quantity_done')),
    #                     'value': sum(move_out_ids.mapped('stock_valuation_layer_ids').mapped('value'))
    #                 })
    #             quant = quants.filtered(lambda q: q.product_id.id == product)
    #             if quant:
    #                 quants_res_new.append({
    #                     'product': product,
    #                     'quantity': sum(quant.mapped('quantity')),
    #                     'value': sum(quant.mapped('value'))
    #                 })
    #         product_in = {item['product']: item for item in move_in_res_new}
    #         product_out = {item['product']: item for item in move_out_res_new}
    #         product_end = {item['product']: item for item in quants_res_new}
    #         for med in meds:
    #             in_qty_new = 0
    #             in_value_new = 0
    #             out_qty_new = 0
    #             out_value_new = 0
    #             end_qty_new = 0
    #             end_value_new = 0
    #             med_data = [index, med.default_code, med.name, med.uom_id.name]
    #             product_in_result = product_in.get(med.product_id.id)
    #             if product_in_result:
    #                 in_qty_new = product_in_result.get('quantity', 0)
    #                 in_value_new = product_in_result.get('value', 0)
    #             product_out_result = product_out.get(med.product_id.id)
    #             if product_out_result:
    #                 out_qty_new = product_out_result.get('quantity', 0)
    #                 out_value_new = product_out_result.get('value', 0)
    #             product_end_result = product_end.get(med.product_id.id)
    #             if product_end_result:
    #                 end_qty_new = product_end_result.get('quantity', 0)
    #                 end_value_new = product_end_result.get('value', 0)
    #             if self.end_datetime < fields.Datetime.now():
    #                 out_qty_past_new = 0
    #                 out_value_past_new = 0
    #                 in_qty_past_new = 0
    #                 in_value_past_new = 0
    #                 moves_in_res_past_new = []
    #                 c_moves_in_domain_past = moves_in_domain + [('date', '>', self.end_datetime),
    #                                                             ('product_id', '=', med.product_id.id)]
    #                 moves_in_past = Move.search(c_moves_in_domain_past)
    #                 if moves_in_past:
    #                     moves_in_res_past_new.append({
    #                         'product': med.product_id.id,
    #                         'quantity': sum(moves_in_past.mapped('quantity_done')),
    #                         'value': sum(moves_in_past.mapped('stock_valuation_layer_ids').mapped('value'))
    #                     })
    #                 product_in_past = {item['product']: item for item in moves_in_res_past_new}
    #                 if product_in_past:
    #                     product_in_result_past = product_in_past.get(med.product_id.id, 0.0)
    #                     if product_in_result_past:
    #                         in_qty_past_new = product_in_result_past.get('quantity', 0)
    #                         in_value_past_new = product_in_result_past.get('value', 0)
    #                 moves_out_res_past_new = []
    #                 c_moves_out_domain_past = moves_out_domain + [('date', '>', self.end_datetime),
    #                                                               ('product_id', '=', med.product_id.id)]
    #                 moves_out_past = Move.search(c_moves_out_domain_past)
    #                 if moves_out_past:
    #                     moves_out_res_past_new.append({
    #                         'product': med.product_id.id,
    #                         'quantity': sum(moves_out_past.mapped('quantity_done')),
    #                         'value': sum(moves_out_past.mapped('stock_valuation_layer_ids').mapped('value'))
    #                     })
    #                 product_out_past = {item['product']: item for item in moves_out_res_past_new}
    #                 if product_out_past:
    #                     product_out_result_past = product_out_past.get(med.product_id.id, 0.0)
    #                     if product_out_result_past:
    #                         out_qty_past_new = product_out_result_past.get('quantity', 0)
    #                         out_value_past_new = product_out_result_past.get('value', 0)
    #                 end_qty_new += (out_qty_past_new - in_qty_past_new)
    #                 end_value_new += (out_value_past_new - in_value_past_new)
    #             if end_qty_new == in_qty_new == out_qty_new == 0:
    #                 continue
    #             begin_qty_new = end_qty_new - in_qty_new + out_qty_new
    #             value_qty_new = abs(end_value_new) - abs(in_value_new) + abs(out_value_new)
    #             inv_data = [begin_qty_new, value_qty_new, in_qty_new, in_value_new, out_qty_new, out_value_new,
    #                         end_qty_new, end_value_new]
    #             med_data += inv_data
    #             for i, j in zip(list(range(4, 12)), inv_data):
    #                 categ_datas[1][i] += j
    #                 data_sum[i] += j
    #             categ_datas.append(med_data)
    #             index += 1
    #         datas += categ_datas
    #     i += 1
    #     datas.append(data_sum)
    #     return datas

    def create_report_new(self):
        if self.env.company.id != 1:
            raise ValidationError('Bạn cần đứng tại chi nhánh SCI để xuất báo cáo !')
        template = self.env['ir.attachment'].browse(self.env.ref('sh_bao_cao.bao_cao_xnt_sci_attachment').id)
        decode = base64.b64decode(template.datas)
        wb = load_workbook(BytesIO(decode))
        ws = wb.active
        thin = borders.Side(style='thin')
        all_border_thin = borders.Border(left=thin, right=thin, top=thin, bottom=thin)
        line_font = Font(name='Times New Roman', size=13)
        ws['a1'].value = 'Công ty: SCI GROUP'
        ws['a1'].font = Font(name='Times New Roman', size=15, bold=True)
        ws['a1'].alignment = Alignment(horizontal='left', vertical='center')
        ws['a2'].value = 'Kho: %s' % self.location.name
        ws['a2'].font = Font(name='Times New Roman', size=15, bold=True)
        ws['a2'].alignment = Alignment(horizontal='left', vertical='center')
        date_range = 'Từ ngày: %s đến ngày: %s' % (
            self.start_datetime.strftime('%d/%m/%Y %H:%M:%S'), self.end_datetime.strftime('%d/%m/%Y %H:%M:%S'))

        ws['a5'].value = date_range
        ws['a5'].font = Font(name='Times New Roman', size=15, bold=True)
        ws['a5'].alignment = Alignment(horizontal='left', vertical='center')
        row = 10
        datas = self.get_data_2()
        for line in datas:
            for col, val in enumerate(line, 1):
                cell = ws.cell(row, col)
                cell.value = val
                cell.border = all_border_thin
                if isinstance(line[0], str):
                    cell.font = Font(name='Times New Roman', size=10, bold=True)
                else:
                    cell.font = Font(name='Times New Roman', size=10)
                if col > 4:
                    cell.number_format = '_(* #,##0_);_(* (#,##0);_(* "-"??_);_(@_)'
                if col == 6:
                    cell.value = "=IF(L{0}>0, L{1} + J{2} - H{3},0)".format(row, row, row, row)
                if col == 5:
                    cell.value = "=IF(K{0}>0, K{1} + I{2} - G{3},0)".format(row, row, row, row)
            row += 1

        row += 1
        ws.cell(row, 9).value = date.today().strftime('Ngày %d tháng %m năm %Y')
        ws.cell(row, 9).font = Font(name='Times New Roman', size=10, italic=True)
        ws.cell(row, 9).alignment = center_alm

        row += 2
        ws.cell(row, 3).value = 'Kế toán'
        ws.cell(row, 9).value = 'Người lập'
        ws.cell(row + 1, 3).value = ws.cell(row + 1, 9).value = '(Ký, họ tên)'
        ws.cell(row, 3).font = ws.cell(row, 9).font = Font(name='Times New Roman', size=10, bold=True)
        ws.cell(row + 1, 3).font = ws.cell(row + 1, 9).font = Font(name='Times New Roman', size=10, italic=True)
        ws.cell(row, 3).alignment = ws.cell(row, 9).alignment = ws.cell(row + 1, 3).alignment = ws.cell(row + 1,
                                                                                                        9).alignment = center_alm

        fp = BytesIO()
        wb.save(fp)
        fp.seek(0)
        report = base64.encodebytes((fp.read()))
        fp.close()
        attachment = self.env['ir.attachment'].sudo().create({
            'name': 'Báo cáo XNT kho SCI.xlsx',
            'datas': report,
            'res_model': 'temp.creation',
            'public': True,
        })
        if self.preview:
            return {'name': 'Báo cáo XNT SCI.xlsx',
                    'type': 'ir.actions.act_window',
                    'res_model': 'temp.wizard',
                    'view_mode': 'form',
                    'view_type': 'form',
                    'target': 'inline',
                    'view_id': self.env.ref('ms_templates.report_wizard').id,
                    'context': {'attachment_id': attachment.id}}
        else:
            url = "/web/content/?model=ir.attachment&id=%s&filename_field=name&field=datas&download=true" \
                  % attachment.id
            return {'name': 'Báo cáo xuất/nhập tồn SCI.xlsx',
                    'type': 'ir.actions.act_url',
                    'url': url,
                    'target': 'self',
                    }

    def get_data_2(self):
        datas = [('', '', '', '', '', '', '', '', '', '', '', '')]
        data_dict = {}
        # TODO: Lấy danh sách sản phẩm theo nhóm sản phẩm
        categories = self.env['sh.medical.medicines.category'].search([], order='type')
        medicines = self.env['sh.medical.medicines'].with_context(active_test=False).search_read(
            [('medicine_category_id', 'in', categories.ids)], ['name', 'id', 'product_id', 'medicine_category_id'])
        products = self.env['sh.medical.medicines'].with_context(active_test=False).search(
            [('medicine_category_id', 'in', categories.ids)]).mapped('product_id')

        product_categrory_dict = {}
        for medicine in medicines:
            if str(medicine['product_id'][0]) not in product_categrory_dict:
                product_categrory_dict[str(medicine['product_id'][0])] = {}
            product_categrory_dict[str(medicine['product_id'][0])]['name'] = medicine['medicine_category_id'][1]
            product_categrory_dict[str(medicine['product_id'][0])]['id'] = medicine['medicine_category_id'][0]

        # TODO: Lấy giá trị và số lượng tồn cuối theo sản phẩm và địa điểm
        quants = self.env['stock.quant'].sudo().search(
            [('product_id', 'in', products.ids), ('location_id', '=', self.location.id)], order='id desc')

        for quant in quants:
            if product_categrory_dict[str(quant.product_id.id)]['id'] not in data_dict:
                data_dict[product_categrory_dict[str(quant.product_id.id)]['id']] = {
                    'id': product_categrory_dict[str(quant.product_id.id)]['id'],
                    'name': product_categrory_dict[str(quant.product_id.id)]['name'],
                    'total_sl_ton_dau': 0,
                    'total_gt_ton_dau': 0,
                    'total_sl_nhap': 0,
                    'total_gt_nhap': 0,
                    'total_sl_xuat': 0,
                    'total_gt_xuat': 0,
                    'data': {}
                }
            if 'total_sl_ton_cuoi' not in data_dict[product_categrory_dict[str(quant.product_id.id)]['id']]:
                data_dict[product_categrory_dict[str(quant.product_id.id)]['id']]['total_sl_ton_cuoi'] = 0
            if 'total_gt_ton_cuoi' not in data_dict[product_categrory_dict[str(quant.product_id.id)]['id']]:
                data_dict[product_categrory_dict[str(quant.product_id.id)]['id']]['total_gt_ton_cuoi'] = 0

            data_dict[product_categrory_dict[str(quant.product_id.id)]['id']]['total_sl_ton_cuoi'] += quant.quantity
            data_dict[product_categrory_dict[str(quant.product_id.id)]['id']]['total_gt_ton_cuoi'] += quant.value
            data_dict[product_categrory_dict[str(quant.product_id.id)]['id']]['data'][str(quant.product_id.id)] = {
                'ton_cuoi': quant.quantity,
                'gt_ton_cuoi': quant.value,
                'sl_ton_dau': 0,
                'gt_ton_dau': 0,
                'sl_nhap': 0,
                'gt_nhap': 0,
                'sl_xuat': 0,
                'gt_xuat': 0,
                'code': quant.product_id.default_code,
                'unit': quant.product_id.uom_id.name,
                'name': quant.product_id.name
            }
        # TODO: Lấy giá trị- số lượng nhập và xuất theo ngày và sản phẩm
        valuations = self.env['stock.valuation.layer'].search(
            [('product_id', 'in', products.ids), ('date', '>=', self.start_datetime),
             ('date', '<=', self.end_datetime), ('quantity', '!=', 0),
             ('value', '!=', 0), '|', ('stock_move_id.location_id', '=', self.location.id),
             ('stock_move_id.location_dest_id', '=', self.location.id)], order='id desc')

        for valuation in valuations:

            if product_categrory_dict[str(valuation.product_id.id)]['id'] not in data_dict:
                data_dict[product_categrory_dict[str(valuation.product_id.id)]['id']] = {
                    'id': product_categrory_dict[str(valuation.product_id.id)]['id'],
                    'name': product_categrory_dict[str(valuation.product_id.id)]['name'],
                    'total_sl_ton_dau': 0,
                    'total_gt_ton_dau': 0,
                    'total_sl_ton_cuoi': 0,
                    'total_gt_ton_cuoi': 0,
                    'total_sl_nhap': 0,
                    'total_gt_nhap': 0,
                    'total_sl_xuat': 0,
                    'total_gt_xuat': 0,
                    'data': {str(valuation.product_id.id): {
                        'ton_cuoi': 0,
                        'gt_ton_cuoi': 0,
                        'sl_ton_dau': 0,
                        'gt_ton_dau': 0,
                        'sl_nhap': 0,
                        'gt_nhap': 0,
                        'sl_xuat': 0,
                        'gt_xuat': 0,
                        'code': valuation.product_id.default_code,
                        'unit': valuation.product_id.uom_id.name,
                        'name': valuation.product_id.name
                    }}
                }
            if str(valuation.product_id.id) not in \
                    data_dict[product_categrory_dict[str(valuation.product_id.id)]['id']]['data']:
                data_dict[product_categrory_dict[str(valuation.product_id.id)]['id']]['data'][
                    str(valuation.product_id.id)] = {
                    'ton_cuoi': 0,
                    'gt_ton_cuoi': 0,
                    'sl_ton_dau': 0,
                    'gt_ton_dau': 0,
                    'sl_nhap': 0,
                    'gt_nhap': 0,
                    'sl_xuat': 0,
                    'gt_xuat': 0,
                    'code': valuation.product_id.default_code,
                    'unit': valuation.product_id.uom_id.name,
                    'name': valuation.product_id.name
                }

            if valuation.quantity >= 0 and valuation.value >= 0:
                data_dict[product_categrory_dict[str(valuation.product_id.id)]['id']][
                    'total_sl_nhap'] += valuation.quantity
                data_dict[product_categrory_dict[str(valuation.product_id.id)]['id']][
                    'total_gt_nhap'] += valuation.value
                data_dict[product_categrory_dict[str(valuation.product_id.id)]['id']]['data'][
                    str(valuation.product_id.id)]['sl_nhap'] += valuation.quantity
                data_dict[product_categrory_dict[str(valuation.product_id.id)]['id']]['data'][
                    str(valuation.product_id.id)]['gt_nhap'] += valuation.value
                #######################################
                # Tính tồn đầu theo gía trị và số lượng nhập
                #######################################



            else:
                data_dict[product_categrory_dict[str(valuation.product_id.id)]['id']][
                    'total_sl_xuat'] += valuation.quantity
                data_dict[product_categrory_dict[str(valuation.product_id.id)]['id']][
                    'total_gt_xuat'] += valuation.value
                data_dict[product_categrory_dict[str(valuation.product_id.id)]['id']]['data'][
                    str(valuation.product_id.id)]['sl_xuat'] += valuation.quantity
                data_dict[product_categrory_dict[str(valuation.product_id.id)]['id']]['data'][
                    str(valuation.product_id.id)]['gt_xuat'] += valuation.value

        for data in data_dict:
            datas.append(
                ('Nhóm hàng: ' + data_dict[data]['name'].upper(),
                 '',
                 '',
                 '',
                 data_dict[data]['total_sl_ton_dau'],
                 data_dict[data]['total_gt_ton_dau'],
                 data_dict[data]['total_sl_nhap'],
                 data_dict[data]['total_gt_nhap'],
                 abs(data_dict[data]['total_sl_xuat']),
                 abs(data_dict[data]['total_gt_xuat']),
                 data_dict[data]['total_sl_ton_cuoi'],
                 data_dict[data]['total_gt_ton_cuoi']))
            index = 0
            if data_dict[data]['data']:
                for data_in_dict in data_dict[data]['data']:
                    datas.append(
                        (index + 1, data_dict[data]['data'][str(data_in_dict)]['code'],
                         data_dict[data]['data'][str(data_in_dict)]['name'],
                         data_dict[data]['data'][str(data_in_dict)]['unit'],
                         '',
                         data_dict[data]['data'][str(data_in_dict)]['gt_ton_dau'],
                         data_dict[data]['data'][str(data_in_dict)]['sl_nhap'],
                         data_dict[data]['data'][str(data_in_dict)]['gt_nhap'],
                         abs(data_dict[data]['data'][str(data_in_dict)]['sl_xuat']),
                         abs(data_dict[data]['data'][str(data_in_dict)]['gt_xuat']),
                         data_dict[data]['data'][str(data_in_dict)]['ton_cuoi'],
                         data_dict[data]['data'][str(data_in_dict)]['gt_ton_cuoi']))
                    index += 1
        return datas

    # def create_report_valuation_layer(self):
    #     if self.env.company.id != 1:
    #         raise ValidationError('Bạn cần đứng tại chi nhánh SCI để xuất báo cáo !')
    #     template = self.env['ir.attachment'].browse(self.env.ref('sh_bao_cao.bao_cao_xnt_sci_attachment').id)
    #     decode = base64.b64decode(template.datas)
    #     wb = load_workbook(BytesIO(decode))
    #     ws = wb.active
    #     thin = borders.Side(style='thin')
    #     all_border_thin = borders.Border(left=thin, right=thin, top=thin, bottom=thin)
    #     line_font = Font(name='Times New Roman', size=13)
    #     ws['a1'].value = 'Công ty: SCI GROUP'
    #     ws['a1'].font = Font(name='Times New Roman', size=15, bold=True)
    #     ws['a1'].alignment = Alignment(horizontal='left', vertical='center')
    #     ws['a2'].value = 'Kho: %s' % self.location.name
    #     ws['a2'].font = Font(name='Times New Roman', size=15, bold=True)
    #     ws['a2'].alignment = Alignment(horizontal='left', vertical='center')
    #     date_range = 'Từ ngày: %s đến ngày: %s' % (
    #         self.start_datetime.strftime('%d/%m/%Y %H:%M:%S'), self.end_datetime.strftime('%d/%m/%Y %H:%M:%S'))
    #
    #     ws['a5'].value = date_range
    #     ws['a5'].font = Font(name='Times New Roman', size=15, bold=True)
    #     ws['a5'].alignment = Alignment(horizontal='left', vertical='center')
    #     row = 10
    #     datas = self.get_data_valuation_layer()
    #     for line in datas:
    #         for col, val in enumerate(line, 1):
    #             cell = ws.cell(row, col)
    #             cell.value = val
    #             cell.border = all_border_thin
    #             if isinstance(line[0], str):
    #                 cell.font = Font(name='Times New Roman', size=10, bold=True)
    #             else:
    #                 cell.font = Font(name='Times New Roman', size=10)
    #             if col > 4:
    #                 cell.number_format = '_(* #,##0_);_(* (#,##0);_(* "-"??_);_(@_)'
    #         row += 1
    #
    #     row += 1
    #     ws.cell(row, 9).value = date.today().strftime('Ngày %d tháng %m năm %Y')
    #     ws.cell(row, 9).font = Font(name='Times New Roman', size=10, italic=True)
    #     ws.cell(row, 9).alignment = center_alm
    #
    #     row += 2
    #     ws.cell(row, 3).value = 'Kế toán'
    #     ws.cell(row, 9).value = 'Người lập'
    #     ws.cell(row + 1, 3).value = ws.cell(row + 1, 9).value = '(Ký, họ tên)'
    #     ws.cell(row, 3).font = ws.cell(row, 9).font = Font(name='Times New Roman', size=10, bold=True)
    #     ws.cell(row + 1, 3).font = ws.cell(row + 1, 9).font = Font(name='Times New Roman', size=10, italic=True)
    #     ws.cell(row, 3).alignment = ws.cell(row, 9).alignment = ws.cell(row + 1, 3).alignment = ws.cell(row + 1,
    #                                                                                                     9).alignment = center_alm
    #
    #     fp = BytesIO()
    #     wb.save(fp)
    #     fp.seek(0)
    #     report = base64.encodebytes((fp.read()))
    #     fp.close()
    #     attachment = self.env['ir.attachment'].sudo().create({
    #         'name': 'Báo cáo XNT kho SCI.xlsx',
    #         'datas': report,
    #         'res_model': 'temp.creation',
    #         'public': True,
    #     })
    #     if self.preview:
    #         return {'name': 'Báo cáo XNT SCI.xlsx',
    #                 'type': 'ir.actions.act_window',
    #                 'res_model': 'temp.wizard',
    #                 'view_mode': 'form',
    #                 'view_type': 'form',
    #                 'target': 'inline',
    #                 'view_id': self.env.ref('ms_templates.report_wizard').id,
    #                 'context': {'attachment_id': attachment.id}}
    #     else:
    #         url = "/web/content/?model=ir.attachment&id=%s&filename_field=name&field=datas&download=true" \
    #               % attachment.id
    #         return {'name': 'Báo cáo xuất/nhập tồn SCI.xlsx',
    #                 'type': 'ir.actions.act_url',
    #                 'url': url,
    #                 'target': 'self',
    #                 }
    #
    # def get_data_valuation_layer(self):
    #     datas = []
    #     main_location = self.location
    #     i = 1
    #     categs = self.env['sh.medical.medicines.category'].search([], order='type')
    #     for categ in categs:
    #         datas.append((categ.name, '', '', '', '', '', '', '', '', '', '', ''))
    #         products = self.env['sh.medical.medicines'].with_context(active_test=False).search(
    #             [('medicine_category_id', '=', categ.id)]).mapped('product_id')
    #         for product in products:
    #             i += 1
    #             ton_cuoi = 0
    #             gia_tri_cuoi = 0
    #             quant = self.env['stock.quant'].sudo().search(
    #                 [('product_id', '=', product.id), ('location_id', '=', main_location.id)])
    #             if quant:
    #                 ton_cuoi = quant.quantity
    #                 gia_tri_cuoi = quant.value
    #             ############################# Nhập trong kỳ trích xuất
    #             valuations_in = self.env['stock.valuation.layer'].search(
    #                 [('product_id', '=', product.id), ('date', '>=', self.start_datetime), '|', ('quantity', '>=', 0),
    #                  ('value', '>=', 0)])
    #             valuations_in = valuations_in.filtered(lambda v: (v.stock_move_id.location_id == main_location) or (
    #                     v.stock_move_id.location_dest_id == main_location))
    #             valuation_in = valuations_in.filtered(lambda vi: vi.date <= self.end_datetime)
    #             qty_in = abs(sum(valuation_in.mapped('quantity')))
    #             valuation_in = abs(sum(valuation_in.mapped('value')))
    #             ############################# Xuất trong kỳ trích xuất
    #             valuations_out = self.env['stock.valuation.layer'].search(
    #                 [('product_id', '=', product.id), ('date', '>=', self.start_datetime), '|', ('quantity', '<', 0),
    #                  ('value', '<', 0)])
    #             valuations_out = valuations_out.filtered(lambda v: (v.stock_move_id.location_id == main_location) or (
    #                     v.stock_move_id.location_dest_id == main_location))
    #             valuation_out = valuations_out.filtered(lambda vi: vi.date <= self.end_datetime)
    #             qty_out = abs(sum(valuation_out.mapped('quantity')))
    #             valuation_out = abs(sum(valuation_out.mapped('value')))
    #             qty_in_past = 0
    #             valuation_in_past = 0
    #             qty_out_past = 0
    #             valuation_out_past = 0
    #             ############################### Chênh lệch giá trị hiện tại với thời điểm trích xuất
    #             if self.end_datetime < fields.Datetime.now():
    #                 valuations_in_past = valuations_in.filtered(lambda vi: vi.date > self.end_datetime)
    #                 qty_in_past = abs(sum(valuations_in_past.mapped('quantity')))
    #                 valuation_in_past = abs(sum(valuations_in_past.mapped('value')))
    #                 valuations_out_past = valuations_out.filtered(lambda vo: vo.date > self.end_datetime)
    #                 qty_out_past = abs(sum(valuations_out_past.mapped('quantity')))
    #                 valuation_out_past = abs(sum(valuations_out_past.mapped('value')))
    #             ton_cuoi += (qty_out_past - qty_in_past)
    #             gia_tri_cuoi += (valuation_out_past - valuation_in_past)
    #             ton_dau = ton_cuoi + qty_out - qty_in
    #             gia_tri_dau = gia_tri_cuoi + valuation_out - valuation_in
    #             if ton_cuoi == qty_in == qty_out:
    #                 continue
    #             datas.append((product.default_code, product.name, product.uom_id.name, ton_dau, gia_tri_dau, qty_in,
    #                           valuation_in, qty_out, valuation_out, ton_cuoi, gia_tri_cuoi))
    #     return datas
