import base64
from datetime import date, datetime
from io import BytesIO

from openpyxl import load_workbook
from openpyxl.styles import Font, borders, Alignment
from pytz import timezone
from calendar import monthrange

from odoo import fields, api, models
from odoo.exceptions import ValidationError

thin = borders.Side(style='thin')
thick = borders.Side(style='medium')
dotted = borders.Side(style='hair')
gray_thin = borders.Side(style='thin', color='808080')
all_border_thin = borders.Border(left=thin, right=thin, top=thin, bottom=thin)
all_border_gray = borders.Border(left=gray_thin, right=gray_thin, top=gray_thin, bottom=gray_thin)
dotted_top_bot = borders.Border(left=thin, right=thin, top=dotted, bottom=dotted)
thick_dotted_top_bot = borders.Border(left=thick, right=thick, top=dotted, bottom=dotted)
all_border_thick = borders.Border(left=thick, right=thick, top=thick, bottom=thick)
center_alm = Alignment(horizontal='center', vertical='center')


class XNTReport(models.TransientModel):
    _name = "bc.xnt"
    _description = "Báo cáo xuất/nhập tồn"

    @api.model
    def _default_warehouse_id(self):
        stock_warehouse = self.env['stock.warehouse'].search([('company_id', '=', self.env.user.company_id.id)], limit=1)
        return stock_warehouse

    today = date.today()
    month = today.month if today.month != 0 else 1
    year = today.year
    _, last_day_of_month = monthrange(year, month)
    last_day = date(int(year), int(month), last_day_of_month)

    company = fields.Many2one('res.company', default=lambda self: self.env.user.company_id.id)
    warehouse_id = fields.Many2one('stock.warehouse', default=_default_warehouse_id)
    start_date = fields.Date('Ngày bắt đầu', default=today.replace(day=1))
    end_date = fields.Date('Ngày kết thúc', default=last_day)
    start_datetime = fields.Datetime('Start datetime', compute='_compute_datetime')
    end_datetime = fields.Datetime('End datetime', compute='_compute_datetime')
    preview = fields.Boolean('Bạn có muốn xem trước ?')
    location = fields.Many2many('stock.location', string='Tủ thuốc', default=None, domain="[('usage', '=', 'internal')]")

    @api.onchange('warehouse_id')
    def _onchange_warehouse_id(self):
        if self.warehouse_id:
            self.company = self.warehouse_id.company_id

    @api.depends('start_date', 'end_date')
    def _compute_datetime(self):
        self.start_datetime = False
        self.end_datetime = False
        if self.start_date and self.end_date:
            local_tz = timezone(self.env.user.tz or 'Etc/GMT+7')
            start_datetime = datetime(self.start_date.year, self.start_date.month, self.start_date.day, 0, 0, 0)
            end_datetime = datetime(self.end_date.year, self.end_date.month, self.end_date.day, 23, 59, 59)
            self.start_datetime = start_datetime
            self.end_datetime = end_datetime

    def create_report_new(self):
        template = self.env['ir.attachment'].browse(self.env.ref('sh_bao_cao.bao_cao_xnt_attachment').id)
        decode = base64.b64decode(template.datas)
        wb = load_workbook(BytesIO(decode))
        ws = wb.active
        thin = borders.Side(style='thin')
        all_border_thin = borders.Border(left=thin, right=thin, top=thin, bottom=thin)
        line_font = Font(name='Times New Roman', size=13)
        ws['a1'].value = 'Công ty: ' + self.company.name
        ws['a1'].font = Font(name='Times New Roman', size=15, bold=True)
        ws['a1'].alignment = Alignment(horizontal='left', vertical='center')
        date_range = 'Từ ngày: %s đến ngày: %s' % (
            self.start_datetime.strftime('%d/%m/%Y %H:%M:%S'), self.end_datetime.strftime('%d/%m/%Y %H:%M:%S'))

        ws['a5'].value = date_range
        ws['a5'].font = Font(name='Times New Roman', size=15, bold=True)
        ws['a5'].alignment = Alignment(horizontal='left', vertical='center')
        row = 10
        datas = self.get_data_2()
        for line in datas:
            for col, val in enumerate(line, 1):
                cell = ws.cell(row, col)
                cell.value = val
                cell.border = all_border_thin
                if isinstance(line[0], str):
                    cell.font = Font(name='Times New Roman', size=10, bold=True)
                else:
                    cell.font = Font(name='Times New Roman', size=10)
                if col > 4:
                    cell.number_format = '_(* #,##0_);_(* (#,##0);_(* "-"??_);_(@_)'
                if col == 12:
                    cell.value = "=IF(1>0, F{1} + H{2} - J{3},0)".format(row, row, row, row)
                if col == 13:
                    cell.value = "=IF(1>0, G{1} + I{2} - K{3},0)".format(row, row, row, row)
            row += 1

        row += 1
        ws.cell(row, 9).value = date.today().strftime('Ngày %d tháng %m năm %Y')
        ws.cell(row, 9).font = Font(name='Times New Roman', size=10, italic=True)
        ws.cell(row, 9).alignment = center_alm

        row += 2
        ws.cell(row, 3).value = 'Kế toán'
        ws.cell(row, 9).value = 'Người lập'
        ws.cell(row + 1, 3).value = ws.cell(row + 1, 9).value = '(Ký, họ tên)'
        ws.cell(row, 3).font = ws.cell(row, 9).font = Font(name='Times New Roman', size=10, bold=True)
        ws.cell(row + 1, 3).font = ws.cell(row + 1, 9).font = Font(name='Times New Roman', size=10, italic=True)
        ws.cell(row, 3).alignment = ws.cell(row, 9).alignment = ws.cell(row + 1, 3).alignment = ws.cell(row + 1,
                                                                                                        9).alignment = center_alm

        fp = BytesIO()
        wb.save(fp)
        fp.seek(0)
        report = base64.encodebytes((fp.read()))
        fp.close()
        attachment = self.env['ir.attachment'].sudo().create({
            'name': 'Báo cáo XNT kho.xlsx',
            'datas': report,
            'res_model': 'temp.creation',
            'public': True,
        })
        url = "/web/content/?model=ir.attachment&id=%s&filename_field=name&field=datas&download=true" \
              % attachment.id
        return {'name': 'Báo cáo xuất/nhập tồn.xlsx',
                'type': 'ir.actions.act_url',
                'url': url,
                'target': 'self',
                }

    def get_data_2(self):
        datas = []
        data_dict = {}
        # TODO: Lấy danh sách sản phẩm theo location
        if self.location:
            location_ids = self.location
        else:
            location_ids = self.env['stock.location'].sudo().search([('company_id', '=', self.company.id), ('usage', '=', 'internal')])

        for location in location_ids:
            stock_moves = self.env['stock.move'].sudo().search(
                ['|', ('location_id', '=', location.id), ('location_dest_id', '=', location.id)]).mapped('id')
            stock_valuation_layer = self.env['stock.valuation.layer'].sudo().search(
                [('stock_move_id', 'in', stock_moves), ('quantity', '!=', 0),
                 ('value', '!=', 0), ('date', '>=', str(self.start_datetime)),
                 ('date', '<=', str(self.end_datetime))])
            products = stock_valuation_layer.mapped('product_id')


            medicines = self.env['sh.medical.medicines'].search_read([('product_id', 'in', products.ids)], ['name', 'id', 'product_id', 'medicine_category_id'])
            product_categrory_dict = {}
            for medicine in medicines:
                if str(medicine['product_id'][0]) not in product_categrory_dict:
                    product_categrory_dict[str(medicine['product_id'][0])] = {}
                product_categrory_dict[str(medicine['product_id'][0])]['name'] = medicine['medicine_category_id'][1]
                product_categrory_dict[str(medicine['product_id'][0])]['id'] = medicine['medicine_category_id'][0]

            data_dict[str(location.id)] = {
                'id': location.id,
                'name': location.display_name,
                'total_sl_ton_dau': 0,
                'total_gt_ton_dau': 0,
                'total_sl_ton_cuoi': 0,
                'total_gt_ton_cuoi': 0,
                'total_sl_nhap': 0,
                'total_gt_nhap': 0,
                'total_sl_xuat': 0,
                'total_gt_xuat': 0,
                'data': {}
            }

            # TODO: Lấy giá trị và số lượng tồn đầu theo sản phẩm và địa điểm
            quants = self.env['stock.quant'].sudo().search(
                [('product_id', 'in', products.ids), ('location_id', '=', location.id), ('in_date', '<=', str(self.start_datetime))], order='id desc')
            for quant in quants:
                data_dict[str(location.id)]['total_sl_ton_dau'] += quant.quantity
                data_dict[str(location.id)]['total_gt_ton_dau'] += quant.value
                medicines_cate = ''
                if str(quant.product_id.id) in product_categrory_dict:
                    medicines_cate = product_categrory_dict[str(quant.product_id.id)]['name']
                data_dict[str(location.id)]['data'][str(quant.product_id.id)] = {
                    'sl_ton_dau': quant.quantity,
                    'gt_ton_dau': quant.value,
                    'sl_nhap': 0,
                    'gt_nhap': 0,
                    'sl_xuat': 0,
                    'gt_xuat': 0,
                    'sl_ton_cuoi': 0,
                    'gt_ton_cuoi': 0,
                    'code': quant.product_id.default_code,
                    'unit': quant.product_id.uom_id.name,
                    'nhom_hang': medicines_cate,
                    'name': quant.product_id.name
                }

            # TODO: Lấy giá trị- số lượng nhập và xuất theo ngày và sản phẩm
            for valuation in stock_valuation_layer:
                medicines_cate = ''
                if str(valuation.product_id.id) in product_categrory_dict:
                    medicines_cate = product_categrory_dict[str(valuation.product_id.id)]['name']
                if str(valuation.product_id.id) not in data_dict[str(location.id)]['data']:
                    data_dict[str(location.id)]['data'][str(valuation.product_id.id)] = {
                        'sl_ton_dau': 0,
                        'gt_ton_dau': 0,
                        'sl_ton_cuoi': 0,
                        'gt_ton_cuoi': 0,
                        'sl_nhap': 0,
                        'gt_nhap': 0,
                        'sl_xuat': 0,
                        'gt_xuat': 0,
                        'code': valuation.product_id.default_code,
                        'unit': valuation.product_id.uom_id.name,
                        'nhom_hang': medicines_cate,
                        'name': valuation.product_id.name
                    }

                if valuation.quantity >= 0 and valuation.value >= 0:
                    data_dict[str(location.id)]['total_sl_nhap'] += valuation.quantity
                    data_dict[str(location.id)]['total_gt_nhap'] += valuation.value
                    data_dict[str(location.id)]['data'][str(valuation.product_id.id)]['sl_nhap'] += valuation.quantity
                    data_dict[str(location.id)]['data'][str(valuation.product_id.id)]['gt_nhap'] += valuation.value
                else:
                    data_dict[str(location.id)]['total_sl_xuat'] += valuation.quantity
                    data_dict[str(location.id)]['total_gt_xuat'] += valuation.value
                    data_dict[str(location.id)]['data'][str(valuation.product_id.id)]['sl_xuat'] += valuation.quantity
                    data_dict[str(location.id)]['data'][str(valuation.product_id.id)]['gt_xuat'] += valuation.value

        for data in data_dict:
            datas.append(
                (data_dict[data]['name'],
                 '',
                 '',
                 '',
                 '',
                 data_dict[data]['total_sl_ton_dau'],
                 data_dict[data]['total_gt_ton_dau'],
                 data_dict[data]['total_sl_nhap'],
                 data_dict[data]['total_gt_nhap'],
                 abs(data_dict[data]['total_sl_xuat']),
                 abs(data_dict[data]['total_gt_xuat']),
                 data_dict[data]['total_sl_ton_cuoi'],
                 data_dict[data]['total_gt_ton_cuoi']))
            index = 0
            if data_dict[data]['data']:
                for data_in_dict in data_dict[data]['data']:
                    datas.append(
                        (index + 1, data_dict[data]['data'][str(data_in_dict)]['code'],
                         data_dict[data]['data'][str(data_in_dict)]['name'],
                         data_dict[data]['data'][str(data_in_dict)]['nhom_hang'],
                         data_dict[data]['data'][str(data_in_dict)]['unit'],
                         data_dict[data]['data'][str(data_in_dict)]['sl_ton_dau'],
                         data_dict[data]['data'][str(data_in_dict)]['gt_ton_dau'],
                         data_dict[data]['data'][str(data_in_dict)]['sl_nhap'],
                         data_dict[data]['data'][str(data_in_dict)]['gt_nhap'],
                         abs(data_dict[data]['data'][str(data_in_dict)]['sl_xuat']),
                         abs(data_dict[data]['data'][str(data_in_dict)]['gt_xuat']),
                         data_dict[data]['data'][str(data_in_dict)]['sl_ton_cuoi'],
                         data_dict[data]['data'][str(data_in_dict)]['gt_ton_cuoi']))
                    index += 1

        datas.append(('', '', '', '', '', '', '', '', '', '', '', '', ''))
        return datas


