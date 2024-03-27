# -*- coding: utf-8 -*-
import base64
import logging
from io import BytesIO

from openpyxl import load_workbook
from openpyxl.styles import Font, borders, PatternFill

from odoo import models, fields

_logger = logging.getLogger(__name__)

thin = borders.Side(style='thin')
thick = borders.Side(style='medium')
dotted = borders.Side(style='hair')
gray_thin = borders.Side(style='thin', color='808080')
all_border_thin = borders.Border(left=thin, right=thin, top=thin, bottom=thin)
all_border_gray = borders.Border(left=gray_thin, right=gray_thin, top=gray_thin, bottom=gray_thin)
dotted_top_bot = borders.Border(left=thin, right=thin, top=dotted, bottom=dotted)


class SHInventoryReport(models.TransientModel):
    _inherit = 'sh.inventory.report'

    is_query = fields.Boolean('Sử dụng truy vấn', default=True, help='Sử dụng truy vấn tăng tốc độ tạo báo cáo')

    def _get_stock_out_data_v2(self):
        result = []
        if self.report_type == 'customer':
            query = """select   rp.id,
                                pp.default_code, 
                                smm.product_id, 
                                pt.name, 
                                uu.name, 
                                smwm.init_quantity, 
                                smwm.quantity, 
                                rp.code_customer,
                                rp.name,
                                rp.year_of_birth
                                from sh_medical_walkin_material smwm
                                left join sh_medical_medicines smm on smm.id = smwm.product_id 
                                left join product_product pp on pp.id = smm.product_id 
                                left join product_template pt on pt.id = pp.product_tmpl_id 
                                left join uom_uom uu on uu.id = pt.uom_id 
                                left join sh_medical_patient smp on smp.id = smwm.patient 
                                left join res_partner rp on rp.id = smp.partner_id
                                where smwm.institution = %s 
                                    and smwm.date_out >= %s
                                    and smwm.date_out <= %s
                                Order by rp.code_customer, pp.default_code
                                    """
            self.env.cr.execute(query, (self.institution.id, self.start_datetime, self.end_datetime))
            datas = self.env.cr.fetchall()

            result = []
            patients = {}
            for line in datas:
                if line[0] in patients:
                    index = len(patients[line[0]][1]) + 1
                    # Lấy dict của vật tư
                    if line[2] in patients[line[0]][1]:
                        patients[line[0]][1][line[2]][4] += line[5]
                        patients[line[0]][1][line[2]][5] += line[6]
                        patients[line[0]][1][line[2]][6] += patients[line[0]][1][line[2]][4] - \
                                                            patients[line[0]][1][line[2]][5]
                    else:
                        patients[line[0]][1][line[2]] = [index, line[1], line[3], line[4], line[5], line[6],
                                                         line[5] - line[6]]
                else:
                    # Lấy tên bệnh nhân
                    # Nếu có ngày sinh
                    if line[9]:
                        name = "[%s] %s - %s" % (line[7], line[8], line[9])
                    else:
                        name = "[%s] %s" % (line[7], line[8])
                    patients[line[0]] = [name,
                                         {line[2]: [1, line[1], line[3], line[4], line[5], line[6], line[5] - line[6]]}]
                # result.append(line)
            for p in patients:
                lines = [[patients[p][0]]]
                for key in patients[p][1]:
                    lines.extend([patients[p][1][key]])
                result.append(lines)

        elif self.report_type == 'service':
            query = """select           smwm.id,
                                        pp.default_code, 
                                        smm.name_use, 
                                        pt.name, 
                                        uu.name,
                                        smwm.init_quantity,
                                        smwm.quantity,
                                        smhcw.name,
                                        smhcs.id,
                                        pp2.default_code,
                                        pt2.name,
                                        pp.id, 
                                        smwm.walkin
                                from sh_medical_walkin_material smwm
                                left join sh_medical_medicines smm on smm.id = smwm.product_id 
                                left join product_product pp on pp.id = smm.product_id 
                                left join product_template pt on pt.id = pp.product_tmpl_id 
                                left join uom_uom uu on uu.id = pt.uom_id 
                                left join sh_medical_health_center_ward smhcw on smhcw.id = smwm.department 
                                left join sh_medical_health_center_service_sh_medical_walkin_material_rel smhcssmwmr on smhcssmwmr.sh_medical_walkin_material_id = smwm.id 
                                left join sh_medical_health_center_service smhcs on smhcssmwmr.sh_medical_health_center_service_id = smhcs.id
                                left join product_product pp2 on pp2.id = smhcs.product_id 
                                left join product_template pt2 on pt2.id = pp2.product_tmpl_id 
                                where smwm.institution = %s 
                                    and smwm.date_out >= %s
                                    and smwm.date_out <= %s
                                Order by pp2.default_code, pp.default_code
                                                """
            self.env.cr.execute(query, (self.institution.id, self.start_datetime, self.end_datetime))
            datas = self.env.cr.fetchall()
            result = []
            services = {}
            vattu_dv = {}
            for line in datas:
                if line[0] in vattu_dv:
                    vattu_dv[line[0]] += 1
                else:
                    vattu_dv[line[0]] = 1

            for line in datas:
                total_service = 0
                if line[0] in vattu_dv:
                    total_service = vattu_dv[line[0]]

                if total_service:
                    sl_thuc_xuat = line[6] / total_service
                else:
                    sl_thuc_xuat = line[6]
                chenh_lech = line[5] - sl_thuc_xuat
                # Id dịch vụ
                if line[8] in services:
                    index = len(services[line[8]][2]) + 1
                    # Lấy dict của vật tư
                    if line[11] in services[line[8]][2]:
                        services[line[8]][2][line[11]][4] += line[5]
                        services[line[8]][2][line[11]][5] += sl_thuc_xuat
                        services[line[8]][2][line[11]][6] = services[line[8]][2][line[11]][4] - \
                                                            services[line[8]][2][line[11]][5]
                    else:
                        services[line[8]][2][line[11]] = [index, line[1], line[3], line[4], line[5], sl_thuc_xuat,
                                                          chenh_lech, line[7]]
                    if line[12] not in services[line[8]][1]:
                        services[line[8]][1].append(line[12])

                else:
                    # Lấy tên dịch vụ
                    name = '[%s] %s' % (line[9], line[10])

                    # Tên dịch vụ, số lượng, line vt
                    services[line[8]] = [name, [line[12]], {
                        line[11]: [1, line[1], line[3], line[4], line[5], sl_thuc_xuat, chenh_lech, line[7]]}, line[7]]

            for p in services:
                lines = [['%s - %sDV' % (services[p][0], len(services[p][1])), services[p][3]]]
                for key in services[p][2]:
                    lines.extend([services[p][2][key]])
                result.append(lines)
        return result

    def report_stock_out(self):
        _logger.info("XUATBAOCAO KHO VTTH")
        inventory_attachment = self.env['ir.attachment'].sudo().browse(
            self.env.ref('shealth_all_in_one.stock_out_report_attachment').id)
        decode = base64.b64decode(inventory_attachment.datas)
        wb = load_workbook(BytesIO(decode))
        ws = wb.active

        if self.is_query:
            datas = self._get_stock_out_data_v2()
        else:
            datas = self._get_stock_out_data()

        if self.report_type == 'service':
            ws['b1'].value = 'THỐNG KÊ XUẤT VẬT TƯ / THUỐC THEO DỊCH VỤ'
            ws['b5'].value = 'Dịch vụ'
            ws['j5'].value, ws['j5'].border = 'Dịch vụ', all_border_thin

        ws['b2'].value = self.institution.name
        ws['b3'].value = 'Từ 00:00 %s đến 23:59 %s' % (
            self.start_date.strftime('%d/%m/%Y'), self.end_date.strftime('%d/%m/%Y'))
        line_font = Font(name='Times New Roman', size=12)
        row = 6
        end_col = 11 if self.report_type == 'service' else 10
        for group in datas:
            ws.cell(row, 2).value, ws.cell(row, 2).font = group[0][0], line_font
            if self.report_type == 'service':
                ws.cell(row, 10).value, ws.cell(row, 10).font = group[0][1], line_font
            for col in range(2, end_col):
                ws.cell(row, col).fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
                ws.cell(row, col).border = all_border_thin
            row += 1
            for i in range(1, len(group)):
                for col in range(3, end_col):
                    ws.cell(row, col).value, ws.cell(row, col).font = group[i][col - 3], line_font
                for col in range(2, end_col):
                    ws.cell(row, col).border = all_border_thin
                row += 1

        fp = BytesIO()
        wb.save(fp)
        fp.seek(0)
        report = base64.encodebytes((fp.read()))
        fp.close()
        attachment = self.env['ir.attachment'].with_user(1).create({'name': 'Stock out report.xlsx',
                                                                    'datas': report,
                                                                    'res_model': 'temp.creation',
                                                                    'public': True})

        url = "/web/content/?model=ir.attachment&id=%s&filename_field=name&field=datas&download=true" \
              % attachment.id

        return {
            'name': 'ĐÁNH GIÁ CHẤT LƯỢNG DỊCH VỤ',
            'type': 'ir.actions.act_url',
            'url': url,
            'target': 'self',
        }
