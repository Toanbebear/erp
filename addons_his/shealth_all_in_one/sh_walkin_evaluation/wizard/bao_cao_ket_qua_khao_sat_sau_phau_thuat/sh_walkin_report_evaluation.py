from odoo import fields, api, models, _
from odoo.exceptions import ValidationError, UserError
from datetime import date, datetime, time
from calendar import monthrange
from openpyxl import load_workbook
from openpyxl.styles import Font, borders, Alignment
import base64
from io import BytesIO
from dateutil.relativedelta import relativedelta
from pytz import timezone, utc
from collections import defaultdict

thin = borders.Side(style='thin')
double = borders.Side(style='double')
all_border_thin = borders.Border(thin, thin, thin, thin)





class shWalkinReportEvaluation(models.TransientModel):
    _name = 'sh.walkin.report.evaluation'
    _description = 'Báo cáo khảo sát sau phẫu thuật'

    company_id = fields.Many2one('res.company', string='Công ty', default=lambda self: self.env.company,
                                 domain=lambda self: [('id', 'in', self.env.companies.ids)])
    # using date to simplify user input, and to extract date to report
    start_date = fields.Date('Ngày bắt đầu', default=date.today().replace(day=1))
    end_date = fields.Date('Ngày kết thúc')
    # convert date to datetime for search domain, should be removed if using datetime directly
    start_datetime = fields.Datetime('Start datetime', compute='_compute_datetime')
    end_datetime = fields.Datetime('End datetime', compute='_compute_datetime')

    # check ngày
    @api.depends('start_date', 'end_date')
    def _compute_datetime(self):
        self.start_datetime = False
        self.end_datetime = False
        if self.start_date and self.end_date:
            local_tz = timezone(self.env.user.tz or 'Etc/GMT+7')
            start_datetime = datetime(self.start_date.year, self.start_date.month, self.start_date.day, 0, 0, 0)
            end_datetime = datetime(self.end_date.year, self.end_date.month, self.end_date.day, 23, 59, 59)
            start_datetime = local_tz.localize(start_datetime, is_dst=None)
            end_datetime = local_tz.localize(end_datetime, is_dst=None)
            self.start_datetime = start_datetime.astimezone(utc).replace(tzinfo=None)
            self.end_datetime = end_datetime.astimezone(utc).replace(tzinfo=None)

    @api.onchange('start_date')
    def _onchange_start_date(self):
        if self.start_date:
            if self.start_date.month == fields.date.today().month:
                self.end_date = fields.date.today()
            else:
                self.end_date = date(self.start_date.year, self.start_date.month,
                                     monthrange(self.start_date.year, self.start_date.month)[1])

    @api.constrains('start_date', 'end_date')
    def check_dates(self):
        for record in self:
            start_date = fields.Date.from_string(record.start_date)
            end_date = fields.Date.from_string(record.end_date)
            if start_date > end_date:
                raise ValidationError(
                    _("Ngày kết thúc không thể ở trước ngày bắt đầu khi xuất báo cáo!!!"))

    def _get_data_report(self):
        ret_data = []
        cata = self.env['sh.medical.evaluation'].search(
            [('check_state', '=', '1'), ('institution.his_company', '=', self.company_id.id),
             ('evaluation_start_date', '>=', self.start_date),
             ('evaluation_end_date', '<=', self.end_date), ('state', '=', 'Completed')], )
        # Phân loại đánh giá
        dict_doctor_attitude = {
            '1': 'Rất hài lòng',
            '2': 'Hài lòng',
            '3': 'Bình thường',
            '4': 'Không hài lòng',
            '5': 'Rất không hài lòng',
            '6': '',
        }

        # Phan loai kết quả
        dict_satisfaction_level = {
            '1': 'Rất hài lòng',
            '2': 'Hài lòng',
            '3': 'Bình Thường',
            '4': 'Không hài lòng',
            '5': 'Rất không hài lòng',
            '6': 'Bảo hành',
            '7': '',
        }
        stt = 1
        for rec in cata:
            for line in rec.surgery_history_survey_ids:
                for doct in line.main_doctor:
                    ret_data.append({
                        "stt": stt,
                        "ngay_tai_kham": rec.evaluation_start_date.strftime('%d/%m/%Y'),
                        "ho_ten": rec.patient.name,
                        "ma_booking": rec.walkin.booking_id.name,
                        "ma_tai_kham": rec.name,
                        "nhom_dich_vu": line.service_performances.service_category.name,
                        "dich_vu_phau_thuat": line.service_performances.name,
                        "ngay_phau_thuat": line.surgery_date.strftime('%d/%m/%Y') if line.surgery_date else '',
                        "bac_si_phau_thuat": doct.name,
                        "bac_si_tai_kham": rec.doctor.name if rec.doctor.name else '',
                        "phan_loai": ','.join(rec.evaluation_services.mapped('name')),
                        "ket_qua": dict_satisfaction_level[
                        '%s' % line.satisfaction_level] if line.satisfaction_level else '',
                        "danh_gia": dict_doctor_attitude['%s' % line.doctor_attitude] if line.doctor_attitude else '',
                    })
                    stt += 1
        return ret_data

    def create_report(self):
        datas = self._get_data_report()
        report_brand_overview_attachment = self.env['ir.attachment'].browse(
            self.env.ref('shealth_all_in_one.bao_cao_ket_qua_khao_sat_sau_phau_thuat_attachment').id)
        decode = base64.b64decode(report_brand_overview_attachment.datas)
        wb = load_workbook(BytesIO(decode))
        ws = wb.active
        line_font = Font(name='Times New Roman', size=12)

        ws['E4'].value += self.start_date.strftime('%d/%m/%Y')
        ws['F4'].value += self.end_datetime.strftime('%d/%m/%Y')
        ws['F6'].value = self.company_id.name

        key_col_list = list(range(1, 14))
        key_list = [
            "stt",
            "ngay_tai_kham",
            "ho_ten",
            "ma_booking",
            "ma_tai_kham",
            "nhom_dich_vu",
            "dich_vu_phau_thuat",
            "ngay_phau_thuat",
            "bac_si_phau_thuat",
            "bac_si_tai_kham",
            "phan_loai",
            "ket_qua",
            "danh_gia",

        ]
        row = 10
        for data in datas:
            for col, k in zip(key_col_list, key_list):
                beforeCell = ws.cell(1, col)
                beforeCell.font = Font(name='Times New Roman', size=12, color='FFFFFF')
                cell = ws.cell(row, col)
                cell.value = data[k]
                cell.font = line_font
                cell.border = all_border_thin
            row += 1

        fp = BytesIO()
        wb.save(fp)
        fp.seek(0)
        report = base64.encodebytes((fp.read()))
        fp.close()
        attachment = self.env['ir.attachment'].sudo().create({
            'name': 'bao_cao_ket_qua_khao_sat_sau_phau_thuat.xlsx',
            'datas': report,
            'res_model': 'temp.creation',
            'public': True,
        })
        url = "/web/content/?model=ir.attachment&id=%s&filename_field=name&field=datas&download=true" \
              % attachment.id
        return {'name': 'BÁO CÁO KHẢO SÁT SAU PHẪU THUẬT',
                'type': 'ir.actions.act_url',
                'url': url,
                'target': 'self',
                }