# -*- coding: utf-8 -*-
import base64
import logging
from calendar import monthrange
from datetime import date, datetime, timedelta
from io import BytesIO

import openpyxl
from odoo import fields, api, models, _
from odoo.exceptions import ValidationError
from odoo.modules.module import get_module_resource
from openpyxl import load_workbook
from openpyxl.styles import Font, borders, Alignment, PatternFill
from pytz import timezone, utc

_logger = logging.getLogger(__name__)


class ThemeReport:
    thin = borders.Side(style='thin')
    double = borders.Side(style='double')
    all_border_thin = borders.Border(thin, thin, thin, thin)
    line_font = Font(name='Times New Roman', size=14)
    da_fill = PatternFill(start_color='0e7661', end_color='0e7661', fill_type='solid')
    kn_fill = PatternFill(start_color='003471', end_color='003471', fill_type='solid')
    pr_fill = PatternFill(start_color='012C5F', end_color='012C5F', fill_type='solid')
    hh_fill = PatternFill(start_color='053D7C', end_color='053D7C', fill_type='solid')
    sci_fill = PatternFill(start_color='003471', end_color='003471', fill_type='solid')


class SurgicalServiceListReport(models.TransientModel):
    _name = 'report.surgical.service.plus'
    _description = 'Báo cáo phẫu thuật bổ sung'

    start_date = fields.Date('Start date', default=date.today().replace(day=1))
    end_date = fields.Date('End date')
    company_id = fields.Many2one('res.company', string='Company', default=lambda self: self.env.company,
                                 domain=lambda self: [('id', 'in', self.env.companies.ids)])

    # convert date to datetime for search domain, should be removed if using datetime directly
    start_datetime = fields.Datetime('Start datetime', compute='_compute_datetime')
    end_datetime = fields.Datetime('End datetime', compute='_compute_datetime')

    # convert date to datetime for search domain, should be removed if using datetime directly
    @api.depends('start_date', 'end_date')
    def _compute_datetime(self):
        self.start_datetime = False
        self.end_datetime = False
        if self.start_date and self.end_date:
            local_tz = timezone(self.env.user.tz or 'Asia/Ho_Chi_Minh')
            start_datetime = datetime(self.start_date.year, self.start_date.month, self.start_date.day, 0, 0, 0)
            end_datetime = datetime(self.end_date.year, self.end_date.month, self.end_date.day, 23, 59, 59)
            start_datetime = local_tz.localize(start_datetime, is_dst=None)
            end_datetime = local_tz.localize(end_datetime, is_dst=None)
            self.start_datetime = start_datetime.astimezone(utc).replace(tzinfo=None)
            self.end_datetime = end_datetime.astimezone(utc).replace(tzinfo=None)

    @api.onchange('start_date')
    def _onchange_start_date(self):
        if self.start_date:
            if self.start_date.month == fields.date.today().month:
                self.end_date = fields.date.today()
            else:
                self.end_date = date(self.start_date.year, self.start_date.month,
                                     monthrange(self.start_date.year, self.start_date.month)[1])

    @api.constrains('start_date', 'end_date')
    def check_dates(self):
        for record in self:
            start_date = fields.Date.from_string(record.start_date)
            end_date = fields.Date.from_string(record.end_date)
            if start_date > end_date:
                raise ValidationError(
                    _("End Date cannot be set before Start Date."))

    def _get_data_report_danh_sach_dich_vu_phau_thuat(self):
        ret_data = []
        datas = self.env['sh.medical.surgery'].search(
            [('institution.his_company', '=', self.company_id.id), ('surgery_date', '>=', self.start_date),
             ('surgery_date', '<=', self.end_date), ('state', '=', 'Done')], order='surgery_date asc')

        # gender
        gender = {
            'male': 'Nam',
            'female': 'Nữ',
            'transguy': 'Transguy',
            'transgirl': 'Transgirl',
            'other': 'Khác'
        }

        # pp vô cảm
        anesthetist_type = {
            'te': 'Gây tê',
            'tien_me': 'Tiền mê',
            'me': 'Gây mê'
        }

        # loại pt
        surgery_type = {
            'DB': 'Loại đặc biệt',
            '1': 'I',
            '2': 'II',
            '3': 'III'
        }
        for data in datas:
            street = data.patient.street if data.patient.street else ''
            state_id = data.patient.state_id.name if data.patient.state_id else ''
            technical_name = []
            for rec_technical_name in data.services:
                if rec_technical_name.technical_name:
                    technical_name.append(rec_technical_name.technical_name)
            for service in data.services:
                doctor_main = ''
                doctor_assist = ''
                doctor_anesthesia = ''
                for member in data.surgery_team:
                    if service in member.service_performances:
                        if member.role.id == 1:
                            if member.team_member.name not in doctor_main:
                                if doctor_main:
                                    continue
                                else:
                                    doctor_main = member.team_member.name
                            else:
                                doctor_main += ', ' + member.team_member.name
                        elif member.role.id == 2:
                            if member.team_member.name not in doctor_assist:
                                if doctor_assist:
                                    continue
                                else:
                                    doctor_assist = member.team_member.name
                            else:
                                doctor_assist += ', ' + member.team_member.name
                        elif member.role.id == 18:
                            if member.team_member.name not in doctor_anesthesia:
                                if doctor_anesthesia:
                                    continue
                                else:
                                    doctor_anesthesia = member.team_member.name
                            else:
                                doctor_anesthesia += ', ' + member.team_member.name
                value = {
                    'patient_name': data.patient.name,
                    'booking_code': data.booking_id.name,
                    'hsba': '',
                    'year_of_birth': data.patient.partner_id.year_of_birth if data.patient.partner_id.year_of_birth else '',
                    'gender': gender['%s' % data.patient.gender],
                    'country': data.patient.partner_id.country_id.name if data.patient.partner_id.country_id.name else '',
                    'pasport': data.patient.partner_id.pass_port if data.patient.pass_port else '',
                    'pasport_date': data.patient.partner_id.pass_port_date if data.patient.partner_id.pass_port_date else '',
                    'pasport_address': data.patient.partner_id.pass_port_issue_by if data.patient.partner_id.pass_port_issue_by else '',
                    'address': street + (',' + state_id) if state_id else '',
                    'function': data.patient.function if data.patient.function else '',
                    'info_diagnosis': service.diagnose if service.diagnose else '',
                    'services': service.name if service.name else '',
                    'surgical_method': service.surgical_method if service.surgical_method else '',
                    'anesthetist_type': anesthetist_type['%s' % data.anesthetist_type] if data.anesthetist_type else '',
                    'hour_walkin': (data.date_requested + timedelta(hours=7)).strftime(
                        '%Hh%M') if data.date_requested else '',
                    'date_walkin': data.date_requested.strftime('%d-%m-%Y') if data.date_requested else '',
                    'hour_surgery': (data.surgery_date + timedelta(hours=7)).strftime(
                        '%Hh%M') if data.surgery_date else '',
                    'date_surgery': data.surgery_date.strftime('%d-%m-%Y') if data.surgery_date else '',
                    'hour_surgery_done': (data.surgery_end_date + timedelta(hours=7)).strftime(
                        '%Hh%M') if data.surgery_end_date else '',
                    'date_surgery_done': data.surgery_end_date.strftime('%d-%m-%Y') if data.surgery_end_date else '',
                    'hour_out': (data.walkin.date_out + timedelta(hours=7)).strftime(
                        '%Hh%M') if data.walkin.date_out else '',
                    'date_out': data.walkin.date_out.strftime('%d-%m-%Y') if data.walkin.date_out else '',
                    'surgery_type': surgery_type['%s' % data.surgery_type] if data.surgery_type else '',
                    'doctor_main': doctor_main,
                    'doctor_assist': doctor_assist,
                    'doctor_anesthesia': doctor_anesthesia,
                    'note': ''
                }
                ret_data.append(value)
        return ret_data

    # Xử lý bằng SQL
    # def _get_data(self):
        # ret_data = []
        # surgery_query = """
        #     SELECT
        #         sms.id,
        #         rp."name" as patient_name,
        #         cl."name" as booking_code,
        #         rp.year_of_birth,
        #         rp.gender,
        #         rc."name" as country,
        #         rp.pass_port as pasport,
        #         rp.pass_port_date as pass_port_date,
        #         rp.pass_port_issue_by as pass_port_issue_by,
        #         rcs."name" as state,
        #         rp.street,
        #         smp.function,
        #         smarw.info_diagnosis,
        #         sms.anesthetist_type,
        #         cl.arrival_date,
        #         sms.date_requested,
        #         sms.surgery_end_date,
        #         smarw.date_out,
        #         sms.surgery_type,
        #         sms.surgery_date,
        #         smlt.date_analysis,
        #         smlt.date_done,
        #         sms.surgeon,
        #         sms.anesthetist
        # FROM sh_medical_surgery sms
        # LEFT JOIN sh_medical_patient smp ON smp.id = sms.patient
        # LEFT JOIN res_partner rp ON rp.id = smp.partner_id
        # LEFT JOIN sh_medical_appointment_register_walkin smarw ON smarw.id = sms.walkin
        # LEFT JOIN crm_lead cl ON cl.id = smarw.booking_id
        # LEFT JOIN res_country rc ON rc.id = rp.country_id
        # LEFT JOIN res_country_state rcs ON rcs.id = rp.state_id
        # LEFT JOIN sh_medical_health_center smhc ON smhc.id = sms.institution
        # LEFT JOIN sh_medical_lab_test smlt ON smarw.id = smlt.walkin
        #         AND smlt.id = (SELECT id FROM sh_medical_lab_test t1 WHERE t1.walkin = smarw.id LIMIT 1)
        # WHERE smhc.his_company = %s
        #     AND sms.surgery_date >= '%s'
        #     AND sms.surgery_date <= '%s'
        #     AND sms.state = 'Done'
        # ORDER BY sms.surgery_date ASC
        # """ % (self.company_id.id, self.start_datetime, self.end_datetime)
        # self.env.cr.execute(surgery_query)
        # surgeries = self.env.cr.dictfetchall()
        # surgery_ids = set()
        # for surgery in surgeries:
        #     surgery_ids.add(surgery['id'])
        # service_query = """
        #     SELECT
        #         sssr.surgery_id,
        #         pt.name,
        #         smhcs.technical_name
        #     FROM sh_surgery_service_rel sssr
        #     LEFT JOIN sh_medical_health_center_service smhcs on smhcs.id = sssr.service_id
        #     LEFT JOIN service_departments_rel sdr  on sdr.sh_medical_health_center_service_id = smhcs.id
        #     LEFT JOIN sh_medical_health_center_ward smhcw  on smhcw.id = sdr.sh_medical_health_center_ward_id
        #     LEFT JOIN sh_medical_health_center smhc on smhc.id = smhcw.institution
        #     LEFT JOIN product_product pp on smhcs.product_id = pp.id
        #     LEFT JOIN product_template pt on pp.product_tmpl_id = pt.id
        #     WHERE smhcw.type = '%s' and sssr.surgery_id in %s
        #     GROUP BY  sssr.surgery_id, pt.name, smhcs.technical_name
        # """ % ('Surgery', tuple(surgery_ids))
        #
        # self.env.cr.execute(service_query)
        # services = self.env.cr.dictfetchall()
        #
        # surgery_services = {}
        # for service in services:
        #     if service['surgery_id'] in surgery_services:
        #         if service['name']:
        #             surgery_services[service['surgery_id']]['name'].append(service['name'])
        #         if service['technical_name']:
        #             surgery_services[service['surgery_id']]['technical_name'].append(service['technical_name'])
        #     else:
        #         data_surgery_id = {}
        #         if service['name']:
        #             data_surgery_id['name'] = [service['name']]
        #         else:
        #             data_surgery_id['name'] = []
        #
        #         if service['technical_name']:
        #             data_surgery_id['technical_name'] = [service['technical_name']]
        #         else:
        #             data_surgery_id['technical_name'] = []
        #
        #         surgery_services[service['surgery_id']] = data_surgery_id
        #
        # # gender
        # gender = {
        #     'male': 'Nam',
        #     'female': 'Nữ',
        #     'transguy': 'Transguy',
        #     'transgirl': 'Transgirl',
        #     'other': 'Khác'
        # }
        #
        # # pp vô cảm
        # anesthetist_type = {
        #     'te': 'Gây tê',
        #     'tien_me': 'Tiền mê',
        #     'me': 'Gây mê'
        # }
        #
        # # loại pt
        # surgery_type = {
        #     'DB': 'Loại đặc biệt',
        #     '1': 'I',
        #     '2': 'II',
        #     '3': 'III'
        # }
        # for data in surgeries:
        #     # Tên dịch vụ
        #     if data['id'] in surgery_services:
        #         surgery_service = surgery_services[data['id']]
        #         services = ', '.join(surgery_service['name'])
        #         surgical_method = ', '.join(surgery_service['technical_name'])
        #     else:
        #         services = ''
        #         surgical_method = ''
        #
        #     # Địa chỉ
        #     if data['street']:
        #         if data['state']:
        #             address = data['street'] + ', ' + data['state']
        #         else:
        #             address = data['street']
        #     else:
        #         if data['state']:
        #             address = data['state']
        #         else:
        #             address = ''
        #
        #     # Ngày giờ
        #     if data['arrival_date']:
        #         arival_hour = (data['arrival_date'] + timedelta(hours=7)).strftime('%Hh%M')
        #         arival_date = data['arrival_date'].strftime('%d-%m-%Y')
        #     else:
        #         arival_hour = ''
        #         arival_date = ''
        #
        #     if data['date_analysis']:
        #         hour_test = (data['date_analysis'] + timedelta(hours=7)).strftime('%Hh%M')
        #         date_test = data['date_analysis'].strftime('%d-%m-%Y')
        #     else:
        #         hour_test = ''
        #         date_test = ''
        #
        #     if data['date_done']:
        #         hour_test_done = (data['date_done'] + timedelta(hours=7)).strftime('%Hh%M')
        #         date_test_done = data['date_done'].strftime('%d-%m-%Y')
        #     else:
        #         hour_test_done = ''
        #         date_test_done = ''
        #
        #     value = {
        #         'patient_name': data['patient_name'],
        #         'booking_code': data['booking_code'],
        #         'hsba': '',
        #         'year_of_birth': data['year_of_birth'] if data['year_of_birth'] else '',
        #         'gender': gender[data['gender']],
        #         'country': data['country'] if data['country'] else '',
        #         'pasport': data['pasport'] if data['pasport'] else '',
        #         'pass_port_date': data['pass_port_date'].strftime('%d-%m-%Y') if data['pass_port_date'] else '',
        #         'pass_port_issue_by': data['pass_port_issue_by'] if data['pass_port_issue_by'] else '',
        #         'address': address,
        #         'function': data['function'] if data['function'] else '',
        #         'info_diagnosis': data['info_diagnosis'] if data['info_diagnosis'] else '',
        #         'services': services,
        #         'surgical_method': surgical_method,
        #         'anesthetist_type': anesthetist_type[data['anesthetist_type']] if data['anesthetist_type'] else '',
        #         'arival_hour': arival_hour,
        #         'arival_date': arival_date,
        #         'hour_test': hour_test,
        #         'date_test': date_test,
        #         'hour_test_done': hour_test_done,
        #         'date_test_done': date_test_done,
        #         'hour_walkin': (data['date_requested'] + timedelta(hours=7)).strftime('%Hh%M') if data[
        #             'date_requested'] else '',
        #         'date_walkin': data['date_requested'].strftime('%d-%m-%Y') if data['date_requested'] else '',
        #         'hour_surgery': (data['surgery_date'] + timedelta(hours=7)).strftime('%Hh%M') if data[
        #             'surgery_date'] else '',
        #         'date_surgery': data['surgery_date'].strftime('%d-%m-%Y') if data['surgery_date'] else '',
        #         'hour_surgery_done': (data['surgery_end_date'] + timedelta(hours=7)).strftime('%Hh%M') if data[
        #             'surgery_end_date'] else '',
        #         'date_surgery_done': data['surgery_end_date'].strftime('%d-%m-%Y') if data['surgery_end_date'] else '',
        #         'hour_out': (data['date_out'] + timedelta(hours=7)).strftime('%Hh%M') if data['date_out'] else '',
        #         'date_out': data['date_out'].strftime('%d-%m-%Y') if data['date_out'] else '',
        #         'surgery_type': surgery_type[data['surgery_type']] if data['surgery_type'] else '',
        #         'surgeon': self.env['sh.medical.physician'].sudo().search([('id', '=', data['surgeon'])]).name if data['surgeon'] else '',
        #         'anesthetist': self.env['sh.medical.physician'].sudo().search([('id', '=', data['anesthetist'])]).name if data['anesthetist'] else '',
        #         'note': ''
        #     }
        #     ret_data.append(value)
        # return ret_data

    def create_report_danh_sach_dich_vu_phau_thuat(self):
        daily_sales_attachment = self.env['ir.attachment'].browse(
            self.env.ref('shealth_all_in_one.report_surgical_service_plus_attachment').id)
        decode = base64.b64decode(daily_sales_attachment.datas)
        wb = load_workbook(BytesIO(decode))
        ws = wb.active

        # datas = self._get_data_report_danh_sach_dich_vu_phau_thuat()
        datas = self._get_data_report_danh_sach_dich_vu_phau_thuat()

        key_list = [
            'patient_name',
            'booking_code',
            'hsba',
            'year_of_birth',
            'gender',
            'country',
            'pasport',
            'pasport_date',
            'pasport_address',
            'address',
            'function',
            'info_diagnosis',
            'services',
            'surgical_method',
            'anesthetist_type',
            'hour_walkin',
            'date_walkin',
            'hour_surgery',
            'date_surgery',
            'hour_surgery_done',
            'date_surgery_done',
            'hour_out',
            'date_out',
            'surgery_type',
            'doctor_main',
            'doctor_assist',
            'doctor_anesthesia',
            'note'
        ]
        key_col_list = list(range(2, len(key_list) + 2))
        row = 8
        index_row = 0
        for line_data in datas:
            ws.cell(row, 1).border = ThemeReport.all_border_thin
            ws.cell(row, 1).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(row, 1).value = index_row + 1
            for col, k in zip(key_col_list, key_list):
                cell = ws.cell(row, col)
                cell.value = line_data[k]
                cell.font = ThemeReport.line_font
                cell.border = ThemeReport.all_border_thin
                if col in [9, 11, 12, 13]:
                    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                else:
                    cell.alignment = Alignment(horizontal='left', vertical='center')
            row += 1
            index_row += 1

        ws['H3'].value = self.start_date.strftime('%d/%m/%Y')
        ws['H4'].value = self.end_date.strftime('%d/%m/%Y')
        ws['C1'].value = self.company_id.name
        ws['C2'].value = 'Địa chỉ: %s' % (
                self.company_id.street + (',' + self.company_id.street2) + (',' + self.company_id.state_id.name))
        ws['C3'].value = 'MST: %s' % (self.company_id.vat if self.company_id.vat else '')
        code_brand = self.company_id.brand_id.code.lower()
        image_path = get_module_resource('report_sale', 'static/img', 'icon_%s.png' % code_brand)
        img = openpyxl.drawing.image.Image(image_path)
        img.anchor = 'A1'
        ws.add_image(img)

        fp = BytesIO()
        wb.save(fp)
        fp.seek(0)
        report = base64.encodebytes((fp.read()))
        fp.close()
        attachment = self.env['ir.attachment'].with_user(1).create({'name': 'bao_cao_phau_thuat_%s.xlsx'% datetime.now().strftime('%d_%m_%Y'),
                                                                    'datas': report,
                                                                    'res_model': 'temp.creation',
                                                                    'public': True})
        url = "/web/content/?model=ir.attachment&id=%s&filename_field=name&field=datas&download=true" \
              % attachment.id
        return {'name': 'Báo cáo phẫu thuật bổ sung',
                'type': 'ir.actions.act_url',
                'url': url,
                'target': 'self',
                }
