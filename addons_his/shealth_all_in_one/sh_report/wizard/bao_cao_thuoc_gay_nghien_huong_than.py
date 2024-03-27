# -*- coding: utf-8 -*-
import openpyxl

from odoo import fields, api, models, _
from odoo.exceptions import ValidationError, UserError
from datetime import date, datetime, time
from calendar import monthrange
from openpyxl import load_workbook
from openpyxl.styles import Font, borders, Alignment, PatternFill
from odoo.modules.module import get_module_resource
import base64
from io import BytesIO
from dateutil.relativedelta import relativedelta
from pytz import timezone, utc

import logging

_logger = logging.getLogger(__name__)


class ThemeReport:
    thin = borders.Side(style='thin')
    double = borders.Side(style='double')
    all_border_thin = borders.Border(thin, thin, thin, thin)
    line_font = Font(name='Times New Roman', size=14)
    da_fill = PatternFill(start_color='0e7661', end_color='0e7661', fill_type='solid')
    kn_fill = PatternFill(start_color='003471', end_color='003471', fill_type='solid')
    pr_fill = PatternFill(start_color='012C5F', end_color='012C5F', fill_type='solid')
    hh_fill = PatternFill(start_color='053D7C', end_color='053D7C', fill_type='solid')
    sci_fill = PatternFill(start_color='003471', end_color='003471', fill_type='solid')


class WalkinReport(models.TransientModel):
    _name = 'report.drug.addiction.psychotropic'
    _description = 'Report drug addiction psychotropic'

    start_date = fields.Date('Start date', default=date.today().replace(day=1))
    end_date = fields.Date('End date')
    company_id = fields.Many2one('res.company', string='Company', default=lambda self: self.env.company,
                                 domain=lambda self: [('id', 'in', self.env.companies.ids)])
    type = fields.Selection([('pttt', 'Phẫu thuật thủ thuật'), ('spa', 'Spa'), ('laser', 'Laser'), ('nha', 'Nha'),
                             ('cshp', 'Chăm sóc hậu phẫu')], string='Loại', default='pttt')
    type_pttt = fields.Selection([
        ('ttdp','Tủ thuốc đại phẫu'),('tttp','Tủ thuốc tiểu phẫu')
    ], string='Tủ', default='ttdp')
    # convert date to datetime for search domain, should be removed if using datetime directly
    start_datetime = fields.Datetime('Start datetime', compute='_compute_datetime')
    end_datetime = fields.Datetime('End datetime', compute='_compute_datetime')

    # convert date to datetime for search domain, should be removed if using datetime directly
    @api.depends('start_date', 'end_date')
    def _compute_datetime(self):
        self.start_datetime = False
        self.end_datetime = False
        if self.start_date and self.end_date:
            local_tz = timezone(self.env.user.tz or 'Etc/GMT+7')
            start_datetime = datetime(self.start_date.year, self.start_date.month, self.start_date.day, 0, 0, 0)
            end_datetime = datetime(self.end_date.year, self.end_date.month, self.end_date.day, 23, 59, 59)
            start_datetime = local_tz.localize(start_datetime, is_dst=None)
            end_datetime = local_tz.localize(end_datetime, is_dst=None)
            self.start_datetime = start_datetime.astimezone(utc).replace(tzinfo=None)
            self.end_datetime = end_datetime.astimezone(utc).replace(tzinfo=None)

    @api.onchange('start_date')
    def _onchange_start_date(self):
        if self.start_date:
            if self.start_date.month == fields.date.today().month:
                self.end_date = fields.date.today()
            else:
                self.end_date = date(self.start_date.year, self.start_date.month,
                                     monthrange(self.start_date.year, self.start_date.month)[1])

    @api.constrains('start_date', 'end_date')
    def check_dates(self):
        for record in self:
            start_date = fields.Date.from_string(record.start_date)
            end_date = fields.Date.from_string(record.end_date)
            if start_date > end_date:
                raise ValidationError(
                    _("End Date cannot be set before Start Date."))

    def get_data_report_bao_cao_thuoc_gay_nghien_huong_than(self, medicine_addictive, addictive_categ):
        ret_data = []
        if self.type == 'pttt':
            # datas = self.env['sh.medical.surgery'].search(
            #     [('institution.his_company', '=', self.company_id.id), ('surgery_date', '>=', self.start_date),
            #      ('surgery_date', '<=', self.end_date), ('state', '=', 'Done')], order='surgery_date asc')
            if self.type_pttt == 'ttdp':
                datas = self.env['sh.medical.surgery'].search(
                    [('institution.his_company', '=', self.company_id.id), ('operating_room.room_type','=','MajorSurgery'), ('surgery_date', '>=', self.start_date),
                     ('surgery_date', '<=', self.end_date), ('state', '=', 'Done')], order='surgery_date asc')
            else:
                datas = self.env['sh.medical.surgery'].search(
                    [('institution.his_company', '=', self.company_id.id), ('operating_room.room_type','=','MinorSurgery'), ('surgery_date', '>=', self.start_date),
                     ('surgery_date', '<=', self.end_date), ('state', '=', 'Done')], order='surgery_date asc')
        elif self.type == 'spa':
            datas = self.env['sh.medical.specialty'].search(
                [('institution.his_company', '=', self.company_id.id), ('services_date', '>=', self.start_date),
                 ('services_date', '<=', self.end_date), ('state', '=', 'Done'), ('department_type', '=', 'Spa')],
                order='services_date asc')
        elif self.type == 'laser':
            datas = self.env['sh.medical.specialty'].search(
                [('institution.his_company', '=', self.company_id.id), ('services_date', '>=', self.start_date),
                 ('services_date', '<=', self.end_date), ('state', '=', 'Done'), ('department_type', '=', 'Laser')],
                order='services_date asc')
        elif self.type == 'nha':
            datas = self.env['sh.medical.specialty'].search(
                [('institution.his_company', '=', self.company_id.id), ('services_date', '>=', self.start_date),
                 ('services_date', '<=', self.end_date), ('state', '=', 'Done'),
                 ('department_type', '=', 'Odontology')],
                order='services_date asc')
        else:
            datas = self.env['sh.medical.patient.rounding'].search(
                [('inpatient_id.institution.his_company', '=', self.company_id.id), ('evaluation_start_date', '>=', self.start_date),
                 ('evaluation_start_date', '<=', self.end_date), ('state', '=', 'Completed')],
                order='evaluation_start_date asc')
        datas_name = datas.mapped('name')
        total_value = {'surgery_date': '', 'stock': '',
                       'patient_name': '', 'booking_id': '',
                       'surgery_name': '', 'has_data': 1}
        for medicine_addictive_item in medicine_addictive:
            total_value['%s' % medicine_addictive_item.default_code] = 0
        ret_data.append(total_value)
        for data_name in datas_name:
            for data_item in datas:
                if data_name == data_item.name:
                    if self.type == 'pttt':
                        value = {'surgery_date': data_item.surgery_date.strftime('%d/%m/%Y'), 'stock': data_item.operating_room.location_medicine_stock.name,
                                 'patient_name': data_item.patient.name, 'booking_id': data_item.booking_id.name,
                                 'surgery_name': data_item.name, 'has_data': 0}
                    elif self.type in ['spa', 'laser', 'nha']:
                        value = {'surgery_date': data_item.services_date.strftime('%d/%m/%Y'), 'stock': data_item.operating_room.location_medicine_stock.name,
                                 'patient_name': data_item.patient.name, 'booking_id': data_item.booking_id.name,
                                 'surgery_name': data_item.name, 'has_data': 0}
                    else:
                        value = {'surgery_date': data_item.evaluation_start_date.strftime('%d/%m/%Y'), 'stock': data_item.inpatient_id.room.location_medicine_stock.name,
                                 'patient_name': data_item.patient.name, 'booking_id': data_item.inpatient_id.walkin.booking_id.name,
                                 'surgery_name': data_item.name, 'has_data': 0}
                    for medicine_addictive_item in medicine_addictive:
                        value['%s' % medicine_addictive_item.default_code] = 0
                    if self.type == 'cshp':
                        for medicine_item in data_item.medicaments:
                            if medicine_item.medicine.medicine_category_id.id in addictive_categ.ids:
                                code_item = medicine_item.medicine.default_code
                                qty = value['%s' % medicine_item.medicine.default_code]
                                qty_item = medicine_item.qty
                                new_qty = qty_item + qty
                                value['%s' % code_item] = new_qty
                                # total_value['%s' % code_item] += value['%s' % code_item]
                                total_value['%s' % code_item] += value['%s' % code_item]
                                value['has_data'] = 1
                        ret_data.append(value)
                    else:
                        for medicine_item in data_item.supplies:
                            if medicine_item.supply.medicine_category_id.id in addictive_categ.ids:
                                code_item = medicine_item.supply.default_code
                                qty = value['%s' % medicine_item.supply.default_code]
                                qty_item = medicine_item.qty_used
                                new_qty = qty_item + qty
                                value['%s' % code_item] = new_qty
                                # total_value['%s' % code_item] += value['%s' % code_item]
                                total_value['%s' % code_item] += value['%s' % code_item]
                                value['has_data'] = 1
                        ret_data.append(value)
        return ret_data

    def create_report_bao_cao_thuoc_gay_nghien_huong_than(self):
        daily_sales_attachment = self.env['ir.attachment'].browse(
            self.env.ref('shealth_all_in_one.bao_cao_thuoc_gay_nghien_huong_than_report_attachment').id)
        decode = base64.b64decode(daily_sales_attachment.datas)
        wb = load_workbook(BytesIO(decode))
        ws = wb.active

        # search thuốc gây nghiện - hướng thần
        addictive_categ = self.env['sh.medical.medicines.category'].search(
            [('name', 'in', ('THUỐC GÂY NGHIỆN', 'THUỐC HƯỚNG TÂM THẦN'))])
        medicine_addictive = self.env['sh.medical.medicines'].search(
            [('medicine_category_id', 'in', addictive_categ.ids)])
        # lấy data
        data = self.get_data_report_bao_cao_thuoc_gay_nghien_huong_than(medicine_addictive=medicine_addictive,
                                                                        addictive_categ=addictive_categ)
        key_list = [
            'surgery_date',
            'stock',
            'patient_name',
            'booking_id',
            'surgery_name',
        ]
        if self.type == 'pttt':
            key_list_title = [
                'STT',
                'Ngày phẫu thuật',
                'Tủ',
                'Họ tên khách hàng',
                'Mã Booking',
                'Mã phiếu PTTT',
            ]
        elif self.type in ['spa', 'laser', 'nha']:
            key_list_title = [
                'STT',
                'Ngày thực hiện',
                'Tủ',
                'Họ tên khách hàng',
                'Mã Booking',
                'Mã phiếu PTTT',
            ]
        else:
            key_list_title = [
                'STT',
                'Ngày bắt đầu',
                'Tủ',
                'Họ tên khách hàng',
                'Mã Booking',
                'Mã phiếu PTTT',
            ]
        uom = {}
        key_list_check = []
        for medicine_addictive_item in medicine_addictive:
            key_list.append(medicine_addictive_item.default_code)
            key_list_check.append(medicine_addictive_item.default_code)
            key_list_title.append(medicine_addictive_item.name)
            uom['%s' % medicine_addictive_item.default_code] = medicine_addictive_item.uom_id.name.lower()
        key_col_list = list(range(2, len(key_list) + 2))
        key_col_list_title = list(range(1, len(key_list) + 2))

        code_brand = self.company_id.brand_id.code.lower()
        if code_brand == 'kn':
            header_fill = ThemeReport.kn_fill
        elif code_brand == 'da':
            header_fill = ThemeReport.da_fill
        elif code_brand == 'pr':
            header_fill = ThemeReport.pr_fill
        elif code_brand == 'hh':
            header_fill = ThemeReport.hh_fill
        else:
            header_fill = ThemeReport.sci_fill

        # in tiêu đề
        row = 5
        index_row = -1
        for col, k in zip(key_col_list_title, key_list_title):
            cell = ws.cell(row, col)
            cell.fill = header_fill
            cell.value = k
            cell.font = Font(name='Times New Roman', size=12, color='FFFFFF')
            cell.border = ThemeReport.all_border_thin
            if col < 7:
                cell.alignment = Alignment(horizontal='center', vertical='center')
            else:
                cell.alignment = Alignment(textRotation=180, horizontal='center', vertical='center')

        # in data
        row = 6
        for line_data in data:
            if line_data['has_data'] == 1:
                ws.cell(row, 1).border = ThemeReport.all_border_thin
                ws.cell(row, 1).alignment = Alignment(horizontal='center', vertical='center')
                ws.cell(row, 1).value = index_row + 1
                for col, k in zip(key_col_list, key_list):
                    cell = ws.cell(row, col)
                    if row == 6:
                        if k in key_list_check:
                            cell.value = round(float(line_data[k]), 1)
                            cell.fill = PatternFill(start_color='BDD7EE', end_color='BDD7EE', fill_type='solid')
                        else:
                            cell.value = line_data[k]
                    else:
                        cell.value = line_data[k]
                    cell.font = ThemeReport.line_font
                    cell.border = ThemeReport.all_border_thin
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                    if col > 5:
                        if k in key_list_check:
                            value = str(round(float(line_data[k]), 1)) + ' ' + uom[k] if line_data[k] != 0 else '-'
                        else:
                            value = line_data[k]
                        cell.value = value
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                row += 1
                index_row += 1

        # in thông tin chung
        cell = ws.cell(4, 7)
        cell.fill = header_fill
        cell.value = 'THÔNG TIN THUỐC (TÊN THUỐC, NỒNG ĐỘ, HÀM LƯỢNG)'
        cell.font = Font(name='Times New Roman', size=18, color='FFFFFF')
        cell.border = ThemeReport.all_border_thin
        cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells(start_row=4, start_column=7, end_row=4, end_column=7 + len(key_list_check) - 1)

        cell = ws.cell(6, 1)
        cell.value = 'TỔNG CỘNG'
        cell.font = Font(name='Times New Roman', size=12)
        cell.border = ThemeReport.all_border_thin
        cell.fill = PatternFill(start_color='BDD7EE', end_color='BDD7EE', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells(start_row=6, start_column=1, end_row=6, end_column=6)

        cell = ws.cell(4, 2)
        cell.value = 'THÔNG TIN KHÁCH HÀNG'
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.font = Font(name='Times New Roman', size=18, color='FFFFFF')
        cell.fill = header_fill
        ws.merge_cells(start_row=4, start_column=2, end_row=4, end_column=6)

        ws['J2'].value = self.start_date.strftime('%d/%m/%Y')
        ws['J2'].font = Font(name='Times New Roman', size=14)
        ws['L2'].value = self.end_date.strftime('%d/%m/%Y')
        ws['L2'].font = Font(name='Times New Roman', size=14)
        ws['C1'].value = self.company_id.name
        ws['C1'].font = Font(name='Times New Roman', size=16)

        code_brand = self.company_id.brand_id.code.lower()
        image_path = get_module_resource('report_sale', 'static/img', 'icon_%s.png' % code_brand)
        img = openpyxl.drawing.image.Image(image_path)
        img.anchor = 'A1'
        ws.add_image(img)

        fp = BytesIO()
        wb.save(fp)
        fp.seek(0)
        report = base64.encodebytes((fp.read()))
        fp.close()
        attachment = self.env['ir.attachment'].with_user(1).create({'name': 'bao_cao_thuoc_gay_nghien_huong_than.xlsx',
                                                                    'datas': report,
                                                                    'res_model': 'temp.creation',
                                                                    'public': True})
        url = "/web/content/?model=ir.attachment&id=%s&filename_field=name&field=datas&download=true" \
              % attachment.id
        return {'name': 'Báo cáo thuốc gây nghiện hướng thần',
                'type': 'ir.actions.act_url',
                'url': url,
                'target': 'self',
                }
