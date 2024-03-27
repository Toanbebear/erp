# -*- coding: utf-8 -*-

import base64
import time as tm
import calendar
from pytz import timezone, utc
from ast import literal_eval
from calendar import monthrange
from copy import copy
from datetime import date, datetime, time, timedelta
from dateutil.relativedelta import relativedelta
from io import BytesIO
import openpyxl
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, borders, Alignment, PatternFill
from openpyxl.worksheet.pagebreak import Break
from operator import itemgetter
from odoo import fields, api, models, _
from odoo.exceptions import ValidationError
import pandas as pd

from odoo.modules.module import get_module_resource
# from odoo.tools import pycompat
from ...sh_surgery.models.mailmerge import MailMerge
# from ...sh_walkins.models.sh_medical_register_for_walkin import num2words_vnm

import random

import logging

_logger = logging.getLogger(__name__)

thin = borders.Side(style='thin')
thick = borders.Side(style='medium')
dotted = borders.Side(style='hair')
gray_thin = borders.Side(style='thin', color='808080')
all_border_thin = borders.Border(left=thin, right=thin, top=thin, bottom=thin)
all_border_gray = borders.Border(left=gray_thin, right=gray_thin, top=gray_thin, bottom=gray_thin)
dotted_top_bot = borders.Border(left=thin, right=thin, top=dotted, bottom=dotted)
thick_dotted_top_bot = borders.Border(left=thick, right=thick, top=dotted, bottom=dotted)
all_border_thick = borders.Border(left=thick, right=thick, top=thick, bottom=thick)
center_alm = Alignment(horizontal='center', vertical='center')

original_report_type = [('supply', 'Supply inventory'), ('medicine', 'Medicine inventory'),
                        ('customer', 'Customer stock out'), ('service', 'Service stock out'),
                        ('antibiotic', 'Antibiotic'), ('addictive', 'Addictive'), ('chemical', 'Chemical report'),
                        ('daily_stock_out', 'Daily stock out'), ('medicine_stock_out', 'Medicine stock out'),
                        ('simple_inventory', 'Total inventory'), ('out_dated', 'Out dated products')]

dict_id = {}
supply = {
    'doctor_appointed': 'Theo y lệnh bác sĩ',
    'not_in_the_bom': 'Không có trong BOM',
    'other': 'Khác'
}
class SHMaterialReport(models.TransientModel):
    _name = 'sh.material.report'
    _description = 'BÁO CÁO VTTH VƯỢT ĐỊNH MỨC'
    _inherit = 'money.mixin'

    institution = fields.Many2one('sh.medical.health.center', string='Cơ sở y tế',
                                  domain=lambda self: [('his_company', 'in', self.env.companies.ids)],
                                  default=lambda self: self.env['sh.medical.health.center'].search(
                                      [('his_company', '=', self.env.companies.ids[0])], limit=1))
    # using date to simplify user input, and to extract date to report
    start_date = fields.Date('Start date', default=date.today() - timedelta(days=1))
    end_date = fields.Date('End date')
    # convert date to datetime for search domain, should be removed if using datetime directly
    start_datetime = fields.Datetime('Start datetime', compute='_compute_datetime')
    end_datetime = fields.Datetime('End datetime', compute='_compute_datetime')
    # products and locations filter for inventory report
    # products = fields.Many2many('product.product', string='Sản phẩm')
    locations = fields.Many2many('stock.location', string='Địa điểm')
    print_excel = fields.Boolean('IN EXCEL', default=True)
    # convert date to datetime for search domain, should be removed if using datetime directly
    @api.depends('start_date', 'end_date')
    def _compute_datetime(self):
        self.start_datetime = False
        self.end_datetime = False
        if self.start_date and self.end_date:
            local_tz = timezone(self.env.user.tz or 'Etc/GMT+7')
            start_datetime = datetime(self.start_date.year, self.start_date.month, self.start_date.day, 0, 0, 0)
            end_datetime = datetime(self.end_date.year, self.end_date.month, self.end_date.day, 23, 59, 59)
            start_datetime = local_tz.localize(start_datetime, is_dst=None)
            end_datetime = local_tz.localize(end_datetime, is_dst=None)
            self.start_datetime = start_datetime.astimezone(utc).replace(tzinfo=None)
            self.end_datetime = end_datetime.astimezone(utc).replace(tzinfo=None)

    @api.onchange('start_date')
    def _onchange_start_date(self):
        if self.start_date:
            self.end_date = self.start_date
            if self.start_date.month == fields.Date.today().month:
                self.end_date = fields.Date.today()
            else:
                self.end_date = date(self.start_date.year, self.start_date.month,
                                     monthrange(self.start_date.year, self.start_date.month)[1])


    @api.constrains('start_date', 'end_date')
    def check_dates(self):
        for record in self:
            start_date = fields.Date.from_string(record.start_date)
            end_date = fields.Date.from_string(record.end_date)
            if start_date > end_date:
                raise ValidationError(
                    _("End Date cannot be set before Start Date."))

    @api.onchange('institution')
    def _onchange_report_institution(self):
        return {'domain': {'locations': [('id', 'in', self._get_locations_by_rights().ids)]}}

    def _get_locations_by_rights(self):
        # Lấy location domain theo quyền, hiện tại chỉ áp dụng cho loại báo cáo là XNT thuốc / vật tư
        locations = self.env['stock.location'].browse()
        reception_location = self.env['stock.location'].search(
            [('company_id', '=', self.institution.his_company.id), ('name', 'ilike', 'lễ tân')], limit=1)
        if self.env.user.has_group('stock.group_stock_manager'):
            wards = self.env['sh.medical.health.center.ward'].search(
                [('institution', '=', self.institution.id), ('type', '!=', False)])
            locations = self.env['stock.location'].search([('location_id', 'in', wards.mapped('location_id.id')),
                                                           ('company_id', '=', self.institution.his_company.id)])
        else:
            ward_grp_dict = {'Laboratory': 'shealth_all_in_one.group_sh_medical_physician_subclinical',
                             'Surgery': 'shealth_all_in_one.group_sh_medical_physician_surgery',
                             'Spa': 'shealth_all_in_one.group_sh_medical_physician_spa',
                             'Laser': 'shealth_all_in_one.group_sh_medical_physician_laser',
                             'Odontology': 'shealth_all_in_one.group_sh_medical_physician_odontology'}
            for w, grp in ward_grp_dict.items():
                if self.env.user.has_group(grp):
                    ward = self.env['sh.medical.health.center.ward'].search(
                        [('institution', '=', self.institution.id), ('type', '=', w)], limit=1)
                    w_locations = self.env['stock.location'].search([('location_id', '=', ward.location_id.id),
                                                                     ('company_id', '=',
                                                                      self.institution.his_company.id)])
                    locations = locations.union(w_locations) or w_locations
        if self.env.user.has_group('shealth_all_in_one.group_sh_medical_receptionist'):
            locations = locations.union(reception_location) or reception_location
        return locations

    def _get_material_data(self):
        # HÀM LẤY DATA VTTH VƯỢT ĐỊNH MỨC

        select = """
        select id from sh_medical_walkin_material
        where ref_id not like %s and ref_id not like %s and date_out >= %s and date_out <= %s
        and institution = %s and location_id in %s and interner_note is not null
        """
        self.env.cr.execute(select, ['sh.medical.surgery%','sh.medical.specialty%',self.start_datetime.strftime('%Y-%m-%d %H:%M:%S'),self.end_datetime.strftime('%Y-%m-%d %H:%M:%S'),
                                     self.institution.id, tuple(self.locations.ids)])
        data_select = self.env.cr.fetchall()
        for data in data_select:
            dict_id[data[0]] = ''
        select_surgery = """
        select smwm.id, smss.explanation_supply from sh_medical_walkin_material smwm
        left join sh_medical_surgery sms on sms.id = cast(replace(smwm.ref_id, 'sh.medical.surgery,', '') as int)
        left join sh_medical_surgery_supply smss on smss.name = sms.id and smss.supply = smwm.product_id
        where smwm.ref_id like %s and smwm.date_out >= %s and smwm.date_out <= %s
        and smwm.institution = %s and smwm.location_id in %s and (smss.explanation_supply is not null or smwm.interner_note is not null)
        """
        self.env.cr.execute(select_surgery, ['sh.medical.surgery%',self.start_datetime.strftime('%Y-%m-%d %H:%M:%S'),
                                     self.end_datetime.strftime('%Y-%m-%d %H:%M:%S'),
                                     self.institution.id, tuple(self.locations.ids)])
        data_select_2 = self.env.cr.fetchall()
        for data in data_select_2:
            dict_id[data[0]] = data[1]
        select_specialty = """
                select smwm.id, smss.explanation_supply from sh_medical_walkin_material smwm
                left join sh_medical_specialty sms on sms.id = cast(replace(smwm.ref_id, 'sh.medical.specialty,', '') as int)
                left join sh_medical_specialty_supply smss on smss.name = sms.id and smss.supply = smwm.product_id
                where smwm.ref_id like %s and smwm.date_out >= %s and smwm.date_out <= %s
                and smwm.institution = %s and smwm.location_id in %s and (smss.explanation_supply is not null or smwm.interner_note is not null)
                """
        self.env.cr.execute(select_specialty, ['sh.medical.specialty%', self.start_datetime.strftime('%Y-%m-%d %H:%M:%S'),
                                             self.end_datetime.strftime('%Y-%m-%d %H:%M:%S'),
                                             self.institution.id, tuple(self.locations.ids)])
        data_select_3 = self.env.cr.fetchall()
        for data in data_select_3:
            dict_id[data[0]] = data[1]
        list_id = []
        for id in dict_id:
            list_id.append(id)
        domain = [('institution', '=', self.institution.id), ('date_out', '>=', self.start_datetime),
                  ('date_out', '<=', self.end_datetime), ('location_id', 'in', self.locations.ids), ('id', 'in', list_id)]
        material_data = []
        # material_data = self.env['sh.medical.walkin.material'].sudo().search(domain)

        group_list_mats = self.env['sh.medical.walkin.material'].read_group(domain, ['id'], ['department'],
                                                                            orderby='department desc')
        for group_lab in group_list_mats:
            dep_detail = self.env['sh.medical.health.center.ward'].browse(group_lab['department'][0])
            # print(dep_detail)
            mats = self.env['sh.medical.walkin.material'].search(group_lab['__domain']).mapped('id')
            material_data.append(
                {'dep_name': dep_detail.name, 'mats': self.env['sh.medical.walkin.material'].browse(mats)})
            # material_data.append(self.env['sh.medical.walkin.material'].browse(mats))
        return material_data

    def report_material(self):
        if not self.print_excel:
            return self.env.ref('shealth_all_in_one.action_report_over_material').report_action(self)
        else:
            template = self.env['ir.attachment'].browse(
                self.env.ref('shealth_all_in_one.report_over_material_attachment').id)
            decode = base64.b64decode(template.datas)
            wb = load_workbook(BytesIO(decode))
            ws = wb.active
            date_range = 'Từ ngày %s-%s-%s đến ngày %s-%s-%s' % (
                self.start_date.day, self.start_date.month, self.start_date.year, self.end_date.day,
                self.end_date.month,
                self.end_date.year)
            image_path = get_module_resource('shealth_all_in_one', 'static/img', 'icon_%s.png' %self.institution.brand.code.lower())
            img = openpyxl.drawing.image.Image(image_path)
            img.anchor = 'a1'
            ws.add_image(img)

            ws['a3'].value = 'Chi nhánh: ' + self.institution.name
            ws['a3'].font = Font(name='Times New Roman', size=13, bold=True)
            ws['a3'].alignment = Alignment(horizontal='center', vertical='center')

            ws['a4'].value = date_range
            ws['a4'].font = Font(name='Times New Roman', size=13, bold=True)
            ws['a4'].alignment = Alignment(horizontal='center', vertical='center')
            thin = borders.Side(style='thin')
            all_border_thin = borders.Border(left=thin, right=thin, top=thin, bottom=thin)
            line_font = Font(name='Times New Roman', size=13)

            datas = self._get_material_data()

            value_col = [
                'stt',
                'ngay_xuat',
                'ten_kh',
                'dich_vu',
                'ten_thuoc',
                'don_vi',
                'sl_dinh_muc',
                'sl_dung_tt',
                'tu_xuat',
                'giai_trinh',
                'ghi_chu_noi_bo'
            ]
            row = 6
            for data in datas:
                index = 1
                cell_dep = ws.cell(row, 1)
                cell_dep.value = data['dep_name']
                cell_dep.font = Font(name='Times New Roman', size=14, bold=True)
                cell_dep.border = all_border_thin
                cell_dep.fill = PatternFill(start_color='ffffba', end_color='ffffba', fill_type='solid')
                ws.merge_cells(start_row=row, start_column=1, end_row=row,end_column=6)
                row += 1
                if data['mats']:
                    for line in data['mats']:
                        col = 1
                        code = line.product_id.code if line.product_id.code else ''
                        name = line.product_id.name if line.product_id.name else ''
                        ten_thuoc = '[' + code + '] ' + name
                        values = {'stt': index,
                                  'ngay_xuat': line.date_out.strftime('%d/%m/%Y'),
                                  'ten_kh': '[' + line.patient.code_customer + '] ' + line.patient.name,
                                  'dich_vu': '; '.join(map(str, line.services.mapped('name'))),
                                  'ten_thuoc': ten_thuoc,
                                  'don_vi': line.uom_id.name,
                                  'sl_dinh_muc': line.init_quantity,
                                  'sl_dung_tt': line.quantity,
                                  'tu_xuat': line.location_id.name,
                                  'giai_trinh': supply[dict_id[line.id]] if dict_id[line.id] in supply else '',
                                  'ghi_chu_noi_bo': line.interner_note if line.interner_note else '-'
                                  }
                        for value in value_col:
                            cell = ws.cell(row, col)
                            cell.value = values[value]
                            cell.font = line_font
                            cell.border = all_border_thin
                            cell.alignment = Alignment(wrap_text=True)
                            if col in [1, 2, 6, 7, 8, 10]:
                                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                            col += 1
                        index += 1
                        row += 1

            fp = BytesIO()
            wb.save(fp)
            fp.seek(0)
            report = base64.encodebytes((fp.read()))
            fp.close()
            attachment = self.env['ir.attachment'].with_user(1).create({
                'name': 'bao_cao_vt_vuot_muc.xlsx',
                'datas': report,
                'res_model': 'temp.creation',
                'public': True,
            })
            url = "/web/content/?model=ir.attachment&id=%s&filename_field=name&field=datas&download=true" \
                  % attachment.id

            return {
                'name': 'Báo cáo VT vượt định mức',
                'type': 'ir.actions.act_url',
                'url': url,
                'target': 'self',
            }
