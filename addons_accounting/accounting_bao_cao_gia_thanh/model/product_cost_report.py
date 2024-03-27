from odoo import fields, models, api, _
from random import randint
from .constant import MRP_PRODUCTION_STATE
from odoo.exceptions import UserError, ValidationError
import base64
import os
from pathlib import Path
import tempfile
from openpyxl import load_workbook


class TasProductCostReport(models.Model):
    _name = "tas.product.cost.report"
    _description = " Product Cost report "

    name = fields.Char('Name', default="Bao Cao Gia Thanh")
    date_from = fields.Date("From", required=True)
    date_end = fields.Date("End", required=True)
    line_ids = fields.One2many("tas.product.cost.report.line", "report_id", string="Lines")
    production_cost_line_ids = fields.One2many("tas.mrp.production.cost.line", "product_cost_report_id", string="Cost Lines")
    budgets = fields.Many2many('crossovered.budget', 'tas_product_cost_report_budget_rel', 'report_id', 'budget_id')
    company_id = fields.Many2one('res.company', "Company", default=lambda self: self.env.company)

    # @api.onchange('date_from', 'date_end')
    # def onchange_date_budget(self):
    #     if self.date_from and self.date_end:
    #         mrp_productions = self.env['sh.medical.specialty'].search([('services_date', '>=', self.date_from), ('services_date', '<=', self.date_end), ('state', 'in', ('Confirmed', 'In Progress', 'Done'))])
    #
    #         line_ids = [(5, 0, 0)]
    #         for mo in mrp_productions:
    #             line_ids.append((0, 0, {
    #                 'mrp_production_id': mo.id
    #             }))
    #         self.line_ids = line_ids

    def compute_mrp_process_cost(self):
        # search trong cac vong lap co the toi uu
        for report in self:
            budgets = self.env['crossovered.budget'].search([
                ('date_from', '>=', report.date_from),
                ('date_to', '<=', report.date_end),
                ('company_id', '=', report.company_id.id)
            ])

            production_cost_line_ids = []
            bom_dict = {}
            cost_line_count = 0

            for budget in budgets:
                for budget_line in budget.crossovered_budget_line:
                    # sum all actual count
                    total_actual_count = 0

                    cost_lines = self.env['tas.mrp.production.cost.line'].search([('date_planned_start', '>=', report.date_from),
                                                                                  ('date_planned_start', '<=', report.date_end),
                                                                                  ('sci_company_id.id', '=', report.company_id.id),
                                                                                    '|',
                                                                                    ('cost_driver_id', '=', budget_line.cost_driver_id.id),
                                                                                    ('cost_driver_id.code', '=', 'nvl')])

                    for cost_line in cost_lines:
                        # compute nvl truoc
                        if cost_line.cost_driver_id.code != 'nvl':
                            total_actual_count += cost_line.actual_count
                        else:
                            if cost_line.loai_phieu == 'sh.medical.lab.test':
                                bom_key = str(cost_line.service_id) + '_' + cost_line.bom_object + '_' + str(cost_line.bom_id) + '_' + str(cost_line.bom_cost_driver_id.id)
                            else:
                                bom_key = cost_line.bom_object + '_' + str(cost_line.bom_id) + '_' + str(cost_line.bom_cost_driver_id.id)

                            if bom_key not in bom_dict:
                                bom_dict[bom_key] = {
                                    'cost_line_count': cost_line.cost_driver_ratio,
                                    'bom_cost_driver_id': cost_line.bom_cost_driver_id.id,
                                    'actual_count': cost_line.actual_count,
                                    'actual_cost_per_unit': cost_line.actual_cost_per_uom_unit
                                }
                            else:
                                cost_line_count = bom_dict[bom_key]['cost_line_count'] + cost_line.cost_driver_ratio
                                actual_count = (bom_dict[bom_key]['actual_count'] * bom_dict[bom_key]['cost_line_count']) + cost_line.actual_count
                                actual_cost_per_unit = (bom_dict[bom_key]['actual_cost_per_unit'] * bom_dict[bom_key]['cost_line_count']) + cost_line.actual_cost_per_uom_unit
                                bom_dict[bom_key] = {
                                    'cost_line_count': cost_line_count,
                                    'bom_cost_driver_id': cost_line.bom_cost_driver_id.id,
                                    'actual_count': actual_count / cost_line_count,
                                    'actual_cost_per_unit': actual_cost_per_unit / cost_line_count
                                }
                    if total_actual_count <= 0 and len(bom_dict) == 0:
                        continue

                    # recomputed_actual_cost_per_uom_unit
                    # for test
                    # budget_line.practical_amount = randint(1000000, 10000000)

                    for cost_line in cost_lines:
                        if cost_line.loai_phieu == 'sh.medical.lab.test':
                            bom_key = str(cost_line.service_id) + '_' + cost_line.bom_object + '_' + str(cost_line.bom_id) + '_' + str(cost_line.bom_cost_driver_id.id)
                        else:
                            bom_key = cost_line.bom_object + '_' + str(cost_line.bom_id) + '_' + str(cost_line.bom_cost_driver_id.id)

                        if cost_line.cost_driver_id.code == 'nvl':
                            recomputed_actual_cost_per_uom_unit = bom_dict[bom_key]['actual_cost_per_unit']
                            cost_line.update({
                                'recomputed_actual_cost_per_uom_unit': recomputed_actual_cost_per_uom_unit,
                                'recomputed_actual_cost_per_bom_unit': recomputed_actual_cost_per_uom_unit * cost_line.actual_count,
                                'recomputed_actual_cost_of_activity': recomputed_actual_cost_per_uom_unit * cost_line.actual_count * cost_line.mrp_production_id.uom_price,
                                'product_cost_report_id': report.id,
                                'is_recompute': True
                            })
                            production_cost_line_ids.append(cost_line.id)
                        else:
                            recomputed_actual_cost_per_uom_unit = abs(budget_line.practical_amount) / total_actual_count

                            cost_line.update({
                                'recomputed_actual_cost_per_uom_unit': recomputed_actual_cost_per_uom_unit,
                                'recomputed_actual_cost_per_bom_unit': recomputed_actual_cost_per_uom_unit * cost_line.actual_count,
                                'recomputed_actual_cost_of_activity': recomputed_actual_cost_per_uom_unit * cost_line.actual_count * cost_line.mrp_production_id.uom_price,
                                'product_cost_report_id': report.id,
                                'is_recompute': True
                            })
                            production_cost_line_ids.append(cost_line.id)

                            if bom_key not in bom_dict:
                                bom_dict[bom_key] = {
                                    'cost_line_count': 1,
                                    'bom_cost_driver_id': cost_line.bom_cost_driver_id.id,
                                    'actual_count': cost_line.actual_count,
                                    'actual_cost_per_unit': cost_line.recomputed_actual_cost_per_uom_unit
                                }
                            else:
                                cost_line_count = bom_dict[bom_key]['cost_line_count'] + 1
                                actual_count = (bom_dict[bom_key]['actual_count'] * bom_dict[bom_key]['cost_line_count']) + cost_line.actual_count
                                actual_cost_per_unit = (bom_dict[bom_key]['actual_cost_per_unit'] * bom_dict[bom_key]['cost_line_count']) + cost_line.recomputed_actual_cost_per_uom_unit
                                bom_dict[bom_key] = {
                                    'cost_line_count': cost_line_count,
                                    'bom_cost_driver_id': cost_line.bom_cost_driver_id.id,
                                    'actual_count': actual_count / cost_line_count,
                                    'actual_cost_per_unit': actual_cost_per_unit / cost_line_count
                                }

            report.budgets = budgets
            report.update({
                'production_cost_line_ids': [(6, 0, production_cost_line_ids)]
            })

            for actual_bom_cost_driver in bom_dict:
                bom_dict[actual_bom_cost_driver].pop('cost_line_count')
                actual_bom_cost_driver = self.env['tas.actual.cost.driver'].create(bom_dict[actual_bom_cost_driver])
                actual_bom_cost_driver.bom_cost_driver_id.update({
                    'actual_count': actual_bom_cost_driver.actual_count,
                    'actual_cost_per_unit': actual_bom_cost_driver.actual_cost_per_unit,
                })

    def get_actual_other_amount(self):
        budgets = self.env['crossovered.budget'].search(
            [('date_from', '>=', self.date_from),
             ('date_to', '<=', self.date_end),
             ('company_id', '=', self.company_id.id)])

        analytic_account_ids = []
        acc_ids = []
        for budget in budgets:
            for budget_line in budget.crossovered_budget_line:
                analytic_account_ids.append(budget_line.analytic_account_id.id)
                acc_ids += budget_line.general_budget_id.account_ids.ids

        analytic_line_obj = self.env['account.analytic.line']
        domain = [
            ('account_id', 'in', analytic_account_ids),
            ('date', '>=', self.date_from),
            ('date', '<=', self.date_end),
            ('company_id', '=', self.company_id.id)
        ]
        if acc_ids:
            domain += [('general_account_id', 'in', acc_ids)]

        account_analytic_lines = analytic_line_obj.read_group(domain, fields=['general_account_id', 'amount'],
                                                              groupby=['general_account_id'])

        return account_analytic_lines

    total_actual_154_without_nvl = fields.Float()

    def create_entry_debit_154_debit_other(self, journal, account_154, other_lines):
        lines = []

        total_credit = 0
        for other_line in other_lines:
            total_credit += abs(other_line.get('amount'))
            lines.append((0, None, {
                'name': 'Kết chuyển chi phí thực tế trong kì sang 154 ERP',
                'account_id': other_line.get('general_account_id')[0],
                'debit': 0.0,
                'credit': abs(other_line.get('amount')),
            }))

        if total_credit > 0:
            lines.append((0, None, {
                'name': 'Kết chuyển chi phí thực tế trong kì sang 154 ERP',
                'account_id': account_154.id,
                'debit': total_credit,
                'credit': 0.0,
            }))

            self.total_actual_154_without_nvl = total_credit

            move = self.env['account.move'].sudo().create({
                'type': 'entry',
                'date': fields.Date.today(),
                'journal_id': journal.id,
                'ref': 'Giá thành',
                'line_ids': lines
            })
            move.post()
            if move.state == 'posted':
                return True
        else:
            return True
        return False

    step_8 = fields.Boolean()
    step_9 = fields.Boolean()

    def create_entry_154_cl_154_erp(self, journal, account_154, account_154_cl, account_632, total_debit_154, total_credit_154):
        if abs(total_debit_154) != abs(total_credit_154):
            if abs(total_debit_154) > abs(total_credit_154):
                move1 = self.env['account.move'].sudo().create({
                    'type': 'entry',
                    'date': fields.Date.today(),
                    'journal_id': journal.id,
                    'ref': 'Giá thành',
                    'line_ids': [
                        (0, None, {
                            'name': 'Xử lý chênh lệch của 154ERP',
                            'account_id': account_154_cl.id,
                            'debit': abs(total_debit_154 - total_credit_154),
                            'credit': 0.0,
                        }),
                        (0, None, {
                            'name': 'Xử lý chênh lệch của 154ERP',
                            'account_id': account_154.id,
                            'debit': 0.0,
                            'credit': abs(total_debit_154 - total_credit_154),
                        }),
                    ]
                })
                move1.post()
                if move1.state == 'posted':
                    # step 10
                    move2 = self.env['account.move'].sudo().create({
                        'type': 'entry',
                        'date': fields.Date.today(),
                        'journal_id': journal.id,
                        'ref': 'Giá thành',
                        'line_ids': [
                            (0, None, {
                                'name': 'Xử lý với tài khoản 632',
                                'account_id': account_632.id,
                                'debit': abs(total_debit_154 - total_credit_154),
                                'credit': 0.0,
                            }),
                            (0, None, {
                                'name': 'Xử lý với tài khoản 632',
                                'account_id': account_154_cl.id,
                                'debit': 0.0,
                                'credit': abs(total_debit_154 - total_credit_154),
                            }),
                        ]
                    })
                    move2.post()
                    if move2.state == 'posted':
                        return True
            else:
                move1 = self.env['account.move'].sudo().create({
                    'type': 'entry',
                    'date': fields.Date.today(),
                    'journal_id': journal.id,
                    'ref': 'Giá thành',
                    'line_ids': [
                        (0, None, {
                            'name': 'Xử lý chênh lệch của 154ERP',
                            'account_id': account_154.id,
                            'debit': abs(total_debit_154 - total_credit_154),
                            'credit': 0.0,
                        }),
                        (0, None, {
                            'name': 'Xử lý chênh lệch của 154ERP',
                            'account_id': account_154_cl.id,
                            'debit': 0.0,
                            'credit': abs(total_debit_154 - total_credit_154),
                        }),
                    ]
                })
                move1.post()
                if move1.state == 'posted':
                    # step 10
                    move2 = self.env['account.move'].sudo().create({
                        'type': 'entry',
                        'date': fields.Date.today(),
                        'journal_id': journal.id,
                        'ref': 'Giá thành',
                        'line_ids': [
                            (0, None, {
                                'name': 'Xử lý với tài khoản 632',
                                'account_id': account_154_cl.id,
                                'debit': abs(total_debit_154 - total_credit_154),
                                'credit': 0.0,
                            }),
                            (0, None, {
                                'name': 'Xử lý với tài khoản 632',
                                'account_id': account_632.id,
                                'debit': 0.0,
                                'credit': abs(total_debit_154 - total_credit_154),
                            }),
                        ]
                    })
                    move2.post()
                    if move2.state == 'posted':
                        return True
        else:
            return True
        return False

    def generate_journal_entry(self):
        account_154 = self.env['account.account'].search([('is_154_erp', '=', True)], limit=1)
        if not account_154:
            raise ValidationError("Bạn chưa cấu hình tài khoản 154 ERP")
        account_154_cl = self.env['account.account'].search([('is_154_cl', '=', True)], limit=1)
        if not account_154_cl:
            raise ValidationError("Bạn chưa cấu hình tài khoản 154 CL")

        account_632 = self.env['account.account'].search([('is_632', '=', True)], limit=1)
        if not account_632:
            raise ValidationError("Bạn chưa cấu hình tài khoản 632")
        # get journal
        journal = self.env['account.journal'].search([('is_gia_thanh', '=', True)], limit=1)
        if not journal:
            raise ValidationError("Bạn chưa cấu hình sổ nhật ký giá thành")

        for report in self:
            account_analytic_lines = report.get_actual_other_amount()
            # region step 8: No 154, có other
            if not report.step_8:
                report.step_8 = report.create_entry_debit_154_debit_other(journal, account_154, account_analytic_lines)
            # endregion step 8

            # region step 9:
            cost_lines = self.env['tas.mrp.production.cost.line'].search(
                [('date_planned_start', '>=', report.date_from),
                 ('date_planned_start', '<=', report.date_end),
                 ('sci_company_id.id', '=', report.company_id.id)])
            total_material_amount = 0
            total_other_amount = 0
            for cost_line in cost_lines:
                if cost_line.cost_driver_id.code == 'nvl':
                    total_material_amount += cost_line.actual_cost_of_activity
                else:
                    total_other_amount += cost_line.actual_cost_of_activity

            if report.step_8 and not report.step_9:
                report.step_9 = report.create_entry_154_cl_154_erp(journal, account_154, account_154_cl, account_632, report.total_actual_154_without_nvl, total_other_amount)
            # endregion step 9

    def _generate_data(self):
        cost_lines = {}

        # production_cost_line_ids = self.env['tas.mrp.production.cost.line'].sudo().search([])

        count = 0
        for production_cost_line in self.production_cost_line_ids:
            key = str(production_cost_line.partner_id.id if production_cost_line.partner_id else False) + '_' + production_cost_line.booking + '_' + str(production_cost_line.service_id.id)

            doanh_thu = 100000
            doanh_thu_thuan = 100000
            giam_tru_doanh_thu = doanh_thu - doanh_thu_thuan

            planned_nvl = production_cost_line.planned_cost_of_activity if production_cost_line.cost_driver_id.code == 'nvl' or production_cost_line.cost_driver_id.name == 'nvl' else 0
            planned_nhan_cong = production_cost_line.planned_cost_of_activity if production_cost_line.cost_driver_id.code == 'cpnc' or production_cost_line.cost_driver_id.name == 'cpnc' else 0
            planned_sxc = production_cost_line.planned_cost_of_activity if production_cost_line.cost_driver_id.code == 'sxc' or production_cost_line.cost_driver_id.name == 'sxc' else 0
            planned_sum = planned_nvl + planned_nhan_cong + planned_sxc

            actual_nvl = production_cost_line.recomputed_actual_cost_of_activity if production_cost_line.cost_driver_id.code == 'nvl' or production_cost_line.cost_driver_id.name == 'nvl' else 0
            actual_nhan_cong = production_cost_line.recomputed_actual_cost_of_activity if production_cost_line.cost_driver_id.code == 'cpnc' or production_cost_line.cost_driver_id.name == 'cpnc' else 0
            actual_sxc = production_cost_line.recomputed_actual_cost_of_activity if production_cost_line.cost_driver_id.code == 'sxc' or production_cost_line.cost_driver_id.name == 'sxc' else 0
            actual_sum = actual_nvl + actual_nhan_cong + actual_sxc

            cost_lines.setdefault(key, {
                'customer_name': production_cost_line.partner_id.name,
                'work_center_id': production_cost_line.work_center_id,
                'booking': production_cost_line.booking,
                'ten_phieu_kham': production_cost_line.ten_phieu_kham,
                'service_id': production_cost_line.service_id.name,
                'service_category_id': production_cost_line.service_category_id.name,
                'sale_order': production_cost_line.sale_order,
                'doanh_thu': doanh_thu,
                'giam_tru_doanh_thu': giam_tru_doanh_thu,
                'doanh_thu_thuan': doanh_thu_thuan,
                'planned_nvl': 0,
                'planned_nhan_cong': 0,
                'planned_sxc': 0,
                'planned_sum': 0,
                'actual_nvl': 0,
                'actual_nhan_cong': 0,
                'actual_sxc': 0,
                'actual_sum': 0,
                'cost_gap': 0,
                'cost': 0,
                'profit_loss': 0,
                'profit_loss_percentage': 0,
                'branch': production_cost_line.sci_company_id.name,
                'sci_company_id': production_cost_line.sci_company_id.name,
            })
            cost_lines[key]['planned_nvl'] += planned_nvl
            cost_lines[key]['planned_nhan_cong'] += planned_nhan_cong
            cost_lines[key]['planned_sxc'] += planned_sxc
            cost_lines[key]['planned_sum'] += planned_sum

            cost_lines[key]['actual_nvl'] += actual_nvl
            cost_lines[key]['actual_nhan_cong'] += actual_nhan_cong
            cost_lines[key]['actual_sxc'] += actual_sxc
            cost_lines[key]['actual_sum'] += actual_sum

            cost_lines[key]['cost_gap'] = (cost_lines[key]['actual_sum'] - cost_lines[key]['planned_sum'])
            cost_lines[key]['cost'] = cost_lines[key]['actual_sum']
            cost_lines[key]['profit_loss'] = cost_lines[key]['doanh_thu_thuan'] - cost_lines[key]['cost']
            cost_lines[key]['profit_loss_percentage'] = cost_lines[key]['cost'] / cost_lines[key]['doanh_thu_thuan'] if cost_lines[key]['doanh_thu_thuan'] != 0 else 0

            # count += 1
            # if count == 10:
            #     break

        return cost_lines



    def _generate_export_file(self, path_template, path_dir, file_name):
        count = 0
        write_start = 16
        _fields_request = self.env["tas.product.cost.report"]._fields
        path_file = os.path.join(path_dir, file_name)

        if not os.path.isfile(path_template):
            raise ValidationError("Path file template not exits")

        # open file
        work_book = load_workbook(path_template)
        work_book.active = True
        work_sheet = work_book["report"]

        # export
        cost_lines = self._generate_data()
        work_sheet.insert_rows(write_start, len(cost_lines))

        for line in cost_lines:
            row = write_start + count
            count += 1

            col = 1
            work_sheet.cell(row, col, str(cost_lines.get(line).get('customer_name')))

            col += 1
            work_sheet.cell(row, col, str(cost_lines.get(line).get('work_center_id')))

            col += 1
            work_sheet.cell(row, col, str(cost_lines.get(line).get('booking')))

            col += 1
            work_sheet.cell(row, col, str(cost_lines.get(line).get('ten_phieu_kham')))

            col += 1
            work_sheet.cell(row, col, str(cost_lines.get(line).get('service_id')))

            col += 1
            work_sheet.cell(row, col, str(cost_lines.get(line).get('service_category_id')))

            col += 1
            work_sheet.cell(row, col, str(cost_lines.get(line).get('sale_order')))

            col += 1
            work_sheet.cell(row, col, str(cost_lines.get(line).get('doanh_thu')))

            col += 1
            work_sheet.cell(row, col, str(cost_lines.get(line).get('giam_tru_doanh_thu')))

            col += 1
            work_sheet.cell(row, col, str(cost_lines.get(line).get('doanh_thu_thuan')))

            col += 1
            work_sheet.cell(row, col, str(cost_lines.get(line).get('planned_nvl')))

            col += 1
            work_sheet.cell(row, col, str(cost_lines.get(line).get('planned_nhan_cong')))

            col += 1
            work_sheet.cell(row, col, str(cost_lines.get(line).get('planned_sxc')))

            col += 1
            work_sheet.cell(row, col, str(cost_lines.get(line).get('planned_sum')))

            col += 1
            work_sheet.cell(row, col, str(cost_lines.get(line).get('actual_nvl')))

            col += 1
            work_sheet.cell(row, col, str(cost_lines.get(line).get('actual_nhan_cong')))

            col += 1
            work_sheet.cell(row, col, str(cost_lines.get(line).get('actual_sxc')))

            col += 1
            work_sheet.cell(row, col, str(cost_lines.get(line).get('actual_sum')))

            col += 1
            work_sheet.cell(row, col, str(cost_lines.get(line).get('cost_gap')))

            col += 1
            work_sheet.cell(row, col, str(cost_lines.get(line).get('cost')))

            col += 1
            work_sheet.cell(row, col, str(cost_lines.get(line).get('profit_loss')))

            col += 1
            work_sheet.cell(row, col, str(cost_lines.get(line).get('profit_loss_percentage')))

            col += 1
            work_sheet.cell(row, col, str(cost_lines.get(line).get('branch')))

            col += 1
            print(col)
            work_sheet.cell(row, col, str(cost_lines.get(line).get('sci_company_id')))

        # last row format not true delete
        work_sheet.delete_rows(write_start + count)
        work_book.save(path_file)

    def _create_attachment_file(self, path_dir: str, file_name: str):
        path_file = os.path.join(path_dir, file_name)
        data = open(path_file, "rb").read()
        base64_encoded = base64.b64encode(data)

        attachment_data = {
            "datas": base64_encoded,
            "type": "binary",
            "name": file_name,
            "mimetype": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        }
        attachment = self.env["ir.attachment"].create(attachment_data)
        return attachment

    def action_download_excel(self):
        """download excel"""
        try:
            file_name = "bc_gia_thanh.xlsx"
            path_model = os.path.dirname(os.path.realpath(__file__))
            path_module = Path(path_model).parent
            path_template = os.path.join(
                path_module,
                "static/templates/bc_gia_thanh.xlsx",
            )
            if not os.path.isfile(path_template):
                raise FileNotFoundError("Path file template not exits")
            with tempfile.TemporaryDirectory() as tmpdir:
                self._generate_export_file(path_template, tmpdir, file_name)
                attachment = self._create_attachment_file(tmpdir, file_name)
                return {
                    "type": "ir.actions.act_url",
                    "url": attachment.local_url,
                    "target": "self",
                }
        except Exception as err:
            raise ValidationError(_("Export dowload all error %s", str(err))) from err


class TasProductCostReportLine(models.Model):
    _name = "tas.product.cost.report.line"
    _description = " Product Cost report line "

    report_id = fields.Many2one('tas.product.cost.report', string="Report", ondelete='cascade')
    # bom_id = fields.Many2one('sh.medical.product.bundle', string="Bom", related='mrp_production_id.bom_id', store=True)
    mrp_production_id = fields.Many2one('sh.medical.specialty', string="Phiếu chuyên khoa", required=True)
    # total_amount = fields.Float("Total amount", related='mrp_production_id.uom_price', store=True)
    # complete_amount = fields.Float("Complete", compute='_compute_operation_quantity', store=True)
    # scrap_amount = fields.Float("Scrap", compute='_compute_operation_quantity', store=True)
    # wip_amount = fields.Float("Wip", compute='_compute_operation_quantity', store=True)
    # cost_of_structure = fields.Float("Cost of structure", compute='_compute_cost_of_structure_operation', store=True)
    # cost_of_operation = fields.Float("Cost of operation", compute='_compute_cost_of_structure_operation', store=True)
    # activity_ids = fields.One2many("tas.product.cost.report.line.activity", "line_id", string="Lines", compute='_compute_activity_ids', store=True)
    # total_cost_of_process = fields.Float("Total cost of process", compute='_total_cost_of_process')
    # actual_cost_per_unit = fields.Float("Actual Cost Per Unit")

    # @api.depends('activity_ids')
    # def _total_cost_of_process(self):
    #     for record in self:
    #         total_cost_of_process = record.cost_of_structure
    #         for activity in record.activity_ids:
    #             total_cost_of_process += activity.cost_of_activity
    #         record.total_cost_of_process = total_cost_of_process
    #
    # @api.depends('mrp_production_id')
    # def _compute_operation_quantity(self):
    #     for record in self:
    #         record.mrp_production_id._compute_operation_quantity()
    #         record.complete_amount = record.mrp_production_id.complete_amount
    #         record.scrap_amount = record.mrp_production_id.scrap_amount
    #         record.wip_amount = record.mrp_production_id.wip_amount
    #
    # @api.depends('mrp_production_id')
    # def _compute_cost_of_structure_operation_bak(self):
    #     for record in self:
    #         # include cost of structure and cost of operation
    #         total_cost = 0
    #         stock_valuation_layer_ids = record.mrp_production_id.move_finished_ids.stock_valuation_layer_ids
    #         cost_of_component = 0
    #         cost_of_operation = 0
    #         for line in stock_valuation_layer_ids:
    #             cost_of_component += line.value
    #
    #         record.cost_of_structure = cost_of_component
    #         record.cost_of_operation = cost_of_operation
    #
    # @api.depends('mrp_production_id')
    # def _compute_cost_of_structure_operation(self):
    #     for record in self:
    #         lines = self.env['report.mrp_account_enterprise.mrp_cost_structure'].get_lines(record.mrp_production_id)
    #         cost_of_component = 0
    #         cost_of_operation = 0
    #         if len(lines) == 1:
    #             cost_of_component = lines[0].get('total_cost')
    #             for operation in lines[0].get('operations'):
    #                 cost_of_operation += operation[3] * operation[4]
    #
    #         record.cost_of_structure = cost_of_component
    #         record.cost_of_operation = cost_of_operation
    #
    # @api.depends('mrp_production_id')
    # def _compute_activity_ids(self):
    #     cost_driver = {}
    #     for record in self:
    #         for budget in record.report_id.budgets:
    #             for budget_line in budget.crossovered_budget_line:
    #                 if budget_line.cost_driver_id:
    #                     cost_driver[budget_line.cost_driver_id.code] = {
    #                         'id': budget_line.cost_driver_id.id,
    #                         'data': [],
    #                         'cost_of_activity': abs(budget_line.practical_amount),
    #                         'complete_amount': 0,
    #                         'complete_unit_amount': 0,
    #                         'wip_amount': 0,
    #                         'wip_unit_amount': 0
    #                     }
    #
    #             for mo in record.report_id.line_ids:
    #                 for bom_cost_driver in mo.mrp_production_id.production_cost_line_ids:
    #                     if bom_cost_driver.cost_driver_id.code in cost_driver:
    #                         cost_driver[bom_cost_driver.cost_driver_id.code]['data'].append(mo.mrp_production_id)
    #                         cost_driver[bom_cost_driver.cost_driver_id.code]['complete_amount'] += mo.mrp_production_id.complete_amount
    #                         cost_driver[bom_cost_driver.cost_driver_id.code]['complete_unit_amount'] += (mo.mrp_production_id.complete_amount * bom_cost_driver.actual_count)
    #                         cost_driver[bom_cost_driver.cost_driver_id.code]['wip_amount'] += mo.mrp_production_id.wip_amount
    #                         cost_driver[bom_cost_driver.cost_driver_id.code]['wip_unit_amount'] += (mo.mrp_production_id.wip_amount * bom_cost_driver.actual_count)
    #
    #     for record in self:
    #         activity_lines = [(5, 0, 0)]
    #         if record.bom_id:
    #             # value = (0, 0, {
    #             #     'activity_type': self.env.ref('tas_bao_cao_gia_thanh.tas_cost_driver_working_hour').id,
    #             #     'cost_of_activity': record.cost_of_operation,
    #             #     # 'cost_per_activity': record.cost_of_operation / record.complete_amount,
    #             #     'cost_per_unit': record.cost_of_operation / record.complete_amount if record.complete_amount > 0 else 0
    #             # })
    #             # activity_lines.append(value)
    #
    #             value = (0, 0, {
    #                 'activity_type': self.env.ref('accounting_bao_cao_gia_thanh.tas_cost_driver_nvl').id,
    #                 'cost_of_activity': record.cost_of_structure,
    #                 # 'cost_per_activity': record.cost_of_structure / record.complete_amount,
    #                 'cost_per_unit': record.cost_of_structure / record.complete_amount if record.complete_amount > 0 else 0
    #             })
    #             activity_lines.append(value)
    #
    #             for bom_cost_driver in record.mrp_production_id.production_cost_line_ids:
    #                 if bom_cost_driver.cost_driver_id.code in cost_driver:
    #                     cost_per_uom_unit = (cost_driver[bom_cost_driver.cost_driver_id.code]['cost_of_activity'] / (cost_driver[bom_cost_driver.cost_driver_id.code]['complete_unit_amount'])) if (cost_driver[bom_cost_driver.cost_driver_id.code]['complete_unit_amount']) > 0 else 0
    #                     cost_per_bom_unit = bom_cost_driver.actual_count * cost_per_uom_unit
    #                     cost_of_activity = record.complete_amount * bom_cost_driver.actual_count * cost_per_bom_unit
    #                     # cost_of_activity = record.complete_amount * cost_per_activity
    #                     value = (0, 0, {
    #                         'activity_type': cost_driver[bom_cost_driver.cost_driver_id.code]['id'],
    #                         'cost_of_activity': cost_of_activity,
    #                         # 'cost_per_activity': cost_per_activity,
    #                         'cost_per_unit': cost_per_bom_unit
    #                     })
    #                     activity_lines.append(value)
    #         record.update({'activity_ids': activity_lines})




class TasProductCostReportLineActivity(models.Model):
    _name = "tas.product.cost.report.line.activity"
    _description = 'TAS product cost report line activity'

    line_id = fields.Many2one('tas.product.cost.report.line', string="Line")
    activity_type = fields.Many2one('tas.cost.driver', string="Cost Driver", required=True)
    cost_of_activity = fields.Float("Cost of activity")
    cost_per_activity = fields.Float("Cost per activity")
    cost_per_unit = fields.Float("Cost per unit")